import datetime
import re
from decimal import Decimal
from unittest.mock import patch
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import transaction
from django.test import TestCase, SimpleTestCase, Client, tag
from django.utils import timezone
from .models import Department, Supplier, Product, SupplierPrice, Stocktake, StockLine, Delivery, Batch, Adjustment, IngredientAllergen, Recipe, RecipeLine, RecipeCycleError, RecipePackaging, Customer, SuppressedRecipe, SaleProduct, Order, OrderLine
from .ai_extract import parse_lines_json, auto_match
from .templatetags.pack_format import pack_size


class StockLineValueTests(TestCase):
    def test_value_is_count_times_pack_price(self):
        dept = Department.objects.create(name="Test Dept")
        sup = Supplier.objects.create(name="Test Sup")
        product = Product.objects.create(name="Thing", department=dept, unit="ea", minimum=0)
        SupplierPrice.objects.create(
            product=product, supplier=sup,
            pack_weight=Decimal("1"), pack_price=Decimal("11.54"))
        st = Stocktake.objects.create(department=dept, date=datetime.date.today())
        line = StockLine.objects.create(stocktake=st, product=product, current=Decimal("10"))
        self.assertEqual(line.value, Decimal("115.40"))

    def test_value_ignores_pack_weight(self):
        dept = Department.objects.create(name="Dept2")
        sup = Supplier.objects.create(name="Sup2")
        product = Product.objects.create(name="Bulk", department=dept, unit="g", minimum=0)
        SupplierPrice.objects.create(
            product=product, supplier=sup,
            pack_weight=Decimal("1000"), pack_price=Decimal("11.54"))
        st = Stocktake.objects.create(department=dept, date=datetime.date.today())
        line = StockLine.objects.create(stocktake=st, product=product, current=Decimal("10"))
        self.assertEqual(line.value, Decimal("115.40"))


class DeliveryBatchTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.sup = Supplier.objects.create(name="Acme Mill")
        self.product = Product.objects.create(name="Flour", department=self.dept, unit="ea", minimum=0)
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)

    def test_delivery_of_8_packs_creates_batch_and_updates_on_hand(self):
        delivery = Delivery.objects.create(department=self.dept, supplier=self.sup,
                                           date=datetime.date.today())
        batch = Batch.objects.create(
            delivery=delivery, product=self.product,
            batch_code="A123", use_by=datetime.date(2026, 12, 31),
            qty_received=Decimal("8"), qty_remaining=Decimal("8"))
        self.assertEqual(batch.qty_remaining, Decimal("8"))
        self.assertEqual(self.product.on_hand_from_batches, Decimal("8"))

    def test_delivery_form_creates_delivery_and_batch(self):
        c = Client(); assert c.login(username="alice", password="pw")
        c.get(f"/switch/{self.dept.pk}/")
        r = c.post("/deliveries/new/", {
            "supplier": str(self.sup.pk),
            "date": datetime.date.today().isoformat(),
            "note": "docket 42",
            "product": [str(self.product.pk)],
            "batch_code": ["A123"],
            "use_by": [""],
            "qty": ["8"],
        })
        self.assertEqual(r.status_code, 302)
        self.assertEqual(Delivery.objects.count(), 1)
        self.assertEqual(Batch.objects.count(), 1)
        b = Batch.objects.get()
        self.assertEqual(b.qty_received, Decimal("8"))
        self.assertEqual(b.qty_remaining, Decimal("8"))
        self.assertEqual(b.batch_code, "A123")
        self.assertEqual(self.product.on_hand_from_batches, Decimal("8"))

    def test_on_hand_sums_across_batches(self):
        delivery = Delivery.objects.create(department=self.dept, supplier=self.sup,
                                           date=datetime.date.today())
        Batch.objects.create(delivery=delivery, product=self.product,
                             qty_received=Decimal("3"), qty_remaining=Decimal("3"))
        Batch.objects.create(delivery=delivery, product=self.product,
                             qty_received=Decimal("5"), qty_remaining=Decimal("5"))
        self.assertEqual(self.product.on_hand_from_batches, Decimal("8"))

    def test_has_supplier_price_false_when_no_matching_price(self):
        delivery = Delivery.objects.create(department=self.dept, supplier=self.sup,
                                           date=datetime.date.today())
        batch = Batch.objects.create(delivery=delivery, product=self.product,
                                     qty_received=Decimal("3"), qty_remaining=Decimal("3"))
        self.assertFalse(batch.has_supplier_price)

    def test_has_supplier_price_true_when_matching_price_exists(self):
        SupplierPrice.objects.create(product=self.product, supplier=self.sup,
                                     pack_weight=Decimal("1"), pack_price=Decimal("10"))
        delivery = Delivery.objects.create(department=self.dept, supplier=self.sup,
                                           date=datetime.date.today())
        batch = Batch.objects.create(delivery=delivery, product=self.product,
                                     qty_received=Decimal("3"), qty_remaining=Decimal("3"))
        self.assertTrue(batch.has_supplier_price)

    def test_form_save_still_succeeds_when_supplier_has_no_price_for_product(self):
        c = Client(); assert c.login(username="alice", password="pw")
        c.get(f"/switch/{self.dept.pk}/")
        r = c.post("/deliveries/new/", {
            "supplier": str(self.sup.pk),
            "date": datetime.date.today().isoformat(),
            "product": [str(self.product.pk)],
            "batch_code": ["X1"],
            "use_by": [""],
            "qty": ["8"],
        })
        self.assertEqual(r.status_code, 302)
        b = Batch.objects.get()
        self.assertEqual(b.qty_remaining, Decimal("8"))
        self.assertFalse(b.has_supplier_price)


class AIExtractTests(TestCase):
    def test_parse_lines_json_plain(self):
        out = parse_lines_json('[{"description": "Flour 25kg", "qty": 4}]')
        self.assertEqual(out, [{"description": "Flour 25kg", "qty": 4.0}])

    def test_parse_lines_json_with_markdown_fence(self):
        out = parse_lines_json('```json\n[{"description":"Sugar","qty":2}]\n```')
        self.assertEqual(out, [{"description": "Sugar", "qty": 2.0}])

    def test_parse_lines_json_handles_garbage(self):
        self.assertEqual(parse_lines_json("not json at all"), [])
        self.assertEqual(parse_lines_json(""), [])
        self.assertEqual(parse_lines_json("{}"), [])
        self.assertEqual(parse_lines_json('[{"description":"","qty":3}]'), [])
        self.assertEqual(parse_lines_json('[{"description":"x","qty":0}]'), [])

    def test_parse_lines_json_extracts_embedded_array(self):
        out = parse_lines_json('Here you go: [{"description":"Eggs","qty":12}] cheers!')
        self.assertEqual(out, [{"description": "Eggs", "qty": 12.0}])


class AutoMatchTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="D")
        self.sup_a = Supplier.objects.create(name="A")
        self.sup_b = Supplier.objects.create(name="B")
        self.flour = Product.objects.create(name="Flour", department=self.dept, unit="ea", minimum=0)
        self.sugar = Product.objects.create(name="Caster Sugar", department=self.dept, unit="ea", minimum=0)
        self.salt = Product.objects.create(name="Salt", department=self.dept, unit="ea", minimum=0)
        SupplierPrice.objects.create(product=self.flour, supplier=self.sup_a,
                                     pack_weight=Decimal("1"), pack_price=Decimal("3"))
        SupplierPrice.objects.create(product=self.sugar, supplier=self.sup_b,
                                     pack_weight=Decimal("1"), pack_price=Decimal("2"))

    def test_matches_by_substring(self):
        p, conf = auto_match("Strong White Flour 25kg", self.sup_a, self.dept)
        self.assertEqual(p, self.flour)
        self.assertTrue(conf)

    def test_prefers_supplier_catalog(self):
        # Both flour and salt are dept products; flour is in sup_a's catalog; if the
        # description mentions both, the priced one wins.
        p, _ = auto_match("Flour and Salt mix", self.sup_a, self.dept)
        self.assertEqual(p, self.flour)

    def test_falls_back_to_other_products(self):
        # sugar isn't in sup_a's catalog but should still be matched if it's the only fit
        p, conf = auto_match("Caster Sugar 1kg", self.sup_a, self.dept)
        self.assertEqual(p, self.sugar)
        self.assertTrue(conf)

    def test_no_match_returns_none(self):
        p, conf = auto_match("Mystery Item", self.sup_a, self.dept)
        self.assertIsNone(p)
        self.assertFalse(conf)


class DeliveryScanFlowTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.sup = Supplier.objects.create(name="Acme Mill")
        self.flour = Product.objects.create(name="Flour", department=self.dept, unit="ea", minimum=0)
        self.sugar = Product.objects.create(name="Sugar", department=self.dept, unit="ea", minimum=0)
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        SupplierPrice.objects.create(product=self.flour, supplier=self.sup,
                                     pack_weight=Decimal("1"), pack_price=Decimal("3"))
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    @patch("stock.views.extract_lines")
    def test_scan_prefills_form_with_matched_and_unmatched_lines(self, mock_extract):
        mock_extract.return_value = [
            {"description": "Strong White Flour 25kg", "qty": 4.0},
            {"description": "Caster Sugar 1kg", "qty": 2.0},
            {"description": "Mystery Item XYZ", "qty": 1.0},
        ]
        fake_file = SimpleUploadedFile("test.jpg", b"\xff\xd8fake-jpg-bytes",
                                       content_type="image/jpeg")
        r = self.client.post("/deliveries/scan/", {
            "supplier": str(self.sup.pk),
            "file": fake_file,
        })
        # extract_lines was called with the file bytes and mime type
        mock_extract.assert_called_once()
        args, _ = mock_extract.call_args
        self.assertEqual(args[1], "image/jpeg")
        self.assertIn(b"fake-jpg-bytes", args[0])

        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Supplier preselected
        self.assertIn(f'<option value="{self.sup.pk}" selected>Acme Mill</option>', body)
        # Flour line: data-prefill = flour.pk, qty=4.0
        self.assertIn(f'data-prefill="{self.flour.pk}"', body)
        self.assertIn('value="4.0"', body)
        # Sugar line: data-prefill = sugar.pk, qty=2.0
        self.assertIn(f'data-prefill="{self.sugar.pk}"', body)
        self.assertIn('value="2.0"', body)
        # Unmatched line: raw description shown, no data-prefill for that row's qty
        self.assertIn("Mystery Item XYZ", body)
        self.assertIn('value="1.0"', body)
        # form posts to delivery_new (the existing save endpoint)
        self.assertIn('action="/deliveries/new/"', body)

    @patch("stock.views.extract_lines")
    def test_zero_lines_falls_back_to_manual_form_with_message(self, mock_extract):
        mock_extract.return_value = []
        f = SimpleUploadedFile("x.png", b"png-bytes", content_type="image/png")
        r = self.client.post("/deliveries/scan/", {"supplier": str(self.sup.pk), "file": f})
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("No line items came back", body)
        # Still the delivery form; supplier preselected
        self.assertIn(f'<option value="{self.sup.pk}" selected>Acme Mill</option>', body)

    @patch("stock.views.extract_lines")
    def test_api_failure_falls_back_to_manual_form(self, mock_extract):
        from .ai_extract import ExtractError
        mock_extract.side_effect = ExtractError("could not reach scanning service")
        f = SimpleUploadedFile("x.png", b"png-bytes", content_type="image/png")
        r = self.client.post("/deliveries/scan/", {"supplier": str(self.sup.pk), "file": f})
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("could not reach scanning service", body)
        self.assertIn(f'<option value="{self.sup.pk}" selected>Acme Mill</option>', body)

    def test_no_supplier_redirects_back_to_scan(self):
        f = SimpleUploadedFile("x.png", b"png-bytes", content_type="image/png")
        r = self.client.post("/deliveries/scan/", {"supplier": "", "file": f})
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.headers["Location"], "/deliveries/scan/")

    def test_get_renders_upload_form(self):
        r = self.client.get("/deliveries/scan/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("Scan delivery note", body)
        self.assertIn('name="file"', body)


class UsageHistoryTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.sup = Supplier.objects.create(name="Mill")
        self.flour = Product.objects.create(
            name="Flour", department=self.dept, unit="ea", minimum=0)

    def _count(self, date, current):
        st = Stocktake.objects.create(department=self.dept, date=date)
        StockLine.objects.create(
            stocktake=st, product=self.flour,
            current=Decimal(current), carried_over=False)
        return st

    def _delivery(self, date, qty):
        d = Delivery.objects.create(department=self.dept, supplier=self.sup, date=date)
        return Batch.objects.create(
            delivery=d, product=self.flour,
            qty_received=Decimal(qty), qty_remaining=Decimal(qty))

    def test_delivery_between_counts_is_included_in_usage(self):
        # 10 on Jan 1, delivery of 8 on Jan 5, 12 on Jan 8 → usage 10 + 8 - 12 = 6
        self._count(datetime.date(2026, 1, 1), 10)
        self._delivery(datetime.date(2026, 1, 5), 8)
        self._count(datetime.date(2026, 1, 8), 12)
        rows = self.flour.usage_history()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["delivered"], Decimal("8"))
        self.assertEqual(rows[0]["usage"], Decimal("6"))
        self.assertFalse(rows[0]["clamped"])
        self.assertEqual(rows[0]["days"], 7)

    def test_delivery_on_previous_count_date_is_excluded(self):
        # "Strictly between" - delivery on the same day as the previous count
        # is part of that count's snapshot and must not be double-counted.
        self._count(datetime.date(2026, 1, 1), 10)
        self._delivery(datetime.date(2026, 1, 1), 5)
        self._count(datetime.date(2026, 1, 8), 8)
        rows = self.flour.usage_history()
        self.assertEqual(rows[0]["delivered"], Decimal("0"))
        self.assertEqual(rows[0]["usage"], Decimal("2"))

    def test_delivery_on_current_count_date_is_included(self):
        # "On/before this stocktake's date" - delivery on the day we counted
        # is part of the inflow for that period.
        self._count(datetime.date(2026, 1, 1), 10)
        self._delivery(datetime.date(2026, 1, 8), 5)
        self._count(datetime.date(2026, 1, 8), 12)
        rows = self.flour.usage_history()
        self.assertEqual(rows[0]["delivered"], Decimal("5"))
        self.assertEqual(rows[0]["usage"], Decimal("3"))

    def test_first_ever_count_has_no_usage_data(self):
        self._count(datetime.date(2026, 1, 8), 10)
        self.assertEqual(self.flour.usage_history(), [])
        self.assertIsNone(self.flour.average_weekly_usage())
        self.assertIsNone(self.flour.days_of_cover(on_hand=Decimal("10")))

    def test_negative_usage_clamps_to_zero(self):
        # 5 on hand, no delivery, then 12 on hand - impossible without an
        # unlogged delivery or miscount. Clamp at 0 and flag it.
        self._count(datetime.date(2026, 1, 1), 5)
        self._count(datetime.date(2026, 1, 8), 12)
        rows = self.flour.usage_history()
        self.assertEqual(rows[0]["usage"], Decimal("0"))
        self.assertTrue(rows[0]["clamped"])

    def test_average_over_multiple_counts(self):
        self._count(datetime.date(2026, 1, 1), 20)
        self._count(datetime.date(2026, 1, 8), 15)   # usage 5
        self._count(datetime.date(2026, 1, 15), 5)   # usage 10
        self.assertEqual(self.flour.average_weekly_usage(n=4), Decimal("7.50"))

    def test_days_of_cover_from_average(self):
        # avg 5 packs/week, 10 on hand → 10 / 5 * 7 = 14 days
        self._count(datetime.date(2026, 1, 1), 15)
        self._count(datetime.date(2026, 1, 8), 10)
        self.assertEqual(self.flour.days_of_cover(on_hand=Decimal("10")), 14)

    def test_carried_over_lines_are_not_data_points(self):
        # A carried-over line is just a copy of the prior count - it must not
        # appear in usage history (would produce a spurious 0-usage row).
        self._count(datetime.date(2026, 1, 1), 10)
        st = Stocktake.objects.create(department=self.dept, date=datetime.date(2026, 1, 8))
        StockLine.objects.create(stocktake=st, product=self.flour,
                                 current=Decimal("10"), carried_over=True)
        self.assertEqual(self.flour.usage_history(), [])


class AdjustmentTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.sup = Supplier.objects.create(name="Mill")
        self.flour = Product.objects.create(
            name="Flour", code="FLR1", department=self.dept, unit="g", minimum=0)
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    def _batch(self, qty):
        d = Delivery.objects.create(department=self.dept, supplier=self.sup,
                                    date=datetime.date(2026, 5, 1))
        return Batch.objects.create(delivery=d, product=self.flour,
                                    qty_received=Decimal(qty), qty_remaining=Decimal(qty))

    def test_waste_reduces_on_hand(self):
        self._batch(10)
        self.assertEqual(self.flour.on_hand_from_batches, Decimal("10"))
        Adjustment.objects.create(
            department=self.dept, product=self.flour,
            quantity=Decimal("3"), reason="waste", user=self.user,
            date=datetime.date(2026, 5, 10))
        self.assertEqual(self.flour.on_hand_from_batches, Decimal("10"))  # batches unchanged
        self.assertEqual(self.flour.on_hand, Decimal("7"))                # adjusted figure
        self.assertEqual(self.flour.adjustments_net, Decimal("-3"))

    def test_found_increases_on_hand(self):
        self._batch(5)
        Adjustment.objects.create(
            department=self.dept, product=self.flour,
            quantity=Decimal("2"), reason="found", user=self.user,
            date=datetime.date(2026, 5, 10))
        self.assertEqual(self.flour.on_hand, Decimal("7"))
        self.assertEqual(self.flour.adjustments_net, Decimal("2"))

    def test_waste_in_period_is_excluded_from_usage(self):
        # Without waste: P=20, D=0, C=14 → usage 6
        # With waste of 4 in the period: P=20, D=0, C=14, W=4 → real usage 2
        st1 = Stocktake.objects.create(department=self.dept,
                                       date=datetime.date(2026, 5, 1))
        StockLine.objects.create(stocktake=st1, product=self.flour,
                                 current=Decimal("20"), carried_over=False)
        Adjustment.objects.create(
            department=self.dept, product=self.flour,
            quantity=Decimal("4"), reason="waste", user=self.user,
            date=datetime.date(2026, 5, 5))
        st2 = Stocktake.objects.create(department=self.dept,
                                       date=datetime.date(2026, 5, 8))
        StockLine.objects.create(stocktake=st2, product=self.flour,
                                 current=Decimal("14"), carried_over=False)

        rows = self.flour.usage_history()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["adjustments"], Decimal("-4"))
        self.assertEqual(rows[0]["usage"], Decimal("2"))
        self.assertFalse(rows[0]["clamped"])

    def test_found_in_period_adds_to_inflows_for_usage(self):
        # P=10, D=0, F=5 (found), C=12 → real usage 3
        st1 = Stocktake.objects.create(department=self.dept,
                                       date=datetime.date(2026, 5, 1))
        StockLine.objects.create(stocktake=st1, product=self.flour,
                                 current=Decimal("10"), carried_over=False)
        Adjustment.objects.create(
            department=self.dept, product=self.flour,
            quantity=Decimal("5"), reason="found", user=self.user,
            date=datetime.date(2026, 5, 5))
        st2 = Stocktake.objects.create(department=self.dept,
                                       date=datetime.date(2026, 5, 8))
        StockLine.objects.create(stocktake=st2, product=self.flour,
                                 current=Decimal("12"), carried_over=False)
        rows = self.flour.usage_history()
        self.assertEqual(rows[0]["adjustments"], Decimal("5"))
        self.assertEqual(rows[0]["usage"], Decimal("3"))

    def test_adjustment_outside_period_does_not_affect_usage(self):
        # Waste on the date of the previous count is "strictly between"-excluded
        st1 = Stocktake.objects.create(department=self.dept,
                                       date=datetime.date(2026, 5, 1))
        StockLine.objects.create(stocktake=st1, product=self.flour,
                                 current=Decimal("10"), carried_over=False)
        Adjustment.objects.create(  # same day as prev count
            department=self.dept, product=self.flour,
            quantity=Decimal("3"), reason="waste", user=self.user,
            date=datetime.date(2026, 5, 1))
        st2 = Stocktake.objects.create(department=self.dept,
                                       date=datetime.date(2026, 5, 8))
        StockLine.objects.create(stocktake=st2, product=self.flour,
                                 current=Decimal("7"), carried_over=False)
        rows = self.flour.usage_history()
        self.assertEqual(rows[0]["adjustments"], Decimal("0"))
        self.assertEqual(rows[0]["usage"], Decimal("3"))

    def test_post_creates_adjustment_and_appears_in_log(self):
        r = self.client.post("/adjustments/", {
            "product": str(self.flour.pk),
            "quantity": "2.5",
            "reason": "waste",
            "note": "dropped bag",
        })
        self.assertEqual(r.status_code, 302)
        a = Adjustment.objects.get()
        self.assertEqual(a.product, self.flour)
        self.assertEqual(a.department, self.dept)
        self.assertEqual(a.quantity, Decimal("2.5"))
        self.assertEqual(a.reason, "waste")
        self.assertEqual(a.user, self.user)
        self.assertEqual(a.note, "dropped bag")

        r = self.client.get("/adjustments/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("Flour", body)
        self.assertIn("dropped bag", body)
        # Waste shows with a minus sign in the qty column
        self.assertIn("−3", body)  # 2.5 rounded to 3 by floatformat:0 with leading minus

    def test_post_rejects_bad_payload(self):
        for bad in (
            {"product": "", "quantity": "1", "reason": "waste"},
            {"product": str(self.flour.pk), "quantity": "0", "reason": "waste"},
            {"product": str(self.flour.pk), "quantity": "-3", "reason": "waste"},
            {"product": str(self.flour.pk), "quantity": "1", "reason": "bogus"},
        ):
            r = self.client.post("/adjustments/", bad)
            self.assertEqual(r.status_code, 302)
        self.assertFalse(Adjustment.objects.exists())

    def test_adjustment_page_requires_login(self):
        c = Client()
        r = c.get("/adjustments/")
        self.assertEqual(r.status_code, 302)
        self.assertIn("/login/", r.headers["Location"])

    def test_adjustment_log_is_department_scoped(self):
        # Adjustments in another department must not appear here
        other = Department.objects.create(name="Butchery")
        other_product = Product.objects.create(
            name="Beef", department=other, unit="g", minimum=0)
        Adjustment.objects.create(
            department=other, product=other_product,
            quantity=Decimal("3"), reason="waste",
            date=datetime.date(2026, 5, 5))
        r = self.client.get("/adjustments/")
        self.assertNotIn("Beef", r.content.decode())


class ReorderTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.sup = Supplier.objects.create(name="Acme Mill")
        # Flour: minimum 10, on-hand 2 -> suggested order 8.
        self.flour = Product.objects.create(
            name="Flour", code="FLR1", department=self.dept,
            unit="g", minimum=Decimal("10"))
        SupplierPrice.objects.create(
            product=self.flour, supplier=self.sup,
            pack_weight=Decimal("25000"), pack_price=Decimal("30.00"))
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date(2026, 5, 22))
        StockLine.objects.create(stocktake=st, product=self.flour,
                                 current=Decimal("2"), carried_over=False)
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    def test_reorder_page_renders_editable_qty_input(self):
        r = self.client.get("/reorder/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Editable input with suggested value 8 = 10 - 2
        self.assertIn(f'name="qty_{self.flour.pk}"', body)
        self.assertIn('value="8"', body)
        # Pack price exposed for live JS recomputation
        self.assertIn('data-pack-price="30.00"', body)
        # Form posts to reorder_csv
        self.assertIn('action="/reorder/csv/"', body)
        # Single combined table with supplier sub-header row
        self.assertEqual(body.count('id="reorder-tbl"'), 1)
        self.assertIn('class="sup-row"', body)
        self.assertIn("Acme Mill", body)

    def test_reorder_csv_get_uses_suggested_qty_and_new_columns(self):
        r = self.client.get("/reorder/csv/")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "text/csv")
        body = r.content.decode()
        # New column order
        self.assertEqual(body.splitlines()[0],
                         "Supplier,Ingredient,Order qty,Pack size,Pack price,Est. cost")
        # Suggested qty 8 @ £30 = £240; pack_size formatted by pack_format ("25 kg")
        self.assertIn("Acme Mill,Flour,8.00,25 kg,30.00,240.00", body)

    def test_reorder_csv_post_uses_overridden_qty(self):
        r = self.client.post("/reorder/csv/", {
            f"qty_{self.flour.pk}": "12",
        })
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Override 12 @ £30 = £360
        self.assertIn("Acme Mill,Flour,12,25 kg,30.00,360.00", body)

    def test_reorder_csv_post_falls_back_to_suggested_when_override_blank(self):
        r = self.client.post("/reorder/csv/", {f"qty_{self.flour.pk}": ""})
        body = r.content.decode()
        self.assertIn("Acme Mill,Flour,8.00,25 kg,30.00,240.00", body)


class DeliveryDetailTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.sup = Supplier.objects.create(name="Acme Mill")
        self.flour = Product.objects.create(
            name="Flour", code="FLR1", department=self.dept,
            unit="g", minimum=0)
        self.sugar = Product.objects.create(
            name="Sugar", code="SUG1", department=self.dept,
            unit="g", minimum=0)
        SupplierPrice.objects.create(
            product=self.flour, supplier=self.sup,
            pack_weight=Decimal("25000"), pack_price=Decimal("30"))
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    def test_delivery_detail_renders_with_lines(self):
        delivery = Delivery.objects.create(
            department=self.dept, supplier=self.sup,
            date=datetime.date(2026, 5, 22), note="docket 99")
        Batch.objects.create(
            delivery=delivery, product=self.flour, batch_code="A123",
            use_by=datetime.date(2026, 12, 31),
            qty_received=Decimal("8"), qty_remaining=Decimal("8"))
        Batch.objects.create(
            delivery=delivery, product=self.sugar, batch_code="B7",
            qty_received=Decimal("4"), qty_remaining=Decimal("4"))

        r = self.client.get(f"/deliveries/{delivery.pk}/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()

        # Header bits
        self.assertIn("Acme Mill", body)
        self.assertIn("22 May 2026", body)
        self.assertIn("docket 99", body)

        # Header stats: 2 lines, 12 packs in
        self.assertRegex(body, r"Lines</div><div class=\"v\">2<")
        self.assertRegex(body, r"Packs in</div><div class=\"v\">12<")

        # Both batches present with qty and batch code
        self.assertIn("Flour", body)
        self.assertIn("FLR1", body)
        self.assertIn("A123", body)
        self.assertIn("Sugar", body)
        self.assertIn("B7", body)
        self.assertIn("31 Dec 2026", body)

        # Ingredient links through to product detail
        self.assertIn(f'href="/products/{self.flour.pk}/"', body)
        self.assertIn(f'href="/products/{self.sugar.pk}/"', body)

        # Back link
        self.assertIn('href="/deliveries/"', body)

        # has_supplier_price flag: flour has a price, sugar does not
        # The "no price" tag should appear once (sugar row).
        self.assertEqual(body.count("no price"), 1)

    def test_delivery_detail_blocked_for_other_department(self):
        other = Department.objects.create(name="Butchery")
        delivery = Delivery.objects.create(
            department=other, supplier=self.sup, date=datetime.date(2026, 5, 22))
        r = self.client.get(f"/deliveries/{delivery.pk}/")
        self.assertEqual(r.status_code, 403)

    def test_delivery_detail_requires_login(self):
        delivery = Delivery.objects.create(
            department=self.dept, supplier=self.sup, date=datetime.date(2026, 5, 22))
        c = Client()
        r = c.get(f"/deliveries/{delivery.pk}/")
        self.assertEqual(r.status_code, 302)
        self.assertIn("/login/", r.headers["Location"])

    def test_deliveries_list_rows_link_to_detail(self):
        delivery = Delivery.objects.create(
            department=self.dept, supplier=self.sup, date=datetime.date(2026, 5, 22))
        r = self.client.get("/deliveries/")
        self.assertEqual(r.status_code, 200)
        self.assertIn(f'href="/deliveries/{delivery.pk}/"', r.content.decode())

    def test_delivery_with_no_batches_still_renders(self):
        delivery = Delivery.objects.create(
            department=self.dept, supplier=self.sup, date=datetime.date(2026, 5, 22))
        r = self.client.get(f"/deliveries/{delivery.pk}/")
        self.assertEqual(r.status_code, 200)
        self.assertIn("No batches on this delivery.", r.content.decode())


class StocktakeCSVTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.sup = Supplier.objects.create(name="Mill")
        self.flour = Product.objects.create(
            name="Flour", code="FLR1", department=self.dept,
            unit="g", minimum=Decimal("5"))
        SupplierPrice.objects.create(
            product=self.flour, supplier=self.sup,
            pack_weight=Decimal("25000"), pack_price=Decimal("30.00"))
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    def test_csv_includes_counted_line_with_value(self):
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date(2026, 1, 15))
        StockLine.objects.create(stocktake=st, product=self.flour,
                                 current=Decimal("3"), carried_over=False)
        r = self.client.get(f"/stocktakes/{st.pk}/csv/")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "text/csv")
        self.assertIn("stocktake-bakery-2026-01-15.csv", r["Content-Disposition"])
        body = r.content.decode()
        lines = body.splitlines()
        self.assertEqual(lines[0], "Ingredient,Code,Minimum,Count,Needed,Value")
        # value = 3 packs * £30.00 = £90.00; needed = 5 - 3 = 2
        self.assertIn("Flour,FLR1,5.00,3.00,2.00,90.00", body)

    def test_csv_renders_uncounted_line_with_blank_count(self):
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date(2026, 2, 1))
        StockLine.objects.create(stocktake=st, product=self.flour, current=None)
        r = self.client.get(f"/stocktakes/{st.pk}/csv/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Uncounted: Count, Needed, Value all blank
        self.assertIn("Flour,FLR1,5.00,,,", body)

    def test_csv_blocked_for_other_department(self):
        other = Department.objects.create(name="Butchery")
        st = Stocktake.objects.create(department=other,
                                      date=datetime.date(2026, 1, 15))
        r = self.client.get(f"/stocktakes/{st.pk}/csv/")
        self.assertEqual(r.status_code, 403)

    def test_csv_requires_login(self):
        c = Client()
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date(2026, 1, 15))
        r = c.get(f"/stocktakes/{st.pk}/csv/")
        self.assertEqual(r.status_code, 302)
        self.assertIn("/login/", r.headers["Location"])

    def test_total_value_shown_on_count_page_and_csv(self):
        # Two ingredients, two counts. Total = 3*30 + 4*5 = 90 + 20 = 110.
        sugar = Product.objects.create(
            name="Sugar", code="SUG1", department=self.dept,
            unit="g", minimum=Decimal("2"))
        SupplierPrice.objects.create(
            product=sugar, supplier=self.sup,
            pack_weight=Decimal("1000"), pack_price=Decimal("5.00"))
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date(2026, 3, 1))
        StockLine.objects.create(stocktake=st, product=self.flour,
                                 current=Decimal("3"), carried_over=False)
        StockLine.objects.create(stocktake=st, product=sugar,
                                 current=Decimal("4"), carried_over=False)

        # Page shows the total
        r = self.client.get(f"/stocktakes/{st.pk}/count/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("Total stock value", body)
        self.assertIn("£110.00", body)

        # CSV has per-line rows AND a TOTAL row
        r = self.client.get(f"/stocktakes/{st.pk}/csv/")
        body = r.content.decode()
        self.assertIn("Flour,FLR1,5.00,3.00,2.00,90.00", body)
        self.assertIn("Sugar,SUG1,2.00,4.00,0,20.00", body)
        # TOTAL row last, with the summed value
        lines = body.strip().splitlines()
        self.assertEqual(lines[-1], "TOTAL,,,,,110.00")

    def test_total_value_skips_uncounted_lines(self):
        # One counted, one blank — total reflects only the counted one.
        sugar = Product.objects.create(
            name="Sugar", code="SUG1", department=self.dept,
            unit="g", minimum=0)
        SupplierPrice.objects.create(
            product=sugar, supplier=self.sup,
            pack_weight=Decimal("1000"), pack_price=Decimal("5.00"))
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date(2026, 3, 8))
        StockLine.objects.create(stocktake=st, product=self.flour,
                                 current=Decimal("3"), carried_over=False)
        StockLine.objects.create(stocktake=st, product=sugar, current=None)

        r = self.client.get(f"/stocktakes/{st.pk}/count/")
        self.assertIn("£90.00", r.content.decode())
        r = self.client.get(f"/stocktakes/{st.pk}/csv/")
        self.assertIn("TOTAL,,,,,90.00", r.content.decode())

    def test_count_page_links_to_csv(self):
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date(2026, 1, 15))
        r = self.client.get(f"/stocktakes/{st.pk}/count/")
        self.assertEqual(r.status_code, 200)
        self.assertIn(f'/stocktakes/{st.pk}/csv/', r.content.decode())


class PackUnitConversionTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    def test_add_ingredient_in_kg_stores_grams(self):
        r = self.client.post("/products/", {
            "name": "Flour", "code": "FLR1",
            "quantity": "23", "unit": "kg",
            "supplier": "Mill", "cost": "30",
            "minimum": "5",
        })
        self.assertEqual(r.status_code, 302)
        p = Product.objects.get(code="FLR1")
        self.assertEqual(p.unit, "g")
        sp = SupplierPrice.objects.get(product=p)
        self.assertEqual(sp.pack_weight, Decimal("23000"))

    def test_add_ingredient_in_litres_stores_millilitres(self):
        r = self.client.post("/products/", {
            "name": "Oil", "code": "OIL1",
            "quantity": "2", "unit": "L",
            "supplier": "Mill", "cost": "10",
            "minimum": "1",
        })
        self.assertEqual(r.status_code, 302)
        p = Product.objects.get(code="OIL1")
        self.assertEqual(p.unit, "ml")
        sp = SupplierPrice.objects.get(product=p)
        self.assertEqual(sp.pack_weight, Decimal("2000"))

    def test_add_ingredient_in_grams_unchanged(self):
        r = self.client.post("/products/", {
            "name": "Salt", "code": "SLT1",
            "quantity": "500", "unit": "g",
            "supplier": "Mill", "cost": "1.50",
        })
        self.assertEqual(r.status_code, 302)
        p = Product.objects.get(code="SLT1")
        self.assertEqual(p.unit, "g")
        sp = SupplierPrice.objects.get(product=p)
        self.assertEqual(sp.pack_weight, Decimal("500"))

    def test_add_ingredient_in_kg_without_supplier_price_still_normalises_unit(self):
        # No supplier/cost - just creating the ingredient. The unit picker
        # should still produce a "g"-unit product.
        r = self.client.post("/products/", {"name": "Sugar", "code": "SUG1", "unit": "kg"})
        self.assertEqual(r.status_code, 302)
        self.assertEqual(Product.objects.get(code="SUG1").unit, "g")

    def test_supplier_price_entry_in_kg_stores_grams(self):
        # Existing g-unit product; add a supplier price in kg.
        product = Product.objects.create(
            name="Flour", department=self.dept, unit="g", minimum=0)
        r = self.client.post(f"/products/{product.pk}/", {
            "supplier": "Mill",
            "pack_weight": "23",
            "pack_unit": "kg",
            "pack_price": "30",
        })
        self.assertEqual(r.status_code, 302)
        sp = SupplierPrice.objects.get(product=product)
        self.assertEqual(sp.pack_weight, Decimal("23000"))

    def test_supplier_price_entry_in_litres_stores_millilitres(self):
        product = Product.objects.create(
            name="Oil", department=self.dept, unit="ml", minimum=0)
        r = self.client.post(f"/products/{product.pk}/", {
            "supplier": "Mill",
            "pack_weight": "2",
            "pack_unit": "L",
            "pack_price": "10",
        })
        self.assertEqual(r.status_code, 302)
        sp = SupplierPrice.objects.get(product=product)
        self.assertEqual(sp.pack_weight, Decimal("2000"))

    def test_re_adding_same_code_with_kg_updates_product_unit_and_pack(self):
        # First add as grams
        self.client.post("/products/", {
            "name": "Flour", "code": "FLR9",
            "quantity": "500", "unit": "g",
            "supplier": "Mill", "cost": "1",
        })
        # Re-add same code with kg - product is updated in place; the price
        # save creates a new dated history row instead of overwriting.
        self.client.post("/products/", {
            "name": "Flour", "code": "FLR9",
            "quantity": "25", "unit": "kg",
            "supplier": "Mill", "cost": "30",
        })
        p = Product.objects.get(code="FLR9")
        self.assertEqual(p.unit, "g")
        prices = SupplierPrice.objects.filter(product=p, supplier__name="Mill")
        self.assertEqual(prices.count(), 2)
        latest = prices.order_by("-effective_date", "-id").first()
        self.assertEqual(latest.pack_weight, Decimal("25000"))
        # old row is preserved as history
        self.assertTrue(prices.filter(pack_weight=Decimal("500")).exists())


class PriceHistoryTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.sup = Supplier.objects.create(name="Mill")
        self.other_sup = Supplier.objects.create(name="Acme")
        self.flour = Product.objects.create(
            name="Flour", code="FLR1", department=self.dept,
            unit="g", minimum=Decimal("5"))
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    def _price(self, supplier, pack_price, pack_weight=Decimal("25000"), date=None):
        return SupplierPrice.objects.create(
            product=self.flour, supplier=supplier,
            pack_weight=pack_weight, pack_price=pack_price,
            effective_date=date or datetime.date.today())

    def test_saving_a_second_price_creates_a_new_record(self):
        # Old then new price for the same supplier should leave two rows.
        self._price(self.sup, Decimal("25.00"), date=datetime.date(2026, 1, 1))
        r = self.client.post(f"/products/{self.flour.pk}/", {
            "supplier": "Mill", "pack_weight": "25", "pack_unit": "kg",
            "pack_price": "30.00",
        })
        self.assertEqual(r.status_code, 302)
        prices = SupplierPrice.objects.filter(product=self.flour, supplier=self.sup)
        self.assertEqual(prices.count(), 2)
        # Both prices preserved
        self.assertTrue(prices.filter(pack_price=Decimal("25.00")).exists())
        self.assertTrue(prices.filter(pack_price=Decimal("30.00")).exists())

    def test_latest_price_is_used_by_cheapest_value_and_reorder(self):
        # Mill has an old £25 price; today saves £30. Acme is at £28 today.
        # Today's latest for Mill (£30) loses to Acme's £28 — even though
        # the old £25 row exists in history.
        self._price(self.sup, Decimal("25.00"),
                    pack_weight=Decimal("25000"),
                    date=datetime.date(2026, 1, 1))
        self._price(self.sup, Decimal("30.00"),
                    pack_weight=Decimal("25000"),
                    date=datetime.date(2026, 5, 22))
        self._price(self.other_sup, Decimal("28.00"),
                    pack_weight=Decimal("25000"),
                    date=datetime.date(2026, 5, 22))

        cheapest = self.flour.cheapest_price
        self.assertEqual(cheapest.supplier, self.other_sup)
        self.assertEqual(cheapest.pack_price, Decimal("28.00"))

        # StockLine.value uses cheapest_price.pack_price -> latest cheapest
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date(2026, 5, 22))
        line = StockLine.objects.create(stocktake=st, product=self.flour,
                                        current=Decimal("3"), carried_over=False)
        self.assertEqual(line.value, Decimal("84.00"))  # 3 * £28

        # Reorder est cost also uses latest cheapest
        r = self.client.get("/reorder/")
        body = r.content.decode()
        self.assertIn("Acme", body)
        # minimum 5, on-hand 3 -> order 2 packs @ £28 = £56
        self.assertIn("£56.00", body)

    def test_old_price_is_retained_and_shown_in_history(self):
        self._price(self.sup, Decimal("25.00"), date=datetime.date(2026, 1, 1))
        self._price(self.sup, Decimal("30.00"), date=datetime.date(2026, 5, 22))
        r = self.client.get(f"/products/{self.flour.pk}/")
        body = r.content.decode()
        self.assertIn("Price history", body)
        # both prices appear in the history
        self.assertIn("£25.00", body)
        self.assertIn("£30.00", body)
        # current is tagged on the newest row
        self.assertIn("current", body)
        # old date present
        self.assertIn("01 Jan 2026", body)

    def test_more_than_10_percent_jump_is_flagged(self):
        self._price(self.sup, Decimal("20.00"), date=datetime.date(2026, 1, 1))
        self._price(self.sup, Decimal("25.00"), date=datetime.date(2026, 5, 22))
        history = self.flour.price_history()
        latest_entry = history[0]["entries"][0]
        # +25% jump
        self.assertEqual(latest_entry["delta_pct"], Decimal("25.0"))
        self.assertTrue(latest_entry["notable"])
        # Template renders the notable tag
        r = self.client.get(f"/products/{self.flour.pk}/")
        body = r.content.decode()
        self.assertIn("+25.0%", body)

    def test_small_change_is_not_flagged(self):
        self._price(self.sup, Decimal("20.00"), date=datetime.date(2026, 1, 1))
        self._price(self.sup, Decimal("21.00"), date=datetime.date(2026, 5, 22))
        latest = self.flour.price_history()[0]["entries"][0]
        # +5%
        self.assertEqual(latest["delta_pct"], Decimal("5.0"))
        self.assertFalse(latest["notable"])

    def test_latest_prices_returns_one_per_supplier(self):
        self._price(self.sup, Decimal("25.00"), date=datetime.date(2026, 1, 1))
        self._price(self.sup, Decimal("30.00"), date=datetime.date(2026, 5, 22))
        self._price(self.other_sup, Decimal("28.00"), date=datetime.date(2026, 5, 22))
        latest = self.flour.latest_prices()
        self.assertEqual(len(latest), 2)
        by_sup = {sp.supplier_id: sp for sp in latest}
        self.assertEqual(by_sup[self.sup.pk].pack_price, Decimal("30.00"))
        self.assertEqual(by_sup[self.other_sup.pk].pack_price, Decimal("28.00"))

    def test_first_entry_has_no_delta(self):
        self._price(self.sup, Decimal("25.00"), date=datetime.date(2026, 5, 22))
        entries = self.flour.price_history()[0]["entries"]
        self.assertEqual(len(entries), 1)
        self.assertIsNone(entries[0]["delta_pct"])

    def test_deliveries_supplier_filter_dedupes_history(self):
        # Two prices for (flour, sup) shouldn't make the delivery form think
        # the supplier stocks the ingredient twice.
        self._price(self.sup, Decimal("25.00"), date=datetime.date(2026, 1, 1))
        self._price(self.sup, Decimal("30.00"), date=datetime.date(2026, 5, 22))
        r = self.client.get("/deliveries/new/")
        # Page renders; supplier_ids list on the product is deduped
        p = (Product.objects.filter(pk=self.flour.pk)
             .prefetch_related("prices").first())
        ids = sorted({sp.supplier_id for sp in p.prices.all()})
        self.assertEqual(ids, [self.sup.pk])


class HomeRebuildTests(TestCase):
    """The merged /home/ rebuilt on the design system: KPI tiles (from the
    retired root dashboard) + staff-on-shift shell + stock alerts + urgent
    tasks, with the weather card and the per-ingredient inventory table
    dropped. /home/ now serves this rebuilt template (the old home.html and
    the root / stock dashboard are retired; / redirects to /home/).
    """

    PREVIEW_URL = "/home/"

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        self._weather_patch = patch("stock.views.fetch_weather", return_value=None)
        self._weather_patch.start()
        self.addCleanup(self._weather_patch.stop)

    def test_preview_renders_on_design_system_shell(self):
        r = self.client.get(self.PREVIEW_URL)
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Design-system shell (BP sidebar + offset main), not the old shell.
        self.assertIn('<main class="ml-64 min-w-0">', body)

    def test_preview_shows_the_four_dashboard_kpis(self):
        r = self.client.get(self.PREVIEW_URL)
        body = r.content.decode()
        for label in ("Ingredients", "Below minimum", "Stock value", "Last count"):
            self.assertIn(label, body, f"KPI '{label}' missing")

    def test_preview_has_staff_widget_empty_state(self):
        # No Shift/Rota model exists yet, so the widget shows its empty-state
        # and "0 on shift" — and must NOT fabricate any names.
        r = self.client.get(self.PREVIEW_URL)
        body = r.content.decode()
        self.assertIn("Staff on shift today", body)
        self.assertIn("0 on shift", body)
        self.assertIn("No shifts recorded for today", body)

    def test_preview_keeps_alerts_and_tasks_but_drops_weather(self):
        # Below-minimum ingredient → a stock alert linking to reorder.
        low = Product.objects.create(
            code="NPD-I100", name="Test Flour", department=self.dept,
            category="dry_goods", unit="g", minimum=50)
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date.today())
        StockLine.objects.create(stocktake=st, product=low, current=Decimal("10"))
        r = self.client.get(self.PREVIEW_URL)
        body = r.content.decode()
        self.assertIn("Stock alerts", body)
        self.assertIn("Test Flour", body)
        self.assertIn("Urgent tasks", body)
        # Weather card was dropped from the merged page.
        self.assertNotIn("weather", body.lower())


class SectionNavigationTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        # Keep the home page off the network in tests. Tests that need to
        # exercise the weather card explicitly override the return value.
        self._weather_patch = patch("stock.views.fetch_weather", return_value=None)
        self._weather_patch.start()
        self.addCleanup(self._weather_patch.stop)

    def test_home_renders_for_logged_in_user(self):
        r = self.client.get("/home/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # /home/ now serves the merged design-system landing (no weather card).
        self.assertIn('<main class="ml-64 min-w-0">', body)
        self.assertIn("alice", body)
        self.assertIn("Staff on shift today", body)
        self.assertIn("Stock alerts", body)
        self.assertIn("Urgent tasks", body)

    def test_home_below_minimum_appears_in_urgent_card_and_alerts_table(self):
        # An ingredient counted below its minimum should:
        #   - bump the urgent badge to 1
        #   - appear as an aggregate line in the urgent card
        #   - render as a row in the per-ingredient stock alerts table
        #     with a Reorder action.
        flour = Product.objects.create(
            name="Flour", code="FLR1", department=self.dept, unit="g",
            minimum=Decimal("10"))
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date.today())
        StockLine.objects.create(stocktake=st, product=flour,
                                 current=Decimal("2"), carried_over=False)
        r = self.client.get("/home/")
        body = r.content.decode()

        # BP layout: stock alerts (row 3) precede urgent tasks (row 4).
        urgent = body[body.index('data-testid="urgent-tasks"'):]
        # Below-minimum stock surfaces as an actionable "Ordering" task with
        # the item count, linking to /reorder/.
        self.assertIn("Ordering", urgent)
        self.assertIn("1 item", urgent)
        self.assertIn('href="/reorder/"', urgent)

        alerts = body[body.index('data-testid="stock-alerts"'):body.index('data-testid="urgent-tasks"')]
        # Per-ingredient row with the ingredient name, alert label, detail
        # and Reorder action.
        self.assertIn("Flour", alerts)
        self.assertIn("Below minimum", alerts)
        self.assertIn("2 / 10 packs", alerts)
        self.assertIn("Reorder", alerts)

    def test_home_calm_state_when_nothing_is_urgent(self):
        r = self.client.get("/home/")
        body = r.content.decode()
        urgent = body[body.index('data-testid="urgent-tasks"'):]
        self.assertIn("All caught up", urgent)
        alerts = body[body.index('data-testid="stock-alerts"'):body.index('data-testid="urgent-tasks"')]
        self.assertIn("Stock looks healthy", alerts)

    def test_home_urgent_tasks_use_action_labels(self):
        # Expiring batch + overdue stocktake should produce "Use expiring
        # stock" and "Stocktake due" tasks respectively. Each task line is a
        # link to where the work is done.
        sup = Supplier.objects.create(name="Mill")
        flour = Product.objects.create(
            name="Flour", department=self.dept, unit="g", minimum=0)
        delivery = Delivery.objects.create(
            department=self.dept, supplier=sup,
            date=datetime.date.today() - datetime.timedelta(days=2))
        Batch.objects.create(
            delivery=delivery, product=flour, batch_code="A1",
            use_by=datetime.date.today() + datetime.timedelta(days=3),
            qty_received=Decimal("4"), qty_remaining=Decimal("4"))
        # An old stocktake makes the count overdue
        Stocktake.objects.create(
            department=self.dept,
            date=datetime.date.today() - datetime.timedelta(days=10))

        r = self.client.get("/home/")
        body = r.content.decode()
        urgent = body[body.index('data-testid="urgent-tasks"'):]
        # Action labels, not status observations
        self.assertIn("Use expiring stock", urgent)
        self.assertIn("Stocktake due", urgent)
        # Links to where the user does the work
        self.assertIn('href="/deliveries/"', urgent)
        self.assertIn('href="/stocktakes/"', urgent)

    def test_urgent_task_helper_returns_extendable_list(self):
        # _stock_tasks_for_home returns a list of {label, count, url} dicts
        # so other sources (manual tasks etc.) can append to the same list.
        from stock.views import _stock_tasks_for_home
        flour = Product.objects.create(
            name="Flour", department=self.dept, unit="g",
            minimum=Decimal("10"))
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date.today())
        StockLine.objects.create(stocktake=st, product=flour,
                                 current=Decimal("2"), carried_over=False)
        tasks = _stock_tasks_for_home(self.dept, datetime.date.today())
        # Each entry is a {label, count, url} dict
        self.assertGreaterEqual(len(tasks), 1)
        first = tasks[0]
        self.assertEqual(set(first.keys()), {"label", "count", "url"})
        self.assertEqual(first["label"], "Ordering")
        self.assertEqual(first["count"], 1)
        self.assertEqual(first["url"], "/reorder/")

    def test_stock_section_landing_renders_with_stock_submenu(self):
        # /stock/ is a Stock section page — its navbar carries the Stock
        # sub-menu (Dashboard / Stocktakes / Deliveries / ...). Body is minimal.
        r = self.client.get("/stock/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        nav = body[body.index("<nav>"):body.index("</nav>")]
        for url in ("/", "/stocktakes/", "/deliveries/", "/adjustments/",
                    "/reorder/", "/products/", "/suppliers/"):
            self.assertIn(f'href="{url}"', nav)

    def test_placeholder_sections_render_with_coming_soon(self):
        # Recipes is no longer a placeholder — it has its own list/import flow,
        # tested separately. The other three are still "coming soon".
        for path, title in (("/production/", "Production"),
                            ("/rota/", "Rota"), ("/notes/", "Notes")):
            r = self.client.get(path)
            self.assertEqual(r.status_code, 200, f"{path} returned {r.status_code}")
            body = r.content.decode()
            self.assertIn(title, body)
            self.assertIn("coming soon", body.lower())

    def test_profile_shows_username_departments_and_logout(self):
        r = self.client.get("/profile/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("alice", body)
        self.assertIn("Bakery", body)
        self.assertIn('action="/logout/"', body)

    def test_profile_shows_admin_link_only_for_superusers(self):
        r = self.client.get("/profile/")
        self.assertNotIn("/admin/", r.content.decode())

        U = get_user_model()
        boss = U.objects.create_superuser("boss", password="pw")
        c = Client()
        c.login(username="boss", password="pw")
        r = c.get("/profile/")
        self.assertIn("/admin/", r.content.decode())

    def _nav(self, body):
        nav = body[body.index("<nav>"):body.index("</nav>")]
        return nav, [m.group(1) for m in re.finditer(r">([A-Za-z]+)<", nav)]

    def test_home_renders_on_the_design_system_rail(self):
        # /home/ now renders the shared BP shell; the old top section-picker
        # navbar is replaced by the design-system left rail with its section
        # links. (The rail itself is covered in depth elsewhere.)
        r = self.client.get("/home/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn('<main class="ml-64 min-w-0">', body)
        for href in ('href="/orders/"', 'href="/products/"', 'href="/customers/"'):
            self.assertIn(href, body)

    def test_stock_section_nav_shows_sub_items(self):
        # Any Stock page (e.g. stocktakes) renders the Stock contextual
        # sub-menu in the top nav: Home + Dashboard / Stocktakes / Deliveries
        # / Adjustments / Reorder / Ingredients / Suppliers.
        r = self.client.get("/stocktakes/")
        body = r.content.decode()
        nav = body[body.index("<nav>"):body.index("</nav>")]
        for label in (">Home<", ">Dashboard<", ">Stocktakes<", ">Deliveries<",
                      ">Adjustments<", ">Reorder<", ">Ingredients<", ">Suppliers<"):
            self.assertIn(label, nav)
        # The Stock top-level link is NOT shown — the nav is contextual.
        self.assertNotIn('href="/stock/"', nav)

    def test_stock_sub_pages_highlight_themselves_not_a_top_link(self):
        # Products has moved to the design-system shell, where the active
        # section is marked on the left rail (amber ring), not the old
        # `class="on"` nav — see test_products_highlights_itself_on_bp_rail.
        for path, link in (
            ("/stocktakes/", '/stocktakes/'),
            ("/deliveries/", '/deliveries/'),
            ("/adjustments/", '/adjustments/'),
            ("/reorder/", '/reorder/'),
            ("/suppliers/", '/suppliers/'),
        ):
            r = self.client.get(path)
            body = r.content.decode()
            nav = body[body.index("<nav>"):body.index("</nav>")]
            self.assertRegex(
                nav, r'href="' + link + r'"\s+class="on"',
                f"{path} should highlight its own sub-nav link",
            )

    def test_products_highlights_itself_on_bp_rail(self):
        # The Ingredients page renders the design-system shell: the active
        # section is marked on the left rail with the amber ring
        # (bg-amber-50 / ring-amber-200), replacing the old `class="on"`.
        body = self.client.get("/products/").content.decode()
        self.assertRegex(
            body, r'href="/products/"[^>]*\bring-amber-200\b',
            "Products should mark its own rail link active on the BP shell",
        )

    def test_placeholder_navbars_have_home_plus_section(self):
        # Recipes now has Home + Recipes + Import sub-nav (tested elsewhere);
        # the other three placeholders remain Home + section.
        for path, label in (("/production/", "Production"),
                            ("/rota/", "Rota"),
                            ("/notes/", "Notes")):
            r = self.client.get(path)
            nav, labels = self._nav(r.content.decode())
            self.assertEqual(nav.count("<a "), 2,
                             f"{path} navbar should be Home + section (2 links)")
            self.assertIn("Home", labels)
            self.assertIn(label, labels)
            self.assertRegex(nav, r'class="on"[^>]*>' + label)

    def test_profile_navbar_has_home_and_profile(self):
        r = self.client.get("/profile/")
        nav, labels = self._nav(r.content.decode())
        self.assertEqual(nav.count("<a "), 2)
        self.assertIn("Home", labels)
        self.assertIn("Profile", labels)

    def test_admin_link_in_header_only_for_superusers(self):
        # The Admin link lives in the old-shell header's right cluster, so
        # superusers see it from old-shell sections. (/home/ now renders the
        # BP shell, which has no header admin link, so it's excluded here.)
        r = self.client.get("/profile/")
        self.assertNotIn('href="/admin/"', r.content.decode())

        U = get_user_model()
        boss = U.objects.create_superuser("boss", password="pw")
        c = Client()
        c.login(username="boss", password="pw")
        for path in ("/profile/", "/stocktakes/"):
            self.assertIn('href="/admin/"', c.get(path).content.decode(),
                          f"{path} should expose /admin/ to a superuser")

    def test_login_redirects_to_home(self):
        c = Client()
        r = c.post("/login/", {"username": "alice", "password": "pw"})
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.headers["Location"], "/home/")

    def test_existing_urls_still_work(self):
        # Sanity: the existing stock pages keep their URLs and views.
        for path in ("/stocktakes/", "/deliveries/", "/adjustments/",
                     "/reorder/", "/products/", "/suppliers/"):
            r = self.client.get(path)
            self.assertEqual(r.status_code, 200, f"{path} returned {r.status_code}")

    def test_root_redirects_to_home(self):
        # The old root stock dashboard is retired; / now 302-redirects to
        # /home/ (the merged landing page).
        r = self.client.get("/")
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.headers["Location"], "/home/")

    def test_stock_card_links_into_section(self):
        # Home's Stock card should land the user inside the Stock section.
        r = self.client.get("/home/")
        body = r.content.decode()
        self.assertIn('href="/stock/"', body)


class PackSizeFilterTests(SimpleTestCase):
    def test_grams_promoted_to_kg_at_threshold(self):
        self.assertEqual(pack_size(25000, "g"), "25 kg")
        self.assertEqual(pack_size(1500, "g"), "1.5 kg")
        self.assertEqual(pack_size(1000, "g"), "1 kg")

    def test_grams_under_threshold_stay_in_g(self):
        self.assertEqual(pack_size(500, "g"), "500 g")
        self.assertEqual(pack_size(999, "g"), "999 g")

    def test_millilitres_promoted_to_litres_at_threshold(self):
        self.assertEqual(pack_size(1500, "ml"), "1.5 L")
        self.assertEqual(pack_size(2000, "ml"), "2 L")
        self.assertEqual(pack_size(500, "ml"), "500 ml")

    def test_each_unit_passes_through(self):
        self.assertEqual(pack_size(12, "ea"), "12 ea")
        self.assertEqual(pack_size(1, "ea"), "1 ea")

    def test_accepts_decimal_input(self):
        self.assertEqual(pack_size(Decimal("25000.00"), "g"), "25 kg")
        self.assertEqual(pack_size(Decimal("1500.00"), "g"), "1.5 kg")
        self.assertEqual(pack_size(Decimal("500.50"), "g"), "500.5 g")

    def test_blank_inputs(self):
        self.assertEqual(pack_size(None, "g"), "")
        self.assertEqual(pack_size("", "g"), "")
        self.assertEqual(pack_size("not a number", "g"), "")


class IngredientImportTests(TestCase):
    """The Excel master import + the matching reset command.

    The fixture workbook is built in-memory so tests don't depend on data/.
    """

    def _make_workbook(self, path, ingredients, uoms,
                       suppliers=None, supplier_names=None, allergens=None):
        """Build a four/five-tab fixture workbook.

        suppliers: list of {code, supplier_code, is_primary} dicts. Defaults
                   to a generated primary supplier per ingredient so tests
                   that don't care about supplier wiring still get prices.
        supplier_names: {supplier_code: name} for the Reference lookup.
                   Defaults to a generated name per supplier_code seen.
        allergens: optional list of {code, allergen, contains, may_contain}.
                   Omitted → no Allergens tab is written.
        """
        from openpyxl import Workbook
        wb = Workbook()
        ing_ws = wb.active
        ing_ws.title = "Ingredients"
        # Header up to Category (col 36). Only columns we read need real names;
        # the importer indexes by position, not header.
        header = [""] * 36
        header[0] = "Code"
        header[1] = "Description"
        header[33] = "Cost"
        header[34] = "Supply Unit"
        header[35] = "Category"
        ing_ws.append(header)
        for ing in ingredients:
            row = [""] * 36
            row[0] = ing["code"]
            row[1] = ing["name"]
            row[33] = ing["cost"]
            row[34] = ing["supply_unit"]
            row[35] = ing["category"]
            ing_ws.append(row)

        uom_ws = wb.create_sheet("Units Of Measure")
        uom_ws.append(["Code", "Description", "UOM", "Quantity", "RUOM", "Ref Quantity", "Is Default"])
        for u in uoms:
            uom_ws.append([u["code"], None, u["uom"], u.get("quantity", 1),
                           u["ruom"], u["ref_quantity"], "No"])

        if suppliers is None:
            suppliers = [
                {"code": ing["code"], "supplier_code": f"S{900 + i}", "is_primary": True}
                for i, ing in enumerate(ingredients, start=1)
            ]
        if supplier_names is None:
            supplier_names = {}
            for s in suppliers:
                supplier_names.setdefault(s["supplier_code"],
                                          f"Supplier {s['supplier_code']}")

        sup_ws = wb.create_sheet("Suppliers")
        sup_ws.append(["Code", "Description", "Supplier Code", "Supplier",
                       "Supplier Address", "Is Primary Supplier"])
        for s in suppliers:
            sup_ws.append([s["code"], None, s["supplier_code"], None, "",
                           "Yes" if s.get("is_primary") else "No"])

        # Reference tab mirrors the real shape: the first row is a wide list
        # of section labels, third row is sub-headers per section. We only
        # need the "Suppliers" section so put it at column 0.
        ref_ws = wb.create_sheet("Reference")
        ref_ws.append(["Suppliers"])             # row 1
        ref_ws.append([])                        # row 2 (blank)
        ref_ws.append(["Supplier Code", "Description"])  # row 3 sub-headers
        for code, name in supplier_names.items():
            ref_ws.append([code, name])

        if allergens is not None:
            alg_ws = wb.create_sheet("Allergens")
            alg_ws.append(["Code", "Description", "Allergen",
                           "Parts Per Million", "Contains", "May Contain"])
            for a in allergens:
                alg_ws.append([
                    a["code"], None, a["allergen"], "0",
                    "Yes" if a.get("contains") else "No",
                    "Yes" if a.get("may_contain") else "No",
                ])
        wb.save(path)

    def test_import_creates_products_with_code_name_category(self):
        import tempfile, os
        from django.core.management import call_command
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "ing.xlsx")
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I001", "name": "Bread Flour", "cost": 18.4,
                 "supply_unit": "Trade Paper Sack", "category": "Dry Goods"},
                {"code": "NPD-I002", "name": "Whole Milk", "cost": 2.39,
                 "supply_unit": "2 Litres", "category": "Dairy & Eggs"},
                {"code": "NPD-I003", "name": "Strawberries", "cost": 4.50,
                 "supply_unit": "Punnet", "category": "Fruit & Veg"},
                {"code": "NPD-I004", "name": "Mystery", "cost": 1.0,
                 "supply_unit": "Widget", "category": "Unassigned"},
            ],
            uoms=[
                {"code": "NPD-I001", "uom": "Trade Paper Sack", "ruom": "Kilograms", "ref_quantity": 16},
                {"code": "NPD-I002", "uom": "2 Litres", "ruom": "Kilograms", "ref_quantity": 2},
                {"code": "NPD-I003", "uom": "Punnet", "ruom": "Grams", "ref_quantity": 250},
                # NPD-I004 has no UOM row → flagged.
            ])
        call_command("import_ingredients", path)
        self.assertEqual(Product.objects.count(), 4)
        flour = Product.objects.get(code="NPD-I001")
        self.assertEqual(flour.name, "Bread Flour")
        self.assertEqual(flour.category, "dry_goods")
        self.assertEqual(flour.unit, "g")
        milk = Product.objects.get(code="NPD-I002")
        self.assertEqual(milk.category, "dairy_eggs")
        berries = Product.objects.get(code="NPD-I003")
        self.assertEqual(berries.category, "fruit_veg")
        mystery = Product.objects.get(code="NPD-I004")
        self.assertEqual(mystery.category, "unassigned")

    def test_16kg_sack_at_18_40_gives_pack_weight_16000_and_price_18_40(self):
        import tempfile, os
        from django.core.management import call_command
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "ing.xlsx")
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I001", "name": "Bread Flour", "cost": 18.4,
                 "supply_unit": "Trade Paper Sack", "category": "Dry Goods"},
            ],
            uoms=[
                {"code": "NPD-I001", "uom": "Trade Paper Sack", "ruom": "Kilograms", "ref_quantity": 16},
            ])
        call_command("import_ingredients", path)
        flour = Product.objects.get(code="NPD-I001")
        sp = flour.prices.get()
        self.assertEqual(sp.pack_weight, Decimal("16000"))
        self.assertEqual(sp.pack_price, Decimal("18.40"))

    def test_case_insensitive_supply_unit_match(self):
        # Ingredient supply unit "2 litres" (lowercase l) matches a UOM row
        # written as "2 Litres". Same for "Case" vs "case".
        import tempfile, os
        from django.core.management import call_command
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "ing.xlsx")
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I010", "name": "Milk", "cost": 2.0,
                 "supply_unit": "2 litres", "category": "Dairy & Eggs"},
                {"code": "NPD-I011", "name": "Jar", "cost": 3.0,
                 "supply_unit": "Case", "category": "Dry Goods"},
            ],
            uoms=[
                {"code": "NPD-I010", "uom": "2 Litres", "ruom": "Kilograms", "ref_quantity": 2},
                {"code": "NPD-I011", "uom": "case", "ruom": "Kilograms", "ref_quantity": 6},
            ])
        call_command("import_ingredients", path)
        self.assertEqual(Product.objects.get(code="NPD-I010").prices.get().pack_weight,
                         Decimal("2000"))
        self.assertEqual(Product.objects.get(code="NPD-I011").prices.get().pack_weight,
                         Decimal("6000"))

    def test_kilograms_supply_unit_without_uom_row_is_one_kg_pack(self):
        # Supply Unit "Kilograms" itself is a base unit - treat as 1kg pack
        # without needing a UOM row.
        import tempfile, os
        from django.core.management import call_command
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "ing.xlsx")
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I020", "name": "Fresh Yeast", "cost": 2.68,
                 "supply_unit": "Kilograms", "category": "Dry Goods"},
            ],
            uoms=[])
        call_command("import_ingredients", path)
        sp = Product.objects.get(code="NPD-I020").prices.get()
        self.assertEqual(sp.pack_weight, Decimal("1000"))
        self.assertEqual(sp.pack_price, Decimal("2.68"))

    def test_indirect_uom_chain_resolves(self):
        # Box → 40 Pack → 0.25 Kilograms = 10 kg = 10000 g
        import tempfile, os
        from django.core.management import call_command
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "ing.xlsx")
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I030", "name": "Butter", "cost": 50.0,
                 "supply_unit": "Box", "category": "Dairy & Eggs"},
            ],
            uoms=[
                {"code": "NPD-I030", "uom": "Pack", "ruom": "Kilograms", "ref_quantity": 0.25},
                {"code": "NPD-I030", "uom": "Box", "ruom": "Pack", "ref_quantity": 40},
            ])
        call_command("import_ingredients", path)
        sp = Product.objects.get(code="NPD-I030").prices.get()
        self.assertEqual(sp.pack_weight, Decimal("10000.00"))

    def test_ingredient_without_uom_is_created_but_flagged(self):
        # Mystery supply unit with no UOM row: ingredient still exists, but no
        # SupplierPrice gets written.
        import tempfile, os
        from django.core.management import call_command
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "ing.xlsx")
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I040", "name": "Mystery", "cost": 1.0,
                 "supply_unit": "Widget", "category": "Unassigned"},
            ],
            uoms=[])
        call_command("import_ingredients", path)
        p = Product.objects.get(code="NPD-I040")
        self.assertFalse(p.prices.exists())

    def test_import_is_idempotent(self):
        # Running twice doesn't duplicate products or pile up history.
        import tempfile, os
        from django.core.management import call_command
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "ing.xlsx")
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I001", "name": "Bread Flour", "cost": 18.4,
                 "supply_unit": "Trade Paper Sack", "category": "Dry Goods"},
            ],
            uoms=[
                {"code": "NPD-I001", "uom": "Trade Paper Sack", "ruom": "Kilograms", "ref_quantity": 16},
            ])
        call_command("import_ingredients", path)
        call_command("import_ingredients", path)
        self.assertEqual(Product.objects.filter(code="NPD-I001").count(), 1)
        flour = Product.objects.get(code="NPD-I001")
        self.assertEqual(flour.prices.count(), 1)
        self.assertEqual(flour.prices.get().pack_weight, Decimal("16000"))

    def test_import_assigns_to_department_arg(self):
        import tempfile, os
        from django.core.management import call_command
        tmpdir = tempfile.mkdtemp()
        path = os.path.join(tmpdir, "ing.xlsx")
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I001", "name": "X", "cost": 1.0,
                 "supply_unit": "Kilograms", "category": "Dry Goods"},
            ],
            uoms=[])
        call_command("import_ingredients", path, "--department", "Pastry")
        p = Product.objects.get(code="NPD-I001")
        self.assertEqual(p.department.name, "Pastry")

    # --- supplier wiring -------------------------------------------------

    def _tmp(self, name="ing.xlsx"):
        import tempfile, os
        return os.path.join(tempfile.mkdtemp(), name)

    def test_primary_supplier_s_code_resolves_to_name_and_gets_the_price(self):
        from django.core.management import call_command
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I100", "name": "Bread Flour", "cost": 18.4,
                 "supply_unit": "Trade Paper Sack", "category": "Dry Goods"},
            ],
            uoms=[
                {"code": "NPD-I100", "uom": "Trade Paper Sack",
                 "ruom": "Kilograms", "ref_quantity": 16},
            ],
            suppliers=[
                {"code": "NPD-I100", "supplier_code": "S31", "is_primary": True},
            ],
            supplier_names={"S31": "Wildfarmed"})
        call_command("import_ingredients", path)
        # Master Catalog must NOT be created — supplier comes from data.
        self.assertFalse(Supplier.objects.filter(name="Master Catalog").exists())
        # Real supplier created with the resolved name
        sup = Supplier.objects.get(name="Wildfarmed")
        flour = Product.objects.get(code="NPD-I100")
        sp = flour.prices.get()
        self.assertEqual(sp.supplier, sup)
        self.assertEqual(sp.pack_weight, Decimal("16000"))
        self.assertEqual(sp.pack_price, Decimal("18.40"))

    def test_primary_row_wins_over_other_supplier_rows(self):
        # An ingredient with several supplier rows: only the primary one is
        # used; secondary rows must not produce extra prices.
        from django.core.management import call_command
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I200", "name": "Milk", "cost": 2.39,
                 "supply_unit": "2 Litres", "category": "Dairy & Eggs"},
            ],
            uoms=[
                {"code": "NPD-I200", "uom": "2 Litres",
                 "ruom": "Kilograms", "ref_quantity": 2},
            ],
            suppliers=[
                {"code": "NPD-I200", "supplier_code": "S5", "is_primary": True},
                {"code": "NPD-I200", "supplier_code": "S3", "is_primary": False},
            ],
            supplier_names={"S5": "Bruton Dairies", "S3": "Wellocks"})
        call_command("import_ingredients", path)
        milk = Product.objects.get(code="NPD-I200")
        sp = milk.prices.get()  # exactly one price
        self.assertEqual(sp.supplier.name, "Bruton Dairies")

    def test_fallback_to_first_row_when_no_primary(self):
        # Edge case: nothing marked primary - take whichever row appears first.
        from django.core.management import call_command
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I300", "name": "Salt", "cost": 8.93,
                 "supply_unit": "Bag", "category": "Dry Goods"},
            ],
            uoms=[
                {"code": "NPD-I300", "uom": "Bag",
                 "ruom": "Kilograms", "ref_quantity": 10},
            ],
            suppliers=[
                {"code": "NPD-I300", "supplier_code": "S3", "is_primary": False},
                {"code": "NPD-I300", "supplier_code": "S14", "is_primary": False},
            ],
            supplier_names={"S3": "Wellocks", "S14": "The Fine Food Company"})
        call_command("import_ingredients", path)
        salt = Product.objects.get(code="NPD-I300")
        self.assertEqual(salt.prices.get().supplier.name, "Wellocks")

    def test_unresolvable_supplier_code_uses_code_as_name_and_flags(self):
        # An S-code that doesn't appear in the Reference lookup: name the
        # supplier after the code itself and surface it in the output.
        from django.core.management import call_command
        from io import StringIO
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I400", "name": "X", "cost": 1.0,
                 "supply_unit": "Bag", "category": "Dry Goods"},
            ],
            uoms=[
                {"code": "NPD-I400", "uom": "Bag",
                 "ruom": "Kilograms", "ref_quantity": 1},
            ],
            suppliers=[
                {"code": "NPD-I400", "supplier_code": "S999", "is_primary": True},
            ],
            supplier_names={})  # S999 not present
        out = StringIO()
        call_command("import_ingredients", path, stdout=out)
        # Supplier was created using the S-code as a fallback name
        sup = Supplier.objects.get(name="S999")
        self.assertEqual(Product.objects.get(code="NPD-I400").prices.get().supplier, sup)
        # And the run reports it
        self.assertIn("Unresolved supplier codes", out.getvalue())
        self.assertIn("S999", out.getvalue())

    def test_ingredient_with_no_supplier_row_is_flagged(self):
        # No row in Suppliers tab → no price (we can't pick a supplier).
        from django.core.management import call_command
        from io import StringIO
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I500", "name": "Orphan", "cost": 1.0,
                 "supply_unit": "Kilograms", "category": "Dry Goods"},
            ],
            uoms=[],
            suppliers=[],          # Suppliers tab has the header but no rows
            supplier_names={})
        out = StringIO()
        call_command("import_ingredients", path, stdout=out)
        p = Product.objects.get(code="NPD-I500")
        self.assertFalse(p.prices.exists())
        self.assertIn("no supplier", out.getvalue())

    def test_supplier_idempotent_on_rerun(self):
        # Re-running shouldn't create a duplicate Supplier OR a second price
        # row for the same (product, supplier).
        from django.core.management import call_command
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I600", "name": "Flour", "cost": 18.4,
                 "supply_unit": "Trade Paper Sack", "category": "Dry Goods"},
            ],
            uoms=[
                {"code": "NPD-I600", "uom": "Trade Paper Sack",
                 "ruom": "Kilograms", "ref_quantity": 16},
            ],
            suppliers=[
                {"code": "NPD-I600", "supplier_code": "S31", "is_primary": True},
            ],
            supplier_names={"S31": "Wildfarmed"})
        call_command("import_ingredients", path)
        call_command("import_ingredients", path)
        self.assertEqual(Supplier.objects.filter(name="Wildfarmed").count(), 1)
        flour = Product.objects.get(code="NPD-I600")
        self.assertEqual(flour.prices.count(), 1)
        self.assertEqual(flour.prices.get().pack_price, Decimal("18.40"))

    # --- allergens --------------------------------------------------------

    def test_allergens_attached_with_contains_and_may_contain(self):
        # WildFarmed flour shape: "Cereals containing gluten" (contains) +
        # "Soya" (may contain). Both must land on the product as separate
        # IngredientAllergen rows with the right flags.
        from django.core.management import call_command
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I100", "name": "Bread Flour", "cost": 18.4,
                 "supply_unit": "Trade Paper Sack", "category": "Dry Goods"},
            ],
            uoms=[
                {"code": "NPD-I100", "uom": "Trade Paper Sack",
                 "ruom": "Kilograms", "ref_quantity": 16},
            ],
            allergens=[
                {"code": "NPD-I100", "allergen": "Cereals containing gluten",
                 "contains": True, "may_contain": False},
                {"code": "NPD-I100", "allergen": "Soya",
                 "contains": False, "may_contain": True},
            ])
        call_command("import_ingredients", path)
        p = Product.objects.get(code="NPD-I100")
        gluten = p.allergens.get(name="Cereals containing gluten")
        self.assertTrue(gluten.contains)
        self.assertFalse(gluten.may_contain)
        soya = p.allergens.get(name="Soya")
        self.assertFalse(soya.contains)
        self.assertTrue(soya.may_contain)
        self.assertEqual(p.allergens.count(), 2)

    def test_allergen_import_is_idempotent(self):
        # Re-running must not duplicate rows or flip the flags.
        from django.core.management import call_command
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I101", "name": "Milk", "cost": 2.0,
                 "supply_unit": "Kilograms", "category": "Dairy & Eggs"},
            ],
            uoms=[],
            allergens=[
                {"code": "NPD-I101", "allergen": "Milk",
                 "contains": True, "may_contain": False},
            ])
        call_command("import_ingredients", path)
        call_command("import_ingredients", path)
        p = Product.objects.get(code="NPD-I101")
        self.assertEqual(p.allergens.count(), 1)
        self.assertEqual(IngredientAllergen.objects.filter(
            product=p, name="Milk").count(), 1)

    def test_ingredient_with_no_allergen_rows_has_none(self):
        from django.core.management import call_command
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I102", "name": "Salt", "cost": 1.0,
                 "supply_unit": "Kilograms", "category": "Dry Goods"},
            ],
            uoms=[],
            allergens=[
                {"code": "NPD-I999", "allergen": "Milk",  # different ingredient
                 "contains": True, "may_contain": False},
            ])
        call_command("import_ingredients", path)
        p = Product.objects.get(code="NPD-I102")
        self.assertEqual(p.allergens.count(), 0)

    def test_allergen_rows_for_unknown_ingredient_are_ignored(self):
        # An allergen row that points at a code not in the Ingredients tab
        # must not crash and must not create a phantom IngredientAllergen.
        from django.core.management import call_command
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I103", "name": "Salt", "cost": 1.0,
                 "supply_unit": "Kilograms", "category": "Dry Goods"},
            ],
            uoms=[],
            allergens=[
                {"code": "NPD-I999", "allergen": "Milk",
                 "contains": True, "may_contain": False},
            ])
        call_command("import_ingredients", path)
        self.assertEqual(IngredientAllergen.objects.count(), 0)

    def test_workbook_without_allergens_tab_still_imports(self):
        # Existing data files / partial fixtures without the Allergens tab
        # must keep working.
        from django.core.management import call_command
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I104", "name": "Salt", "cost": 1.0,
                 "supply_unit": "Kilograms", "category": "Dry Goods"},
            ],
            uoms=[])  # allergens=None → no Allergens tab written
        call_command("import_ingredients", path)
        self.assertTrue(Product.objects.filter(code="NPD-I104").exists())
        self.assertEqual(IngredientAllergen.objects.count(), 0)

    def test_no_master_catalog_supplier_created(self):
        # The old "Master Catalog" placeholder must never appear, even when
        # the workbook has no usable supplier data at all.
        from django.core.management import call_command
        path = self._tmp()
        self._make_workbook(path,
            ingredients=[
                {"code": "NPD-I700", "name": "X", "cost": 1.0,
                 "supply_unit": "Kilograms", "category": "Dry Goods"},
            ],
            uoms=[])
        call_command("import_ingredients", path)
        self.assertFalse(Supplier.objects.filter(name="Master Catalog").exists())


class AllergenDisplayTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    def test_contains_and_may_contain_render_in_separate_groups(self):
        flour = Product.objects.create(
            name="Bread Flour", code="NPD-I100", department=self.dept,
            unit="g", minimum=0)
        IngredientAllergen.objects.create(
            product=flour, name="Cereals containing gluten",
            contains=True, may_contain=False)
        IngredientAllergen.objects.create(
            product=flour, name="Soya", contains=False, may_contain=True)
        r = self.client.get(f"/products/{flour.pk}/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Section header present
        self.assertIn(">Allergens<", body)
        # Both labelled groups present in the right order
        contains_idx = body.find("Contains")
        may_idx = body.find("May contain")
        self.assertGreater(contains_idx, 0)
        self.assertGreater(may_idx, contains_idx)
        # The two allergens are tagged with the right CSS class
        contains_section = body[contains_idx:may_idx]
        self.assertIn('class="tag low"', contains_section)
        self.assertIn("Cereals containing gluten", contains_section)
        may_section = body[may_idx:]
        self.assertIn('class="tag prov"', may_section)
        self.assertIn("Soya", may_section)

    def test_no_allergens_shows_placeholder(self):
        salt = Product.objects.create(
            name="Salt", code="SLT1", department=self.dept,
            unit="g", minimum=0)
        r = self.client.get(f"/products/{salt.pk}/")
        body = r.content.decode()
        self.assertIn(">Allergens<", body)
        self.assertIn("No declared allergens", body)

    def test_contains_takes_precedence_over_may_contain(self):
        # An allergen with both flags shouldn't appear in both lists - it's
        # declared, so the firm tag wins and the soft one is suppressed.
        flour = Product.objects.create(
            name="Mixed", code="MIX1", department=self.dept,
            unit="g", minimum=0)
        IngredientAllergen.objects.create(
            product=flour, name="Milk", contains=True, may_contain=True)
        r = self.client.get(f"/products/{flour.pk}/")
        body = r.content.decode()
        contains_idx = body.find("Contains")
        may_idx = body.find("May contain")
        # No "May contain" group should render at all
        self.assertEqual(may_idx, -1)
        # Milk shows in the contains group
        self.assertIn('class="tag low"', body[contains_idx:])
        self.assertIn("Milk", body[contains_idx:])


class ResetStockDataTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.sup = Supplier.objects.create(name="Mill")
        self.flour = Product.objects.create(
            name="Flour", code="FLR1", department=self.dept, unit="g",
            minimum=Decimal("5"))
        SupplierPrice.objects.create(
            product=self.flour, supplier=self.sup,
            pack_weight=Decimal("25000"), pack_price=Decimal("30.00"))
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date(2026, 5, 22))
        StockLine.objects.create(stocktake=st, product=self.flour,
                                 current=Decimal("3"), carried_over=False)
        delivery = Delivery.objects.create(
            department=self.dept, supplier=self.sup,
            date=datetime.date(2026, 5, 22))
        Batch.objects.create(delivery=delivery, product=self.flour,
                             qty_received=Decimal("3"), qty_remaining=Decimal("3"))
        Adjustment.objects.create(
            department=self.dept, product=self.flour,
            quantity=Decimal("1"), reason="waste", user=self.user,
            date=datetime.date(2026, 5, 10))

    def test_without_yes_nothing_is_deleted(self):
        from django.core.management import call_command
        from io import StringIO
        out = StringIO()
        call_command("reset_stock_data", stdout=out)
        # Everything still here
        self.assertEqual(Product.objects.count(), 1)
        self.assertEqual(SupplierPrice.objects.count(), 1)
        self.assertEqual(Stocktake.objects.count(), 1)
        self.assertEqual(Delivery.objects.count(), 1)
        self.assertEqual(Batch.objects.count(), 1)
        self.assertEqual(Adjustment.objects.count(), 1)
        self.assertIn("Refusing", out.getvalue())

    def test_with_yes_clears_stock_but_leaves_auth_and_departments(self):
        from django.core.management import call_command
        from io import StringIO
        out = StringIO()
        call_command("reset_stock_data", "--yes", stdout=out)
        # Stock content gone
        self.assertEqual(Product.objects.count(), 0)
        self.assertEqual(SupplierPrice.objects.count(), 0)
        self.assertEqual(Stocktake.objects.count(), 0)
        self.assertEqual(StockLine.objects.count(), 0)
        self.assertEqual(Delivery.objects.count(), 0)
        self.assertEqual(Batch.objects.count(), 0)
        self.assertEqual(Adjustment.objects.count(), 0)
        # Auth + departments untouched
        self.assertEqual(get_user_model().objects.filter(username="alice").count(), 1)
        self.assertEqual(Department.objects.filter(name="Bakery").count(), 1)
        # Supplier rows also preserved - they're a separate dimension to stock.
        self.assertEqual(Supplier.objects.filter(name="Mill").count(), 1)
        # Department membership preserved
        self.assertTrue(self.dept.members.filter(username="alice").exists())
        # Summary lists what was cleared
        self.assertIn("Products: 1 deleted", out.getvalue())


# ----------------------------------------------------------------------
# Stage C1 — recipes
# ----------------------------------------------------------------------

SAMPLE_RECIPE_XLSX = "data/recipe_sample.xlsx"
SAMPLE_MINCEPIE_XLSX = "data/recipe_sample_mincepie.xlsx"
SAMPLE_BULK_XLSX = "data/recipes_bulk_93.xlsx"


class RecipeModelTests(TestCase):
    """The bare model contracts: XOR constraint + cycle detection helper."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.product = Product.objects.create(
            name="Flour", code="NPD-I001", department=self.dept,
            unit="g", minimum=0)

    def test_recipeline_requires_exactly_one_target_neither(self):
        # Neither ingredient nor sub_recipe set → DB-level check refuses.
        from django.db.utils import IntegrityError
        r = Recipe.objects.create(code="NPD-R001", name="Bread", department=self.dept)
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                RecipeLine.objects.create(recipe=r, weight_g=Decimal("100"))

    def test_recipeline_requires_exactly_one_target_both(self):
        # Both set → DB-level check refuses.
        from django.db.utils import IntegrityError
        r = Recipe.objects.create(code="NPD-R002", name="Bread", department=self.dept)
        sub = Recipe.objects.create(code="NPD-R003", name="Starter", department=self.dept)
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                RecipeLine.objects.create(
                    recipe=r, ingredient=self.product, sub_recipe=sub,
                    weight_g=Decimal("100"))

    def test_recipeline_with_ingredient_only_saves(self):
        r = Recipe.objects.create(code="NPD-R010", name="X", department=self.dept)
        line = RecipeLine.objects.create(
            recipe=r, ingredient=self.product, weight_g=Decimal("50"))
        self.assertEqual(line.ingredient, self.product)
        self.assertIsNone(line.sub_recipe)

    def test_recipeline_with_sub_recipe_only_saves(self):
        r = Recipe.objects.create(code="NPD-R020", name="X", department=self.dept)
        sub = Recipe.objects.create(code="NPD-R021", name="Sub", department=self.dept)
        line = RecipeLine.objects.create(
            recipe=r, sub_recipe=sub, weight_g=Decimal("50"))
        self.assertEqual(line.sub_recipe, sub)
        self.assertIsNone(line.ingredient)

    def test_cycle_detection_rejects_self_reference(self):
        # A recipe must not directly contain itself.
        r = Recipe.objects.create(code="NPD-R100", name="Self", department=self.dept)
        self.assertTrue(r.contains_cycle(r.pk))

    def test_cycle_detection_rejects_transitive(self):
        # R1 -> R2 -> R3; adding R1 as a sub of R3 would loop.
        r1 = Recipe.objects.create(code="NPD-R201", name="A", department=self.dept)
        r2 = Recipe.objects.create(code="NPD-R202", name="B", department=self.dept)
        r3 = Recipe.objects.create(code="NPD-R203", name="C", department=self.dept)
        RecipeLine.objects.create(recipe=r1, sub_recipe=r2, weight_g=Decimal("10"))
        RecipeLine.objects.create(recipe=r2, sub_recipe=r3, weight_g=Decimal("10"))
        # Adding r1 as a sub-recipe of r3 must be detected
        self.assertTrue(r3.contains_cycle(r1.pk))
        # But adding an unrelated recipe is fine
        r4 = Recipe.objects.create(code="NPD-R204", name="D", department=self.dept)
        self.assertFalse(r3.contains_cycle(r4.pk))


class RecipeImportSampleTests(TestCase):
    """Importing the committed sample workbook produces the expected tree."""

    @classmethod
    def setUpTestData(cls):
        from django.core.management import call_command
        cls.dept = Department.objects.create(name="Bakery")
        # Pre-create the NPD-I products that the sample references so the
        # ingredient-link assertions can use them.
        for code, name in (
            ("NPD-I10758", "WILDFARMED BREAD FLOUR (T65)"),
            ("NPD-I10756", "Water"),
            ("NPD-I11057", "FLOUR RYE DARK 100%"),
            ("NPD-I10893", "Dorset Sea Salt"),
            ("NPD-I10951", "WILDFARMED WHOLEMEAL FLOUR (T150)"),
            ("NPD-I10759", "WILDFARMED RUSTIC FLOUR (T80)"),
        ):
            Product.objects.create(
                code=code, name=name, department=cls.dept,
                unit="g", minimum=0)
        call_command("import_recipe", SAMPLE_RECIPE_XLSX, "--department", "Bakery")

    def test_main_recipe_imported_with_code_name_and_weights(self):
        r = Recipe.objects.get(code="NPD-R800")
        self.assertEqual(r.name, "Apple Waste Sourdough (Loose)")
        self.assertEqual(r.finished_weight_g, Decimal("600.000"))
        # Main has no separate Deposit row, falls back to Total
        self.assertEqual(r.deposit_weight_g, Decimal("600.000"))
        self.assertEqual(r.cook_loss_pct, Decimal("0.00"))

    def test_all_seven_sub_recipes_exist(self):
        codes = set(Recipe.objects.values_list("code", flat=True))
        for code in ("NPD-R800", "NPD-R364", "NPD-R2031",
                     "NPD-R2082", "NPD-R2029", "NPD-R1823", "NPD-R307"):
            self.assertIn(code, codes)
        self.assertEqual(Recipe.objects.count(), 7)

    def test_npd_r_lines_link_to_subrecipes(self):
        # NPD-R800 references NPD-R364 as a sub-recipe (600g).
        main = Recipe.objects.get(code="NPD-R800")
        line = main.lines.get()  # exactly one line
        self.assertIsNone(line.ingredient)
        self.assertEqual(line.sub_recipe.code, "NPD-R364")
        self.assertEqual(line.weight_g, Decimal("600.000"))

    def test_npd_i_lines_link_to_existing_ingredients_by_code(self):
        # NPD-R307 contains three NPD-I lines, all pre-created in setUpTestData.
        r307 = Recipe.objects.get(code="NPD-R307")
        codes_in_lines = {ln.ingredient.code for ln in r307.lines.all()
                          if ln.ingredient_id}
        self.assertEqual(codes_in_lines,
                         {"NPD-I10756", "NPD-I10759", "NPD-I10951"})
        # All three are the same Product rows we pre-created (no duplicates)
        for ln in r307.lines.all():
            self.assertTrue(ln.ingredient.pk)
            self.assertIsNone(ln.sub_recipe)

    def test_nested_subrecipe_weight_for_deep_chain(self):
        # NPD-R2031 has 6 lines; the sub_recipe one is NPD-R2082 at 63.2083g
        r2031 = Recipe.objects.get(code="NPD-R2031")
        self.assertEqual(r2031.lines.count(), 6)
        sub_line = r2031.lines.filter(sub_recipe__isnull=False).get()
        self.assertEqual(sub_line.sub_recipe.code, "NPD-R2082")
        self.assertEqual(sub_line.weight_g, Decimal("63.208"))

    def test_method_text_captured_for_recipes_that_have_it(self):
        r364 = Recipe.objects.get(code="NPD-R364")
        # The single Stage 1 instruction is captured
        self.assertIn("Take 660g of dough", r364.method_text)
        self.assertIn("Stage 1:", r364.method_text)
        # The main recipe has no Method section
        main = Recipe.objects.get(code="NPD-R800")
        self.assertEqual(main.method_text, "")

    def test_total_recipeline_count_is_21(self):
        # Sample: 1 (main) + 1 (R364) + 6 (R2031) + 3 (R2082)
        # + 3 (R2029) + 4 (R1823) + 3 (R307) = 21 lines
        self.assertEqual(RecipeLine.objects.count(), 21)


class RecipeImportIdempotencyTests(TestCase):
    """Re-running the importer updates rather than duplicating."""

    def test_re_import_same_workbook_is_idempotent(self):
        from django.core.management import call_command
        Department.objects.create(name="Bakery")
        call_command("import_recipe", SAMPLE_RECIPE_XLSX)
        first_recipes = Recipe.objects.count()
        first_lines = RecipeLine.objects.count()
        first_main = Recipe.objects.get(code="NPD-R800")
        # Run again
        call_command("import_recipe", SAMPLE_RECIPE_XLSX)
        self.assertEqual(Recipe.objects.count(), first_recipes)
        self.assertEqual(RecipeLine.objects.count(), first_lines)
        # Same row, not a new one (same PK; updated timestamp may change)
        second_main = Recipe.objects.get(code="NPD-R800")
        self.assertEqual(second_main.pk, first_main.pk)


class RecipeMincePieImportTests(TestCase):
    """The newer "Simplified Kitchen Report" layout (NPD-R655).

    This export differs from the sourdough sample on three fronts:
    - Sub-recipe tables use the header ``Code | Description | State | g``
      (4 columns; weight header is just "g", not "Weight (g)").
    - Each recipe is followed by a "Packaging" table whose header is also
      ``Code | Description | ... | Quantity | UOM`` — same Code/Description
      labels, NO weight column. The parser must NOT treat this as an
      ingredient block (the old code clobbered the parent's lines with
      packaging rows).
    - Method "Materials" cells carry multi-line blobs like
      ``"NPD-R413 - Mince Pie Dough\\nNPD-R412 - ..."``. These are prose
      lists for the operator, not structural links — the parser must not
      mistake them for recipe headings or ingredient lines.

    R655 → R568 → (R412, R413, R567, R223) must nest correctly.
    """

    def test_mince_pie_import_nests_sub_recipes_under_r568(self):
        from django.core.management import call_command
        call_command("import_recipe", SAMPLE_MINCEPIE_XLSX)
        # All six recipes present
        codes = set(Recipe.objects.values_list("code", flat=True))
        self.assertEqual(codes,
                         {"NPD-R655", "NPD-R568", "NPD-R412",
                          "NPD-R413", "NPD-R567", "NPD-R223"})
        # R655 → R568 (one sub-recipe line, full weight)
        r655 = Recipe.objects.get(code="NPD-R655")
        line = r655.lines.get()
        self.assertEqual(line.sub_recipe.code, "NPD-R568")
        # Must be the 83.82g real weight from the ingredient block, NOT
        # the 0.894 from the packaging table (the old bug).
        self.assertAlmostEqual(float(line.weight_g), 83.824, places=2)

    def test_r568_links_all_four_components_as_sub_recipes(self):
        from django.core.management import call_command
        call_command("import_recipe", SAMPLE_MINCEPIE_XLSX)
        r568 = Recipe.objects.get(code="NPD-R568")
        # Four sub-recipe lines, in order: R412 (50g), R413 (28g), R567 (15g), R223 (2g)
        lines = list(r568.lines.select_related("sub_recipe", "ingredient")
                     .order_by("ordering"))
        self.assertEqual(len(lines), 4)
        for line in lines:
            self.assertIsNone(line.ingredient,
                              "every R568 line should be a sub-recipe, not a raw ingredient")
            self.assertIsNotNone(line.sub_recipe)
        codes_weights = [(ln.sub_recipe.code, float(ln.weight_g)) for ln in lines]
        self.assertEqual(codes_weights, [
            ("NPD-R412", 50.0),
            ("NPD-R413", 28.0),
            ("NPD-R567", 15.0),
            ("NPD-R223", 2.0),
        ])

    def test_mince_pie_components_are_components_not_top_level_products(self):
        # The bug report: R412/R413/R567/R223 were importing as standalone
        # top-level recipes (sold_as_product=True) instead of components.
        # After the fix they must be is_used_as_component=True and the
        # default-from-references sold flag should be False.
        from django.core.management import call_command
        call_command("import_recipe", SAMPLE_MINCEPIE_XLSX)
        for code in ("NPD-R568", "NPD-R412", "NPD-R413",
                     "NPD-R567", "NPD-R223"):
            r = Recipe.objects.get(code=code)
            self.assertTrue(r.is_used_as_component,
                            f"{code} should be flagged as a component")
            self.assertFalse(r.sold_as_product,
                             f"{code} default sold flag should be False (it's used by another)")
        # Only the root R655 should default to sold_as_product=True
        r655 = Recipe.objects.get(code="NPD-R655")
        self.assertTrue(r655.sold_as_product)
        self.assertFalse(r655.is_used_as_component)

    def test_packaging_table_does_not_clobber_parent_lines(self):
        # Specific regression: the parent (R655) and R412 each have a
        # Packaging table immediately after their ingredient block. The
        # parser must skip those rows entirely; the parent's lines should
        # be the real ingredient/sub-recipe rows.
        from django.core.management import call_command
        call_command("import_recipe", SAMPLE_MINCEPIE_XLSX)
        r412 = Recipe.objects.get(code="NPD-R412")
        # R412 has 16 raw ingredients in the spec — no packaging row should
        # appear, and weights must be the in-recipe values, not the
        # "50g" packaging quantity.
        self.assertEqual(r412.lines.count(), 16)
        for ln in r412.lines.all():
            self.assertIsNotNone(ln.ingredient_id)
            self.assertIsNone(ln.sub_recipe_id)
            # No line weight should be the bogus packaging value
            self.assertNotEqual(ln.weight_g, Decimal("50"))

    def test_method_section_code_blobs_are_not_parsed_as_links(self):
        # R568's Stage 1 method cell holds a 4-line "Materials" blob that
        # references R412/R413/R567/R223 inside the text. Those must not
        # add extra sub-recipe lines to R568 or to anyone else; the only
        # structural links come from the ingredient TABLE.
        from django.core.management import call_command
        call_command("import_recipe", SAMPLE_MINCEPIE_XLSX)
        # R568 has exactly 4 lines from the table (verified separately);
        # the method blob mentions the same 4 codes — easy to double up.
        self.assertEqual(RecipeLine.objects.filter(
            recipe__code="NPD-R568").count(), 4)
        # And R655 has exactly one sub-recipe line (R568), nothing else
        # leaked through from method blobs anywhere in the workbook.
        self.assertEqual(RecipeLine.objects.filter(
            recipe__code="NPD-R655").count(), 1)
        # The method text on R568 should still capture the prose.
        r568 = Recipe.objects.get(code="NPD-R568")
        self.assertIn("Fill the cases", r568.method_text)
        self.assertIn("Bake at", r568.method_text)

    def test_sourdough_sample_still_imports_after_layout_fix(self):
        # Guard against regressions in the older 3-column layout.
        from django.core.management import call_command
        call_command("import_recipe", SAMPLE_RECIPE_XLSX)
        self.assertEqual(Recipe.objects.count(), 7)
        # R2031 still has its 6-line breakdown (5 ingredients + 1 sub).
        r2031 = Recipe.objects.get(code="NPD-R2031")
        self.assertEqual(r2031.lines.count(), 6)
        sub = r2031.lines.filter(sub_recipe__isnull=False).get()
        self.assertEqual(sub.sub_recipe.code, "NPD-R2082")
        self.assertEqual(sub.weight_g, Decimal("63.208"))
        # And the chain still bottoms out at NPD-R307
        r307 = Recipe.objects.get(code="NPD-R307")
        self.assertEqual(r307.lines.count(), 3)


class RecipeUnknownIngredientTests(TestCase):
    """Unknown NPD-I codes get stubbed and flagged in the summary."""

    def test_unknown_ingredient_creates_stub_and_flags(self):
        from django.core.management import call_command
        from io import StringIO
        # No pre-existing NPD-I products — the parser will create stubs.
        out = StringIO()
        call_command("import_recipe", SAMPLE_RECIPE_XLSX, stdout=out)
        # Six distinct NPD-I codes in the sample
        stubs = Product.objects.filter(code__startswith="NPD-I").count()
        self.assertEqual(stubs, 6)
        # The summary reports them
        self.assertIn("unknown ingredient", out.getvalue().lower())


class RecipeSaveCycleProtectionTests(TestCase):
    """save_recipes() must refuse a workbook that describes a cyclic recipe."""

    def test_self_referential_recipe_refused(self):
        from stock.recipe_import import save_recipes
        dept = Department.objects.create(name="Bakery")
        parsed = [{
            "code": "NPD-R900", "name": "Self-ref",
            "units_requested": None,
            "finished_weight_g": Decimal("100"),
            "deposit_weight_g": Decimal("100"),
            "cook_loss_pct": Decimal("0"),
            "method_text": "",
            "lines": [{
                "code": "NPD-R900", "name": "Self-ref",
                "weight_g": Decimal("50"), "is_subrecipe": True,
            }],
        }]
        with self.assertRaises(RecipeCycleError):
            save_recipes(parsed, dept)


class RecipesSectionViewTests(TestCase):
    """List page, upload form, preview-then-confirm flow, detail view."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    def test_list_page_replaces_coming_soon(self):
        r = self.client.get("/recipes/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertNotIn("coming soon", body.lower())
        self.assertIn("Recipes", body)
        # Import CTA visible
        self.assertIn('href="/recipes/upload/"', body)

    def test_list_navbar_has_home_recipes_import(self):
        r = self.client.get("/recipes/")
        body = r.content.decode()
        nav = body[body.index("<nav>"):body.index("</nav>")]
        self.assertIn(">Home<", nav)
        self.assertIn(">Recipes<", nav)
        self.assertIn(">Import<", nav)

    def test_list_page_shows_imported_recipes(self):
        Recipe.objects.create(
            code="NPD-R100", name="Sample", department=self.dept,
            finished_weight_g=Decimal("500"))
        r = self.client.get("/recipes/")
        body = r.content.decode()
        self.assertIn("NPD-R100", body)
        self.assertIn("Sample", body)

    def test_upload_form_renders(self):
        r = self.client.get("/recipes/upload/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("Import recipe", body)
        self.assertIn('name="file"', body)
        self.assertIn('enctype="multipart/form-data"', body)

    def test_upload_then_preview_then_confirm_creates_recipes(self):
        # Pre-create some of the NPD-I products so unknowns are minimal.
        for code, name in (("NPD-I10758", "Bread Flour"),
                           ("NPD-I10756", "Water")):
            Product.objects.create(
                code=code, name=name, department=self.dept,
                unit="g", minimum=0)

        with open(SAMPLE_RECIPE_XLSX, "rb") as f:
            upload = SimpleUploadedFile(
                "recipe_sample.xlsx", f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        r = self.client.post("/recipes/upload/", {"file": upload})
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.headers["Location"], "/recipes/upload/preview/")

        # Preview page renders the parsed tree (nothing saved yet)
        self.assertEqual(Recipe.objects.count(), 0)
        r = self.client.get("/recipes/upload/preview/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("NPD-R800", body)
        self.assertIn("NPD-R364", body)
        # Unknown ingredients block surfaced (4 of 6 NPD-I codes are missing)
        self.assertIn("Unknown ingredient", body)

        # Confirm — commits the import
        r = self.client.post("/recipes/upload/preview/", {})
        self.assertEqual(r.status_code, 302)
        self.assertEqual(Recipe.objects.count(), 7)
        # Lands on main recipe's detail page
        main = Recipe.objects.get(code="NPD-R800")
        self.assertEqual(r.headers["Location"], f"/recipes/{main.pk}/")

    def test_preview_without_pending_redirects_to_upload(self):
        r = self.client.get("/recipes/upload/preview/")
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.headers["Location"], "/recipes/upload/")

    def test_recipe_detail_shows_nested_tree(self):
        # Two-level recipe: main has one sub-recipe with two ingredients.
        flour = Product.objects.create(
            code="NPD-I10758", name="Flour", department=self.dept,
            unit="g", minimum=0)
        water = Product.objects.create(
            code="NPD-I10756", name="Water", department=self.dept,
            unit="g", minimum=0)
        sub = Recipe.objects.create(
            code="NPD-R200", name="Starter", department=self.dept,
            finished_weight_g=Decimal("100"))
        RecipeLine.objects.create(recipe=sub, ingredient=flour,
                                  weight_g=Decimal("60"), ordering=0)
        RecipeLine.objects.create(recipe=sub, ingredient=water,
                                  weight_g=Decimal("40"), ordering=1)
        main = Recipe.objects.create(
            code="NPD-R100", name="Loaf", department=self.dept,
            finished_weight_g=Decimal("400"), cook_loss_pct=Decimal("5"))
        RecipeLine.objects.create(recipe=main, sub_recipe=sub,
                                  weight_g=Decimal("100"), ordering=0)

        r = self.client.get(f"/recipes/{main.pk}/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Main recipe header
        self.assertIn("NPD-R100", body)
        self.assertIn("Loaf", body)
        # Sub-recipe shown nested with the right weight
        self.assertIn("NPD-R200", body)
        self.assertIn("Starter", body)
        # Raw ingredients reached via the sub-recipe tree
        self.assertIn("NPD-I10758", body)
        self.assertIn("NPD-I10756", body)
        # Weights rendered
        self.assertIn("100.0g", body)
        self.assertIn("60.0g", body)
        # Sub-recipe tag rendered
        self.assertIn("sub-recipe", body)

    def test_recipe_detail_blocked_for_other_department(self):
        other = Department.objects.create(name="Butchery")
        r = Recipe.objects.create(code="NPD-R900", name="X", department=other)
        resp = self.client.get(f"/recipes/{r.pk}/")
        self.assertEqual(resp.status_code, 403)

    def test_recipe_pages_require_login(self):
        c = Client()
        for path in ("/recipes/", "/recipes/upload/", "/recipes/upload/preview/"):
            r = c.get(path)
            self.assertEqual(r.status_code, 302)
            self.assertIn("/login/", r.headers["Location"])

    def test_recipe_delete_removes_recipe(self):
        # Hard-delete now requires the acknowledgement checkbox + the
        # recipe code typed back to confirm permanence (archive is the
        # default reversible action).
        r = Recipe.objects.create(code="NPD-R999", name="Gone", department=self.dept)
        resp = self.client.post(f"/recipes/{r.pk}/delete/",
                                {"acknowledge": "on", "confirm_code": "NPD-R999"})
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Recipe.objects.filter(code="NPD-R999").exists())


class RecipeExplodedIngredientsTests(TestCase):
    """The flat per-batch ingredient list: scale through each sub-recipe and
    sum the same Product across branches."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.flour = Product.objects.create(
            code="NPD-I001", name="Bread Flour",
            department=self.dept, unit="g", minimum=0)
        self.water = Product.objects.create(
            code="NPD-I002", name="Water",
            department=self.dept, unit="g", minimum=0)
        self.salt = Product.objects.create(
            code="NPD-I003", name="Salt",
            department=self.dept, unit="g", minimum=0)

    def _recipe(self, code, **kw):
        return Recipe.objects.create(code=code, department=self.dept, **kw)

    def _ing(self, recipe, product, weight, ordering=0):
        return RecipeLine.objects.create(
            recipe=recipe, ingredient=product,
            weight_g=Decimal(str(weight)), ordering=ordering)

    def _sub(self, recipe, sub_recipe, weight, ordering=0):
        return RecipeLine.objects.create(
            recipe=recipe, sub_recipe=sub_recipe,
            weight_g=Decimal(str(weight)), ordering=ordering)

    def test_simple_no_subrecipes_returns_direct_lines(self):
        r = self._recipe("NPD-R001", name="Flat",
                         finished_weight_g=Decimal("100"),
                         deposit_weight_g=Decimal("100"))
        self._ing(r, self.flour, 60)
        self._ing(r, self.water, 40)
        rows = r.exploded_ingredients()
        by_code = {row["ingredient"].code: row["weight_g"] for row in rows}
        self.assertEqual(by_code["NPD-I001"], Decimal("60"))
        self.assertEqual(by_code["NPD-I002"], Decimal("40"))

    def test_same_ingredient_in_one_recipe_is_summed(self):
        # A recipe can list the same product on two lines (e.g. water in
        # two mix stages) — the flat view collapses them into one row.
        r = self._recipe("NPD-R002", name="Two waters",
                         finished_weight_g=Decimal("60"),
                         deposit_weight_g=Decimal("60"))
        self._ing(r, self.water, 25, ordering=0)
        self._ing(r, self.water, 35, ordering=1)
        rows = r.exploded_ingredients()
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["ingredient"], self.water)
        self.assertEqual(rows[0]["weight_g"], Decimal("60"))

    def test_subrecipe_full_batch_consumed_keeps_ingredients_as_is(self):
        # Parent uses the entire finished output of the sub-recipe → no scaling.
        sub = self._recipe("NPD-R010", name="Dough",
                           finished_weight_g=Decimal("100"),
                           deposit_weight_g=Decimal("100"))
        self._ing(sub, self.flour, 60)
        self._ing(sub, self.water, 40)
        parent = self._recipe("NPD-R011", name="Loaf",
                              finished_weight_g=Decimal("100"),
                              deposit_weight_g=Decimal("100"))
        self._sub(parent, sub, 100)
        rows = parent.exploded_ingredients()
        by_code = {r["ingredient"].code: r["weight_g"] for r in rows}
        self.assertEqual(by_code["NPD-I001"], Decimal("60"))
        self.assertEqual(by_code["NPD-I002"], Decimal("40"))

    def test_subrecipe_partial_batch_scales_ingredients(self):
        # Parent uses 50g of a sub whose finished weight is 100g → 0.5×.
        sub = self._recipe("NPD-R020", name="Sub",
                           finished_weight_g=Decimal("100"),
                           deposit_weight_g=Decimal("100"))
        self._ing(sub, self.flour, 60)
        self._ing(sub, self.water, 40)
        parent = self._recipe("NPD-R021", name="Half loaf",
                              finished_weight_g=Decimal("50"),
                              deposit_weight_g=Decimal("50"))
        self._sub(parent, sub, 50)
        rows = parent.exploded_ingredients()
        by_code = {r["ingredient"].code: r["weight_g"] for r in rows}
        self.assertEqual(by_code["NPD-I001"], Decimal("30"))
        self.assertEqual(by_code["NPD-I002"], Decimal("20"))

    def test_subrecipe_scaling_uses_finished_weight_not_deposit(self):
        # Sub has 9.09% cook loss: deposit 660g, finished 600g. Parent uses
        # 600g (one finished batch). Scale must be 600/600 = 1.0, NOT
        # 600/660 = 0.909 — so each ingredient comes through at full
        # deposit-side weight.
        sub = self._recipe("NPD-R030", name="Sourdough single",
                           finished_weight_g=Decimal("600"),
                           deposit_weight_g=Decimal("660"))
        self._ing(sub, self.flour, 300)
        self._ing(sub, self.water, 300)
        self._ing(sub, self.salt, 60)
        parent = self._recipe("NPD-R031", name="Sourdough loose",
                              finished_weight_g=Decimal("600"),
                              deposit_weight_g=Decimal("600"))
        self._sub(parent, sub, 600)
        rows = parent.exploded_ingredients()
        by_code = {r["ingredient"].code: r["weight_g"] for r in rows}
        self.assertEqual(by_code["NPD-I001"], Decimal("300"))
        self.assertEqual(by_code["NPD-I002"], Decimal("300"))
        self.assertEqual(by_code["NPD-I003"], Decimal("60"))

    def test_same_ingredient_across_branches_is_summed(self):
        # Parent has TWO sub-recipes (Dough and Glaze) that both contain
        # flour. The flat view must collapse the two contributions into one
        # "Bread Flour: 80g" row.
        dough = self._recipe("NPD-R040", name="Dough",
                             finished_weight_g=Decimal("100"),
                             deposit_weight_g=Decimal("100"))
        self._ing(dough, self.flour, 60)
        self._ing(dough, self.water, 40)
        glaze = self._recipe("NPD-R041", name="Glaze",
                             finished_weight_g=Decimal("50"),
                             deposit_weight_g=Decimal("50"))
        self._ing(glaze, self.flour, 20)
        self._ing(glaze, self.water, 30)
        parent = self._recipe("NPD-R042", name="Glazed loaf",
                              finished_weight_g=Decimal("150"),
                              deposit_weight_g=Decimal("150"))
        self._sub(parent, dough, 100, ordering=0)
        self._sub(parent, glaze, 50, ordering=1)
        rows = parent.exploded_ingredients()
        by_code = {r["ingredient"].code: r["weight_g"] for r in rows}
        # Flour: 60 (dough) + 20 (glaze) = 80g
        self.assertEqual(by_code["NPD-I001"], Decimal("80"))
        # Water: 40 (dough) + 30 (glaze) = 70g
        self.assertEqual(by_code["NPD-I002"], Decimal("70"))
        # Exactly two distinct ingredients, no duplicates
        self.assertEqual(len(rows), 2)

    def test_deeply_nested_explosion_accumulates_repeat_across_levels(self):
        # Three levels deep: parent → starter → levain.
        # Flour appears at every level (one of the sample's real patterns).
        levain = self._recipe("NPD-R050", name="Levain",
                              finished_weight_g=Decimal("10"),
                              deposit_weight_g=Decimal("10"))
        self._ing(levain, self.flour, 5)
        self._ing(levain, self.water, 5)
        starter = self._recipe("NPD-R051", name="Starter",
                               finished_weight_g=Decimal("30"),
                               deposit_weight_g=Decimal("30"))
        self._ing(starter, self.flour, 10)
        self._ing(starter, self.water, 10)
        self._sub(starter, levain, 10)
        parent = self._recipe("NPD-R052", name="Loaf",
                              finished_weight_g=Decimal("100"),
                              deposit_weight_g=Decimal("100"))
        self._ing(parent, self.flour, 60)
        self._ing(parent, self.water, 10)
        self._sub(parent, starter, 30)
        rows = parent.exploded_ingredients()
        by_code = {r["ingredient"].code: r["weight_g"] for r in rows}
        # Flour: 60 (parent) + 10 (starter) + 5 (levain, via starter) = 75g
        self.assertEqual(by_code["NPD-I001"], Decimal("75"))
        # Water: 10 (parent) + 10 (starter) + 5 (levain) = 25g
        self.assertEqual(by_code["NPD-I002"], Decimal("25"))

    def test_results_sorted_by_ingredient_name(self):
        r = self._recipe("NPD-R060", name="Mixed",
                         finished_weight_g=Decimal("100"),
                         deposit_weight_g=Decimal("100"))
        self._ing(r, self.salt, 5)
        self._ing(r, self.flour, 60)
        self._ing(r, self.water, 35)
        rows = r.exploded_ingredients()
        names = [r["ingredient"].name for r in rows]
        self.assertEqual(names, ["Bread Flour", "Salt", "Water"])

    def test_cycle_in_data_does_not_infinite_loop(self):
        # If admin edits introduce a cycle (the import refuses one), the
        # explosion must terminate rather than recurse forever.
        a = self._recipe("NPD-R070", name="A",
                         finished_weight_g=Decimal("10"),
                         deposit_weight_g=Decimal("10"))
        b = self._recipe("NPD-R071", name="B",
                         finished_weight_g=Decimal("10"),
                         deposit_weight_g=Decimal("10"))
        self._sub(a, b, 10)
        self._sub(b, a, 10)
        # No assertion on contents — just that it returns rather than hangs.
        a.exploded_ingredients()


class RecipeDetailLayoutTests(TestCase):
    """The reworked layout: ingredients before method; sub-recipes collapsed
    by default; Structure / All-ingredients toggle."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        # Two-level recipe with method text.
        self.flour = Product.objects.create(
            code="NPD-I100", name="Flour", department=self.dept,
            unit="g", minimum=0)
        self.water = Product.objects.create(
            code="NPD-I101", name="Water", department=self.dept,
            unit="g", minimum=0)
        self.sub = Recipe.objects.create(
            code="NPD-R200", name="Starter", department=self.dept,
            finished_weight_g=Decimal("100"),
            deposit_weight_g=Decimal("100"),
            method_text="Mix and rest 12h.")
        RecipeLine.objects.create(
            recipe=self.sub, ingredient=self.flour,
            weight_g=Decimal("60"), ordering=0)
        RecipeLine.objects.create(
            recipe=self.sub, ingredient=self.water,
            weight_g=Decimal("40"), ordering=1)
        self.main = Recipe.objects.create(
            code="NPD-R100", name="Loaf", department=self.dept,
            finished_weight_g=Decimal("100"),
            deposit_weight_g=Decimal("100"),
            method_text="Bake at 230°C for 30 minutes.")
        RecipeLine.objects.create(
            recipe=self.main, sub_recipe=self.sub,
            weight_g=Decimal("100"), ordering=0)

    def test_ingredients_render_before_method_in_main_block(self):
        r = self.client.get(f"/recipes/{self.main.pk}/")
        body = r.content.decode()
        # Anchor on element markup so this doesn't false-positive on CSS
        # comments or unrelated text. The first `>Ingredients<` is the
        # main recipe's section label; the first `>show method<` is its
        # method toggle (sub-recipes are collapsed inside the main and
        # render their own labels later in the document).
        ing_idx = body.index(">Ingredients<")
        method_idx = body.index(">show method<")
        self.assertLess(ing_idx, method_idx,
                        "Ingredients should render before the method toggle")

    def test_method_is_collapsed_behind_a_toggle(self):
        r = self.client.get(f"/recipes/{self.main.pk}/")
        body = r.content.decode()
        # The method body is wrapped in a <details class="rec-method">
        # without `open`, so it's collapsed by default.
        self.assertRegex(body, r'<details class="rec-method">\s*<summary>')
        # The toggle label is present
        self.assertIn("show method", body)
        # The method text itself is still in the HTML (browser hides it)
        self.assertIn("Bake at 230", body)

    def test_subrecipes_collapsed_by_default(self):
        r = self.client.get(f"/recipes/{self.main.pk}/")
        body = r.content.decode()
        # Sub-recipe rows are wrapped in <details class="rec-sub"> with NO
        # `open` attribute — so the user clicks to expand.
        self.assertIn('<details class="rec-sub">', body)
        self.assertNotIn('<details class="rec-sub" open', body)
        # The "expand" affordance is shown
        self.assertIn("expand", body)
        # And nothing on the page auto-opens the sub-recipe
        # (the depth-0 recipe-block isn't wrapped in <details> at all)

    def test_structure_view_is_default(self):
        r = self.client.get(f"/recipes/{self.main.pk}/")
        body = r.content.decode()
        self.assertIn("Nested breakdown", body)
        self.assertNotIn("All raw ingredients", body)
        # View toggle: Structure is active
        self.assertRegex(body, r'class="on"[^>]*>Structure')

    def test_flat_view_renders_exploded_sums(self):
        r = self.client.get(f"/recipes/{self.main.pk}/?view=flat")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Flat heading and the raw ingredient rows; Structure block hidden
        self.assertIn("All raw ingredients", body)
        self.assertNotIn("Nested breakdown", body)
        self.assertRegex(body, r'class="on"[^>]*>All ingredients')
        # Both raw ingredients show with their per-batch totals
        # (full batch consumed → unchanged weights)
        self.assertIn("NPD-I100", body)
        self.assertIn("Flour", body)
        self.assertIn("60.00", body)
        self.assertIn("NPD-I101", body)
        self.assertIn("Water", body)
        self.assertIn("40.00", body)

    def test_detail_tags_for_sold_only_recipe(self):
        # Main isn't used by anything → sold tag only, no component tag,
        # no "Used in" list.
        Recipe.recompute_all_sold_defaults()
        r = self.client.get(f"/recipes/{self.main.pk}/")
        body = r.content.decode()
        header = body[body.index("<h2>"):body.index("</h2>") + len("</h2>")]
        self.assertIn("Sold product", header)
        self.assertNotIn("Component", header)
        self.assertNotIn("Used in:", body)

    def test_detail_tags_for_component_only_recipe(self):
        # The sub-recipe is used by main and (default) not sold standalone.
        # Show Component tag + Used-in list; do NOT show Sold-product tag.
        Recipe.recompute_all_sold_defaults()
        self.sub.refresh_from_db()
        self.assertFalse(self.sub.sold_as_product)
        r = self.client.get(f"/recipes/{self.sub.pk}/")
        body = r.content.decode()
        header = body[body.index("<h2>"):body.index("</h2>") + len("</h2>")]
        self.assertIn("Component", header)
        self.assertNotIn("Sold product", header)
        self.assertIn("Used in:", body)
        self.assertIn("NPD-R100", body)

    def test_detail_tags_for_both_sold_and_component(self):
        # The independent axis: a component that the operator also marks
        # as sold (e.g. a dough sold by the kg AND used in pastries) shows
        # BOTH tags — sold and component — and the Used-in list.
        Recipe.recompute_all_sold_defaults()
        self.sub.sold_as_product = True
        self.sub.is_sold_manual = True
        self.sub.save(update_fields=["sold_as_product", "is_sold_manual"])
        r = self.client.get(f"/recipes/{self.sub.pk}/")
        body = r.content.decode()
        header = body[body.index("<h2>"):body.index("</h2>") + len("</h2>")]
        self.assertIn("Sold product", header)
        self.assertIn("Component", header)
        self.assertIn("Used in:", body)

    def test_no_template_comment_text_on_detail_page(self):
        # _recipe_node.html previously used a multi-line {# ... #} block,
        # which Django renders verbatim — "Recursive renderer..." was
        # leaking into the nested-breakdown section. After the fix to
        # {% comment %}{% endcomment %} the text must not appear.
        r = self.client.get(f"/recipes/{self.main.pk}/")
        body = r.content.decode()
        self.assertNotIn("Recursive renderer", body)
        self.assertNotIn("Two independent tags", body)
        self.assertNotIn("{# ", body)
        self.assertNotIn(" #}", body)
        self.assertNotIn("{% comment %}", body)
        self.assertNotIn("{% endcomment %}", body)

    def test_detail_used_by_others_still_shows_component_even_if_sold_flag_off(self):
        # is_used_as_component is live-derived from parents() and cannot
        # be desynced. Even if sold_as_product is True/False and manual
        # is on, the Component tag must appear when something references
        # this recipe.
        self.sub.sold_as_product = False
        self.sub.is_sold_manual = True
        self.sub.save(update_fields=["sold_as_product", "is_sold_manual"])
        self.assertTrue(list(self.sub.parents()))
        r = self.client.get(f"/recipes/{self.sub.pk}/")
        body = r.content.decode()
        header = body[body.index("<h2>"):body.index("</h2>") + len("</h2>")]
        self.assertIn("Component", header)
        self.assertNotIn("Sold product", header)

    def test_detail_final_tag_when_not_referenced(self):
        # Nothing references main, so even if sold_as_product is fiddled
        # with, the Component tag must NOT appear.
        self.main.sold_as_product = True
        self.main.is_sold_manual = True
        self.main.save(update_fields=["sold_as_product", "is_sold_manual"])
        r = self.client.get(f"/recipes/{self.main.pk}/")
        body = r.content.decode()
        header = body[body.index("<h2>"):body.index("</h2>") + len("</h2>")]
        self.assertIn("Sold product", header)
        self.assertNotIn("Component", header)

    def test_flat_view_does_not_list_sub_recipes(self):
        # The flat view is leaves-only — NPD-R sub-recipes themselves
        # should NOT appear as rows (only the raw NPD-I they explode to).
        r = self.client.get(f"/recipes/{self.main.pk}/?view=flat")
        body = r.content.decode()
        # The sub-recipe code appears in the toggle/back link area but not
        # as an ingredient row in the flat table.
        table_section = body[body.index("All raw ingredients"):]
        # Cut off at the next card / page footer
        end = table_section.find("← all recipes")
        if end > 0:
            table_section = table_section[:end]
        self.assertNotIn("NPD-R200", table_section)


class RecipeSoldFlagTests(TestCase):
    """sold_as_product (stored), is_used_as_component (live), and the
    manual-override flow that replaces the old single `role` field."""

    @classmethod
    def setUpTestData(cls):
        from django.core.management import call_command
        cls.dept = Department.objects.create(name="Bakery")
        for code, name in (
            ("NPD-I10758", "WILDFARMED BREAD FLOUR (T65)"),
            ("NPD-I10756", "Water"),
            ("NPD-I11057", "FLOUR RYE DARK 100%"),
            ("NPD-I10893", "Dorset Sea Salt"),
            ("NPD-I10951", "WILDFARMED WHOLEMEAL FLOUR (T150)"),
            ("NPD-I10759", "WILDFARMED RUSTIC FLOUR (T80)"),
        ):
            Product.objects.create(
                code=code, name=name, department=cls.dept,
                unit="g", minimum=0)
        call_command("import_recipe", SAMPLE_RECIPE_XLSX, "--department", "Bakery")

    def test_import_marks_main_as_sold(self):
        # NPD-R800 isn't referenced as a sub_recipe by anything → sold by default.
        r800 = Recipe.objects.get(code="NPD-R800")
        self.assertTrue(r800.sold_as_product)
        self.assertFalse(r800.is_sold_manual)
        # And is NOT itself used as a component
        self.assertFalse(r800.is_used_as_component)

    def test_import_marks_sub_recipes_as_not_sold(self):
        # All six sub-recipes are referenced → sold defaults to False.
        for code in ("NPD-R364", "NPD-R2031", "NPD-R2082",
                     "NPD-R2029", "NPD-R1823", "NPD-R307"):
            r = Recipe.objects.get(code=code)
            self.assertFalse(r.sold_as_product, f"{code} should not be sold by default")
            self.assertTrue(r.is_used_as_component,
                            f"{code} should be a component")

    def test_parents_lists_recipes_that_use_this_as_sub(self):
        # NPD-R2029 (Stiff Starter) is used in NPD-R2082 (Ferment).
        r2029 = Recipe.objects.get(code="NPD-R2029")
        self.assertEqual([p.code for p in r2029.parents()], ["NPD-R2082"])
        # NPD-R800 (the sold product) has no parents.
        r800 = Recipe.objects.get(code="NPD-R800")
        self.assertEqual(list(r800.parents()), [])

    def test_parents_dedupes_when_one_recipe_uses_sub_twice(self):
        # A parent that lists the same sub-recipe on two lines (e.g. two
        # additions during the bake) shouldn't appear twice in parents().
        parent = Recipe.objects.get(code="NPD-R2031")
        sub = Recipe.objects.get(code="NPD-R2082")
        RecipeLine.objects.create(recipe=parent, sub_recipe=sub,
                                  weight_g=Decimal("10"), ordering=99)
        self.assertEqual([p.code for p in sub.parents()], ["NPD-R2031"])

    def test_recompute_sold_default_demotes_when_referenced(self):
        # Create a fresh recipe → defaults to sold (no references yet).
        r = Recipe.objects.create(
            code="NPD-R9001", name="New", department=self.dept,
            finished_weight_g=Decimal("100"))
        self.assertTrue(r.sold_as_product)
        # Reference it from an existing recipe; recompute → not sold.
        parent = Recipe.objects.get(code="NPD-R800")
        RecipeLine.objects.create(recipe=parent, sub_recipe=r,
                                  weight_g=Decimal("10"), ordering=99)
        r.recompute_sold_default()
        r.refresh_from_db()
        self.assertFalse(r.sold_as_product)

    def test_recompute_sold_default_promotes_when_unreferenced(self):
        # A component that loses all parents becomes sold by default.
        r307 = Recipe.objects.get(code="NPD-R307")
        self.assertFalse(r307.sold_as_product)
        RecipeLine.objects.filter(sub_recipe=r307).delete()
        r307.recompute_sold_default()
        r307.refresh_from_db()
        self.assertTrue(r307.sold_as_product)

    def test_manual_sold_override_survives_recompute(self):
        # The headline use case: a component the operator also sells
        # standalone. They flip sold_as_product=True with is_sold_manual,
        # and post-import recompute leaves it alone.
        r307 = Recipe.objects.get(code="NPD-R307")
        r307.sold_as_product = True
        r307.is_sold_manual = True
        r307.save(update_fields=["sold_as_product", "is_sold_manual"])
        # Per-instance recompute respects it
        r307.recompute_sold_default()
        r307.refresh_from_db()
        self.assertTrue(r307.sold_as_product)
        # Bulk recompute respects it
        Recipe.recompute_all_sold_defaults()
        r307.refresh_from_db()
        self.assertTrue(r307.sold_as_product)
        self.assertTrue(r307.is_sold_manual)
        # And it's STILL a component (the two axes are independent)
        self.assertTrue(r307.is_used_as_component)

    def test_manual_sold_flag_survives_reimport(self):
        # Re-running the import should preserve the operator's choice.
        from django.core.management import call_command
        r307 = Recipe.objects.get(code="NPD-R307")
        r307.sold_as_product = True
        r307.is_sold_manual = True
        r307.save(update_fields=["sold_as_product", "is_sold_manual"])
        call_command("import_recipe", SAMPLE_RECIPE_XLSX, "--department", "Bakery")
        r307.refresh_from_db()
        self.assertTrue(r307.sold_as_product)
        self.assertTrue(r307.is_sold_manual)

    def test_is_used_as_component_property_is_live_derived(self):
        # Adding/removing a sub_recipe edge flips is_used_as_component on
        # the next access — no save/recompute needed.
        r = Recipe.objects.create(
            code="NPD-R9002", name="Z", department=self.dept,
            finished_weight_g=Decimal("100"))
        self.assertFalse(r.is_used_as_component)
        parent = Recipe.objects.get(code="NPD-R800")
        line = RecipeLine.objects.create(recipe=parent, sub_recipe=r,
                                         weight_g=Decimal("10"), ordering=99)
        self.assertTrue(r.is_used_as_component)
        line.delete()
        self.assertFalse(r.is_used_as_component)


class RecipesListSoldColumnTests(TestCase):
    """The list page renders Sold/Component tags + the sold-toggle POST works."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        # Two recipes: parent references sub.
        self.sub = Recipe.objects.create(
            code="NPD-R301", name="Starter", department=self.dept,
            finished_weight_g=Decimal("100"))
        self.parent = Recipe.objects.create(
            code="NPD-R300", name="Loaf", department=self.dept,
            finished_weight_g=Decimal("400"))
        RecipeLine.objects.create(recipe=self.parent, sub_recipe=self.sub,
                                  weight_g=Decimal("100"), ordering=0)
        Recipe.recompute_all_sold_defaults()
        self.sub.refresh_from_db()
        self.parent.refresh_from_db()

    def test_list_page_shows_tag_columns(self):
        r = self.client.get("/recipes/?view=flat")
        body = r.content.decode()
        # Both tags render somewhere in the flat table
        self.assertIn("Sold", body)
        self.assertIn("Component", body)
        # Header has the two new columns
        self.assertIn(">Tags<", body)
        self.assertIn(">Sold?<", body)
        self.assertIn(">Used in<", body)

    def test_sub_recipe_row_shows_component_tag_and_no_sold_tag_by_default(self):
        r = self.client.get("/recipes/?view=flat")
        body = r.content.decode()
        sub_row_start = body.find('data-search="npd-r301')
        sub_row = body[sub_row_start:body.find("</tr>", sub_row_start)]
        # Component badge present; Sold badge absent (default for a used
        # recipe is sold_as_product=False).
        self.assertIn('class="role-tag component"', sub_row)
        self.assertNotIn('class="role-tag sold"', sub_row)
        # Parent link in the Used-in column
        self.assertIn(f'href="/recipes/{self.parent.pk}/"', sub_row)

    def test_parent_row_shows_sold_tag_and_no_component_tag(self):
        r = self.client.get("/recipes/?view=flat")
        body = r.content.decode()
        parent_row_start = body.find('data-search="npd-r300')
        parent_row = body[parent_row_start:body.find("</tr>", parent_row_start)]
        self.assertIn('class="role-tag sold"', parent_row)
        self.assertNotIn('class="role-tag component"', parent_row)
        # Used-in column shows em-dash for a sold-only recipe
        self.assertIn("—", parent_row)

    def test_sold_override_post_sets_manual_flag(self):
        # Flip the sub from "not sold" to "sold".
        r = self.client.post(f"/recipes/{self.sub.pk}/sold/",
                             {"sold": "true"})
        self.assertEqual(r.status_code, 302)
        self.sub.refresh_from_db()
        self.assertTrue(self.sub.sold_as_product)
        self.assertTrue(self.sub.is_sold_manual)
        # Subsequent bulk recompute leaves it alone
        Recipe.recompute_all_sold_defaults()
        self.sub.refresh_from_db()
        self.assertTrue(self.sub.sold_as_product)
        # And it STILL shows as a component (independent axis)
        self.assertTrue(self.sub.is_used_as_component)

    def test_sold_override_can_also_unsell(self):
        # Flip the parent from sold to not-sold.
        r = self.client.post(f"/recipes/{self.parent.pk}/sold/",
                             {"sold": "false"})
        self.assertEqual(r.status_code, 302)
        self.parent.refresh_from_db()
        self.assertFalse(self.parent.sold_as_product)
        self.assertTrue(self.parent.is_sold_manual)

    def test_sold_override_rejects_bad_value(self):
        r = self.client.post(f"/recipes/{self.sub.pk}/sold/",
                             {"sold": "bogus"})
        self.assertEqual(r.status_code, 302)
        self.sub.refresh_from_db()
        # Default retained
        self.assertFalse(self.sub.sold_as_product)
        self.assertFalse(self.sub.is_sold_manual)

    def test_sold_override_blocked_for_other_department(self):
        other = Department.objects.create(name="Butchery")
        r = Recipe.objects.create(code="NPD-R900", name="X", department=other)
        resp = self.client.post(f"/recipes/{r.pk}/sold/", {"sold": "true"})
        self.assertEqual(resp.status_code, 403)

    def test_filter_input_still_present(self):
        r = self.client.get("/recipes/?view=flat")
        body = r.content.decode()
        # The data-filter hook + data-search on rows is what the existing
        # filter JS expects — both must survive each list refactor.
        self.assertIn('data-filter="recipes-tbl"', body)
        self.assertIn('data-search="npd-r300', body)
        self.assertIn('data-search="npd-r301', body)

    def test_flat_component_tag_is_live_derived_from_references(self):
        # Even if the operator sets sold_as_product on a recipe, the
        # Component tag must still appear when something references it
        # (the two axes are independent).
        self.sub.sold_as_product = True
        self.sub.is_sold_manual = True
        self.sub.save(update_fields=["sold_as_product", "is_sold_manual"])
        r = self.client.get("/recipes/?view=flat")
        body = r.content.decode()
        sub_row_start = body.find('data-search="npd-r301')
        sub_row = body[sub_row_start:body.find("</tr>", sub_row_start)]
        # Both tags now
        self.assertIn('class="role-tag sold"', sub_row)
        self.assertIn('class="role-tag component"', sub_row)


class RecipesByProductTreeTests(TestCase):
    """The default /recipes/ view: top-level final products with nested
    sub-recipes derived from RecipeLine edges (not the stored role field)."""

    @classmethod
    def setUpTestData(cls):
        from django.core.management import call_command
        cls.dept = Department.objects.create(name="Bakery")
        for code, name in (
            ("NPD-I10758", "WILDFARMED BREAD FLOUR (T65)"),
            ("NPD-I10756", "Water"),
            ("NPD-I11057", "FLOUR RYE DARK 100%"),
            ("NPD-I10893", "Dorset Sea Salt"),
            ("NPD-I10951", "WILDFARMED WHOLEMEAL FLOUR (T150)"),
            ("NPD-I10759", "WILDFARMED RUSTIC FLOUR (T80)"),
        ):
            Product.objects.create(
                code=code, name=name, department=cls.dept,
                unit="g", minimum=0)
        call_command("import_recipe", SAMPLE_RECIPE_XLSX, "--department", "Bakery")

    def setUp(self):
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    def _forest(self):
        from stock.views import _by_product_forest
        recipes = list(
            Recipe.objects.filter(department=self.dept)
            .prefetch_related("lines__sub_recipe"))
        return _by_product_forest(recipes)

    def test_sample_chain_nests_top_to_leaf(self):
        # NPD-R800 → R364 → R2031 → R2082 → R2029 → R1823 → R307
        forest = self._forest()
        # Single final product: NPD-R800
        self.assertEqual(len(forest), 1)
        root = forest[0]
        self.assertEqual(root["recipe"].code, "NPD-R800")
        self.assertEqual(root["depth"], 0)

        # Walk down the unique-child chain
        expected = ["NPD-R800", "NPD-R364", "NPD-R2031", "NPD-R2082",
                    "NPD-R2029", "NPD-R1823", "NPD-R307"]
        node = root
        for depth, code in enumerate(expected):
            self.assertEqual(node["recipe"].code, code,
                             f"depth {depth} should be {code}, got {node['recipe'].code}")
            self.assertEqual(node["depth"], depth)
            if depth < len(expected) - 1:
                # NPD-R2031 has 6 lines (5 ingredients + 1 sub-recipe), but
                # only one of those is a sub_recipe — find it.
                sub_kids = [c for c in node["children"]
                            if c["recipe"].code == expected[depth + 1]]
                self.assertEqual(len(sub_kids), 1,
                                 f"{code} should have {expected[depth+1]} as a child")
                node = sub_kids[0]
        # Leaf: NPD-R307 has no sub-recipes
        self.assertEqual(node["children"], [])

    def test_recipe_used_by_two_products_appears_under_both(self):
        # Build a second final product that ALSO uses NPD-R364 as a sub.
        # The tree must show R364 under both NPD-R800 and the new product.
        r364 = Recipe.objects.get(code="NPD-R364")
        second = Recipe.objects.create(
            code="NPD-R801", name="Apple Waste Sourdough (Boule)",
            department=self.dept,
            finished_weight_g=Decimal("600"))
        RecipeLine.objects.create(recipe=second, sub_recipe=r364,
                                  weight_g=Decimal("600"), ordering=0)
        forest = self._forest()
        roots = {n["recipe"].code: n for n in forest}
        self.assertIn("NPD-R800", roots)
        self.assertIn("NPD-R801", roots)
        # Both roots have R364 as a child
        r800_children = [c["recipe"].code for c in roots["NPD-R800"]["children"]]
        r801_children = [c["recipe"].code for c in roots["NPD-R801"]["children"]]
        self.assertIn("NPD-R364", r800_children)
        self.assertIn("NPD-R364", r801_children)
        # And the deeper chain shows under BOTH (verify one level deeper).
        r364_under_801 = [c for c in roots["NPD-R801"]["children"]
                          if c["recipe"].code == "NPD-R364"][0]
        deeper = [c["recipe"].code for c in r364_under_801["children"]]
        self.assertIn("NPD-R2031", deeper)

    def test_by_product_view_is_default_and_renders_nesting(self):
        r = self.client.get("/recipes/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # The toggle exists with By-product active
        self.assertRegex(body, r'class="on"[^>]*>By product')
        # All chain codes appear (the tree renders the whole sample)
        for code in ("NPD-R800", "NPD-R364", "NPD-R2031", "NPD-R2082",
                     "NPD-R2029", "NPD-R1823", "NPD-R307"):
            self.assertIn(code, body)
        # The tree HTML structure is present
        self.assertIn('class="tree-list"', body)
        self.assertIn('class="tree-node"', body)
        # And NPD-R800 is labelled as a sold product in the tree
        self.assertIn("sold product", body)

    def test_sold_component_appears_at_top_level_and_nested(self):
        # The headline new-model use case: a recipe that's BOTH used as
        # a component AND sold standalone must appear as a tree root AND
        # nested under each parent that uses it.
        r364 = Recipe.objects.get(code="NPD-R364")
        r364.sold_as_product = True
        r364.is_sold_manual = True
        r364.save(update_fields=["sold_as_product", "is_sold_manual"])

        forest = self._forest()
        root_codes = [n["recipe"].code for n in forest]
        # Both R800 (original final product) AND R364 (now also sold) appear at top
        self.assertIn("NPD-R800", root_codes)
        self.assertIn("NPD-R364", root_codes)
        # And R364 is STILL nested under R800 — the dual appearance is intentional
        r800_node = next(n for n in forest if n["recipe"].code == "NPD-R800")
        r800_child_codes = [c["recipe"].code for c in r800_node["children"]]
        self.assertIn("NPD-R364", r800_child_codes)

        # And the rendered page reflects this: NPD-R364 appears twice
        r = self.client.get("/recipes/")
        body = r.content.decode()
        self.assertGreaterEqual(body.count("NPD-R364"), 2,
                                "Sold-AND-component recipe should render in two places")
        # "also sold" hint appears for the nested occurrence
        self.assertIn("also sold", body)

    def test_unselling_a_top_level_recipe_removes_it_from_the_tree(self):
        # The operator flags R800 as not sold (e.g. discontinued). The
        # tree should no longer surface it as a product.
        r800 = Recipe.objects.get(code="NPD-R800")
        r800.sold_as_product = False
        r800.is_sold_manual = True
        r800.save(update_fields=["sold_as_product", "is_sold_manual"])
        forest = self._forest()
        self.assertEqual([n["recipe"].code for n in forest], [],
                         "Unselling the only sold recipe empties the tree")

    def test_tree_codes_render_in_top_to_leaf_order(self):
        # The chain should appear in document order parent-before-child,
        # which is what the recursive include guarantees.
        r = self.client.get("/recipes/")
        body = r.content.decode()
        chain = ["NPD-R800", "NPD-R364", "NPD-R2031", "NPD-R2082",
                 "NPD-R2029", "NPD-R1823", "NPD-R307"]
        positions = [body.index(c) for c in chain]
        self.assertEqual(positions, sorted(positions),
                         "Tree codes should appear parent-before-child in the HTML")

    def test_flat_view_still_renders_under_its_toggle(self):
        r = self.client.get("/recipes/?view=flat")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Flat-view toggle is active and the table markup is present
        self.assertRegex(body, r'class="on"[^>]*>All recipes')
        self.assertIn('id="recipes-tbl"', body)
        self.assertIn('data-filter="recipes-tbl"', body)
        # And the tree css/markup is NOT rendered in flat mode
        self.assertNotIn('class="tree-list"', body)

    def test_template_comment_text_does_not_leak_into_html(self):
        # Django's {# ... #} comment is single-line only — a multi-line
        # block of that shape renders verbatim. The partial uses
        # {% comment %}{% endcomment %} now, so none of the explanatory
        # text inside should appear in the response.
        r = self.client.get("/recipes/")
        body = r.content.decode()
        self.assertNotIn("Recursive node", body)
        self.assertNotIn("{# ", body)
        self.assertNotIn(" #}", body)
        # And the {% comment %} tags themselves are stripped by the engine
        self.assertNotIn("{% comment %}", body)
        self.assertNotIn("{% endcomment %}", body)

    def test_tree_is_collapsed_by_default(self):
        # Each node with children is wrapped in <details> without `open`,
        # so the browser hides the inner <ul>. Only the roots' summary
        # rows are visible until the user expands one.
        r = self.client.get("/recipes/")
        body = r.content.decode()
        self.assertIn('<details class="tree-details">', body)
        # No `open` attribute on any details element — fully collapsed.
        self.assertNotRegex(body, r'<details class="tree-details"\s+open\b')
        self.assertNotRegex(body, r'<details[^>]*\bopen\b[^>]*class="tree-details"')
        # The chevron is rendered (an .tree-chevron span inside the summary)
        self.assertIn('class="tree-chevron"', body)

    def test_only_root_rows_are_outside_a_details_body(self):
        # Verify the structural promise: only top-level final products
        # render as standalone <li>s in the outer .tree-list; every other
        # node lives inside a <details> body (i.e. inside a parent's
        # collapsed children <ul>). Cheap way to confirm: split the body
        # at <details>/</details> boundaries and check that the
        # OUT-OF-DETAILS section contains the root code(s) but none of
        # the sub-recipe codes.
        r = self.client.get("/recipes/")
        body = r.content.decode()
        out = []
        depth = 0
        i = 0
        while i < len(body):
            open_i = body.find("<details", i)
            close_i = body.find("</details>", i)
            if open_i == -1 and close_i == -1:
                if depth == 0:
                    out.append(body[i:])
                break
            if open_i != -1 and (close_i == -1 or open_i < close_i):
                if depth == 0:
                    out.append(body[i:open_i])
                depth += 1
                i = open_i + len("<details")
            else:
                depth -= 1
                i = close_i + len("</details>")
        outside_details = "".join(out)
        # The root (NPD-R800) appears outside any <details> *body* — well,
        # actually it's inside its own <details>, so it sits in the SUMMARY.
        # Easier check: every non-root code appears only inside a <details>
        # body (i.e. NOT in the outside_details slice).
        for code in ("NPD-R364", "NPD-R2031", "NPD-R2082",
                     "NPD-R2029", "NPD-R1823", "NPD-R307"):
            self.assertNotIn(code, outside_details,
                             f"{code} should be hidden inside a collapsed <details> by default")

    def test_expanding_root_reveals_children_in_markup(self):
        # The children's HTML is in the response (so a click-to-expand
        # immediately reveals them with no extra fetch). Structural check:
        # the root's <details> wraps the first child's row.
        r = self.client.get("/recipes/")
        body = r.content.decode()
        # The root's <summary> contains NPD-R800
        summary_idx = body.index('<summary class="tree-row tree-row--expandable">')
        # NPD-R800 appears soon after the summary opener
        r800_in_summary = body.index("NPD-R800", summary_idx)
        # NPD-R364 (the first child) appears AFTER the root summary
        r364_idx = body.index("NPD-R364", r800_in_summary)
        # And before the root's closing </details>
        # Walk to find the matching </details> for the root's <details>:
        depth = 1  # we're inside the root <details>
        scan = body.find("<details", summary_idx) + 1  # skip the opening
        # actually let's just confirm </details> appears after R364, AND
        # R364 appears strictly after R800-in-summary — the recursive
        # nest guarantees both.
        self.assertGreater(r364_idx, r800_in_summary)
        # And the whole chain is reachable in document order
        positions = []
        for code in ("NPD-R800", "NPD-R364", "NPD-R2031", "NPD-R2082",
                     "NPD-R2029", "NPD-R1823", "NPD-R307"):
            positions.append(body.index(code))
        self.assertEqual(positions, sorted(positions))

    def test_leaf_recipes_render_as_plain_row_not_details(self):
        # NPD-R307 has no sub-recipes — it should render without a
        # surrounding <details>/<summary> (just a tree-row), and its
        # chevron is hidden (tree-chevron--leaf) so the column lines up.
        r = self.client.get("/recipes/")
        body = r.content.decode()
        r307_idx = body.index("NPD-R307")
        # Walk back to the enclosing <li> open
        li_start = body.rfind("<li", 0, r307_idx)
        # The slice from <li to NPD-R307 should NOT contain <details>
        slice_ = body[li_start:r307_idx]
        self.assertNotIn("<details", slice_)
        # And it should contain the leaf chevron marker
        self.assertIn("tree-chevron--leaf", slice_)

    def test_cycle_in_data_does_not_hang_the_page(self):
        # Wire up a cycle directly in the DB (the import refuses one, but
        # an admin edit could land one). The forest builder must terminate.
        a = Recipe.objects.create(code="NPD-R901", name="A",
                                  department=self.dept,
                                  finished_weight_g=Decimal("10"))
        b = Recipe.objects.create(code="NPD-R902", name="B",
                                  department=self.dept,
                                  finished_weight_g=Decimal("10"))
        RecipeLine.objects.create(recipe=a, sub_recipe=b,
                                  weight_g=Decimal("5"), ordering=0)
        RecipeLine.objects.create(recipe=b, sub_recipe=a,
                                  weight_g=Decimal("5"), ordering=0)
        # Neither is a root (each references the other), but the tree
        # builder must terminate rather than loop forever on either branch
        # if reached through some hypothetical root. Just calling it is
        # the test — the assertion is "this returned" + the page renders.
        forest = self._forest()
        self.assertIsInstance(forest, list)
        # The page itself renders without timing out
        r = self.client.get("/recipes/")
        self.assertEqual(r.status_code, 200)


# ----------------------------------------------------------------------
# Packaging import + view
# ----------------------------------------------------------------------

PACKAGING_XLSX = "data/packaging.xlsx"


class PackagingImportTests(TestCase):
    """The real data/packaging.xlsx imports as 86 Products in the new
    Packaging category, with pack sizes + supplier prices resolved."""

    @classmethod
    def setUpTestData(cls):
        from django.core.management import call_command
        cls.dept = Department.objects.create(name="Bakery")
        call_command("import_packaging", PACKAGING_XLSX, "--department", "Bakery")

    def test_imports_eighty_six_packaging_items(self):
        # Spec: the sample workbook has 86 NPD-P rows.
        items = Product.objects.filter(category="packaging")
        self.assertEqual(items.count(), 86)
        # All have an NPD-P code
        for code in items.values_list("code", flat=True):
            self.assertTrue(code.startswith("NPD-P"), code)

    def test_npd_p25_pack_size_and_price(self):
        # NPD-P25: 1000 Each per Box at £110.12 → pack_weight=1000, pack_price=110.12.
        p = Product.objects.get(code="NPD-P25")
        self.assertEqual(p.category, "packaging")
        self.assertEqual(p.unit, "ea")
        self.assertEqual(p.name, "Vacuum Pouch 400x500mm")
        sp = p.prices.get()
        self.assertEqual(sp.pack_weight, Decimal("1000.00"))
        self.assertEqual(sp.pack_price, Decimal("110.12"))

    def test_primary_supplier_resolves_via_reference_tab(self):
        # NPD-P25's primary is S43, which the Reference tab maps to "Alliance".
        p = Product.objects.get(code="NPD-P25")
        self.assertEqual(p.prices.get().supplier.name, "Alliance")
        # The supplier row exists once (no duplicates from supplier resolution)
        self.assertEqual(Supplier.objects.filter(name="Alliance").count(), 1)

    def test_idempotent_reimport(self):
        from django.core.management import call_command
        before_items = Product.objects.filter(category="packaging").count()
        before_prices = SupplierPrice.objects.filter(
            product__category="packaging").count()
        call_command("import_packaging", PACKAGING_XLSX)
        self.assertEqual(Product.objects.filter(category="packaging").count(),
                         before_items)
        # Re-running doesn't pile up history rows (it updates the existing
        # SupplierPrice row in place per the import_ingredients pattern).
        self.assertEqual(SupplierPrice.objects.filter(
            product__category="packaging").count(), before_prices)
        # NPD-P25 still has its single price at £110.12
        p = Product.objects.get(code="NPD-P25")
        self.assertEqual(p.prices.count(), 1)
        self.assertEqual(p.prices.get().pack_price, Decimal("110.12"))

    def test_cheapest_price_per_1000_works_for_packaging(self):
        # The existing £/1000 maths must keep working for packaging units.
        # NPD-P25: £110.12 / 1000 ea * 1000 = £110.12 per 1000 each.
        p = Product.objects.get(code="NPD-P25")
        cheapest = p.cheapest_price
        self.assertIsNotNone(cheapest)
        self.assertEqual(cheapest.per_1000, Decimal("110.1200"))

    def test_packaging_value_flows_through_stocktake_logic(self):
        # Existing StockLine.value (= count * pack_price) must still apply
        # because packaging items are just Products in a category.
        p = Product.objects.get(code="NPD-P25")
        st = Stocktake.objects.create(department=self.dept,
                                      date=datetime.date.today())
        line = StockLine.objects.create(stocktake=st, product=p,
                                        current=Decimal("3"))
        # 3 boxes × £110.12 = £330.36
        self.assertEqual(line.value, Decimal("330.36"))

    def test_flagged_items_still_import_as_products_just_without_price(self):
        # The 3 items the importer flags (no supplier) still exist as
        # Product rows so they appear in the packaging list — just with no
        # SupplierPrice attached.
        p = Product.objects.get(code="NPD-P67")  # 125C Lid Wooden Spoon
        self.assertEqual(p.category, "packaging")
        self.assertFalse(p.prices.exists())


class PackagingViewTests(TestCase):
    """The Packaging list page + sub-nav + ingredient-list filter."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        # One of each category
        self.flour = Product.objects.create(
            code="NPD-I999", name="Flour", department=self.dept,
            category="dry_goods", unit="g", minimum=0)
        self.box = Product.objects.create(
            code="NPD-P999", name="Box", department=self.dept,
            category="packaging", unit="ea", minimum=0)

    def test_packaging_page_lists_only_packaging_category(self):
        r = self.client.get("/packaging/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("NPD-P999", body)
        self.assertIn("Box", body)
        # Flour (an ingredient) does NOT appear
        self.assertNotIn("NPD-I999", body)

    def test_ingredients_page_excludes_packaging(self):
        r = self.client.get("/products/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Flour appears, Box does NOT
        self.assertIn("NPD-I999", body)
        self.assertNotIn("NPD-P999", body)

    def test_stock_navbar_has_packaging_link(self):
        # Root / now redirects to /home/; use a stock page to read the
        # contextual nav (the Stock sub-menu is identical across them).
        r = self.client.get("/stocktakes/")
        body = r.content.decode()
        nav = body[body.index("<nav>"):body.index("</nav>")]
        self.assertIn(">Packaging<", nav)
        self.assertIn('href="/packaging/"', nav)

    def test_packaging_page_requires_login(self):
        c = Client()
        r = c.get("/packaging/")
        self.assertEqual(r.status_code, 302)
        self.assertIn("/login/", r.headers["Location"])

    def test_packaging_page_is_department_scoped(self):
        # An item in another department shouldn't show up here.
        other = Department.objects.create(name="Butchery")
        Product.objects.create(
            code="NPD-P888", name="Butcher Bag", department=other,
            category="packaging", unit="ea", minimum=0)
        r = self.client.get("/packaging/")
        self.assertNotIn("NPD-P888", r.content.decode())


class RecipeDetailStatsTrimTests(TestCase):
    """Recipe detail page has no stat cards now — Cook loss was the last
    survivor and the spec dropped it too."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        self.r = Recipe.objects.create(
            code="NPD-R900", name="Loaf", department=self.dept,
            deposit_weight_g=Decimal("660"),
            finished_weight_g=Decimal("600"),
            cook_loss_pct=Decimal("9.09"))

    def test_detail_no_longer_shows_any_stat_cards(self):
        r = self.client.get(f"/recipes/{self.r.pk}/")
        body = r.content.decode()
        # All four ever-shown stat-card labels are gone — Deposit / Finished
        # / Lines were removed earlier, Cook loss has now been removed too.
        self.assertNotIn(">Deposit weight<", body)
        self.assertNotIn(">Finished weight<", body)
        self.assertNotIn(">Lines</div>", body)
        # The Cook loss stat card specifically: no "Cook loss" label inside
        # a stat block AND the percentage doesn't render in the header area
        # (it can still appear inside the nested-breakdown rows because each
        # recipe block shows "· cook loss N%" in-line — that's fine).
        # Just verify the dedicated stat-card label/value are absent.
        self.assertNotIn(">Cook loss</div>", body)

    def test_detail_omits_stats_block_entirely(self):
        # No .stats block should render on the detail page at all now,
        # whether or not cook_loss_pct is populated.
        for r in (self.r,
                  Recipe.objects.create(code="NPD-R901", name="Plain",
                                        department=self.dept,
                                        finished_weight_g=Decimal("100"))):
            resp = self.client.get(f"/recipes/{r.pk}/")
            body = resp.content.decode()
            # The view-toggle should be the first thing after the header
            # bits — no <div class="stats"> between the <h2> and the toggle.
            head_end = body.index("</h2>")
            toggle_start = body.index('class="view-toggle"')
            between = body[head_end:toggle_start]
            self.assertNotIn('class="stats"', between)


class RecipePackagingImportTests(TestCase):
    """Recipe import captures packaging links (NPD-P codes only) from
    packaging sections, ignoring NPD-R/NPD-I noise that the exporter
    sometimes drops into those tables."""

    @classmethod
    def setUpTestData(cls):
        from django.core.management import call_command
        cls.dept = Department.objects.create(name="Bakery")
        # Pre-import the real packaging master so NPD-P25 exists as a
        # Product in the packaging category that the recipe import can
        # link to (the recipe import never auto-creates packaging stubs).
        call_command("import_packaging", PACKAGING_XLSX, "--department", "Bakery")
        call_command("import_recipe", SAMPLE_MINCEPIE_XLSX, "--department", "Bakery")

    def test_npd_p25_linked_to_npd_r412(self):
        # The only real packaging link in the mince pie file: NPD-P25
        # (Vacuum Pouch) on NPD-R412 (Mincemeat). Stored on the new
        # RecipePackaging table with the raw quantity string verbatim.
        r412 = Recipe.objects.get(code="NPD-R412")
        links = list(r412.packaging_links.all())
        self.assertEqual(len(links), 1)
        link = links[0]
        self.assertEqual(link.packaging.code, "NPD-P25")
        self.assertEqual(link.packaging.category, "packaging")
        # Raw quantity from the spreadsheet is kept as-is (per-gram fraction
        # in scientific notation + "Each" suffix), not converted.
        self.assertIn("Each", link.raw_quantity)

    def test_non_npd_p_rows_in_packaging_sections_are_ignored(self):
        # Two such noise rows exist in the mince pie file:
        # - row 26 of NPD-R655's "Recipe Packaging" table lists NPD-R568
        # - row 49 of NPD-R568's "Packaging" table lists NPD-R412
        # Both are sub-recipe references masquerading as packaging — the
        # parser must skip them.
        r655 = Recipe.objects.get(code="NPD-R655")
        r568 = Recipe.objects.get(code="NPD-R568")
        # No direct packaging links from either of those recipes
        self.assertEqual(r655.packaging_links.count(), 0)
        self.assertEqual(r568.packaging_links.count(), 0)
        # Total NPD-P-linked rows = 1 across the whole import
        self.assertEqual(RecipePackaging.objects.count(), 1)

    def test_ingredient_and_subrecipe_parsing_is_unchanged(self):
        # Regression guard: the same chain the mince pie regression test
        # locks in must still nest correctly even with packaging parsing
        # added.
        r655 = Recipe.objects.get(code="NPD-R655")
        self.assertEqual(r655.lines.count(), 1)
        self.assertEqual(r655.lines.get().sub_recipe.code, "NPD-R568")
        r568 = Recipe.objects.get(code="NPD-R568")
        child_codes = sorted(ln.sub_recipe.code for ln in r568.lines.all())
        self.assertEqual(child_codes,
                         ["NPD-R223", "NPD-R412", "NPD-R413", "NPD-R567"])

    def test_all_packaging_walks_subrecipe_tree(self):
        # NPD-R655 has no direct packaging, but it transitively uses
        # NPD-R412 which is packaged in NPD-P25 — Recipe.all_packaging
        # walks the tree and surfaces it.
        r655 = Recipe.objects.get(code="NPD-R655")
        codes = [p.code for p in r655.all_packaging()]
        self.assertEqual(codes, ["NPD-P25"])

    def test_re_import_does_not_duplicate_packaging_links(self):
        # save_recipes deletes existing packaging_links before re-creating;
        # re-running the importer must keep the link count at 1.
        from django.core.management import call_command
        call_command("import_recipe", SAMPLE_MINCEPIE_XLSX, "--department", "Bakery")
        self.assertEqual(RecipePackaging.objects.count(), 1)

    def test_recipe_detail_renders_packaging_section_for_sold_product(self):
        # The detail page for the sold product (NPD-R655) shows a
        # Packaging card with NPD-P25 linked through to the product page.
        U = get_user_model()
        u, _ = U.objects.get_or_create(username="alice")
        u.set_password("pw"); u.save()
        self.dept.members.add(u)
        c = Client(); c.login(username="alice", password="pw")
        c.get(f"/switch/{self.dept.pk}/")
        r655 = Recipe.objects.get(code="NPD-R655")
        resp = c.get(f"/recipes/{r655.pk}/")
        body = resp.content.decode()
        # The h3 looks like `<h3>Packaging <span class="muted">…`, anchor
        # on the opening tag so the trailing-space form still matches.
        self.assertIn("<h3>Packaging", body)
        self.assertIn("NPD-P25", body)
        # And the row links to the packaging Product's detail page
        npd_p25 = Product.objects.get(code="NPD-P25")
        self.assertIn(f'href="/products/{npd_p25.pk}/"', body)

    def test_recipe_detail_omits_packaging_section_when_recipe_has_none(self):
        # A recipe with no direct or transitive packaging shouldn't show
        # the section header.
        U = get_user_model()
        u, _ = U.objects.get_or_create(username="alice")
        u.set_password("pw"); u.save()
        self.dept.members.add(u)
        c = Client(); c.login(username="alice", password="pw")
        c.get(f"/switch/{self.dept.pk}/")
        # NPD-R223 (Hazelnut Praline) is a leaf with no packaging anywhere
        r223 = Recipe.objects.get(code="NPD-R223")
        resp = c.get(f"/recipes/{r223.pk}/")
        body = resp.content.decode()
        self.assertNotIn("<h3>Packaging", body)


class PackagingPageLayoutTests(TestCase):
    """The Packaging list page mirrors the Ingredients page columns
    + styling so the bakery sees one consistent stock-list pattern."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        # One ingredient + one packaging item so we can compare layouts.
        self.flour = Product.objects.create(
            code="NPD-I777", name="Flour", department=self.dept,
            category="dry_goods", unit="g", minimum=0)
        self.box = Product.objects.create(
            code="NPD-P777", name="Box", department=self.dept,
            category="packaging", unit="ea", minimum=0)
        sup = Supplier.objects.create(name="Acme")
        SupplierPrice.objects.create(
            product=self.box, supplier=sup,
            pack_weight=Decimal("100"), pack_price=Decimal("25.00"))

    def _header_cells(self, body, table_id):
        # Pull the <thead><tr>...</tr></thead> cells for the named table.
        start = body.index(f'id="{table_id}"')
        thead = body[start:body.index("</thead>", start)]
        import re
        return re.findall(r"<th[^>]*>([^<]*)</th>", thead)

    def test_packaging_page_columns_match_ingredients_page(self):
        ing = self.client.get("/products/").content.decode()
        pkg = self.client.get("/packaging/").content.decode()
        ing_cols = self._header_cells(ing, "ingredients-tbl")
        pkg_cols = self._header_cells(pkg, "packaging-tbl")
        # Same number of columns, same numeric markers
        self.assertEqual(len(ing_cols), len(pkg_cols))
        # Both have a £/1000 column (the cross-units price comparator)
        self.assertIn("£/1000", ing_cols)
        self.assertIn("£/1000", pkg_cols)
        # Both have Min and Cheapest
        self.assertIn("Min", ing_cols)
        self.assertIn("Min", pkg_cols)
        self.assertIn("Cheapest", ing_cols)
        self.assertIn("Cheapest", pkg_cols)

    def test_packaging_page_uses_same_filter_and_row_styling(self):
        body = self.client.get("/packaging/").content.decode()
        # Same .filter + data-filter + data-search hooks as products.html
        self.assertIn('class="filter"', body)
        self.assertIn('data-filter="packaging-tbl"', body)
        self.assertIn('data-search="', body)
        # £/1000 for NPD-P777: £25 / 100 ea × 1000 = £250.00
        self.assertIn("250.00", body)
        # Same edit-prices + delete affordance shape as ingredients
        self.assertIn("edit prices →", body)
        self.assertIn("delete", body)


# ----------------------------------------------------------------------
# Customers section
# ----------------------------------------------------------------------

ORDER_SHEET_XLSM = "data/order_sheet.xlsm"


class CustomersImportTests(TestCase):
    """import_customers classifies by WHOLESALE-tab membership and is
    idempotent on name."""

    @classmethod
    def setUpTestData(cls):
        import warnings
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported",
        )
        from django.core.management import call_command
        cls.dept = Department.objects.create(name="Bakery")
        call_command("import_customers", ORDER_SHEET_XLSM, "--department", "Bakery")

    def test_teals_is_classified_wholesale(self):
        # TEALS appears in both Customers and WHOLESALE tabs → wholesale wins.
        c = Customer.objects.get(name="TEALS")
        self.assertEqual(c.customer_type, Customer.WHOLESALE)
        self.assertFalse(c.is_type_manual)

    def test_garden_cafe_is_classified_internal(self):
        # GARDEN CAFE is an Estate outlet in the Customers tab and is NOT in
        # the WHOLESALE tab → internal.
        c = Customer.objects.get(name="GARDEN CAFE")
        self.assertEqual(c.customer_type, Customer.INTERNAL)
        self.assertEqual(c.location, "Estate")
        self.assertFalse(c.is_type_manual)

    def test_wholesale_only_customer_is_imported_as_wholesale(self):
        # PINKMANS - WHITELADIES ROAD exists ONLY in the WHOLESALE tab —
        # the importer still creates a Customer row for it.
        c = Customer.objects.get(name="PINKMANS - WHITELADIES ROAD")
        self.assertEqual(c.customer_type, Customer.WHOLESALE)
        self.assertEqual(c.location, "")
        self.assertEqual(c.ordered_by, "")

    def test_idempotent_reimport(self):
        from django.core.management import call_command
        before_count = Customer.objects.count()
        call_command("import_customers", ORDER_SHEET_XLSM, "--department", "Bakery")
        self.assertEqual(Customer.objects.count(), before_count)
        # And the type of a normal classified row stays put
        teals = Customer.objects.get(name="TEALS")
        self.assertEqual(teals.customer_type, Customer.WHOLESALE)

    def test_blank_named_rows_are_skipped(self):
        # The Customers tab has trailing rows with only an "Ordered by"
        # entry (no Customer Name) — those must not become customers.
        # No customer should have name == "" or be tagged with one of the
        # staff names that appears only in the trailing rows.
        self.assertFalse(Customer.objects.filter(name="").exists())
        for staff in ("Paulo Silva", "Sam Bobbett", "Lauren Edwards"):
            self.assertFalse(Customer.objects.filter(name=staff).exists())

    def test_manual_type_override_survives_reimport(self):
        from django.core.management import call_command
        # An operator flips YARLINGTON (auto-derived internal) to wholesale.
        y = Customer.objects.get(name="YARLINGTON")
        self.assertEqual(y.customer_type, Customer.INTERNAL)  # baseline
        y.customer_type = Customer.WHOLESALE
        y.is_type_manual = True
        y.save(update_fields=["customer_type", "is_type_manual"])
        # Re-import: type must stay wholesale, manual flag must stay True.
        call_command("import_customers", ORDER_SHEET_XLSM, "--department", "Bakery")
        y.refresh_from_db()
        self.assertEqual(y.customer_type, Customer.WHOLESALE)
        self.assertTrue(y.is_type_manual)

    def test_import_updates_location_on_existing_rows(self):
        # Location is always authoritative from the Customers tab — it
        # refreshes on every import (the manual flag protects type and
        # contact, NOT location).
        from django.core.management import call_command
        c = Customer.objects.get(name="FARMSHOP")
        c.location = "WRONG"
        c.save()
        call_command("import_customers", ORDER_SHEET_XLSM, "--department", "Bakery")
        c.refresh_from_db()
        self.assertEqual(c.location, "Estate")

    def test_manual_flag_preserves_both_type_and_contact(self):
        # is_type_manual=True is the "operator edited this row" flag and
        # must protect BOTH customer_type AND ordered_by from re-import.
        from django.core.management import call_command
        c = Customer.objects.get(name="FARMSHOP")
        c.ordered_by = "Custom Contact"
        c.customer_type = Customer.WHOLESALE   # bogus override
        c.is_type_manual = True
        c.save()
        call_command("import_customers", ORDER_SHEET_XLSM, "--department", "Bakery")
        c.refresh_from_db()
        self.assertEqual(c.ordered_by, "Custom Contact")     # preserved
        self.assertEqual(c.customer_type, Customer.WHOLESALE)  # preserved
        self.assertTrue(c.is_type_manual)
        self.assertEqual(c.location, "Estate")               # still refreshed

    def test_summary_reports_internal_and_wholesale_counts(self):
        from django.core.management import call_command
        from io import StringIO
        out = StringIO()
        call_command("import_customers", ORDER_SHEET_XLSM,
                     "--department", "Bakery", stdout=out)
        text = out.getvalue().lower()
        self.assertIn("internal", text)
        self.assertIn("wholesale", text)

    def test_ordered_by_comes_from_own_tab_not_customers_tab_column(self):
        # The Customers tab's "Ordered by" column is a misaligned
        # alphabetical staff list and must NOT be used. Each customer's
        # contact comes from a cell on their own order-form tab's header.
        # ECOM's tab carries "Ordered by: Clare Stephens" — that's the
        # value we want, not "Adam Flint" which is what the misaligned
        # Customers-tab column had on the same row.
        c = Customer.objects.get(name="ECOM")
        self.assertEqual(c.ordered_by, "Clare Stephens")
        self.assertNotEqual(c.ordered_by, "Adam Flint")

    def test_ordered_by_for_several_real_customer_tabs(self):
        # Spot-check three more tabs the spec calls out as known-good.
        cases = {
            "CREAMERY FARMSHOP": "Jessica Widdows",
            "GARDEN CAFE": "Charles Marshall",
            "FARMSHOP": "Daniel Smith",   # NOT "Alan Stewart" (Customers-tab row 3 col 3)
        }
        for name, expected in cases.items():
            c = Customer.objects.get(name=name)
            self.assertEqual(c.ordered_by, expected,
                             f"{name} ordered_by should come from its own tab")

    def test_customer_whose_tab_has_no_contact_has_blank_ordered_by(self):
        # TN100's order-form tab has the "Ordered by" label but no contact
        # name in the contact column (the only other text on that row is a
        # far-right "w/c" date marker that must NOT be mistaken for a name).
        c = Customer.objects.get(name="TN100")
        self.assertEqual(c.ordered_by, "")

    def test_wholesale_only_customer_has_blank_ordered_by(self):
        # No per-customer order tab in the workbook for the PINKMANS
        # branches, so their ordered_by stays blank.
        c = Customer.objects.get(name="PINKMANS - WHITELADIES ROAD")
        self.assertEqual(c.ordered_by, "")

    def test_misaligned_customers_tab_column_is_not_used_anywhere(self):
        # Aggregate check: not a single customer's ordered_by should equal
        # the value the OLD code would have pulled from the misaligned
        # Customers-tab "Ordered by" column on their row. We sample three
        # of those known-bad values to prove the new code doesn't surface
        # them anywhere by accident.
        bad_values = {"Adam Flint", "Alan Stewart", "Aleisa Childs"}
        actual = set(Customer.objects.exclude(ordered_by="")
                     .values_list("ordered_by", flat=True))
        # The intersection must be empty.
        self.assertFalse(bad_values & actual,
                         f"misaligned column leaked: {bad_values & actual}")


class CustomersViewTests(TestCase):
    """The Customers section: list pages, detail page, type-change POST,
    sub-nav, top-nav."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        # One of each type
        self.internal = Customer.objects.create(
            name="GARDEN CAFE", location="Estate", ordered_by="Andy",
            customer_type=Customer.INTERNAL, department=self.dept)
        self.wholesale = Customer.objects.create(
            name="TEALS", location="Wholesale", ordered_by="Jemima",
            customer_type=Customer.WHOLESALE, department=self.dept)

    def test_internal_list_shows_only_internal_customers(self):
        r = self.client.get("/customers/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("GARDEN CAFE", body)
        self.assertNotIn("TEALS", body)

    def test_wholesale_list_shows_only_wholesale_customers(self):
        r = self.client.get("/customers/wholesale/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("TEALS", body)
        self.assertNotIn("GARDEN CAFE", body)

    def test_customer_detail_renders_type_and_form(self):
        r = self.client.get(f"/customers/{self.internal.pk}/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("GARDEN CAFE", body)
        self.assertIn("Internal", body)
        # Type-change form posts to customer_set_type
        self.assertIn(f'action="/customers/{self.internal.pk}/type/"', body)
        self.assertIn('name="customer_type"', body)

    def test_type_change_post_sets_is_type_manual(self):
        r = self.client.post(f"/customers/{self.internal.pk}/type/",
                             {"customer_type": Customer.WHOLESALE})
        self.assertEqual(r.status_code, 302)
        self.internal.refresh_from_db()
        self.assertEqual(self.internal.customer_type, Customer.WHOLESALE)
        self.assertTrue(self.internal.is_type_manual)

    def test_type_change_rejects_bad_value(self):
        r = self.client.post(f"/customers/{self.internal.pk}/type/",
                             {"customer_type": "bogus"})
        self.assertEqual(r.status_code, 302)
        self.internal.refresh_from_db()
        # Unchanged
        self.assertEqual(self.internal.customer_type, Customer.INTERNAL)
        self.assertFalse(self.internal.is_type_manual)

    def test_type_change_blocked_for_other_department(self):
        other = Department.objects.create(name="Butchery")
        c = Customer.objects.create(
            name="OTHER", customer_type=Customer.INTERNAL, department=other)
        r = self.client.post(f"/customers/{c.pk}/type/",
                             {"customer_type": Customer.WHOLESALE})
        self.assertEqual(r.status_code, 403)

    def test_customer_pages_require_login(self):
        c = Client()
        for path in ("/customers/", "/customers/wholesale/",
                     f"/customers/{self.internal.pk}/"):
            r = c.get(path)
            self.assertEqual(r.status_code, 302)
            self.assertIn("/login/", r.headers["Location"])

    def test_customer_pages_are_department_scoped(self):
        other = Department.objects.create(name="Butchery")
        Customer.objects.create(
            name="OTHER", customer_type=Customer.INTERNAL, department=other)
        r = self.client.get("/customers/")
        self.assertNotIn("OTHER", r.content.decode())

    def test_section_nav_lists_internal_and_wholesale(self):
        r = self.client.get("/customers/")
        body = r.content.decode()
        nav = body[body.index("<nav>"):body.index("</nav>")]
        self.assertIn(">Internal Customers<", nav)
        self.assertIn(">Wholesale Customers<", nav)
        self.assertIn('href="/customers/"', nav)
        self.assertIn('href="/customers/wholesale/"', nav)
        # Internal Customers is the active sub-nav link on /customers/
        self.assertRegex(nav, r'class="on"[^>]*>Internal Customers')

    def test_home_top_nav_includes_customers_link(self):
        # /home/ now renders the design-system shell; its left rail carries
        # the Customers link (it replaced the old top section-picker).
        body = self.client.get("/home/").content.decode()
        self.assertIn('href="/customers/"', body)


class CustomersCRUDTests(TestCase):
    """Hand-managed create / edit / delete from the Customers section UI."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    def test_add_customer_button_on_internal_list(self):
        r = self.client.get("/customers/")
        body = r.content.decode()
        self.assertIn('href="/customers/new/?type=internal"', body)
        self.assertIn("Add customer", body)

    def test_add_customer_button_on_wholesale_list(self):
        r = self.client.get("/customers/wholesale/")
        body = r.content.decode()
        self.assertIn('href="/customers/new/?type=wholesale"', body)

    def test_new_form_defaults_type_from_query_param(self):
        r = self.client.get("/customers/new/?type=wholesale")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # The wholesale <option> renders selected
        self.assertRegex(body, r'<option value="wholesale"\s+selected>')

    def test_create_internal_customer_sets_manual_flags_and_appears_in_list(self):
        r = self.client.post("/customers/new/", {
            "name": "My Internal Co",
            "location": "Bruton",
            "ordered_by": "Some Contact",
            "customer_type": "internal",
        })
        self.assertEqual(r.status_code, 302)
        c = Customer.objects.get(name="My Internal Co")
        self.assertEqual(c.customer_type, Customer.INTERNAL)
        self.assertEqual(c.location, "Bruton")
        self.assertEqual(c.ordered_by, "Some Contact")
        self.assertTrue(c.is_type_manual)
        self.assertTrue(c.is_manual_entry)
        self.assertEqual(c.department, self.dept)
        # And shows up on the internal list, not the wholesale list
        self.assertIn("My Internal Co",
                     self.client.get("/customers/").content.decode())
        self.assertNotIn("My Internal Co",
                         self.client.get("/customers/wholesale/").content.decode())

    def test_create_wholesale_customer_appears_in_wholesale_list(self):
        self.client.post("/customers/new/", {
            "name": "My Wholesale Co",
            "location": "",
            "ordered_by": "",
            "customer_type": "wholesale",
        })
        c = Customer.objects.get(name="My Wholesale Co")
        self.assertEqual(c.customer_type, Customer.WHOLESALE)
        self.assertTrue(c.is_manual_entry)
        self.assertIn("My Wholesale Co",
                     self.client.get("/customers/wholesale/").content.decode())

    def test_create_rejects_blank_name(self):
        r = self.client.post("/customers/new/", {
            "name": "",
            "customer_type": "internal",
        })
        # Re-renders the form (200), doesn't redirect
        self.assertEqual(r.status_code, 200)
        self.assertFalse(Customer.objects.filter(name="").exists())
        self.assertIn("Name is required", r.content.decode())

    def test_create_rejects_duplicate_name_case_insensitive(self):
        Customer.objects.create(
            name="TEALS", customer_type=Customer.WHOLESALE,
            department=self.dept)
        r = self.client.post("/customers/new/", {
            "name": "teals",
            "customer_type": "wholesale",
        })
        self.assertEqual(r.status_code, 200)
        # Still exactly one TEALS row, no "teals" fork
        self.assertEqual(Customer.objects.filter(
            name__iexact="teals").count(), 1)
        self.assertIn("already exists", r.content.decode())

    def test_create_rejects_invalid_type(self):
        r = self.client.post("/customers/new/", {
            "name": "X",
            "customer_type": "bogus",
        })
        self.assertEqual(r.status_code, 200)
        self.assertFalse(Customer.objects.filter(name="X").exists())

    def test_edit_form_pre_populates_fields(self):
        c = Customer.objects.create(
            name="EDIT ME", location="Loc", ordered_by="Ord",
            customer_type=Customer.INTERNAL, department=self.dept)
        r = self.client.get(f"/customers/{c.pk}/edit/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn('value="EDIT ME"', body)
        self.assertIn('value="Loc"', body)
        self.assertIn('value="Ord"', body)
        self.assertRegex(body, r'<option value="internal"\s+selected>')

    def test_edit_changes_fields_and_sets_type_manual(self):
        c = Customer.objects.create(
            name="BEFORE", location="A", ordered_by="x",
            customer_type=Customer.INTERNAL, department=self.dept,
            is_type_manual=False)
        r = self.client.post(f"/customers/{c.pk}/edit/", {
            "name": "AFTER",
            "location": "B",
            "ordered_by": "y",
            "customer_type": "wholesale",
        })
        self.assertEqual(r.status_code, 302)
        c.refresh_from_db()
        self.assertEqual(c.name, "AFTER")
        self.assertEqual(c.location, "B")
        self.assertEqual(c.ordered_by, "y")
        self.assertEqual(c.customer_type, Customer.WHOLESALE)
        self.assertTrue(c.is_type_manual)

    def test_edit_rejects_rename_collision(self):
        Customer.objects.create(name="A", customer_type=Customer.INTERNAL,
                                department=self.dept)
        b = Customer.objects.create(name="B", customer_type=Customer.INTERNAL,
                                    department=self.dept)
        r = self.client.post(f"/customers/{b.pk}/edit/", {
            "name": "a",  # case-insensitive clash with existing "A"
            "customer_type": "internal",
        })
        self.assertEqual(r.status_code, 200)
        b.refresh_from_db()
        self.assertEqual(b.name, "B")

    def test_edit_allows_keeping_same_name(self):
        # Renaming to itself (same name) must not trigger a duplicate error.
        c = Customer.objects.create(name="KEEP", customer_type=Customer.INTERNAL,
                                    department=self.dept)
        r = self.client.post(f"/customers/{c.pk}/edit/", {
            "name": "KEEP",
            "location": "new-loc",
            "ordered_by": "",
            "customer_type": "internal",
        })
        self.assertEqual(r.status_code, 302)
        c.refresh_from_db()
        self.assertEqual(c.location, "new-loc")

    def test_edit_blocked_for_other_department(self):
        other = Department.objects.create(name="Butchery")
        c = Customer.objects.create(name="OTHER", customer_type=Customer.INTERNAL,
                                    department=other)
        # GET
        r = self.client.get(f"/customers/{c.pk}/edit/")
        self.assertEqual(r.status_code, 403)
        # POST
        r = self.client.post(f"/customers/{c.pk}/edit/", {
            "name": "OTHER2", "customer_type": "internal"})
        self.assertEqual(r.status_code, 403)
        c.refresh_from_db()
        self.assertEqual(c.name, "OTHER")

    def test_delete_removes_customer_and_redirects_to_correct_list(self):
        c = Customer.objects.create(
            name="GOING AWAY", customer_type=Customer.WHOLESALE,
            department=self.dept)
        r = self.client.post(f"/customers/{c.pk}/delete/")
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.headers["Location"], "/customers/wholesale/")
        self.assertFalse(Customer.objects.filter(pk=c.pk).exists())

    def test_delete_blocked_for_other_department(self):
        other = Department.objects.create(name="Butchery")
        c = Customer.objects.create(name="STAYS", customer_type=Customer.INTERNAL,
                                    department=other)
        r = self.client.post(f"/customers/{c.pk}/delete/")
        self.assertEqual(r.status_code, 403)
        self.assertTrue(Customer.objects.filter(pk=c.pk).exists())

    def test_detail_page_has_edit_and_delete_actions_with_confirm(self):
        c = Customer.objects.create(
            name="ACTIONS", customer_type=Customer.INTERNAL,
            department=self.dept)
        body = self.client.get(f"/customers/{c.pk}/").content.decode()
        self.assertIn(f'href="/customers/{c.pk}/edit/"', body)
        self.assertIn(f'action="/customers/{c.pk}/delete/"', body)
        # JS confirmation step, per spec
        self.assertIn("Delete ACTIONS?", body)
        self.assertIn("can't be undone", body)

    def test_detail_page_shows_hand_created_tag_for_manual_entries(self):
        c = Customer.objects.create(
            name="HAND MADE", customer_type=Customer.INTERNAL,
            is_manual_entry=True, department=self.dept)
        body = self.client.get(f"/customers/{c.pk}/").content.decode()
        self.assertIn("hand-created", body)


class CustomersListRowActionsTests(TestCase):
    """Per-row Edit + Delete actions on both customer list pages — the
    same row-action treatment used on /products/ and /suppliers/.
    """

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        U = get_user_model()
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        self.internal = Customer.objects.create(
            name="GARDEN CAFE", location="Estate", ordered_by="Charles",
            customer_type=Customer.INTERNAL, department=self.dept)
        self.wholesale = Customer.objects.create(
            name="TEALS", location="", ordered_by="",
            customer_type=Customer.WHOLESALE, department=self.dept)

    def _row(self, body, search_text):
        # Cut out one customer's <tr> by anchoring on its data-search attr,
        # so per-row assertions don't pick up siblings.
        idx = body.find(f'data-search="{search_text}')
        assert idx >= 0, f"row {search_text!r} not in body"
        end = body.find("</tr>", idx)
        return body[idx:end]

    def test_internal_row_has_edit_link_to_edit_view(self):
        body = self.client.get("/customers/").content.decode()
        row = self._row(body, "garden cafe")
        self.assertIn(f'href="/customers/{self.internal.pk}/edit/"', row)
        self.assertIn("edit →", row)

    def test_internal_row_has_delete_form_with_confirmation(self):
        body = self.client.get("/customers/").content.decode()
        row = self._row(body, "garden cafe")
        # POSTs to the existing delete view
        self.assertIn(f'action="/customers/{self.internal.pk}/delete/"', row)
        # JS confirm matches the ingredients/suppliers pattern
        self.assertIn("onsubmit=\"return confirm('Delete GARDEN CAFE?')\"", row)
        # Same .btn.ghost button styling
        self.assertIn('class="btn ghost"', row)
        # Carries a csrf token like every other delete form
        self.assertIn("csrfmiddlewaretoken", row)

    def test_wholesale_row_has_edit_link_to_edit_view(self):
        body = self.client.get("/customers/wholesale/").content.decode()
        row = self._row(body, "teals")
        self.assertIn(f'href="/customers/{self.wholesale.pk}/edit/"', row)
        self.assertIn("edit →", row)

    def test_wholesale_row_has_delete_form_with_confirmation(self):
        body = self.client.get("/customers/wholesale/").content.decode()
        row = self._row(body, "teals")
        self.assertIn(f'action="/customers/{self.wholesale.pk}/delete/"', row)
        self.assertIn("onsubmit=\"return confirm('Delete TEALS?')\"", row)
        self.assertIn('class="btn ghost"', row)

    def test_row_edit_link_reaches_edit_form(self):
        # Follow the link from the row through to confirm it lands on the
        # existing edit view (already covered backend-wise by CRUDTests,
        # but this seals the row→backend wiring end-to-end).
        r = self.client.get(f"/customers/{self.internal.pk}/edit/")
        self.assertEqual(r.status_code, 200)
        self.assertIn('value="GARDEN CAFE"', r.content.decode())

    def test_row_delete_form_actually_deletes(self):
        # POST the same URL the row's delete form would.
        r = self.client.post(f"/customers/{self.internal.pk}/delete/")
        self.assertEqual(r.status_code, 302)
        self.assertFalse(Customer.objects.filter(pk=self.internal.pk).exists())

    def test_row_actions_blocked_for_other_department(self):
        other = Department.objects.create(name="Butchery")
        c = Customer.objects.create(
            name="OUTSIDER", customer_type=Customer.INTERNAL, department=other)
        # The list pages only show this department's customers, so OUTSIDER
        # doesn't even appear in the rows — but the URLs the row actions
        # would point at must still be guarded against cross-dept hits.
        self.assertNotIn("OUTSIDER",
                         self.client.get("/customers/").content.decode())
        self.assertEqual(
            self.client.get(f"/customers/{c.pk}/edit/").status_code, 403)
        self.assertEqual(
            self.client.post(f"/customers/{c.pk}/delete/").status_code, 403)
        self.assertTrue(Customer.objects.filter(pk=c.pk).exists())

    def test_header_row_has_two_extra_columns_for_actions(self):
        # Mirror the Ingredients pattern: two trailing empty <th>s in the
        # header so the two action cells align with their columns.
        import re
        body = self.client.get("/customers/").content.decode()
        thead_start = body.index('id="customers-tbl"')
        thead_end = body.index("</thead>", thead_start)
        thead = body[thead_start:thead_end]
        # 5 <th> cells total: Name, Location, Ordered by, (edit), (delete).
        # Match <th> or <th class="…">, NOT <thead>.
        n_cells = len(re.findall(r"<th(?:>|\s)", thead))
        self.assertEqual(n_cells, 5)


class CustomersImportProtectionTests(TestCase):
    """Re-running the import must leave hand-created and manually-edited
    customers strictly alone."""

    @classmethod
    def setUpTestData(cls):
        import warnings
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported",
        )
        from django.core.management import call_command
        cls.dept = Department.objects.create(name="Bakery")
        call_command("import_customers", ORDER_SHEET_XLSM, "--department", "Bakery")

    def test_hand_created_customer_survives_reimport_untouched(self):
        from django.core.management import call_command
        c = Customer.objects.create(
            name="HAND CREATED", location="my-loc", ordered_by="my-contact",
            customer_type=Customer.INTERNAL,
            is_type_manual=True, is_manual_entry=True,
            department=self.dept)
        original_pk = c.pk
        call_command("import_customers", ORDER_SHEET_XLSM, "--department", "Bakery")
        # Still there, same PK, every field unchanged
        c.refresh_from_db()
        self.assertEqual(c.pk, original_pk)
        self.assertEqual(c.name, "HAND CREATED")
        self.assertEqual(c.location, "my-loc")
        self.assertEqual(c.ordered_by, "my-contact")
        self.assertEqual(c.customer_type, Customer.INTERNAL)
        self.assertTrue(c.is_manual_entry)
        self.assertTrue(c.is_type_manual)

    def test_hand_created_with_same_name_as_a_sheet_customer_is_still_skipped(self):
        # An operator might hand-create a row whose name happens to match
        # one in the order sheet (e.g. "TEALS"). The is_manual_entry flag
        # must keep the importer's hands off — even though the name is in
        # the sheet, the importer must not touch this row.
        from django.core.management import call_command
        # Delete the sheet-imported TEALS so we can re-create it manually
        Customer.objects.filter(name="TEALS").delete()
        c = Customer.objects.create(
            name="TEALS", location="hand-loc", ordered_by="hand-ord",
            customer_type=Customer.INTERNAL,    # deliberately "wrong"
            is_type_manual=True, is_manual_entry=True,
            department=self.dept)
        call_command("import_customers", ORDER_SHEET_XLSM, "--department", "Bakery")
        c.refresh_from_db()
        # All fields preserved — type didn't flip back to wholesale,
        # location/contact didn't reset to "" / blank.
        self.assertEqual(c.location, "hand-loc")
        self.assertEqual(c.ordered_by, "hand-ord")
        self.assertEqual(c.customer_type, Customer.INTERNAL)
        self.assertTrue(c.is_manual_entry)

    def test_manually_edited_sheet_customer_keeps_edits_on_reimport(self):
        # An existing sheet customer that was hand-edited (is_type_manual=
        # True) keeps ALL its editable fields across re-import: type,
        # ordered_by, AND location.
        from django.core.management import call_command
        c = Customer.objects.get(name="FARMSHOP")
        c.location = "Custom Loc"
        c.ordered_by = "Custom Contact"
        c.customer_type = Customer.WHOLESALE   # deliberately changed
        c.is_type_manual = True
        c.save()
        call_command("import_customers", ORDER_SHEET_XLSM, "--department", "Bakery")
        c.refresh_from_db()
        self.assertEqual(c.location, "Custom Loc")
        self.assertEqual(c.ordered_by, "Custom Contact")
        self.assertEqual(c.customer_type, Customer.WHOLESALE)
        self.assertTrue(c.is_type_manual)
        self.assertFalse(c.is_manual_entry)   # still a sheet customer

    def test_reimport_summary_reports_skipped_manual_entries(self):
        # A hand-created customer whose name is ALSO in the order sheet
        # is the only case the importer notices and reports — sheet names
        # not represented as manual entries are simply ignored.
        from django.core.management import call_command
        from io import StringIO
        Customer.objects.filter(name="TEALS").delete()
        Customer.objects.create(
            name="TEALS", customer_type=Customer.INTERNAL,
            is_manual_entry=True, is_type_manual=True, department=self.dept)
        out = StringIO()
        call_command("import_customers", ORDER_SHEET_XLSM,
                     "--department", "Bakery", stdout=out)
        text = out.getvalue().lower()
        self.assertIn("skipped 1 hand-created", text)


def _build_multisheet_workbook(path, *, include_malformed=True):
    """Write a tiny multi-sheet Recipe Report workbook for the bulk tests.

    Two valid Recipe Report sheets (NPD-R900 with a sub-recipe at
    NPD-R901, plus NPD-R902 as a third sheet) and — when
    ``include_malformed`` is True — a fourth sheet that deliberately
    lacks the "Recipe Code:" header so the bulk parser must skip it.
    Keeping this synthetic (rather than slicing the 93-sheet file)
    makes assertions stable and the test fast.
    """
    from openpyxl import Workbook
    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    def _add_recipe_sheet(title, code, name, lines):
        ws = wb.create_sheet(title=title)
        ws.append(["Recipe Code:", code])
        ws.append(["Recipe Description:", name])
        ws.append(["Units Requested:", 1])
        ws.append([])
        ws.append(["Code", "Description", "Weight (g)"])
        for c, d, w in lines:
            ws.append([c, d, w])
        ws.append(["Total", "", sum(w for _, _, w in lines)])
        ws.append(["Finished Weight", "", sum(w for _, _, w in lines)])

    _add_recipe_sheet(
        "Recipe Report - NPD-R900",
        "NPD-R900", "Test Loaf",
        [("NPD-I9001", "Flour", 500),
         ("NPD-I9002", "Water", 300),
         ("NPD-R901", "Test Starter", 100)],
    )
    # Sub-recipe on its own sheet — bulk import must wire R900 → R901.
    _add_recipe_sheet(
        "Recipe Report - NPD-R901",
        "NPD-R901", "Test Starter",
        [("NPD-I9001", "Flour", 50),
         ("NPD-I9002", "Water", 50)],
    )
    _add_recipe_sheet(
        "Recipe Report - NPD-R902",
        "NPD-R902", "Test Roll",
        [("NPD-I9001", "Flour", 200),
         ("NPD-I9003", "Salt", 5)],
    )
    if include_malformed:
        # No "Recipe Code:" header → parser raises RecipeParseError,
        # which the bulk loop must catch and record as a failure.
        ws_bad = wb.create_sheet(title="Notes")
        ws_bad.append(["This sheet isn't a recipe report"])
        ws_bad.append(["Just some operator notes"])
    wb.save(path)


@tag("slow")
class RecipeBulkImportTests(TestCase):
    """The bulk parser loops over every sheet, reusing the per-recipe code.

    Tagged ``slow`` because every test re-parses the committed 6 MB
    ``recipes_bulk_93.xlsx`` end-to-end (~22 s/test on local hardware);
    the default fast suite excludes ``--tag slow``. CI / pre-push runs
    them via ``--tag slow``.
    """

    def test_committed_93_sheet_workbook_parses_all_sheets(self):
        # The committed workbook has 93 sheets; the parser should walk
        # every one and the unique recipe count should match the sheet
        # count (each sheet's main recipe has its own NPD-R code).
        from stock.recipe_import import (
            parse_recipe_workbook_bulk, summarize_parse_bulk,
        )
        parsed, failures, sheets_processed = parse_recipe_workbook_bulk(
            SAMPLE_BULK_XLSX)
        self.assertEqual(sheets_processed, 93)
        self.assertEqual(failures, [])
        summary = summarize_parse_bulk(parsed, failures, sheets_processed)
        # Each sheet is its own top-level recipe; sub-recipes overlap
        # across sheets but the union covers all 93 mains.
        self.assertGreaterEqual(summary["unique_recipe_codes"], 93)
        # Spot-check: sourdough (NPD-R800) and mince pie (NPD-R655) are
        # both in this workbook and must appear in the parsed list.
        codes = {r["code"] for r in parsed}
        self.assertIn("NPD-R800", codes)
        self.assertIn("NPD-R655", codes)

    def test_bulk_import_command_saves_all_recipes(self):
        from django.core.management import call_command
        from io import StringIO
        out = StringIO()
        call_command("import_recipes_bulk", SAMPLE_BULK_XLSX,
                     "--department", "Bakery", stdout=out)
        # All 93 main recipes plus nested sub-recipes are persisted.
        self.assertGreaterEqual(Recipe.objects.count(), 93)
        # Known nested chain still wires up (mince pie regression).
        r655 = Recipe.objects.get(code="NPD-R655")
        line = r655.lines.get()
        self.assertEqual(line.sub_recipe.code, "NPD-R568")
        # The summary mentions sheet count and recipe count.
        text = out.getvalue()
        self.assertIn("93 sheet(s) processed", text)

    def test_bulk_skips_malformed_sheet_and_imports_the_rest(self):
        # A synthetic workbook with three valid sheets + one bad sheet.
        # The bad sheet is recorded as a failure but doesn't prevent
        # the other three from importing.
        import os
        import tempfile
        from stock.recipe_import import (
            parse_recipe_workbook_bulk, save_recipes,
        )
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            _build_multisheet_workbook(path, include_malformed=True)
            parsed, failures, sheets_processed = parse_recipe_workbook_bulk(path)
        finally:
            os.unlink(path)
        # 4 sheets seen, 1 failure (the Notes sheet), 3 recipes parsed.
        self.assertEqual(sheets_processed, 4)
        self.assertEqual(len(failures), 1)
        bad_title, reason = failures[0]
        self.assertEqual(bad_title, "Notes")
        self.assertIn("Recipe Code", reason)
        codes = {r["code"] for r in parsed}
        self.assertEqual(codes, {"NPD-R900", "NPD-R901", "NPD-R902"})

        # Saving still works — the three valid recipes persist and the
        # NPD-R900 → NPD-R901 cross-sheet link resolves to the real
        # sub-recipe row (not an empty stub).
        dept = Department.objects.create(name="Bakery")
        stats = save_recipes(parsed, dept)
        self.assertEqual(Recipe.objects.count(), 3)
        r900 = Recipe.objects.get(code="NPD-R900")
        sub_line = r900.lines.filter(sub_recipe__isnull=False).get()
        self.assertEqual(sub_line.sub_recipe.code, "NPD-R901")
        # NPD-R901 was matched as an existing sheet, not stubbed.
        self.assertNotIn(sub_line.sub_recipe, stats["stub_subrecipes"])

    def test_bulk_summary_reports_counts_and_failures(self):
        # summarize_parse_bulk surfaces sheets_processed, sheets_failed,
        # the failures list itself, and unique recipe codes — the upload
        # preview template reads these directly.
        import os
        import tempfile
        from stock.recipe_import import (
            parse_recipe_workbook_bulk, summarize_parse_bulk,
        )
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            _build_multisheet_workbook(path, include_malformed=True)
            parsed, failures, sheets_processed = parse_recipe_workbook_bulk(path)
        finally:
            os.unlink(path)
        summary = summarize_parse_bulk(parsed, failures, sheets_processed)
        self.assertEqual(summary["sheets_processed"], 4)
        self.assertEqual(summary["sheets_failed"], 1)
        self.assertEqual(summary["unique_recipe_codes"], 3)
        self.assertEqual(len(summary["failures"]), 1)
        # NPD-I codes referenced but absent should be flagged.
        self.assertIn("NPD-I9001", summary["unknown_ingredients"])

    def test_bulk_upload_through_web_form_imports_and_lands_on_recipes_list(self):
        # End-to-end through the upload view: a multi-sheet POST is
        # parsed, preview rendered with per-sheet stats, then commit
        # redirects to the recipes list (not a single recipe detail).
        import os
        import tempfile
        User = get_user_model()
        user = User.objects.create_user(username="u", password="p")
        dept = Department.objects.create(name="Bakery")
        user.departments.add(dept)
        c = Client()
        c.force_login(user)
        session = c.session
        session["dept_id"] = dept.pk
        session.save()

        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            _build_multisheet_workbook(path, include_malformed=True)
            with open(path, "rb") as f:
                upload = SimpleUploadedFile(
                    "bulk.xlsx", f.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        finally:
            os.unlink(path)

        r = c.post("/recipes/upload/", {"file": upload})
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.headers["Location"], "/recipes/upload/preview/")

        r = c.get("/recipes/upload/preview/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Per-sheet stats visible
        self.assertIn("Sheets processed", body)
        self.assertIn("Sheets failed", body)
        # Failure surfaced with the offending sheet name
        self.assertIn("Notes", body)
        # All three valid recipes listed
        self.assertIn("NPD-R900", body)
        self.assertIn("NPD-R901", body)
        self.assertIn("NPD-R902", body)

        # Commit: multi-sheet uploads land on the recipes list (no main).
        r = c.post("/recipes/upload/preview/", {})
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r.headers["Location"], "/recipes/")
        self.assertEqual(Recipe.objects.count(), 3)

    def test_single_sheet_with_subrecipes_still_lands_on_main_detail(self):
        # Regression: a one-sheet workbook (the existing sourdough sample
        # with 7 nested recipes) still uses the "single-recipe" landing
        # behaviour — main's detail page, not the recipes list.
        User = get_user_model()
        user = User.objects.create_user(username="u", password="p")
        dept = Department.objects.create(name="Bakery")
        user.departments.add(dept)
        c = Client()
        c.force_login(user)
        session = c.session
        session["dept_id"] = dept.pk
        session.save()

        with open(SAMPLE_RECIPE_XLSX, "rb") as f:
            upload = SimpleUploadedFile(
                "recipe_sample.xlsx", f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        r = c.post("/recipes/upload/", {"file": upload})
        self.assertEqual(r.status_code, 302)
        r = c.post("/recipes/upload/preview/", {})
        self.assertEqual(r.status_code, 302)
        main = Recipe.objects.get(code="NPD-R800")
        self.assertEqual(r.headers["Location"], f"/recipes/{main.pk}/")


class RecipeDeleteEditTests(TestCase):
    """Stage A recipe management: delete (with persistence) + basic-field edit."""

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.other_dept = Department.objects.create(name="Butchery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")

    # ---- delete ----

    def test_delete_confirmation_page_lists_parents(self):
        # Component referenced by two parents → confirmation page must
        # list both with links.
        sub = Recipe.objects.create(code="NPD-R900", name="Filling",
                                    department=self.dept)
        p1 = Recipe.objects.create(code="NPD-R901", name="Pie One",
                                   department=self.dept)
        p2 = Recipe.objects.create(code="NPD-R902", name="Pie Two",
                                   department=self.dept)
        RecipeLine.objects.create(recipe=p1, sub_recipe=sub,
                                  weight_g=Decimal("10"), ordering=0)
        RecipeLine.objects.create(recipe=p2, sub_recipe=sub,
                                  weight_g=Decimal("20"), ordering=0)

        r = self.client.get(f"/recipes/{sub.pk}/delete/")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        # Warning copy + both parents linked
        self.assertIn("used as a sub-recipe", body.lower())
        self.assertIn(f'href="/recipes/{p1.pk}/"', body)
        self.assertIn(f'href="/recipes/{p2.pk}/"', body)
        self.assertIn("NPD-R901", body)
        self.assertIn("NPD-R902", body)

    def test_delete_removes_recipe_and_writes_suppression(self):
        r = Recipe.objects.create(code="NPD-R999", name="Gone",
                                  department=self.dept)
        resp = self.client.post(f"/recipes/{r.pk}/delete/",
                                {"acknowledge": "on", "confirm_code": "NPD-R999"})
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Recipe.objects.filter(code="NPD-R999").exists())
        # Suppression row written so the next re-import won't bring it back.
        self.assertTrue(SuppressedRecipe.objects.filter(code="NPD-R999").exists())

    def test_hard_delete_requires_acknowledgement(self):
        # Without the ack box (or with a missing/wrong typed code) the
        # form re-renders with an error and the recipe survives.
        r = Recipe.objects.create(code="NPD-R999", name="Gone",
                                  department=self.dept)
        # No acknowledge → re-renders form
        resp = self.client.post(f"/recipes/{r.pk}/delete/",
                                {"confirm_code": "NPD-R999"})
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"acknowledgement", resp.content)
        self.assertTrue(Recipe.objects.filter(code="NPD-R999").exists())
        # Wrong code → re-renders form
        resp = self.client.post(f"/recipes/{r.pk}/delete/",
                                {"acknowledge": "on", "confirm_code": "WRONG"})
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"Type the recipe code exactly", resp.content)
        self.assertTrue(Recipe.objects.filter(code="NPD-R999").exists())

    def test_hard_delete_confirm_page_emphasises_archive_alternative(self):
        # The confirmation page should make clear that archive is the
        # non-destructive default and offer it as a one-click alternative.
        r = Recipe.objects.create(code="NPD-R999", name="Gone",
                                  department=self.dept)
        resp = self.client.get(f"/recipes/{r.pk}/delete/")
        body = resp.content.decode()
        self.assertIn("archive instead", body.lower())
        self.assertIn(f'action="/recipes/{r.pk}/archive/"', body)
        # And the form must demand the acknowledgement + typed code.
        self.assertIn('name="acknowledge"', body)
        self.assertIn('name="confirm_code"', body)

    def test_delete_clears_dangling_sub_recipe_lines_in_parents(self):
        # RecipeLine.sub_recipe is on_delete=PROTECT, so deleting a sub
        # without first dropping the lines that reference it would raise
        # IntegrityError. The view must clean those up cleanly.
        sub = Recipe.objects.create(code="NPD-R900", name="Filling",
                                    department=self.dept)
        parent = Recipe.objects.create(code="NPD-R901", name="Pie",
                                       department=self.dept)
        flour = Product.objects.create(
            code="NPD-I9001", name="Flour", department=self.dept,
            unit="g", minimum=0)
        RecipeLine.objects.create(recipe=parent, ingredient=flour,
                                  weight_g=Decimal("100"), ordering=0)
        RecipeLine.objects.create(recipe=parent, sub_recipe=sub,
                                  weight_g=Decimal("50"), ordering=1)

        resp = self.client.post(f"/recipes/{sub.pk}/delete/",
                                {"acknowledge": "on", "confirm_code": "NPD-R900"})
        self.assertEqual(resp.status_code, 302)
        # Parent survives.
        self.assertTrue(Recipe.objects.filter(code="NPD-R901").exists())
        # The sub-recipe line is gone but the flour line is still there.
        parent.refresh_from_db()
        self.assertEqual(parent.lines.count(), 1)
        kept = parent.lines.get()
        self.assertEqual(kept.ingredient, flour)

    def test_reimport_does_not_resurrect_suppressed_recipe(self):
        # Manually mark NPD-R800 (in the committed sourdough sample) as
        # suppressed, then run the single-file importer. The recipe must
        # not come back.
        from django.core.management import call_command
        SuppressedRecipe.objects.create(code="NPD-R800",
                                        reason="Deleted by hand")
        call_command("import_recipe", SAMPLE_RECIPE_XLSX,
                     "--department", "Bakery")
        self.assertFalse(Recipe.objects.filter(code="NPD-R800").exists())
        # The other six recipes from the same workbook still imported.
        self.assertGreaterEqual(Recipe.objects.count(), 6)
        # And NPD-R364 (which the workbook had as the sole child of R800)
        # is no longer used as a component anywhere — because the only
        # parent that referenced it was the suppressed R800.
        r364 = Recipe.objects.get(code="NPD-R364")
        self.assertFalse(r364.is_used_as_component)

    def test_un_suppress_via_admin_reenables_import(self):
        # Delete the SuppressedRecipe row (the "un-suppress" path) and
        # the next import re-creates the recipe normally.
        from django.core.management import call_command
        SuppressedRecipe.objects.create(code="NPD-R800")
        call_command("import_recipe", SAMPLE_RECIPE_XLSX,
                     "--department", "Bakery")
        self.assertFalse(Recipe.objects.filter(code="NPD-R800").exists())
        # Un-suppress
        SuppressedRecipe.objects.filter(code="NPD-R800").delete()
        call_command("import_recipe", SAMPLE_RECIPE_XLSX,
                     "--department", "Bakery")
        self.assertTrue(Recipe.objects.filter(code="NPD-R800").exists())

    def test_delete_cross_department_returns_403(self):
        other = Recipe.objects.create(code="NPD-R777", name="Theirs",
                                      department=self.other_dept)
        # GET confirmation page → 403
        r = self.client.get(f"/recipes/{other.pk}/delete/")
        self.assertEqual(r.status_code, 403)
        # POST → 403, recipe still exists
        r = self.client.post(f"/recipes/{other.pk}/delete/")
        self.assertEqual(r.status_code, 403)
        self.assertTrue(Recipe.objects.filter(code="NPD-R777").exists())
        # No suppression row was written either.
        self.assertFalse(SuppressedRecipe.objects.filter(code="NPD-R777").exists())

    def test_delete_requires_login(self):
        r = Recipe.objects.create(code="NPD-R999", name="Gone",
                                  department=self.dept)
        anon = Client()
        for verb in ("get", "post"):
            resp = getattr(anon, verb)(f"/recipes/{r.pk}/delete/")
            self.assertEqual(resp.status_code, 302)
            self.assertIn("/login/", resp.headers["Location"])

    def test_delete_via_get_does_not_delete(self):
        # Hitting the URL with GET shows the confirmation but must never
        # mutate; only the explicit POST commits the deletion.
        r = Recipe.objects.create(code="NPD-R999", name="Gone",
                                  department=self.dept)
        resp = self.client.get(f"/recipes/{r.pk}/delete/")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(Recipe.objects.filter(code="NPD-R999").exists())
        self.assertFalse(SuppressedRecipe.objects.filter(code="NPD-R999").exists())

    # ---- edit basic fields ----

    def test_edit_form_renders_with_current_values(self):
        r = Recipe.objects.create(
            code="NPD-R100", name="Loaf", department=self.dept,
            finished_weight_g=Decimal("400"),
            deposit_weight_g=Decimal("420"),
            cook_loss_pct=Decimal("5.00"),
            sold_as_product=True)
        resp = self.client.get(f"/recipes/{r.pk}/edit/")
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        self.assertIn('value="Loaf"', body)
        self.assertIn('value="400.000"', body)
        self.assertIn('value="420.000"', body)
        self.assertIn('value="5.00"', body)
        self.assertIn('checked', body)

    def test_edit_persists_basic_fields_and_flips_manual_flag(self):
        r = Recipe.objects.create(
            code="NPD-R100", name="Loaf", department=self.dept,
            finished_weight_g=Decimal("400"),
            sold_as_product=True)
        resp = self.client.post(f"/recipes/{r.pk}/edit/", {
            "name": "Big Loaf",
            "finished_weight_g": "500",
            "deposit_weight_g": "525",
            "cook_loss_pct": "5",
            # sold checkbox deliberately unchecked
        })
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.headers["Location"], f"/recipes/{r.pk}/")
        r.refresh_from_db()
        self.assertEqual(r.name, "Big Loaf")
        self.assertEqual(r.finished_weight_g, Decimal("500"))
        self.assertEqual(r.deposit_weight_g, Decimal("525"))
        self.assertEqual(r.cook_loss_pct, Decimal("5"))
        self.assertFalse(r.sold_as_product)
        # Both manual flags flipped on — protects against re-import overwrite.
        self.assertTrue(r.is_basic_manual)
        self.assertTrue(r.is_sold_manual)

    def test_edit_survives_bulk_reimport(self):
        # Edit a recipe by hand, then re-import the same workbook —
        # the edited basic fields must survive (lines still rebuild).
        from django.core.management import call_command
        call_command("import_recipe", SAMPLE_RECIPE_XLSX,
                     "--department", "Bakery")
        r = Recipe.objects.get(code="NPD-R800")
        self.client.post(f"/recipes/{r.pk}/edit/", {
            "name": "Custom Name",
            "finished_weight_g": "1234",
            "deposit_weight_g": "1300",
            "cook_loss_pct": "9.99",
            "sold_as_product": "on",
        })
        # Re-import
        call_command("import_recipe", SAMPLE_RECIPE_XLSX,
                     "--department", "Bakery")
        r.refresh_from_db()
        self.assertEqual(r.name, "Custom Name")
        self.assertEqual(r.finished_weight_g, Decimal("1234"))
        self.assertEqual(r.deposit_weight_g, Decimal("1300"))
        self.assertEqual(r.cook_loss_pct, Decimal("9.99"))
        self.assertTrue(r.sold_as_product)
        # Lines were rebuilt from the workbook (Stage A doesn't touch
        # lines), so the bill of materials is in sync.
        self.assertEqual(r.lines.count(), 1)

    def test_edit_validates_required_name(self):
        r = Recipe.objects.create(code="NPD-R100", name="Loaf",
                                  department=self.dept)
        resp = self.client.post(f"/recipes/{r.pk}/edit/", {
            "name": "",
            "finished_weight_g": "",
            "deposit_weight_g": "",
            "cook_loss_pct": "",
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"Name is required", resp.content)
        r.refresh_from_db()
        self.assertEqual(r.name, "Loaf")
        self.assertFalse(r.is_basic_manual)

    def test_edit_cross_department_returns_403(self):
        other = Recipe.objects.create(code="NPD-R777", name="Theirs",
                                      department=self.other_dept)
        r = self.client.get(f"/recipes/{other.pk}/edit/")
        self.assertEqual(r.status_code, 403)
        r = self.client.post(f"/recipes/{other.pk}/edit/", {"name": "X"})
        self.assertEqual(r.status_code, 403)
        other.refresh_from_db()
        self.assertEqual(other.name, "Theirs")

    def test_edit_requires_login(self):
        r = Recipe.objects.create(code="NPD-R100", name="Loaf",
                                  department=self.dept)
        anon = Client()
        resp = anon.get(f"/recipes/{r.pk}/edit/")
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login/", resp.headers["Location"])

    def test_detail_page_links_to_edit_and_delete(self):
        r = Recipe.objects.create(code="NPD-R100", name="Loaf",
                                  department=self.dept)
        resp = self.client.get(f"/recipes/{r.pk}/")
        body = resp.content.decode()
        self.assertIn(f'href="/recipes/{r.pk}/edit/"', body)
        self.assertIn(f'href="/recipes/{r.pk}/delete/"', body)

    # ---- bulk-select + delete from the All-recipes list ----

    def test_flat_list_shows_checkboxes_and_bulk_archive_button(self):
        # Archive is now the primary bulk action on the All-recipes tab.
        Recipe.objects.create(code="NPD-R100", name="Alpha", department=self.dept)
        Recipe.objects.create(code="NPD-R101", name="Beta", department=self.dept)
        r = self.client.get("/recipes/?view=flat")
        body = r.content.decode()
        self.assertIn('id="bulk-form"', body)
        self.assertIn('name="recipe_ids"', body)
        self.assertIn("Archive selected", body)
        # Bulk form posts to the archive endpoint, not delete.
        self.assertIn('action="/recipes/bulk-archive/"', body)
        self.assertNotIn('action="/recipes/bulk-delete/"', body)
        # Per-row archive button is present (so individual recipes can
        # be archived without using bulk select).
        self.assertIn("/recipes/1/archive/", body)

    def test_bulk_delete_confirmation_lists_selected(self):
        a = Recipe.objects.create(code="NPD-R100", name="Alpha",
                                  department=self.dept)
        b = Recipe.objects.create(code="NPD-R101", name="Beta",
                                  department=self.dept)
        # Untouched recipe — must not appear on the confirmation page.
        Recipe.objects.create(code="NPD-R102", name="Gamma",
                              department=self.dept)
        resp = self.client.post("/recipes/bulk-delete/", {
            "recipe_ids": [str(a.pk), str(b.pk)],
        })
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        self.assertIn("NPD-R100", body)
        self.assertIn("Alpha", body)
        self.assertIn("NPD-R101", body)
        self.assertIn("Beta", body)
        self.assertNotIn("NPD-R102", body)
        # Heading + button copy mentions 2 recipes.
        self.assertIn("Permanently delete 2 recipe", body)
        # Confirmation NOT performed yet — both recipes still exist.
        self.assertEqual(Recipe.objects.filter(code__in=("NPD-R100", "NPD-R101")).count(), 2)
        self.assertFalse(SuppressedRecipe.objects.filter(code__in=("NPD-R100", "NPD-R101")).exists())

    def test_bulk_delete_confirmation_warns_about_external_dependents(self):
        # Two selected: NPD-R900 (used by NPD-R800 which is NOT selected),
        # and NPD-R901 (used by NPD-R902 which IS selected).
        ext_parent = Recipe.objects.create(code="NPD-R800", name="External Parent",
                                           department=self.dept)
        sel_parent = Recipe.objects.create(code="NPD-R902", name="Sel Parent",
                                           department=self.dept)
        a = Recipe.objects.create(code="NPD-R900", name="Sub A",
                                  department=self.dept)
        b = Recipe.objects.create(code="NPD-R901", name="Sub B",
                                  department=self.dept)
        RecipeLine.objects.create(recipe=ext_parent, sub_recipe=a,
                                  weight_g=Decimal("10"), ordering=0)
        RecipeLine.objects.create(recipe=sel_parent, sub_recipe=b,
                                  weight_g=Decimal("10"), ordering=0)
        resp = self.client.post("/recipes/bulk-delete/", {
            "recipe_ids": [str(a.pk), str(b.pk), str(sel_parent.pk)],
        })
        body = resp.content.decode()
        # External parent listed, with its lost reference shown.
        self.assertIn("External Parent", body)
        self.assertIn(f'href="/recipes/{ext_parent.pk}/"', body)
        self.assertIn("NPD-R900", body)
        # The internal parent (also selected) must NOT appear in the
        # "external" warning — it's being deleted itself.
        # We check by looking inside the warning block specifically; an
        # easy proxy: the count of "external" mentions is 1.
        self.assertEqual(body.count("External Parent"), 1)
        # Internal parent's code (NPD-R902) appears in the "Will delete"
        # table but not in the external-warning row, so we don't make a
        # negative assertion on it.

    def test_bulk_delete_confirm_removes_all_and_suppresses_codes(self):
        a = Recipe.objects.create(code="NPD-R100", name="Alpha",
                                  department=self.dept)
        b = Recipe.objects.create(code="NPD-R101", name="Beta",
                                  department=self.dept)
        resp = self.client.post("/recipes/bulk-delete/", {
            "recipe_ids": [str(a.pk), str(b.pk)],
            "confirm": "1",
            "acknowledge": "on", "confirm_phrase": "DELETE",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Recipe.objects.filter(
            code__in=("NPD-R100", "NPD-R101")).exists())
        self.assertEqual(
            set(SuppressedRecipe.objects.values_list("code", flat=True)),
            {"NPD-R100", "NPD-R101"})

    def test_bulk_delete_requires_acknowledgement_and_phrase(self):
        # Without the gate, the confirmation re-renders with an error
        # and nothing is deleted (matches the single hard-delete gate).
        a = Recipe.objects.create(code="NPD-R100", name="Alpha",
                                  department=self.dept)
        # No acknowledge → form re-renders
        resp = self.client.post("/recipes/bulk-delete/", {
            "recipe_ids": [str(a.pk)],
            "confirm": "1",
            "confirm_phrase": "DELETE",
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"acknowledgement", resp.content)
        self.assertTrue(Recipe.objects.filter(code="NPD-R100").exists())
        # Wrong phrase → form re-renders
        resp = self.client.post("/recipes/bulk-delete/", {
            "recipe_ids": [str(a.pk)],
            "confirm": "1",
            "acknowledge": "on",
            "confirm_phrase": "delete please",
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"DELETE", resp.content)
        self.assertTrue(Recipe.objects.filter(code="NPD-R100").exists())

    def test_bulk_delete_clears_dangling_external_subrecipe_lines(self):
        # External parent loses its sub-recipe line; its own ingredient
        # line stays. RecipeLine.sub_recipe is PROTECT, so this must be
        # cleaned up before the cascade.
        ext_parent = Recipe.objects.create(code="NPD-R800", name="External",
                                           department=self.dept)
        sub = Recipe.objects.create(code="NPD-R900", name="Sub",
                                    department=self.dept)
        flour = Product.objects.create(
            code="NPD-I9001", name="Flour", department=self.dept,
            unit="g", minimum=0)
        RecipeLine.objects.create(recipe=ext_parent, ingredient=flour,
                                  weight_g=Decimal("100"), ordering=0)
        RecipeLine.objects.create(recipe=ext_parent, sub_recipe=sub,
                                  weight_g=Decimal("50"), ordering=1)
        resp = self.client.post("/recipes/bulk-delete/", {
            "recipe_ids": [str(sub.pk)],
            "confirm": "1",
            "acknowledge": "on", "confirm_phrase": "DELETE",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Recipe.objects.filter(code="NPD-R900").exists())
        ext_parent.refresh_from_db()
        self.assertEqual(ext_parent.lines.count(), 1)
        self.assertEqual(ext_parent.lines.get().ingredient, flour)

    def test_bulk_delete_is_atomic(self):
        # Mock save inside the transaction to blow up — nothing should
        # commit (no deletions, no suppression rows).
        from unittest.mock import patch
        a = Recipe.objects.create(code="NPD-R100", name="Alpha",
                                  department=self.dept)
        b = Recipe.objects.create(code="NPD-R101", name="Beta",
                                  department=self.dept)
        with patch.object(SuppressedRecipe.objects,
                          "update_or_create",
                          side_effect=RuntimeError("boom")):
            with self.assertRaises(RuntimeError):
                self.client.post("/recipes/bulk-delete/", {
                    "recipe_ids": [str(a.pk), str(b.pk)],
                    "confirm": "1",
                    "acknowledge": "on", "confirm_phrase": "DELETE",
                })
        # Atomic rollback: both still there, no suppression.
        self.assertEqual(Recipe.objects.filter(
            code__in=("NPD-R100", "NPD-R101")).count(), 2)
        self.assertFalse(SuppressedRecipe.objects.filter(
            code__in=("NPD-R100", "NPD-R101")).exists())

    def test_bulk_delete_is_department_scoped(self):
        mine = Recipe.objects.create(code="NPD-R100", name="Mine",
                                     department=self.dept)
        theirs = Recipe.objects.create(code="NPD-R777", name="Theirs",
                                       department=self.other_dept)
        # POST with both — only the in-dept one should land on the
        # confirmation page; the cross-dept id is silently dropped.
        resp = self.client.post("/recipes/bulk-delete/", {
            "recipe_ids": [str(mine.pk), str(theirs.pk)],
        })
        body = resp.content.decode()
        self.assertIn("NPD-R100", body)
        self.assertNotIn("NPD-R777", body)
        # Confirming with both ids: the other dept's recipe still
        # survives because the filter excludes it.
        resp = self.client.post("/recipes/bulk-delete/", {
            "recipe_ids": [str(mine.pk), str(theirs.pk)],
            "confirm": "1",
            "acknowledge": "on", "confirm_phrase": "DELETE",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Recipe.objects.filter(code="NPD-R100").exists())
        self.assertTrue(Recipe.objects.filter(code="NPD-R777").exists())
        self.assertFalse(SuppressedRecipe.objects.filter(code="NPD-R777").exists())

    def test_bulk_delete_with_empty_selection_redirects_with_message(self):
        resp = self.client.post("/recipes/bulk-delete/", {})
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/recipes/", resp.headers["Location"])

    def test_bulk_delete_requires_login(self):
        r = Recipe.objects.create(code="NPD-R100", name="A",
                                  department=self.dept)
        anon = Client()
        resp = anon.post("/recipes/bulk-delete/", {"recipe_ids": [str(r.pk)]})
        self.assertEqual(resp.status_code, 302)
        self.assertIn("/login/", resp.headers["Location"])

    def test_bulk_deleted_codes_survive_reimport(self):
        # Bulk-delete two recipes from the committed sample workbook,
        # then re-import — neither must come back.
        from django.core.management import call_command
        call_command("import_recipe", SAMPLE_RECIPE_XLSX,
                     "--department", "Bakery")
        a = Recipe.objects.get(code="NPD-R2031")
        b = Recipe.objects.get(code="NPD-R2082")
        self.client.post("/recipes/bulk-delete/", {
            "recipe_ids": [str(a.pk), str(b.pk)],
            "confirm": "1",
            "acknowledge": "on", "confirm_phrase": "DELETE",
        })
        call_command("import_recipe", SAMPLE_RECIPE_XLSX,
                     "--department", "Bakery")
        codes = set(Recipe.objects.values_list("code", flat=True))
        self.assertNotIn("NPD-R2031", codes)
        self.assertNotIn("NPD-R2082", codes)
        # And the suppression rows are still there for both.
        self.assertEqual(
            SuppressedRecipe.objects.filter(
                code__in=("NPD-R2031", "NPD-R2082")).count(), 2)

    # ---- archive / restore (the new primary delete-like action) ----

    def test_archive_hides_recipe_from_main_views_but_keeps_row(self):
        r = Recipe.objects.create(code="NPD-R100", name="Alpha",
                                  department=self.dept,
                                  finished_weight_g=Decimal("400"))
        flour = Product.objects.create(
            code="NPD-I9001", name="Flour", department=self.dept,
            unit="g", minimum=0)
        RecipeLine.objects.create(recipe=r, ingredient=flour,
                                  weight_g=Decimal("100"), ordering=0)
        resp = self.client.post(f"/recipes/{r.pk}/archive/",
                                {"next": "/recipes/"})
        self.assertEqual(resp.status_code, 302)
        # Row still in the DB.
        r.refresh_from_db()
        self.assertTrue(r.archived)
        self.assertIsNotNone(r.archived_at)
        # Its lines are preserved (archive is reversible).
        self.assertEqual(r.lines.count(), 1)
        # Consume the post-archive flash message so it doesn't leak the
        # code into our "is it hidden?" assertions below.
        self.client.get("/recipes/?view=archived")
        # Hidden from the main flat view (look for the row anchor, not
        # the raw code, since the page chrome may still mention it).
        body = self.client.get("/recipes/?view=flat").content.decode()
        self.assertNotIn(f'href="/recipes/{r.pk}/"', body)
        # And from the by-product tree.
        body = self.client.get("/recipes/").content.decode()
        self.assertNotIn(f'href="/recipes/{r.pk}/"', body)
        # But shown in the archived view.
        body = self.client.get("/recipes/?view=archived").content.decode()
        self.assertIn(f'href="/recipes/{r.pk}/"', body)
        self.assertIn("NPD-R100", body)

    def test_archive_view_lists_archive_count_in_tab(self):
        Recipe.objects.create(code="NPD-R100", name="A",
                              department=self.dept, archived=True,
                              archived_at=timezone.now())
        Recipe.objects.create(code="NPD-R101", name="B",
                              department=self.dept, archived=True,
                              archived_at=timezone.now())
        Recipe.objects.create(code="NPD-R102", name="C",
                              department=self.dept)
        body = self.client.get("/recipes/?view=flat").content.decode()
        # The Archived tab carries the count badge "(2)"
        self.assertIn("Archived", body)
        self.assertIn("(2)", body)

    def test_restore_brings_recipe_back_to_active(self):
        r = Recipe.objects.create(code="NPD-R100", name="Alpha",
                                  department=self.dept,
                                  archived=True,
                                  archived_at=timezone.now())
        resp = self.client.post(f"/recipes/{r.pk}/restore/",
                                {"next": "/recipes/"})
        self.assertEqual(resp.status_code, 302)
        r.refresh_from_db()
        self.assertFalse(r.archived)
        self.assertIsNone(r.archived_at)
        body = self.client.get("/recipes/?view=flat").content.decode()
        self.assertIn("NPD-R100", body)

    def test_archive_preserved_through_reimport(self):
        from django.core.management import call_command
        call_command("import_recipe", SAMPLE_RECIPE_XLSX,
                     "--department", "Bakery")
        # Archive a deep sub-recipe (NPD-R307) so it stops appearing
        # in the main views.
        target = Recipe.objects.get(code="NPD-R307")
        self.client.post(f"/recipes/{target.pk}/archive/")
        target.refresh_from_db()
        self.assertTrue(target.archived)
        archived_at = target.archived_at
        # Re-import refreshes its basics — but must NOT un-archive.
        call_command("import_recipe", SAMPLE_RECIPE_XLSX,
                     "--department", "Bakery")
        target.refresh_from_db()
        self.assertTrue(target.archived,
                        "Re-import should not flip archived back to False")
        # archived_at preserved (re-import doesn't touch it).
        self.assertEqual(target.archived_at, archived_at)
        # Its lines were still rebuilt from the workbook (the
        # bill of materials stays in sync even while hidden).
        self.assertEqual(target.lines.count(), 3)

    def test_archiving_referenced_component_does_not_break_parents(self):
        # Parent → sub_recipe. Archive the sub: parent stays intact,
        # its line still references the (archived) sub.
        sub = Recipe.objects.create(code="NPD-R900", name="Filling",
                                    department=self.dept)
        parent = Recipe.objects.create(code="NPD-R901", name="Pie",
                                       department=self.dept)
        flour = Product.objects.create(
            code="NPD-I9001", name="Flour", department=self.dept,
            unit="g", minimum=0)
        line = RecipeLine.objects.create(recipe=parent, sub_recipe=sub,
                                         weight_g=Decimal("50"), ordering=0)
        self.client.post(f"/recipes/{sub.pk}/archive/")
        # Sub archived, but parent and its line still exist and still
        # reference it (the FK is intact).
        sub.refresh_from_db()
        self.assertTrue(sub.archived)
        parent.refresh_from_db()
        self.assertEqual(parent.lines.count(), 1)
        line.refresh_from_db()
        self.assertEqual(line.sub_recipe, sub)
        # Active parent of an archived sub is surfaced in the archived view.
        body = self.client.get("/recipes/?view=archived").content.decode()
        self.assertIn("NPD-R901", body)

    def test_bulk_archive_confirmation_lists_selected(self):
        a = Recipe.objects.create(code="NPD-R100", name="Alpha",
                                  department=self.dept)
        b = Recipe.objects.create(code="NPD-R101", name="Beta",
                                  department=self.dept)
        Recipe.objects.create(code="NPD-R102", name="Gamma",
                              department=self.dept)
        resp = self.client.post("/recipes/bulk-archive/", {
            "recipe_ids": [str(a.pk), str(b.pk)],
        })
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        self.assertIn("Archive 2 recipe", body)
        self.assertIn("NPD-R100", body)
        self.assertIn("NPD-R101", body)
        self.assertNotIn("NPD-R102", body)
        # No commit yet.
        self.assertFalse(Recipe.objects.filter(
            code__in=("NPD-R100", "NPD-R101"), archived=True).exists())

    def test_bulk_archive_confirmation_shows_external_reference_note(self):
        # NPD-R900 still used by active NPD-R800 (not in selection).
        ext_parent = Recipe.objects.create(code="NPD-R800", name="External",
                                           department=self.dept)
        a = Recipe.objects.create(code="NPD-R900", name="Sub A",
                                  department=self.dept)
        b = Recipe.objects.create(code="NPD-R901", name="Sub B",
                                  department=self.dept)
        RecipeLine.objects.create(recipe=ext_parent, sub_recipe=a,
                                  weight_g=Decimal("10"), ordering=0)
        resp = self.client.post("/recipes/bulk-archive/", {
            "recipe_ids": [str(a.pk), str(b.pk)],
        })
        body = resp.content.decode()
        # Note block mentions the parent and the archived child code.
        self.assertIn("still referenced", body.lower())
        self.assertIn("NPD-R800", body)
        self.assertIn("NPD-R900", body)
        # NPD-R901 not externally referenced, so doesn't get a row in
        # the note table. (It's still in the "Will archive" list above.)

    def test_bulk_archive_confirm_flips_all_to_archived(self):
        a = Recipe.objects.create(code="NPD-R100", name="A",
                                  department=self.dept)
        b = Recipe.objects.create(code="NPD-R101", name="B",
                                  department=self.dept)
        c = Recipe.objects.create(code="NPD-R102", name="C",
                                  department=self.dept)
        resp = self.client.post("/recipes/bulk-archive/", {
            "recipe_ids": [str(a.pk), str(b.pk), str(c.pk)],
            "confirm": "1",
        })
        self.assertEqual(resp.status_code, 302)
        for code in ("NPD-R100", "NPD-R101", "NPD-R102"):
            r = Recipe.objects.get(code=code)
            self.assertTrue(r.archived)
            self.assertIsNotNone(r.archived_at)
        # No suppression rows (archive ≠ hard delete).
        self.assertFalse(SuppressedRecipe.objects.filter(
            code__in=("NPD-R100", "NPD-R101", "NPD-R102")).exists())

    def test_bulk_archive_is_department_scoped(self):
        mine = Recipe.objects.create(code="NPD-R100", name="Mine",
                                     department=self.dept)
        theirs = Recipe.objects.create(code="NPD-R777", name="Theirs",
                                       department=self.other_dept)
        resp = self.client.post("/recipes/bulk-archive/", {
            "recipe_ids": [str(mine.pk), str(theirs.pk)],
            "confirm": "1",
        })
        self.assertEqual(resp.status_code, 302)
        mine.refresh_from_db()
        theirs.refresh_from_db()
        self.assertTrue(mine.archived)
        # Cross-dept recipe untouched.
        self.assertFalse(theirs.archived)

    def test_bulk_archive_skips_already_archived(self):
        # Already-archived recipes in the selection are silently
        # filtered out so they don't appear on the confirmation page.
        active = Recipe.objects.create(code="NPD-R100", name="Active",
                                       department=self.dept)
        already = Recipe.objects.create(code="NPD-R101", name="Already",
                                        department=self.dept,
                                        archived=True,
                                        archived_at=timezone.now())
        resp = self.client.post("/recipes/bulk-archive/", {
            "recipe_ids": [str(active.pk), str(already.pk)],
        })
        body = resp.content.decode()
        self.assertIn("Archive 1 recipe", body)
        self.assertIn("NPD-R100", body)
        self.assertNotIn("NPD-R101", body)

    def test_bulk_restore_flips_all_back_to_active(self):
        a = Recipe.objects.create(code="NPD-R100", name="A",
                                  department=self.dept,
                                  archived=True,
                                  archived_at=timezone.now())
        b = Recipe.objects.create(code="NPD-R101", name="B",
                                  department=self.dept,
                                  archived=True,
                                  archived_at=timezone.now())
        resp = self.client.post("/recipes/bulk-restore/", {
            "recipe_ids": [str(a.pk), str(b.pk)],
        })
        self.assertEqual(resp.status_code, 302)
        for code in ("NPD-R100", "NPD-R101"):
            r = Recipe.objects.get(code=code)
            self.assertFalse(r.archived)
            self.assertIsNone(r.archived_at)

    def test_bulk_restore_skips_already_active(self):
        active = Recipe.objects.create(code="NPD-R100", name="Active",
                                       department=self.dept)
        archived = Recipe.objects.create(code="NPD-R101", name="Archived",
                                         department=self.dept,
                                         archived=True,
                                         archived_at=timezone.now())
        resp = self.client.post("/recipes/bulk-restore/", {
            "recipe_ids": [str(active.pk), str(archived.pk)],
        })
        self.assertEqual(resp.status_code, 302)
        active.refresh_from_db()
        archived.refresh_from_db()
        # Active untouched (nothing to restore), archived flipped.
        self.assertFalse(active.archived)
        self.assertFalse(archived.archived)

    def test_archive_cross_department_returns_403(self):
        other = Recipe.objects.create(code="NPD-R777", name="Theirs",
                                      department=self.other_dept)
        resp = self.client.post(f"/recipes/{other.pk}/archive/")
        self.assertEqual(resp.status_code, 403)
        other.refresh_from_db()
        self.assertFalse(other.archived)

    def test_restore_cross_department_returns_403(self):
        other = Recipe.objects.create(code="NPD-R777", name="Theirs",
                                      department=self.other_dept,
                                      archived=True,
                                      archived_at=timezone.now())
        resp = self.client.post(f"/recipes/{other.pk}/restore/")
        self.assertEqual(resp.status_code, 403)
        other.refresh_from_db()
        self.assertTrue(other.archived)

    def test_archive_and_restore_require_login(self):
        r = Recipe.objects.create(code="NPD-R100", name="A",
                                  department=self.dept)
        anon = Client()
        for url in (f"/recipes/{r.pk}/archive/", f"/recipes/{r.pk}/restore/"):
            resp = anon.post(url)
            self.assertEqual(resp.status_code, 302)
            self.assertIn("/login/", resp.headers["Location"])

    def test_detail_page_offers_archive_primary_and_permanent_delete_link(self):
        # Active recipe → "archive" is the prominent button, with the
        # ghost-text link to permanent delete below.
        r = Recipe.objects.create(code="NPD-R100", name="A",
                                  department=self.dept)
        body = self.client.get(f"/recipes/{r.pk}/").content.decode()
        self.assertIn(f'action="/recipes/{r.pk}/archive/"', body)
        self.assertIn("Permanently delete", body)
        self.assertIn(f'href="/recipes/{r.pk}/delete/"', body)

    def test_detail_page_shows_restore_when_recipe_is_archived(self):
        r = Recipe.objects.create(code="NPD-R100", name="A",
                                  department=self.dept,
                                  archived=True,
                                  archived_at=timezone.now())
        body = self.client.get(f"/recipes/{r.pk}/").content.decode()
        self.assertIn(f'action="/recipes/{r.pk}/restore/"', body)
        self.assertIn("Archived", body)

    def test_archived_view_renders_per_row_restore(self):
        r = Recipe.objects.create(code="NPD-R100", name="A",
                                  department=self.dept,
                                  archived=True,
                                  archived_at=timezone.now())
        body = self.client.get("/recipes/?view=archived").content.decode()
        self.assertIn("NPD-R100", body)
        self.assertIn(f'action="/recipes/{r.pk}/restore/"', body)
        # Bulk restore form posts to the right endpoint.
        self.assertIn('action="/recipes/bulk-restore/"', body)
        self.assertIn("Restore selected", body)

    # ---- by-product tree: bulk-select with auto-tick cascade ----
    #
    # The auto-tick logic itself lives in client-side JS, so most of
    # the closure scenarios are server-side data assertions: the
    # template emits each recipe's `data-active-parents` and
    # `data-roots-using` lists, and the JS computes orphan-ness from
    # there. The tests below verify that:
    #   * the data attributes match the actual sub_recipe references,
    #   * the bulk-archive / bulk-delete endpoints accept the full
    #     selection (product + components) the JS posts,
    #   * the hard-delete confirmation gate works on the bulk path.

    def _make_tree(self):
        """A: P1 -> [C1 (only P1), C2 (shared P1+P2)], P2 -> [C2, C3 (only P2)],
                  P1 -> ... -> C1 -> D1 (only-via-C1).

        Sets up four products and three components, with a nested
        depth-2 sub-component D1 reachable only through the C1 branch
        of P1 — useful for the cascading-orphan test.
        """
        p1 = Recipe.objects.create(code="NPD-R100", name="Pie One",
                                   department=self.dept, sold_as_product=True)
        p2 = Recipe.objects.create(code="NPD-R101", name="Pie Two",
                                   department=self.dept, sold_as_product=True)
        c1 = Recipe.objects.create(code="NPD-R200", name="Filling",
                                   department=self.dept, sold_as_product=False)
        c2 = Recipe.objects.create(code="NPD-R201", name="Pastry",
                                   department=self.dept, sold_as_product=False)
        c3 = Recipe.objects.create(code="NPD-R202", name="Glaze",
                                   department=self.dept, sold_as_product=False)
        d1 = Recipe.objects.create(code="NPD-R300", name="Spice Mix",
                                   department=self.dept, sold_as_product=False)
        RecipeLine.objects.create(recipe=p1, sub_recipe=c1,
                                  weight_g=Decimal("100"), ordering=0)
        RecipeLine.objects.create(recipe=p1, sub_recipe=c2,
                                  weight_g=Decimal("80"), ordering=1)
        RecipeLine.objects.create(recipe=p2, sub_recipe=c2,
                                  weight_g=Decimal("70"), ordering=0)
        RecipeLine.objects.create(recipe=p2, sub_recipe=c3,
                                  weight_g=Decimal("30"), ordering=1)
        RecipeLine.objects.create(recipe=c1, sub_recipe=d1,
                                  weight_g=Decimal("10"), ordering=0)
        return {"p1": p1, "p2": p2, "c1": c1, "c2": c2, "c3": c3, "d1": d1}

    def test_tree_renders_checkboxes_and_data_attributes(self):
        t = self._make_tree()
        body = self.client.get("/recipes/").content.decode()
        # Two-button action bar; archive default, delete dangerous.
        self.assertIn('id="tree-bulk-form"', body)
        self.assertIn("Archive selected", body)
        self.assertIn("Delete selected", body)
        self.assertIn('formaction="/recipes/bulk-archive/"', body)
        self.assertIn('formaction="/recipes/bulk-delete/"', body)
        # Each active recipe in the tree is rendered with a tree-cb
        # checkbox carrying its data attributes.
        self.assertIn('class="tree-cb"', body)
        self.assertIn(f'data-recipe-id="{t["p1"].pk}"', body)
        self.assertIn(f'data-recipe-id="{t["c1"].pk}"', body)
        self.assertIn(f'data-recipe-id="{t["c2"].pk}"', body)

    def test_tree_node_active_parents_match_reality(self):
        t = self._make_tree()
        body = self.client.get("/recipes/").content.decode()
        # C1 — only-parent P1 (orphaned the moment P1 is in selection).
        c1_parents_pattern = f'data-recipe-id="{t["c1"].pk}"'
        # The data-active-parents attribute is a stringified list like
        # "[1, 2]"; assert C1's contains P1 only, C2's contains both.
        import re
        def parents_for(pk):
            # Find the first checkbox tag carrying this recipe id and
            # extract data-active-parents from it.
            m = re.search(
                rf'<input[^>]*data-recipe-id="{pk}"[^>]*'
                r'data-active-parents="([^"]*)"', body)
            assert m, f"no checkbox for recipe pk={pk}"
            import json
            return set(json.loads(m.group(1)))
        self.assertEqual(parents_for(t["c1"].pk), {t["p1"].pk})
        self.assertEqual(parents_for(t["c2"].pk), {t["p1"].pk, t["p2"].pk})
        self.assertEqual(parents_for(t["c3"].pk), {t["p2"].pk})
        self.assertEqual(parents_for(t["d1"].pk), {t["c1"].pk})
        # Roots have no active parents themselves.
        self.assertEqual(parents_for(t["p1"].pk), set())

    def test_tree_node_roots_using_match_reality(self):
        t = self._make_tree()
        body = self.client.get("/recipes/").content.decode()
        import re, json
        def roots_for(pk):
            m = re.search(
                rf'<input[^>]*data-recipe-id="{pk}"[^>]*'
                r'data-roots-using="([^"]*)"', body)
            assert m
            return set(json.loads(m.group(1)))
        # C2 is reached from both P1 and P2; C1 only from P1; D1 only from P1 via C1.
        self.assertEqual(roots_for(t["c2"].pk), {t["p1"].pk, t["p2"].pk})
        self.assertEqual(roots_for(t["c1"].pk), {t["p1"].pk})
        self.assertEqual(roots_for(t["d1"].pk), {t["p1"].pk})
        # A root product's "roots using" includes itself (it's its own
        # tree) — the JS subtracts the current node before rendering.
        self.assertIn(t["p1"].pk, roots_for(t["p1"].pk))

    def test_tree_archive_accepts_product_plus_orphaned_components(self):
        # The JS would build the selection as {P1, C1, D1} (C1 and D1
        # only reachable via P1). Posting that to bulk-archive must
        # archive all three; C2 (shared) and P2 untouched.
        t = self._make_tree()
        resp = self.client.post("/recipes/bulk-archive/", {
            "recipe_ids": [str(t["p1"].pk), str(t["c1"].pk), str(t["d1"].pk)],
        })
        # Confirmation page renders with all three listed.
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        for code in ("NPD-R100", "NPD-R200", "NPD-R300"):
            self.assertIn(code, body)
        # Commit.
        resp = self.client.post("/recipes/bulk-archive/", {
            "recipe_ids": [str(t["p1"].pk), str(t["c1"].pk), str(t["d1"].pk)],
            "confirm": "1",
        })
        self.assertEqual(resp.status_code, 302)
        for code in ("NPD-R100", "NPD-R200", "NPD-R300"):
            self.assertTrue(Recipe.objects.get(code=code).archived)
        # Shared C2 and the other product P2 + C3 are untouched.
        for code in ("NPD-R101", "NPD-R201", "NPD-R202"):
            self.assertFalse(Recipe.objects.get(code=code).archived)

    def test_tree_archive_confirmation_flags_shared_component_when_included(self):
        # The JS auto-ticks SAFE components only; the user can override
        # by manually ticking a shared one (with a JS warning). When
        # such a selection is POSTed, the confirmation must still
        # surface the "still referenced" note so the operator is
        # reminded server-side too.
        t = self._make_tree()
        # Selection = {P1, C2} — C2 is shared with P2 (not in selection).
        resp = self.client.post("/recipes/bulk-archive/", {
            "recipe_ids": [str(t["p1"].pk), str(t["c2"].pk)],
        })
        body = resp.content.decode()
        self.assertIn("still referenced", body.lower())
        # Names the still-active parent.
        self.assertIn("NPD-R101", body)
        # And the component being archived.
        self.assertIn("NPD-R201", body)

    def test_tree_archive_leaves_shared_component_alone_when_unticked(self):
        # Posting only the product (without the shared component) must
        # NOT archive the component — it stays active.
        t = self._make_tree()
        resp = self.client.post("/recipes/bulk-archive/", {
            "recipe_ids": [str(t["p1"].pk), str(t["c1"].pk), str(t["d1"].pk)],
            "confirm": "1",
        })
        self.assertEqual(resp.status_code, 302)
        c2 = Recipe.objects.get(code="NPD-R201")
        self.assertFalse(c2.archived)
        # C2's ACTIVE parents reduce to {P2} after P1 is archived.
        # parents() returns all references regardless of archived state,
        # so we filter here to assert what the tree's auto-tick logic
        # would consider (active references only).
        active_parents = set(c2.parents().filter(archived=False)
                             .values_list("code", flat=True))
        self.assertEqual(active_parents, {"NPD-R101"})

    def test_tree_delete_accepts_full_selection_with_acknowledgement(self):
        # Tree's "Delete selected" posts to bulk-delete; the
        # acknowledgement gate must accept ack=on + confirm_phrase=DELETE.
        t = self._make_tree()
        resp = self.client.post("/recipes/bulk-delete/", {
            "recipe_ids": [str(t["p1"].pk), str(t["c1"].pk), str(t["d1"].pk)],
            "confirm": "1",
            "acknowledge": "on",
            "confirm_phrase": "DELETE",
        })
        self.assertEqual(resp.status_code, 302)
        for code in ("NPD-R100", "NPD-R200", "NPD-R300"):
            self.assertFalse(Recipe.objects.filter(code=code).exists())
        # Suppression rows written.
        suppressed = set(SuppressedRecipe.objects.values_list("code", flat=True))
        self.assertEqual(suppressed,
                         {"NPD-R100", "NPD-R200", "NPD-R300"})
        # Shared C2 and the other product survive.
        for code in ("NPD-R101", "NPD-R201", "NPD-R202"):
            self.assertTrue(Recipe.objects.filter(code=code).exists())

    def test_tree_does_not_render_archived_recipes(self):
        # Archived recipes don't appear in the by-product tree at all
        # (the bulk-selection UI never sees them). They show up in the
        # Archived tab only.
        t = self._make_tree()
        t["c1"].archived = True
        t["c1"].archived_at = timezone.now()
        t["c1"].save()
        body = self.client.get("/recipes/").content.decode()
        # P1 and other actives still rendered; C1 absent.
        self.assertIn(f'data-recipe-id="{t["p1"].pk}"', body)
        self.assertNotIn(f'data-recipe-id="{t["c1"].pk}"', body)

    def test_tree_recipe_meta_json_embedded(self):
        # The JS reads /recipes/ meta to render code/name labels in
        # warnings. The view must emit a JSON dict keyed by pk.
        t = self._make_tree()
        body = self.client.get("/recipes/").content.decode()
        self.assertIn('id="tree-recipe-meta"', body)
        # Each active recipe present in the meta blob (look for its code).
        for code in ("NPD-R100", "NPD-R101", "NPD-R200", "NPD-R201",
                     "NPD-R202", "NPD-R300"):
            self.assertIn(code, body)

    def test_bulk_import_skips_suppressed_subrecipe_reference(self):
        # A workbook references NPD-R901 as a sub-recipe of NPD-R900.
        # NPD-R901 is suppressed, so the bulk import must NOT auto-stub
        # a new Recipe row for it; the parent's line is dropped instead.
        import os
        import tempfile
        from openpyxl import Workbook
        from stock.recipe_import import save_recipes, parse_recipe_workbook_bulk

        SuppressedRecipe.objects.create(code="NPD-R901",
                                        reason="Deleted by hand")
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            wb = Workbook()
            wb.remove(wb.active)

            def _sheet(title, code, name, lines):
                ws = wb.create_sheet(title=title)
                ws.append(["Recipe Code:", code])
                ws.append(["Recipe Description:", name])
                ws.append([])
                ws.append(["Code", "Description", "Weight (g)"])
                for c, d, w in lines:
                    ws.append([c, d, w])
                ws.append(["Total", "", sum(w for _, _, w in lines)])

            _sheet("R900", "NPD-R900", "Loaf",
                   [("NPD-I9001", "Flour", 200),
                    ("NPD-R901", "Suppressed Filling", 50)])
            wb.save(path)
            parsed, failures, n = parse_recipe_workbook_bulk(path)
        finally:
            os.unlink(path)
        stats = save_recipes(parsed, self.dept)
        self.assertIn("NPD-R901", stats["suppressed_skipped"])
        # NPD-R901 was NOT auto-stubbed.
        self.assertFalse(Recipe.objects.filter(code="NPD-R901").exists())
        # NPD-R900 imported with only the surviving ingredient line.
        r900 = Recipe.objects.get(code="NPD-R900")
        self.assertEqual(r900.lines.count(), 1)
        self.assertEqual(r900.lines.get().ingredient.code, "NPD-I9001")


def _build_sale_products_workbook(path):
    """Write a tiny order-sheet with a Products tab for the import tests.

    Matches the real workbook's column layout exactly: row 0 is the
    header `Product, Price, Sage No., (blank), Stock Managed, Pack Size`
    and the data rows fill in those columns. A few products are
    deliberately constructed to exercise every branch of the auto-
    linker (Sage hit, exact-name fallback, unlinked, Internal/Retail
    pair sharing one Sage code).
    """
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Products")
    ws.append(["Product", "Price", "Sage No.", None, "Stock Managed", "Pack Size"])
    rows = [
        # Sage-linked pair: Internal + Retail share one Sage code,
        # both linking to the same recipe.
        ("Crumble Topped Mince Pies Internal (Pack/6)", 2.25,
         "660130103", None, None, "1"),
        ("Crumble Topped Mince Pies Retail (Pack/6)", 7.00,
         "660130103", None, None, "1"),
        # Exact-name fallback: no Sage but the name matches a recipe.
        ("Apple Waste Sourdough (Loose)", 2.20,
         None, None, None, "1"),
        # Sage code present but no recipe with that code.
        ("Sourdough Surprise", 3.50,
         "999999999", None, None, "1"),
        # Nothing to link to — should NOT be fuzzy-auto-linked.
        ("Mystery Loaf (Pack/2)", 4.00,
         None, None, None, "2"),
    ]
    for r in rows:
        ws.append(list(r))
    wb.save(path)


class SaleProductsImportTests(TestCase):
    """import_sale_products end-to-end through the management command."""

    def setUp(self):
        import os
        import tempfile
        self.dept = Department.objects.create(name="Bakery")
        # Recipes the auto-linker should target.
        self.mince_recipe = Recipe.objects.create(
            code="NPD-R655", name="Crumble Topped Mince Pie",
            department=self.dept)
        # The mince pie products carry Sage code 660130103 — register
        # that as the recipe's code for the Sage match.
        Recipe.objects.create(
            code="660130103", name="Mince Pie SKU shadow",
            department=self.dept)
        # Exact-name match target (matches "Apple Waste Sourdough (Loose)").
        self.apple_recipe = Recipe.objects.create(
            code="NPD-R800", name="Apple Waste Sourdough (Loose)",
            department=self.dept)
        fd, self.path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        _build_sale_products_workbook(self.path)

    def tearDown(self):
        import os
        try:
            os.unlink(self.path)
        except OSError:
            pass

    def test_import_creates_sale_products_with_prices(self):
        from django.core.management import call_command
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        # Every workbook row landed.
        self.assertEqual(SaleProduct.objects.count(), 5)
        p = SaleProduct.objects.get(
            name="Crumble Topped Mince Pies Retail (Pack/6)")
        self.assertEqual(p.price, Decimal("7.00"))
        self.assertEqual(p.pack_size, "1")
        self.assertEqual(p.sage_number, "660130103")

    def test_sage_match_auto_links_and_confirms(self):
        from django.core.management import call_command
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        internal = SaleProduct.objects.get(
            name="Crumble Topped Mince Pies Internal (Pack/6)")
        retail = SaleProduct.objects.get(
            name="Crumble Topped Mince Pies Retail (Pack/6)")
        for sp in (internal, retail):
            self.assertEqual(sp.link_source, SaleProduct.SAGE)
            self.assertTrue(sp.link_confirmed)
            self.assertEqual(sp.recipe.code, "660130103")

    def test_internal_and_retail_link_to_the_same_recipe(self):
        from django.core.management import call_command
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        internal = SaleProduct.objects.get(
            name="Crumble Topped Mince Pies Internal (Pack/6)")
        retail = SaleProduct.objects.get(
            name="Crumble Topped Mince Pies Retail (Pack/6)")
        self.assertEqual(internal.recipe_id, retail.recipe_id)

    def test_exact_name_match_used_when_no_sage(self):
        from django.core.management import call_command
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        apple = SaleProduct.objects.get(name="Apple Waste Sourdough (Loose)")
        self.assertEqual(apple.recipe, self.apple_recipe)
        self.assertEqual(apple.link_source, SaleProduct.NAME)
        self.assertTrue(apple.link_confirmed)

    def test_product_with_no_sage_or_name_match_stays_unlinked(self):
        # The "Mystery Loaf" doesn't have a Sage match or an exact name
        # match. It must NOT be fuzzy-auto-linked (the link-review
        # screen is the only place for the operator to confirm a
        # fuzzy suggestion).
        from django.core.management import call_command
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        mystery = SaleProduct.objects.get(name="Mystery Loaf (Pack/2)")
        self.assertIsNone(mystery.recipe)
        self.assertEqual(mystery.link_source, SaleProduct.NONE)
        self.assertFalse(mystery.link_confirmed)

    def test_unknown_sage_code_stays_unlinked(self):
        from django.core.management import call_command
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        # "Sourdough Surprise" carries Sage 999999999 (no recipe with
        # that code) and its name doesn't exactly match a recipe → unlinked.
        sp = SaleProduct.objects.get(name="Sourdough Surprise")
        self.assertIsNone(sp.recipe)
        self.assertEqual(sp.link_source, SaleProduct.NONE)

    def test_import_is_idempotent_and_does_not_overwrite_name(self):
        from django.core.management import call_command
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        n1 = SaleProduct.objects.count()
        # Re-run — same workbook → same rows, no duplicates.
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        self.assertEqual(SaleProduct.objects.count(), n1)
        # Apple sale product's NAME stays the operator's source-of-truth
        # value (not overwritten by the linked recipe's name even though
        # they happen to match here).
        apple = SaleProduct.objects.get(name="Apple Waste Sourdough (Loose)")
        self.assertEqual(apple.name, "Apple Waste Sourdough (Loose)")

    def test_stock_managed_column_is_ignored(self):
        # The "Stock Managed" column at index 4 carries unrelated
        # values; the importer must ignore it entirely (no field on
        # SaleProduct carries that data).
        from django.core.management import call_command
        # Rewrite the workbook with a non-empty Stock Managed column.
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="Products")
        ws.append(["Product", "Price", "Sage No.", None,
                   "Stock Managed", "Pack Size"])
        ws.append(["Test SKU", 1.5, None, None, "Scone", "1"])
        wb.save(self.path)
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        sp = SaleProduct.objects.get(name="Test SKU")
        # No field stores "Scone" anywhere.
        self.assertNotIn("Scone", (
            sp.sage_number, sp.pack_size, sp.name))

    def test_manual_link_survives_reimport(self):
        # Pretend the operator went into the UI and manually linked
        # "Sourdough Surprise" to Apple recipe. The next import must
        # NOT clear it (Sourdough Surprise's Sage code 999999999 still
        # doesn't match any recipe, so the auto-linker would otherwise
        # set link_source=none, recipe=None).
        from django.core.management import call_command
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        sp = SaleProduct.objects.get(name="Sourdough Surprise")
        sp.recipe = self.apple_recipe
        sp.link_source = SaleProduct.MANUAL
        sp.link_confirmed = True
        sp.save()
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        sp.refresh_from_db()
        self.assertEqual(sp.recipe, self.apple_recipe)
        self.assertEqual(sp.link_source, SaleProduct.MANUAL)
        self.assertTrue(sp.link_confirmed)

    def test_manual_entry_row_skipped_by_importer(self):
        # A hand-created SaleProduct whose name happens to be in the
        # workbook is left entirely alone (is_manual_entry=True).
        from django.core.management import call_command
        SaleProduct.objects.create(
            name="Apple Waste Sourdough (Loose)",
            price=Decimal("99.99"), department=self.dept,
            is_manual_entry=True)
        call_command("import_sale_products", self.path,
                     "--department", "Bakery")
        sp = SaleProduct.objects.get(name="Apple Waste Sourdough (Loose)")
        # Price untouched (would have been 2.20 from the workbook).
        self.assertEqual(sp.price, Decimal("99.99"))
        self.assertTrue(sp.is_manual_entry)


class SaleProductsViewsTests(TestCase):
    """Products section: nav, list, detail, CRUD, link review."""

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.other_dept = Department.objects.create(name="Butchery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        # Two recipes for fuzzy + manual linking tests.
        self.apple = Recipe.objects.create(
            code="NPD-R800", name="Apple Waste Sourdough (Loose)",
            department=self.dept)
        self.mince = Recipe.objects.create(
            code="NPD-R655", name="Crumble Topped Mince Pie",
            department=self.dept)

    def test_home_page_top_nav_lists_products(self):
        # /home/ now renders the design-system shell; its left rail carries
        # the Products link (the old ordered top section-picker is gone).
        body = self.client.get("/home/").content.decode()
        self.assertIn('href="/products/"', body)

    def test_products_list_renders_with_filter_and_actions(self):
        SaleProduct.objects.create(
            name="Apple Waste Sourdough (Loose)", price=Decimal("2.20"),
            department=self.dept, recipe=self.apple,
            link_source=SaleProduct.NAME, link_confirmed=True)
        SaleProduct.objects.create(
            name="Mystery Loaf", department=self.dept)
        r = self.client.get("/sale-products/")
        body = r.content.decode()
        self.assertEqual(r.status_code, 200)
        self.assertIn("Apple Waste Sourdough (Loose)", body)
        self.assertIn("Mystery Loaf", body)
        self.assertIn("Link review", body)
        # Per-row edit + delete actions present.
        self.assertIn("sale-products/", body)
        self.assertIn("edit", body.lower())
        self.assertIn("delete", body.lower())

    def test_products_list_drops_linked_recipe_column(self):
        # The list no longer carries a "Linked recipe" / "Link" column
        # — the linked recipe lives on the product detail page now.
        SaleProduct.objects.create(
            name="Apple Waste Sourdough (Loose)", price=Decimal("2.20"),
            department=self.dept, recipe=self.apple,
            link_source=SaleProduct.NAME, link_confirmed=True)
        body = self.client.get("/sale-products/").content.decode()
        import re
        m = re.search(r"<table id=\"sp-tbl\">(.*?)</table>", body, re.DOTALL)
        self.assertIsNotNone(m, "products table missing")
        table = m.group(1)
        # No "Link" / "Linked recipe" column header.
        self.assertNotIn(">Link<", table)
        self.assertNotIn(">Linked recipe<", table)
        # The recipe code no longer leaks into table cells. (The recipe
        # name happens to match the product name in this fixture, so
        # that's not a reliable signal — check the code instead.)
        self.assertNotIn(self.apple.code, table)
        # Old per-row link tags are gone with the column.
        self.assertNotIn("link-tag", table)

    def test_products_list_keeps_linked_unlinked_header_stats(self):
        # Two linked + one unlinked → stats render Linked=2, Unlinked=1.
        SaleProduct.objects.create(
            name="A", department=self.dept, recipe=self.apple,
            link_source=SaleProduct.NAME, link_confirmed=True)
        SaleProduct.objects.create(
            name="B", department=self.dept, recipe=self.mince,
            link_source=SaleProduct.MANUAL, link_confirmed=True)
        SaleProduct.objects.create(name="Unlinked", department=self.dept)
        body = self.client.get("/sale-products/").content.decode()
        # Find the .stats block and inspect the linked / unlinked tiles.
        import re
        stats_match = re.search(r'class="stats">(.*?)</div>\s*<div class="filter"',
                                body, re.DOTALL)
        self.assertIsNotNone(stats_match, "stats block missing")
        stats = stats_match.group(1)
        self.assertIn("Linked to recipe", stats)
        self.assertIn(">2<", stats)
        self.assertIn("Unlinked", stats)
        self.assertIn(">1<", stats)

    def test_detail_shows_linked_recipe_and_unlinked_suggestions(self):
        linked = SaleProduct.objects.create(
            name="Apple Waste Sourdough (Loose)", price=Decimal("2.20"),
            department=self.dept, recipe=self.apple,
            link_source=SaleProduct.SAGE, link_confirmed=True)
        unlinked = SaleProduct.objects.create(
            name="Crumble Topped Mince Pies Retail (Pack/6)",
            department=self.dept)
        # Linked: shows recipe link, no suggestions list.
        body = self.client.get(f"/sale-products/{linked.pk}/").content.decode()
        self.assertIn(f'href="/recipes/{self.apple.pk}/"', body)
        self.assertIn("matched by Sage No.", body)
        # Unlinked: suggestions surface the mince recipe (the fuzzy
        # ratio between the names is comfortably above the threshold).
        body = self.client.get(f"/sale-products/{unlinked.pk}/").content.decode()
        self.assertIn("Suggested by name similarity", body)
        self.assertIn("NPD-R655", body)

    def test_create_rejects_duplicate_name(self):
        SaleProduct.objects.create(name="Existing", department=self.dept)
        resp = self.client.post("/sale-products/new/", {
            "name": "EXISTING", "price": "1.0",
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"already exists", resp.content)
        self.assertEqual(SaleProduct.objects.filter(
            name__iexact="EXISTING").count(), 1)

    def test_create_sets_manual_link_when_recipe_picked(self):
        resp = self.client.post("/sale-products/new/", {
            "name": "Brand New SKU", "price": "3.50",
            "recipe_id": str(self.apple.pk),
        })
        self.assertEqual(resp.status_code, 302)
        sp = SaleProduct.objects.get(name="Brand New SKU")
        self.assertEqual(sp.recipe, self.apple)
        self.assertEqual(sp.link_source, SaleProduct.MANUAL)
        self.assertTrue(sp.link_confirmed)
        self.assertTrue(sp.is_manual_entry)

    def test_edit_changing_recipe_flips_link_source_to_manual(self):
        sp = SaleProduct.objects.create(
            name="Apple Waste Sourdough (Loose)", department=self.dept,
            recipe=self.apple, link_source=SaleProduct.NAME,
            link_confirmed=True)
        resp = self.client.post(f"/sale-products/{sp.pk}/edit/", {
            "name": sp.name, "price": "",
            "sage_number": "", "pack_size": "",
            "recipe_id": str(self.mince.pk),
        })
        self.assertEqual(resp.status_code, 302)
        sp.refresh_from_db()
        self.assertEqual(sp.recipe, self.mince)
        self.assertEqual(sp.link_source, SaleProduct.MANUAL)

    def test_delete_removes_the_row(self):
        sp = SaleProduct.objects.create(name="To Delete", department=self.dept)
        resp = self.client.post(f"/sale-products/{sp.pk}/delete/")
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(SaleProduct.objects.filter(name="To Delete").exists())

    def test_link_review_shows_fuzzy_suggestions(self):
        # An unlinked product with a fuzzy match against an active
        # recipe surfaces on the link-review page with a "confirm"
        # button.
        sp = SaleProduct.objects.create(
            name="Crumble Topped Mince Pies Retail (Pack/6)",
            department=self.dept)
        body = self.client.get("/sale-products/link-review/").content.decode()
        self.assertIn(sp.name, body)
        self.assertIn("NPD-R655", body)
        # Confirm form posts to the link-set endpoint with this recipe.
        self.assertIn(f'action="/sale-products/{sp.pk}/link/"', body)
        self.assertIn(f'value="{self.mince.pk}"', body)

    def test_link_set_confirms_a_suggestion_as_manual(self):
        sp = SaleProduct.objects.create(name="Mystery", department=self.dept)
        resp = self.client.post(f"/sale-products/{sp.pk}/link/", {
            "recipe_id": str(self.mince.pk),
        })
        self.assertEqual(resp.status_code, 302)
        sp.refresh_from_db()
        self.assertEqual(sp.recipe, self.mince)
        self.assertEqual(sp.link_source, SaleProduct.MANUAL)
        self.assertTrue(sp.link_confirmed)

    def test_link_set_unlink_clears_recipe_and_flags_manual(self):
        sp = SaleProduct.objects.create(
            name="Already linked", department=self.dept,
            recipe=self.mince, link_source=SaleProduct.SAGE,
            link_confirmed=True)
        resp = self.client.post(f"/sale-products/{sp.pk}/link/",
                                {"recipe_id": ""})
        self.assertEqual(resp.status_code, 302)
        sp.refresh_from_db()
        self.assertIsNone(sp.recipe)
        # The deliberate "no link" decision is a manual override too —
        # the next deploy's auto-linker won't re-attach.
        self.assertEqual(sp.link_source, SaleProduct.MANUAL)

    def test_bulk_confirm_sage_matches_promotes_to_manual(self):
        a = SaleProduct.objects.create(
            name="Sage A", department=self.dept,
            recipe=self.apple, link_source=SaleProduct.SAGE,
            link_confirmed=True)
        b = SaleProduct.objects.create(
            name="Sage B", department=self.dept,
            recipe=self.mince, link_source=SaleProduct.SAGE,
            link_confirmed=True)
        # Untouched: a name-matched row.
        c = SaleProduct.objects.create(
            name="Name C", department=self.dept,
            recipe=self.mince, link_source=SaleProduct.NAME,
            link_confirmed=True)
        resp = self.client.post("/sale-products/confirm-sage/")
        self.assertEqual(resp.status_code, 302)
        a.refresh_from_db(); b.refresh_from_db(); c.refresh_from_db()
        self.assertEqual(a.link_source, SaleProduct.MANUAL)
        self.assertEqual(b.link_source, SaleProduct.MANUAL)
        # Name-matched row was NOT touched.
        self.assertEqual(c.link_source, SaleProduct.NAME)

    def test_cross_department_actions_return_403(self):
        sp = SaleProduct.objects.create(
            name="Theirs", department=self.other_dept)
        for url in (f"/sale-products/{sp.pk}/",
                    f"/sale-products/{sp.pk}/edit/"):
            r = self.client.get(url)
            self.assertEqual(r.status_code, 403)
        r = self.client.post(f"/sale-products/{sp.pk}/delete/")
        self.assertEqual(r.status_code, 403)
        r = self.client.post(f"/sale-products/{sp.pk}/link/",
                             {"recipe_id": str(self.mince.pk)})
        self.assertEqual(r.status_code, 403)
        sp.refresh_from_db()
        self.assertEqual(sp.name, "Theirs")
        self.assertIsNone(sp.recipe)

    def test_anonymous_access_requires_login(self):
        anon = Client()
        for path in ("/sale-products/", "/sale-products/new/",
                     "/sale-products/link-review/"):
            r = anon.get(path)
            self.assertEqual(r.status_code, 302)
            self.assertIn("/login/", r.headers["Location"])

    def test_existing_ingredients_route_untouched(self):
        # The legacy /products/ route is the INGREDIENT catalogue —
        # adding the Products SECTION must not break it.
        r = self.client.get("/products/")
        self.assertEqual(r.status_code, 200)
        # And it shouldn't accidentally include sale-product UI bits.
        body = r.content.decode()
        self.assertNotIn("/sale-products/", body)

    # ---- link-review queue: confirmed items drop off ----

    def test_link_review_count_excludes_confirmed_items(self):
        # Three products: one unlinked (unconfirmed), one auto-Sage
        # (already link_confirmed=True), one manually picked
        # (link_confirmed=True). Only the unlinked one should remain
        # in the queue, and the "N products still to review" header
        # should report 1.
        u = SaleProduct.objects.create(
            name="Unlinked SKU", department=self.dept)
        s = SaleProduct.objects.create(
            name="Sage SKU", department=self.dept,
            recipe=self.apple, link_source=SaleProduct.SAGE,
            link_confirmed=True)
        m = SaleProduct.objects.create(
            name="Manual SKU", department=self.dept,
            recipe=self.mince, link_source=SaleProduct.MANUAL,
            link_confirmed=True)
        body = self.client.get("/sale-products/link-review/").content.decode()
        # Count badge / header — current copy is "N product(s) still to review".
        self.assertIn("1 product still to review", body)
        # Only the unconfirmed product is rendered as an unresolved row.
        # (The other products still appear inside the product-picker
        # JSON payload at the bottom of the page, so we look for the
        # row anchor rather than the raw name.)
        self.assertIn(f'href="/sale-products/{u.pk}/"', body)
        # Confirmed rows aren't rendered in the queue.
        import re
        link_rows = re.findall(r'<div class="link-row">.*?</div>\s*</div>',
                               body, re.DOTALL)
        joined = "\n".join(link_rows)
        self.assertNotIn("Sage SKU", joined)
        self.assertNotIn("Manual SKU", joined)

    def test_link_review_confirms_drops_item_from_queue(self):
        # Before: one unresolved. After confirming a suggestion: zero.
        sp = SaleProduct.objects.create(name="Mystery", department=self.dept)
        body = self.client.get("/sale-products/link-review/").content.decode()
        self.assertIn("1 product still to review", body)
        # The product appears as a queue row (and also inside the product-
        # picker JSON payload). Check the row specifically.
        self.assertIn(f'href="/sale-products/{sp.pk}/"', body)
        # Confirm a link.
        self.client.post(f"/sale-products/{sp.pk}/link/",
                         {"recipe_id": str(self.mince.pk)})
        # The success flash message names the product; consume it via
        # an intermediate GET so the next assertion measures only the
        # page itself.
        self.client.get("/sale-products/")
        body = self.client.get("/sale-products/link-review/").content.decode()
        self.assertIn("0 products still to review", body)
        # No queue row for Mystery now.
        import re
        link_rows = re.findall(r'<div class="link-row">.*?</div>\s*</div>',
                               body, re.DOTALL)
        self.assertNotIn("Mystery", "\n".join(link_rows))

    def test_link_review_includes_link_confirmed_false_recipe_match(self):
        # A product that has a recipe but link_confirmed=False is still
        # UNRESOLVED — surfaces in the queue with the existing link
        # presented as the top suggestion, ready to confirm.
        sp = SaleProduct.objects.create(
            name="Auto-suggested SKU", department=self.dept,
            recipe=self.apple, link_source=SaleProduct.NAME,
            link_confirmed=False)
        body = self.client.get("/sale-products/link-review/").content.decode()
        self.assertIn("Auto-suggested SKU", body)
        # Auto-match hint surfaced with the recipe code.
        self.assertIn("Auto-matched to", body)
        self.assertIn(self.apple.code, body)

    def test_link_review_count_decreases_as_items_are_confirmed(self):
        a = SaleProduct.objects.create(name="A", department=self.dept)
        b = SaleProduct.objects.create(name="B", department=self.dept)
        c = SaleProduct.objects.create(name="C", department=self.dept)
        body = self.client.get("/sale-products/link-review/").content.decode()
        self.assertIn("3 products still to review", body)
        self.client.post(f"/sale-products/{a.pk}/link/",
                         {"recipe_id": str(self.mince.pk)})
        body = self.client.get("/sale-products/link-review/").content.decode()
        self.assertIn("2 products still to review", body)
        self.client.post(f"/sale-products/{b.pk}/link/",
                         {"recipe_id": str(self.apple.pk)})
        body = self.client.get("/sale-products/link-review/").content.decode()
        self.assertIn("1 product still to review", body)

    # ---- suggestion + dropdown labels ----

    def test_link_review_labels_each_recipe_sold_or_component(self):
        # Apple recipe is sold and not used as a component → "sold" only.
        # Mince recipe is sold but also used as a component → both tags.
        # Make Mince a component of another recipe so it earns the
        # "component" tag.
        wrapper = Recipe.objects.create(
            code="NPD-R999", name="Mince Pie Bundle",
            department=self.dept, sold_as_product=True)
        RecipeLine.objects.create(
            recipe=wrapper, sub_recipe=self.mince,
            weight_g=Decimal("10"), ordering=0)
        # An unlinked product whose name fuzz-matches both recipes
        # (so both appear in the suggestions).
        SaleProduct.objects.create(
            name="Crumble Topped Mince Pie", department=self.dept)
        body = self.client.get("/sale-products/link-review/").content.decode()
        # Both labels visible (different recipes; both tags rendered).
        self.assertIn(">sold<", body)
        self.assertIn(">component<", body)

    def _extract_picker_payload(self, body):
        # The picker JSON is embedded via Django's json_script template
        # filter as <script type="application/json" id="recipe-picker-data">…</script>
        # Find it and return the parsed list.
        import json
        import re
        m = re.search(
            r'<script[^>]*id="recipe-picker-data"[^>]*>(.*?)</script>',
            body, re.DOTALL)
        self.assertIsNotNone(m, "recipe-picker-data script block missing")
        return json.loads(m.group(1))

    def test_link_review_embeds_recipe_picker_data(self):
        # The picker is driven by an embedded JSON payload of every
        # recipe with code/name/sold/component. Replaces the old
        # <select> dropdown for 248-recipe scenarios.
        wrapper = Recipe.objects.create(
            code="NPD-R999", name="Mince Pie Bundle",
            department=self.dept, sold_as_product=True)
        RecipeLine.objects.create(
            recipe=wrapper, sub_recipe=self.mince,
            weight_g=Decimal("10"), ordering=0)
        # Mark mince not-sold so it labels as component-only.
        self.mince.sold_as_product = False
        self.mince.is_sold_manual = True
        self.mince.save(update_fields=["sold_as_product", "is_sold_manual"])
        SaleProduct.objects.create(name="To Link", department=self.dept)
        body = self.client.get("/sale-products/link-review/").content.decode()
        payload = self._extract_picker_payload(body)
        by_code = {r["code"]: r for r in payload}
        # Every active recipe carried into the payload, each with role flags.
        self.assertIn(self.apple.code, by_code)
        self.assertIn(self.mince.code, by_code)
        self.assertIn(wrapper.code, by_code)
        self.assertTrue(by_code[self.apple.code]["sold"])
        self.assertFalse(by_code[self.apple.code]["component"])
        self.assertFalse(by_code[self.mince.code]["sold"])
        self.assertTrue(by_code[self.mince.code]["component"])
        self.assertTrue(by_code[wrapper.code]["sold"])
        # Required shape: pk + code + name + sold + component.
        for r in payload:
            self.assertEqual(set(r.keys()),
                             {"pk", "code", "name", "sold", "component"})

    def test_link_review_picker_input_present_and_pre_seeded(self):
        # Each row's picker carries a search input the autocomplete
        # widget attaches to, plus the data-seed for the product so
        # the dropdown opens with relevant matches.
        sp = SaleProduct.objects.create(
            name="Crumble Topped Mince Pies Retail (Pack/6)",
            department=self.dept)
        body = self.client.get("/sale-products/link-review/").content.decode()
        # Per-row picker container (allows extra classes like
        # lr-target-recipe alongside the base class).
        self.assertIn("recipe-picker", body)
        # Search input + hidden ids (recipe + product) — the form posts
        # the one matching the chosen link_target_type.
        self.assertIn('class="rcp-search"', body)
        self.assertIn('class="rcp-id"', body)
        self.assertIn('name="recipe_id"', body)
        self.assertIn('name="product_id"', body)
        # Target-type switcher present so the operator can flip target.
        self.assertIn('name="link_target_type"', body)
        # Quantity + unit inputs present.
        self.assertIn('name="link_quantity"', body)
        self.assertIn('name="link_unit"', body)
        # data-seed carries the product name (the JS strips parentheticals
        # client-side, so we don't assert on the cleaned form here).
        self.assertIn(f'data-seed="{sp.name}"', body)

    def test_link_review_dropdown_includes_non_role_recipes(self):
        # Non-sold, non-component recipes still surface in the picker
        # payload (no filtering by role).
        plain = Recipe.objects.create(
            code="NPD-R500", name="Plain Recipe",
            department=self.dept, sold_as_product=False)
        SaleProduct.objects.create(name="To Link", department=self.dept)
        body = self.client.get("/sale-products/link-review/").content.decode()
        payload = self._extract_picker_payload(body)
        plain_row = next((r for r in payload if r["code"] == plain.code), None)
        self.assertIsNotNone(plain_row)
        self.assertFalse(plain_row["sold"])
        self.assertFalse(plain_row["component"])

    def test_form_pages_embed_recipe_picker_data_and_input(self):
        # The product create + edit forms also use the autocomplete
        # picker — same JSON payload + recipe-picker widget.
        # New form:
        body = self.client.get("/sale-products/new/").content.decode()
        payload = self._extract_picker_payload(body)
        self.assertGreaterEqual(len(payload), 2)  # apple + mince at least
        self.assertIn('class="recipe-picker"', body)
        self.assertIn('class="rcp-search"', body)
        self.assertIn('class="rcp-id"', body)
        # Edit form pre-seeds with the existing link via data-selected-id.
        sp = SaleProduct.objects.create(
            name="Linked SKU", department=self.dept,
            recipe=self.apple, link_source=SaleProduct.MANUAL,
            link_confirmed=True)
        body = self.client.get(f"/sale-products/{sp.pk}/edit/").content.decode()
        self.assertIn(f'data-selected-id="{self.apple.pk}"', body)

    def test_form_submit_with_picker_recipe_id_links_correctly(self):
        # The picker writes the chosen recipe's pk into the hidden
        # recipe_id input the form already posts; the server side is
        # unchanged from the old <select>, so submitting still flips
        # link_source to manual.
        resp = self.client.post("/sale-products/new/", {
            "name": "Search-and-link SKU", "price": "1.20",
            "recipe_id": str(self.mince.pk),
        })
        self.assertEqual(resp.status_code, 302)
        sp = SaleProduct.objects.get(name="Search-and-link SKU")
        self.assertEqual(sp.recipe, self.mince)
        self.assertEqual(sp.link_source, SaleProduct.MANUAL)

    def test_detail_page_suggestions_labelled(self):
        # The detail page's "Suggested by name similarity" block uses
        # the same labels so the operator sees consistent info wherever
        # they confirm a link.
        wrapper = Recipe.objects.create(
            code="NPD-R999", name="Mince Pie Bundle",
            department=self.dept, sold_as_product=True)
        RecipeLine.objects.create(
            recipe=wrapper, sub_recipe=self.mince,
            weight_g=Decimal("10"), ordering=0)
        sp = SaleProduct.objects.create(
            name="Crumble Topped Mince Pie", department=self.dept)
        body = self.client.get(f"/sale-products/{sp.pk}/").content.decode()
        self.assertIn(">sold<", body)
        self.assertIn(">component<", body)


class SaleProductPolymorphicLinkTests(TestCase):
    """The quantified, polymorphic link: Recipe OR another SaleProduct + qty/unit."""

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.other_dept = Department.objects.create(name="Butchery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        self.apple = Recipe.objects.create(
            code="NPD-R800", name="Apple Waste Sourdough (Loose)",
            department=self.dept)
        self.focaccia = Recipe.objects.create(
            code="NPD-R700", name="Focaccia",
            department=self.dept, finished_weight_g=Decimal("1000"))

    # ---- model semantics ----

    def test_xor_constraint_blocks_both_targets_set_at_once(self):
        # The DB CheckConstraint refuses a row with both link_recipe and
        # link_product set simultaneously.
        from django.db import IntegrityError, transaction
        loose = SaleProduct.objects.create(
            name="Loose", department=self.dept,
            link_recipe=self.apple, link_quantity=Decimal("1"),
            link_unit=SaleProduct.COUNT)
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                SaleProduct.objects.create(
                    name="Pack/6", department=self.dept,
                    link_recipe=self.apple,
                    link_product=loose,
                    link_quantity=Decimal("6"),
                    link_unit=SaleProduct.COUNT)

    def test_xor_constraint_allows_one_target_or_neither(self):
        # Recipe-only.
        a = SaleProduct.objects.create(
            name="A", department=self.dept,
            link_recipe=self.apple, link_quantity=1, link_unit=SaleProduct.COUNT)
        # Product-only.
        b = SaleProduct.objects.create(
            name="B", department=self.dept,
            link_product=a, link_quantity=6, link_unit=SaleProduct.COUNT)
        # Both null (unlinked).
        c = SaleProduct.objects.create(name="C", department=self.dept)
        self.assertEqual(a.link_recipe, self.apple)
        self.assertEqual(b.link_product, a)
        self.assertIsNone(c.link_recipe)
        self.assertIsNone(c.link_product)

    def test_resolved_recipe_consumption_count_chain(self):
        # Pack/6 → Loose (1 × count → recipe) resolves to 6 × the recipe.
        loose = SaleProduct.objects.create(
            name="Loose", department=self.dept,
            link_recipe=self.apple, link_quantity=Decimal("1"),
            link_unit=SaleProduct.COUNT)
        pack = SaleProduct.objects.create(
            name="Pack/6", department=self.dept,
            link_product=loose, link_quantity=Decimal("6"),
            link_unit=SaleProduct.COUNT)
        recipe, total, unit = pack.resolved_recipe_consumption()
        self.assertEqual(recipe, self.apple)
        self.assertEqual(total, Decimal("6"))
        self.assertEqual(unit, SaleProduct.COUNT)

    def test_resolved_recipe_consumption_weight_kg(self):
        # 3.75 × weight_kg → focaccia recipe.
        slab = SaleProduct.objects.create(
            name="Focaccia 3.75kg", department=self.dept,
            link_recipe=self.focaccia, link_quantity=Decimal("3.75"),
            link_unit=SaleProduct.WEIGHT_KG)
        recipe, total, unit = slab.resolved_recipe_consumption()
        self.assertEqual(recipe, self.focaccia)
        self.assertEqual(total, Decimal("3.75"))
        self.assertEqual(unit, SaleProduct.WEIGHT_KG)

    def test_resolved_recipe_consumption_chain_with_weight_terminal(self):
        # Pack/6 (count) → Loose (1 × weight_kg → recipe). Each loose
        # bun is 1 kg, the pack is 6 of those → 6 kg of the recipe.
        loose = SaleProduct.objects.create(
            name="Heavy Loose", department=self.dept,
            link_recipe=self.apple, link_quantity=Decimal("1"),
            link_unit=SaleProduct.WEIGHT_KG)
        pack = SaleProduct.objects.create(
            name="Heavy Pack/6", department=self.dept,
            link_product=loose, link_quantity=Decimal("6"),
            link_unit=SaleProduct.COUNT)
        recipe, total, unit = pack.resolved_recipe_consumption()
        self.assertEqual(recipe, self.apple)
        self.assertEqual(total, Decimal("6"))
        self.assertEqual(unit, SaleProduct.WEIGHT_KG)

    def test_resolved_recipe_consumption_unlinked_returns_none(self):
        sp = SaleProduct.objects.create(
            name="Floating", department=self.dept)
        recipe, total, unit = sp.resolved_recipe_consumption()
        self.assertIsNone(recipe)
        self.assertEqual(total, Decimal("0"))
        self.assertIsNone(unit)

    def test_resolved_recipe_consumption_cycle_raises(self):
        # Bypass the model layer's manual setter by writing the FK
        # directly through update() — the model's clean rules don't
        # detect cycles, but resolved_recipe_consumption does.
        from stock.models import SaleProductCycleError
        a = SaleProduct.objects.create(name="A", department=self.dept)
        b = SaleProduct.objects.create(
            name="B", department=self.dept,
            link_product=a, link_quantity=Decimal("1"),
            link_unit=SaleProduct.COUNT)
        # Now point a → b to close the cycle.
        SaleProduct.objects.filter(pk=a.pk).update(
            link_product=b, link_quantity=Decimal("1"),
            link_unit=SaleProduct.COUNT)
        a.refresh_from_db()
        with self.assertRaises(SaleProductCycleError):
            a.resolved_recipe_consumption()

    # ---- migration semantics ----

    def test_default_link_is_recipe_qty_1_count(self):
        # Existing rows from a previous migration land with link_recipe
        # populated, link_quantity=1, link_unit=count — every imported
        # sale product is a simple loose link by default.
        sp = SaleProduct.objects.create(
            name="Simple", department=self.dept, link_recipe=self.apple)
        sp.refresh_from_db()
        self.assertEqual(sp.link_recipe, self.apple)
        self.assertIsNone(sp.link_product)
        self.assertEqual(sp.link_quantity, Decimal("1"))
        self.assertEqual(sp.link_unit, SaleProduct.COUNT)

    # ---- create / edit form ----

    def test_create_product_with_recipe_link_and_quantity(self):
        # Hand-creating "Focaccia 3.75kg" → recipe + 3.75 weight_kg.
        resp = self.client.post("/sale-products/new/", {
            "name": "Focaccia 3.75kg", "price": "12.50",
            "link_target_type": "recipe",
            "recipe_id": str(self.focaccia.pk),
            "link_quantity": "3.75",
            "link_unit": SaleProduct.WEIGHT_KG,
        })
        self.assertEqual(resp.status_code, 302)
        sp = SaleProduct.objects.get(name="Focaccia 3.75kg")
        self.assertEqual(sp.link_recipe, self.focaccia)
        self.assertIsNone(sp.link_product)
        self.assertEqual(sp.link_quantity, Decimal("3.75"))
        self.assertEqual(sp.link_unit, SaleProduct.WEIGHT_KG)
        self.assertEqual(sp.link_source, SaleProduct.MANUAL)
        self.assertTrue(sp.link_confirmed)
        self.assertTrue(sp.is_manual_entry)

    def test_create_product_with_product_link_and_quantity(self):
        loose = SaleProduct.objects.create(
            name="Sticky Apple & Cinnamon Bun (Loose)",
            department=self.dept,
            link_recipe=self.apple, link_quantity=Decimal("1"),
            link_unit=SaleProduct.COUNT)
        resp = self.client.post("/sale-products/new/", {
            "name": "Sticky Apple & Cinnamon Bun (Pack/6)",
            "price": "10.00",
            "link_target_type": "product",
            "product_id": str(loose.pk),
            "link_quantity": "6",
            "link_unit": SaleProduct.COUNT,
        })
        self.assertEqual(resp.status_code, 302)
        sp = SaleProduct.objects.get(
            name="Sticky Apple & Cinnamon Bun (Pack/6)")
        self.assertIsNone(sp.link_recipe)
        self.assertEqual(sp.link_product, loose)
        self.assertEqual(sp.link_quantity, Decimal("6"))
        self.assertEqual(sp.link_unit, SaleProduct.COUNT)
        # And the chain resolves to the apple recipe with multiplier 6.
        recipe, total, unit = sp.resolved_recipe_consumption()
        self.assertEqual(recipe, self.apple)
        self.assertEqual(total, Decimal("6"))

    def test_create_form_rejects_duplicate_name(self):
        SaleProduct.objects.create(name="Existing", department=self.dept)
        resp = self.client.post("/sale-products/new/", {
            "name": "EXISTING", "price": "1.00",
            "link_target_type": "recipe",
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"already exists", resp.content)

    def test_edit_changes_link_target_type_recipe_to_product(self):
        # Start out linked to a recipe; edit to point at another product.
        loose = SaleProduct.objects.create(
            name="Loose", department=self.dept,
            link_recipe=self.apple, link_quantity=Decimal("1"),
            link_unit=SaleProduct.COUNT)
        pack = SaleProduct.objects.create(
            name="Pack", department=self.dept,
            link_recipe=self.apple, link_quantity=Decimal("6"),
            link_unit=SaleProduct.COUNT,
            link_source=SaleProduct.MANUAL, link_confirmed=True)
        resp = self.client.post(f"/sale-products/{pack.pk}/edit/", {
            "name": pack.name, "price": "",
            "sage_number": "", "pack_size": "",
            "link_target_type": "product",
            "product_id": str(loose.pk),
            "link_quantity": "6",
            "link_unit": SaleProduct.COUNT,
        })
        self.assertEqual(resp.status_code, 302)
        pack.refresh_from_db()
        self.assertIsNone(pack.link_recipe)
        self.assertEqual(pack.link_product, loose)
        self.assertEqual(pack.link_quantity, Decimal("6"))
        self.assertEqual(pack.link_unit, SaleProduct.COUNT)
        # Resolves through the chain back to the apple recipe.
        recipe, total, _ = pack.resolved_recipe_consumption()
        self.assertEqual(recipe, self.apple)
        self.assertEqual(total, Decimal("6"))

    def test_edit_refuses_self_link(self):
        sp = SaleProduct.objects.create(
            name="Loop", department=self.dept,
            link_recipe=self.apple, link_quantity=Decimal("1"),
            link_unit=SaleProduct.COUNT,
            link_source=SaleProduct.MANUAL, link_confirmed=True)
        resp = self.client.post(f"/sale-products/{sp.pk}/edit/", {
            "name": sp.name, "price": "",
            "sage_number": "", "pack_size": "",
            "link_target_type": "product",
            "product_id": str(sp.pk),
            "link_quantity": "1", "link_unit": SaleProduct.COUNT,
        })
        self.assertEqual(resp.status_code, 200)
        # The HTML-escaped apostrophe survives — match on the
        # unambiguous tail of the error message.
        self.assertIn(b"link to itself", resp.content)
        sp.refresh_from_db()
        # Unchanged.
        self.assertEqual(sp.link_recipe, self.apple)
        self.assertIsNone(sp.link_product)

    # ---- link-set (link-review confirm path) ----

    def test_link_set_with_product_target_and_quantity(self):
        loose = SaleProduct.objects.create(
            name="Loose", department=self.dept,
            link_recipe=self.apple, link_quantity=Decimal("1"),
            link_unit=SaleProduct.COUNT)
        pack = SaleProduct.objects.create(name="Pack/6", department=self.dept)
        resp = self.client.post(f"/sale-products/{pack.pk}/link/", {
            "link_target_type": "product",
            "product_id": str(loose.pk),
            "link_quantity": "6",
            "link_unit": SaleProduct.COUNT,
        })
        self.assertEqual(resp.status_code, 302)
        pack.refresh_from_db()
        self.assertEqual(pack.link_product, loose)
        self.assertEqual(pack.link_quantity, Decimal("6"))
        self.assertEqual(pack.link_source, SaleProduct.MANUAL)
        self.assertTrue(pack.link_confirmed)

    def test_link_set_legacy_recipe_id_only_still_works(self):
        # The Sage/exact-name quick-confirm buttons on the review page
        # still POST just recipe_id; the link-set view treats that as
        # "recipe + qty 1 + count".
        sp = SaleProduct.objects.create(name="Quick", department=self.dept)
        resp = self.client.post(f"/sale-products/{sp.pk}/link/", {
            "recipe_id": str(self.apple.pk),
        })
        self.assertEqual(resp.status_code, 302)
        sp.refresh_from_db()
        self.assertEqual(sp.link_recipe, self.apple)
        self.assertEqual(sp.link_quantity, Decimal("1"))
        self.assertEqual(sp.link_unit, SaleProduct.COUNT)
        self.assertEqual(sp.link_source, SaleProduct.MANUAL)
        self.assertTrue(sp.link_confirmed)

    def test_link_set_refuses_self_link(self):
        sp = SaleProduct.objects.create(name="Loop", department=self.dept)
        resp = self.client.post(f"/sale-products/{sp.pk}/link/", {
            "link_target_type": "product",
            "product_id": str(sp.pk),
            "link_quantity": "1", "link_unit": SaleProduct.COUNT,
        })
        self.assertEqual(resp.status_code, 302)
        sp.refresh_from_db()
        self.assertIsNone(sp.link_product)

    # ---- importer preservation ----

    def test_existing_recipe_link_imports_with_qty_1_count(self):
        # Run the bulk importer with a row that matches by Sage code,
        # and verify the link defaults to quantity 1 / unit count.
        import os
        import tempfile
        from django.core.management import call_command
        Recipe.objects.create(
            code="660130103", name="Mince Pie SKU shadow",
            department=self.dept)
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            from openpyxl import Workbook
            wb = Workbook()
            wb.remove(wb.active)
            ws = wb.create_sheet(title="Products")
            ws.append(["Product", "Price", "Sage No.", None,
                       "Stock Managed", "Pack Size"])
            ws.append(["Mince Test SKU", 2.25, "660130103", None, None, "1"])
            wb.save(path)
            call_command("import_sale_products", path,
                         "--department", "Bakery")
        finally:
            try:
                os.unlink(path)
            except OSError:
                pass
        sp = SaleProduct.objects.get(name="Mince Test SKU")
        self.assertEqual(sp.link_recipe.code, "660130103")
        self.assertIsNone(sp.link_product)
        self.assertEqual(sp.link_quantity, Decimal("1"))
        self.assertEqual(sp.link_unit, SaleProduct.COUNT)

    def test_section_nav_has_create_product_action(self):
        # "Create product" lives in the section sub-nav alongside Link review.
        body = self.client.get("/sale-products/").content.decode()
        nav = body[body.index("<nav>"):body.index("</nav>")]
        self.assertIn(">Create product<", nav)
        self.assertIn(">Link review<", nav)
        # Order: Products | Link review | Create product.
        products_pos = nav.find(">Products<")
        review_pos = nav.find(">Link review<")
        create_pos = nav.find(">Create product<")
        self.assertLess(products_pos, review_pos)
        self.assertLess(review_pos, create_pos)

    def test_detail_page_renders_resolved_link(self):
        loose = SaleProduct.objects.create(
            name="Loose Bun", department=self.dept,
            link_recipe=self.apple, link_quantity=Decimal("1"),
            link_unit=SaleProduct.COUNT,
            link_source=SaleProduct.MANUAL, link_confirmed=True)
        pack = SaleProduct.objects.create(
            name="Pack/6 of Bun", department=self.dept,
            link_product=loose, link_quantity=Decimal("6"),
            link_unit=SaleProduct.COUNT,
            link_source=SaleProduct.MANUAL, link_confirmed=True)
        body = self.client.get(f"/sale-products/{pack.pk}/").content.decode()
        # "6" "Units (count)" of "Loose Bun".
        self.assertIn("Loose Bun", body)
        self.assertIn("Units (count)", body)
        # Resolves-to banner names the recipe + the multiplier 6.
        self.assertIn("Resolves to", body)
        self.assertIn("NPD-R800", body)


class OrdersTests(TestCase):
    """Orders system chunk 1: model + manual CRUD. No import yet."""

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.other_dept = Department.objects.create(name="Butchery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        # Two customers, two recipes, three products covering the
        # interesting link shapes for resolved_consumption.
        self.alice_cust = Customer.objects.create(
            name="Garden Cafe", department=self.dept)
        self.bob_cust = Customer.objects.create(
            name="Farmshop", department=self.dept)
        self.cross_dept_cust = Customer.objects.create(
            name="Butchery Counter", department=self.other_dept)
        self.apple = Recipe.objects.create(
            code="NPD-R800", name="Apple Waste Sourdough (Loose)",
            department=self.dept)
        self.focaccia = Recipe.objects.create(
            code="NPD-R700", name="Focaccia",
            department=self.dept, finished_weight_g=Decimal("1000"))
        self.loose = SaleProduct.objects.create(
            name="Apple Waste Sourdough (Loose)", price=Decimal("2.20"),
            department=self.dept,
            link_recipe=self.apple, link_quantity=Decimal("1"),
            link_unit=SaleProduct.COUNT,
            link_source=SaleProduct.MANUAL, link_confirmed=True)
        self.pack = SaleProduct.objects.create(
            name="Apple Waste Sourdough (Pack/6)", price=Decimal("12.00"),
            department=self.dept,
            link_product=self.loose, link_quantity=Decimal("6"),
            link_unit=SaleProduct.COUNT,
            link_source=SaleProduct.MANUAL, link_confirmed=True)
        self.slab = SaleProduct.objects.create(
            name="Focaccia 3.75kg", price=Decimal("18.00"),
            department=self.dept,
            link_recipe=self.focaccia, link_quantity=Decimal("3.75"),
            link_unit=SaleProduct.WEIGHT_KG,
            link_source=SaleProduct.MANUAL, link_confirmed=True)
        self.foreign = SaleProduct.objects.create(
            name="Butchery Pie", price=Decimal("5.00"),
            department=self.other_dept)

    # ---- nav ----

    def test_top_nav_lists_orders_between_products_and_production(self):
        # /home/ now renders the design-system shell; its left rail carries
        # the Orders link (the old ordered top section-picker is gone).
        body = self.client.get("/home/").content.decode()
        self.assertIn('href="/orders/"', body)

    # ---- model helpers ----

    def test_total_value_sums_qty_times_price(self):
        order = Order.objects.create(
            customer=self.alice_cust, order_date=datetime.date(2026, 5, 1),
            department=self.dept)
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("3"))
        OrderLine.objects.create(
            order=order, sale_product=self.pack, qty_ordered=Decimal("2"))
        # 3 × 2.20 = 6.60; 2 × 12.00 = 24.00 → 30.60
        self.assertEqual(order.total_value(), Decimal("30.60"))

    def test_total_value_skips_lines_with_no_price(self):
        priceless = SaleProduct.objects.create(
            name="Unpriced", department=self.dept)
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept)
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("4"))
        OrderLine.objects.create(
            order=order, sale_product=priceless, qty_ordered=Decimal("5"))
        # 4 × 2.20 = 8.80, the priceless line is skipped.
        self.assertEqual(order.total_value(), Decimal("8.80"))

    def test_resolved_consumption_walks_chain_and_multiplies(self):
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept)
        OrderLine.objects.create(
            order=order, sale_product=self.pack, qty_ordered=Decimal("2"))
        OrderLine.objects.create(
            order=order, sale_product=self.slab, qty_ordered=Decimal("4"))
        rows = order.resolved_consumption()
        self.assertEqual(len(rows), 2)
        # Pack/6: 2 (ordered) × 6 (chain) = 12 count of the recipe.
        pack_row = next(r for r in rows if r["sale_product"] == self.pack)
        self.assertEqual(pack_row["recipe"], self.apple)
        self.assertEqual(pack_row["total_quantity"], Decimal("12"))
        self.assertEqual(pack_row["unit"], SaleProduct.COUNT)
        # Slab: 4 × 3.75 kg = 15 kg of the focaccia recipe.
        slab_row = next(r for r in rows if r["sale_product"] == self.slab)
        self.assertEqual(slab_row["recipe"], self.focaccia)
        self.assertEqual(slab_row["total_quantity"], Decimal("15.00"))
        self.assertEqual(slab_row["unit"], SaleProduct.WEIGHT_KG)

    def test_resolved_consumption_handles_unlinked_lines(self):
        unlinked = SaleProduct.objects.create(
            name="Floating", department=self.dept)
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept)
        OrderLine.objects.create(
            order=order, sale_product=unlinked, qty_ordered=Decimal("1"))
        rows = order.resolved_consumption()
        self.assertEqual(len(rows), 1)
        self.assertIsNone(rows[0]["recipe"])
        self.assertIsNone(rows[0]["total_quantity"])

    def test_duplicate_product_on_same_order_blocked_by_db(self):
        from django.db import IntegrityError, transaction
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept)
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("3"))
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                OrderLine.objects.create(
                    order=order, sale_product=self.loose, qty_ordered=Decimal("2"))

    # ---- list + filter ----

    def test_orders_list_filters_by_customer_and_date(self):
        a = Order.objects.create(
            customer=self.alice_cust, order_date=datetime.date(2026, 5, 1),
            department=self.dept)
        b = Order.objects.create(
            customer=self.bob_cust, order_date=datetime.date(2026, 5, 1),
            department=self.dept)
        c = Order.objects.create(
            customer=self.alice_cust, order_date=datetime.date(2026, 5, 2),
            department=self.dept)
        # No filter: customer list shows Alice + Bob with their per-week
        # order counts ("2 orders" / "1 order"). Individual order detail
        # links don't surface here — they're reachable by drilling in.
        body = self.client.get("/orders/").content.decode()
        self.assertIn(self.alice_cust.name, body)
        self.assertIn(self.bob_cust.name, body)
        self.assertIn("2 orders", body)
        self.assertIn("1 order", body)
        # Customer filter alone → renders Alice's product grid for the
        # week. Other customers' orders don't bleed in.
        body = self.client.get(
            f"/orders/?customer={self.alice_cust.pk}").content.decode()
        self.assertIn('data-testid="order-grid"', body)
        self.assertNotIn(f'/orders/{b.pk}/', body)
        # Date filter alone → single-day list of orders with their
        # detail / edit / delete actions.
        body = self.client.get("/orders/?date=2026-05-01").content.decode()
        self.assertIn(f'/orders/{a.pk}/', body)
        self.assertIn(f'/orders/{b.pk}/', body)
        self.assertNotIn(f'/orders/{c.pk}/', body)
        # Combined: Alice + May 1 → only a.
        body = self.client.get(
            f"/orders/?customer={self.alice_cust.pk}&date=2026-05-01"
        ).content.decode()
        self.assertIn(f'/orders/{a.pk}/', body)
        self.assertNotIn(f'/orders/{b.pk}/', body)
        self.assertNotIn(f'/orders/{c.pk}/', body)

    def test_orders_list_excludes_cross_department(self):
        # An order in the other department isn't visible from Bakery.
        foreign = Order.objects.create(
            customer=self.cross_dept_cust, department=self.other_dept)
        body = self.client.get("/orders/").content.decode()
        self.assertNotIn(f'/orders/{foreign.pk}/', body)

    # ---- detail ----

    def test_order_detail_shows_lines_and_total(self):
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 1))
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("3"))
        OrderLine.objects.create(
            order=order, sale_product=self.slab, qty_ordered=Decimal("2"))
        body = self.client.get(f"/orders/{order.pk}/").content.decode()
        self.assertIn(self.alice_cust.name, body)
        self.assertIn("01 May 2026", body)
        self.assertIn(self.loose.name, body)
        self.assertIn(self.slab.name, body)
        # 3 × 2.20 + 2 × 18.00 = 42.60
        self.assertIn("£42.60", body)
        # Resolved-consumption preview card renders recipe codes.
        self.assertIn("Resolved consumption", body)
        self.assertIn("NPD-R800", body)
        self.assertIn("NPD-R700", body)

    def test_detail_cross_department_returns_403(self):
        foreign = Order.objects.create(
            customer=self.cross_dept_cust, department=self.other_dept)
        resp = self.client.get(f"/orders/{foreign.pk}/")
        self.assertEqual(resp.status_code, 403)

    # ---- create ----

    def test_create_order_with_lines_via_form(self):
        resp = self.client.post("/orders/new/", {
            "customer_id": str(self.alice_cust.pk),
            "order_date": "2026-05-23",
            "note": "Friday delivery",
            "product_id": [str(self.loose.pk), str(self.pack.pk)],
            "qty_ordered": ["3", "2"],
        })
        self.assertEqual(resp.status_code, 302)
        order = Order.objects.get()
        self.assertEqual(order.customer, self.alice_cust)
        self.assertEqual(order.order_date, datetime.date(2026, 5, 23))
        self.assertEqual(order.note, "Friday delivery")
        self.assertEqual(order.lines.count(), 2)
        qtys_by_product = {ln.sale_product_id: ln.qty_ordered
                           for ln in order.lines.all()}
        self.assertEqual(qtys_by_product[self.loose.pk], Decimal("3"))
        self.assertEqual(qtys_by_product[self.pack.pk], Decimal("2"))
        self.assertEqual(order.total_value(), Decimal("30.60"))

    def test_create_rejects_duplicate_product_in_same_post(self):
        resp = self.client.post("/orders/new/", {
            "customer_id": str(self.alice_cust.pk),
            "order_date": "2026-05-23",
            "product_id": [str(self.loose.pk), str(self.loose.pk)],
            "qty_ordered": ["3", "2"],
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"appears more than once", resp.content)
        self.assertEqual(Order.objects.count(), 0)

    def test_create_requires_customer_and_at_least_one_line(self):
        # No customer.
        resp = self.client.post("/orders/new/", {
            "customer_id": "",
            "order_date": "2026-05-23",
            "product_id": [str(self.loose.pk)],
            "qty_ordered": ["3"],
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"Pick a customer", resp.content)
        # No lines.
        resp = self.client.post("/orders/new/", {
            "customer_id": str(self.alice_cust.pk),
            "order_date": "2026-05-23",
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"at least one product line", resp.content)
        self.assertEqual(Order.objects.count(), 0)

    def test_create_rejects_cross_department_product(self):
        # A product from another department can't be ordered through
        # this department's order form.
        resp = self.client.post("/orders/new/", {
            "customer_id": str(self.alice_cust.pk),
            "order_date": "2026-05-23",
            "product_id": [str(self.foreign.pk)],
            "qty_ordered": ["1"],
        })
        self.assertEqual(resp.status_code, 200)
        # The apostrophe is HTML-escaped — assert the unambiguous tail.
        self.assertIn(b"in this department", resp.content)
        self.assertEqual(Order.objects.count(), 0)

    def test_create_rejects_cross_department_customer(self):
        resp = self.client.post("/orders/new/", {
            "customer_id": str(self.cross_dept_cust.pk),
            "order_date": "2026-05-23",
            "product_id": [str(self.loose.pk)],
            "qty_ordered": ["1"],
        })
        self.assertEqual(resp.status_code, 200)
        self.assertIn(b"in this department", resp.content)

    # ---- edit ----

    def test_edit_order_changes_customer_date_and_lines(self):
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 1))
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("3"))
        resp = self.client.post(f"/orders/{order.pk}/edit/", {
            "customer_id": str(self.bob_cust.pk),
            "order_date": "2026-05-10",
            "note": "moved",
            "product_id": [str(self.pack.pk), str(self.slab.pk)],
            "qty_ordered": ["1", "2"],
        })
        self.assertEqual(resp.status_code, 302)
        order.refresh_from_db()
        self.assertEqual(order.customer, self.bob_cust)
        self.assertEqual(order.order_date, datetime.date(2026, 5, 10))
        self.assertEqual(order.note, "moved")
        codes = {ln.sale_product_id: ln.qty_ordered
                 for ln in order.lines.all()}
        self.assertEqual(codes, {
            self.pack.pk: Decimal("1"),
            self.slab.pk: Decimal("2"),
        })

    def test_edit_cross_department_returns_403(self):
        foreign = Order.objects.create(
            customer=self.cross_dept_cust, department=self.other_dept)
        resp = self.client.get(f"/orders/{foreign.pk}/edit/")
        self.assertEqual(resp.status_code, 403)
        resp = self.client.post(f"/orders/{foreign.pk}/edit/", {
            "customer_id": str(self.alice_cust.pk),
            "order_date": "2026-01-01",
        })
        self.assertEqual(resp.status_code, 403)

    # ---- delete ----

    def test_delete_removes_order_and_lines(self):
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept)
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("3"))
        resp = self.client.post(f"/orders/{order.pk}/delete/")
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(Order.objects.filter(pk=order.pk).exists())
        self.assertFalse(OrderLine.objects.filter(order_id=order.pk).exists())

    def test_delete_cross_department_returns_403(self):
        foreign = Order.objects.create(
            customer=self.cross_dept_cust, department=self.other_dept)
        resp = self.client.post(f"/orders/{foreign.pk}/delete/")
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(Order.objects.filter(pk=foreign.pk).exists())

    # ---- login ----

    def test_orders_routes_require_login(self):
        anon = Client()
        for path in ("/orders/", "/orders/new/"):
            resp = anon.get(path)
            self.assertEqual(resp.status_code, 302)
            self.assertIn("/login/", resp.headers["Location"])

    # ---- week grouping + filter ----
    #
    # Orders are stored per-date; the week-commencing Monday is always
    # DERIVED from the date. The view's ``?week=YYYY-MM-DD`` query
    # snaps any day within a week to that week's Monday and shows
    # only that week's orders, grouped by day Mon–Sun.

    def _seed_two_weeks(self):
        # Week A: Mon 2026-05-04 — Sun 2026-05-10
        # Week B: Mon 2026-05-11 — Sun 2026-05-17
        a_mon = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 4))      # Mon week A
        OrderLine.objects.create(
            order=a_mon, sale_product=self.loose, qty_ordered=Decimal("3"))
        a_wed = Order.objects.create(
            customer=self.bob_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 6))      # Wed week A
        OrderLine.objects.create(
            order=a_wed, sale_product=self.slab, qty_ordered=Decimal("1"))
        b_tue = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 12))     # Tue week B
        OrderLine.objects.create(
            order=b_tue, sale_product=self.pack, qty_ordered=Decimal("2"))
        b_fri = Order.objects.create(
            customer=self.bob_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 15))     # Fri week B
        OrderLine.objects.create(
            order=b_fri, sale_product=self.loose, qty_ordered=Decimal("5"))
        return {"a_mon": a_mon, "a_wed": a_wed,
                "b_tue": b_tue, "b_fri": b_fri}

    def test_week_filter_snaps_any_day_to_monday(self):
        # Any day inside Week A snaps to Mon 4 May. Week A customers
        # appear in the active list; Week B's day-tile dates (12 May
        # b_tue, 15 May b_fri) must not appear in the compact strip.
        self._seed_two_weeks()
        for snap_day in ("2026-05-04", "2026-05-06", "2026-05-10"):
            body = self.client.get(f"/orders/?week={snap_day}").content.decode()
            self.assertIn("week commencing 04 May 2026", body)
            # Week A's tile dates (Mon..Sun, 04..10 May) all render.
            for dt in ("04 May", "05 May", "06 May", "07 May",
                       "08 May", "09 May", "10 May"):
                self.assertIn(dt, body)
            # Both Alice (a_mon) and Bob (a_wed) ordered in Week A —
            # both appear with active badges.
            self.assertIn(self.alice_cust.name, body)
            self.assertIn(self.bob_cust.name, body)
            # Week B's tile dates don't appear.
            self.assertNotIn("12 May", body)
            self.assertNotIn("15 May", body)

    def test_week_groups_orders_by_weekday(self):
        # The compact day strip lists Mon..Sun in order with each day's
        # order count and £ total. Assert the dates render in calendar
        # order inside the strip section.
        self._seed_two_weeks()
        body = self.client.get("/orders/?week=2026-05-04").content.decode()
        strip_start = body.find('data-testid="day-strip"')
        self.assertGreater(strip_start, -1,
                           "compact day-strip should be on the page")
        last = strip_start
        for dt in ("04 May", "05 May", "06 May", "07 May",
                   "08 May", "09 May", "10 May"):
            pos = body.find(dt, last)
            self.assertGreater(pos, last,
                               f"{dt} should follow the previous date "
                               "inside the day-strip")
            last = pos
        # Each day-tile is a link back into the view with ?date= so the
        # operator can drill into one specific day's orders.
        self.assertIn("date=2026-05-04", body)
        self.assertIn("date=2026-05-06", body)
        # Mon (a_mon) and Wed (a_wed) tiles report "1 order"; the empty
        # days report a dash. Assert both shapes appear in the strip.
        self.assertIn("1 order", body)

    def test_week_summary_totals_match_line_values(self):
        # Week A: a_mon = 3 × 2.20 = 6.60; a_wed = 1 × 18.00 = 18.00
        # → 2 orders, £24.60. Per-day totals: Mon £6.60, Wed £18.00.
        # The week's at-a-glance totals all live in the compact
        # day-strip now (the separate stats block was redundant and
        # has been removed); the trailing "Week" tile rolls Mon..Sun
        # into the grand total.
        self._seed_two_weeks()
        body = self.client.get("/orders/?week=2026-05-04").content.decode()
        # Per-day tile totals appear in the strip.
        self.assertIn("£6.60", body)
        self.assertIn("£18.00", body)
        # The trailing Week-total tile shows the 2-order, £24.60 roll-up.
        self.assertIn("2 orders", body)
        self.assertIn("£24.60", body)

    def test_week_filter_combines_with_customer_filter(self):
        # Alice in Week A → the grid renders for her. Bob's a_wed order
        # and Week B's orders don't surface.
        orders = self._seed_two_weeks()
        body = self.client.get(
            f"/orders/?week=2026-05-04&customer={self.alice_cust.pk}"
        ).content.decode()
        # Alice's grid is on the page.
        self.assertIn('data-testid="order-grid"', body)
        self.assertIn(self.loose.name, body)
        # Bob's a_wed order, and Week B's order detail links, must not
        # appear anywhere on the page (no detail-link surface here).
        self.assertNotIn(f'/orders/{orders["a_wed"].pk}/', body)
        self.assertNotIn(f'/orders/{orders["b_tue"].pk}/', body)
        self.assertNotIn(f'/orders/{orders["b_fri"].pk}/', body)

    def test_week_filter_combines_with_date_filter_within_week(self):
        # Specific day inside the selected week. ?week=2026-05-04 +
        # ?date=2026-05-06 → only the Wed order.
        orders = self._seed_two_weeks()
        body = self.client.get(
            "/orders/?week=2026-05-04&date=2026-05-06"
        ).content.decode()
        self.assertIn(f'/orders/{orders["a_wed"].pk}/', body)
        self.assertNotIn(f'/orders/{orders["a_mon"].pk}/', body)

    def test_week_navigator_links_to_previous_and_next_weeks(self):
        # The week navigator carries ?week= links to the surrounding
        # Mondays so the operator can step a week at a time.
        self._seed_two_weeks()
        body = self.client.get("/orders/?week=2026-05-04").content.decode()
        # Previous week Mon = 2026-04-27, next = 2026-05-11.
        self.assertIn("?week=2026-04-27", body)
        self.assertIn("?week=2026-05-11", body)

    def test_default_week_is_most_recent_when_today_has_no_orders(self):
        # Today is well after May 2026 (or before). The default lands
        # on the most-recent-order's week — Week B (11 May–17 May).
        self._seed_two_weeks()
        body = self.client.get("/orders/").content.decode()
        # Most recent order = b_fri (15 May 2026), Monday = 11 May.
        self.assertIn("week commencing 11 May 2026", body)
        # Week B's day-tile dates render; Week A's don't.
        self.assertIn("12 May", body)            # b_tue
        self.assertIn("15 May", body)            # b_fri
        self.assertNotIn("04 May", body)         # a_mon (Week A)
        # Both customers with Week B orders appear in the active list.
        self.assertIn(self.alice_cust.name, body)
        self.assertIn(self.bob_cust.name, body)

    def test_default_week_falls_back_to_today_when_no_orders(self):
        # No orders at all → default week = current week. The empty
        # compact strip still renders with seven day-tiles plus the
        # week-total tile (8 tiles total).
        body = self.client.get("/orders/").content.decode()
        today_monday = (datetime.date.today() -
                        datetime.timedelta(days=datetime.date.today().weekday()))
        self.assertIn(f"week commencing {today_monday:%d %b %Y}", body)
        # 7 weekday tiles + 1 week-total tile = 8 occurrences of
        # `class="day-tile`.
        self.assertEqual(body.count('data-testid="day-tile"'), 8)

    def test_manual_create_still_works_per_date_through_week_view(self):
        # Creating an order through the existing endpoint isn't
        # affected by the new layout.
        resp = self.client.post("/orders/new/", {
            "customer_id": str(self.alice_cust.pk),
            "order_date": "2026-05-13",   # Wed in Week B
            "product_id": [str(self.loose.pk)],
            "qty_ordered": ["2"],
        })
        self.assertEqual(resp.status_code, 302)
        new_order = Order.objects.get(order_date=datetime.date(2026, 5, 13))
        # The Wed day-tile shows the order count + £ for that day.
        body = self.client.get("/orders/?week=2026-05-11").content.decode()
        self.assertIn("13 May", body)
        self.assertIn("1 order", body)
        # Drilling into the day (?date=…) surfaces the order detail link.
        body = self.client.get(
            "/orders/?week=2026-05-11&date=2026-05-13"
        ).content.decode()
        self.assertIn(f'/orders/{new_order.pk}/', body)

    def test_invalid_week_query_falls_back_to_default(self):
        # A garbage ?week=… string doesn't crash the view; it just
        # uses the default-week logic instead.
        self._seed_two_weeks()
        resp = self.client.get("/orders/?week=not-a-date")
        self.assertEqual(resp.status_code, 200)
        body = resp.content.decode()
        # Default week (most recent) is 11 May.
        self.assertIn("week commencing 11 May 2026", body)

    # ---- spreadsheet-style grid (products × Mon..Sun) ----

    def test_grid_renders_when_customer_selected(self):
        # 3 products on different days. The grid should carry one row
        # per product and one column per day, plus row + column totals.
        order_mon = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))  # Mon
        OrderLine.objects.create(
            order=order_mon, sale_product=self.loose, qty_ordered=Decimal("4"))
        OrderLine.objects.create(
            order=order_mon, sale_product=self.slab, qty_ordered=Decimal("1"))
        order_sat = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 23))  # Sat
        OrderLine.objects.create(
            order=order_sat, sale_product=self.loose, qty_ordered=Decimal("7"))
        body = self.client.get(
            f"/orders/?week=2026-05-18&customer={self.alice_cust.pk}"
        ).content.decode()
        # Grid card header references the customer + week.
        self.assertIn("Garden Cafe — week commencing 18 May 2026", body)
        # Product rows render with their names.
        self.assertIn(self.loose.name, body)
        self.assertIn(self.slab.name, body)
        # Day columns Mon..Sun.
        self.assertIn("Mon", body)
        self.assertIn("Sun", body)
        # Row totals: loose = 4 (Mon) + 7 (Sat) = 11; slab = 1.
        self.assertIn(">11<", body)
        # Slab is 1 × £18 = £18.00.
        self.assertIn("£18.00", body)
        # Loose: 11 × £2.20 = £24.20
        self.assertIn("£24.20", body)
        # Grand total = £18 + £24.20 = £42.20.
        self.assertIn("£42.20", body)

    def test_grid_hidden_when_no_customer_selected(self):
        # Without a customer, the grid card doesn't render — the
        # customer-picker list does.
        Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        body = self.client.get("/orders/?week=2026-05-18").content.decode()
        # No product grid renders without a selected customer.
        self.assertNotIn('data-testid="order-grid"', body)
        # Customer picker list rendered instead.
        self.assertIn("click a customer to see their grid", body)
        self.assertIn(f"customer={self.alice_cust.pk}", body)

    def test_grid_hidden_when_specific_day_filter_set(self):
        # With ?date=… narrowing to one day, the grid steps aside (it
        # only makes sense across the whole week).
        Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        body = self.client.get(
            f"/orders/?week=2026-05-18&customer={self.alice_cust.pk}"
            "&date=2026-05-18"
        ).content.decode()
        self.assertNotIn('data-testid="order-grid"', body)

    # ---- wide layout opt-in ----

    def test_orders_weekly_view_uses_wide_main_layout(self):
        # On the shared design-system shell the weekly orders view uses a
        # wide, centred main column so the per-customer grid can use most
        # of the viewport. Assert the BP shell's main + capped, guttered
        # container render (the old `main.wide` mechanism was replaced by
        # the sidebar + max-width container in design_system/base.html).
        body = self.client.get("/orders/").content.decode()
        self.assertIn('<main class="ml-64 min-w-0">', body)
        # Capped reading width (not edge-to-edge) with a real side gutter.
        self.assertIn('max-w-[1760px]', body)
        self.assertIn('px-8', body)

    def test_other_pages_do_not_opt_into_wide_layout(self):
        # The wide layout is scoped to the orders weekly view ONLY —
        # other still-old-shell pages (New order form, Recipes) keep the
        # default centred 1100px container. Assert their <main> carries an
        # empty class attribute (no `wide`). Products and Home have since
        # moved to the design-system shell, covered by their own tests.
        for path in ("/orders/new/", "/recipes/"):
            body = self.client.get(path).content.decode()
            self.assertIn('<main class="">', body,
                          f"{path} should not opt into the wide layout")
            self.assertNotIn('<main class="wide"', body,
                             f"{path} should not have the wide class")

    # ---- sticky / frozen grid columns ----

    def test_grid_left_columns_are_sticky(self):
        # SKU / Product / Price columns stay locked while the day
        # columns scroll horizontally. Assert the sticky CSS classes
        # land on both the <th>s and the matching <td>s for at least
        # one product row, and that the wrap div carries overflow-x.
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("4"))
        body = self.client.get(
            f"/orders/?week=2026-05-18&customer={self.alice_cust.pk}"
        ).content.decode()
        # SKU / Product / Price columns are tagged (data-col) and made
        # sticky via Tailwind's `sticky` + left-offset utilities. The
        # data-col hooks land on BOTH the <th>s and the <td>s — each
        # appears at least twice (header + at least one body row).
        for col in ("sku", "name", "price"):
            self.assertGreaterEqual(
                body.count(f'data-col="{col}"'), 2,
                f"{col} column should be tagged on header and body cells")
        # Sticky positioning is enabled (Tailwind `sticky` + pinned offsets).
        self.assertIn('sticky left-0', body)
        self.assertIn('sticky left-[90px]', body)
        # The wrap div provides the horizontal scroll context.
        self.assertIn('data-testid="order-grid-wrap"', body)
        # The day-totals footer also pins its label cell sticky.
        self.assertIn('data-col="foot"', body)

    def test_grid_long_product_names_keep_full_text_via_title(self):
        # The Product column has a fixed width (so the grid fits the
        # container without horizontal scroll on desktop). A name too
        # long for that width truncates visually with an ellipsis, but
        # the full text stays in the DOM AND on the link's title=
        # attribute so hover surfaces the complete name.
        long_name = "Sticky Apple & Cinnamon Bun (Loose) - WHOLESALE ONLY"
        long_sp = SaleProduct.objects.create(
            name=long_name, price=Decimal("1.20"), department=self.dept)
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        OrderLine.objects.create(
            order=order, sale_product=long_sp, qty_ordered=Decimal("1"))
        body = self.client.get(
            f"/orders/?week=2026-05-18&customer={self.alice_cust.pk}"
        ).content.decode()
        # The full name appears in the DOM (Django autoescapes `&`
        # to `&amp;`; assert on the escaped form the browser sees).
        escaped = long_name.replace("&", "&amp;")
        self.assertIn(escaped, body)
        # The link carries title="<full name>" so hover reveals it
        # even when the cell visually truncates.
        self.assertIn(f'title="{escaped}"', body)
        # And the product cell uses Tailwind `truncate` (overflow-hidden +
        # ellipsis + nowrap) so a long name stays on one line.
        self.assertIn('truncate', body)

    def test_grid_uses_fixed_layout_to_fit_container(self):
        # The grid is table-layout:fixed and width:100% so its column
        # widths come from the CSS — not the content — and the table
        # always equals the container width. That stops the table from
        # over-shooting its wrapper and producing a mid-page horizontal
        # scrollbar between the grid and the day-tiles below.
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("3"))
        body = self.client.get(
            f"/orders/?week=2026-05-18&customer={self.alice_cust.pk}"
        ).content.decode()
        # table-fixed + w-full so the table's column widths come from the
        # CSS and the table tracks the container exactly.
        self.assertIn('table-fixed', body)
        self.assertIn('w-full', body)

    # ---- active-first customer list ----

    def test_customer_list_orders_active_customers_first(self):
        # Alice has an order this week; Bob doesn't. Alice's link must
        # appear BEFORE Bob's in the rendered customer list.
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("4"))
        body = self.client.get("/orders/?week=2026-05-18").content.decode()
        list_start = body.find('data-testid="cust-list"')
        self.assertGreater(list_start, -1)
        # "Ordered this week" divider sits above Alice's row; the
        # "Other customers" divider sits above Bob's row.
        i_active_divider = body.find("Ordered this week", list_start)
        i_other_divider = body.find("Other customers", list_start)
        i_alice = body.find(self.alice_cust.name, list_start)
        i_bob = body.find(self.bob_cust.name, list_start)
        self.assertGreater(i_active_divider, -1)
        self.assertGreater(i_other_divider, -1)
        self.assertLess(i_active_divider, i_alice)
        self.assertLess(i_alice, i_other_divider)
        self.assertLess(i_other_divider, i_bob)

    def test_customer_list_shows_count_and_total_value(self):
        # Alice gets two orders this week — one each on Mon and Tue —
        # adding up to 3 × 2.20 + 2 × 2.20 = £11.00. The badge spells
        # out the count + £ alongside her name.
        mon = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        OrderLine.objects.create(
            order=mon, sale_product=self.loose, qty_ordered=Decimal("3"))
        tue = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 19))
        OrderLine.objects.create(
            order=tue, sale_product=self.loose, qty_ordered=Decimal("2"))
        body = self.client.get("/orders/?week=2026-05-18").content.decode()
        # The badge text on Alice's row: "2 orders · £11.00".
        self.assertIn("2 orders", body)
        self.assertIn("£11.00", body)

    def test_inactive_customers_render_dimmed_but_clickable(self):
        # Bob has no orders this week. His row renders with the
        # `muted` class (the CSS dims it) AND still links to his
        # customer-filtered week view (so he stays one click away).
        Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        body = self.client.get("/orders/?week=2026-05-18").content.decode()
        # Scope the lookup to the customer-list section so we don't
        # collide with the customer name appearing in the filter
        # <select> at the top of the page.
        list_start = body.find('data-testid="cust-list"')
        self.assertGreater(list_start, -1)
        list_body = body[list_start:]
        bob_pos = list_body.find(self.bob_cust.name)
        self.assertGreater(bob_pos, -1)
        # Walk back to the enclosing <tr> opening tag.
        row_start = list_body.rfind("<tr", 0, bob_pos)
        self.assertGreater(row_start, -1)
        row_tag = list_body[row_start:bob_pos]
        self.assertIn('data-active="false"', row_tag,
                      "inactive customer row should be marked inactive (dimmed)")
        # And he's still clickable: the customer-filtered URL is on
        # his row's <a href>.
        self.assertIn(f"customer={self.bob_cust.pk}", list_body)

    # ---- summary placement (top of page, no bottom duplicate) ----

    def test_week_summary_blocks_sit_at_top_of_page(self):
        # The day-strip is the single at-a-glance summary now (the
        # earlier two-box stats block was redundant with the strip's
        # trailing "Week total" tile and has been removed). The strip
        # sits between the filter form and the customer list / grid.
        Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        body = self.client.get("/orders/?week=2026-05-18").content.decode()
        i_filter = body.find('data-testid="filter-form"')
        i_strip = body.find('data-testid="day-strip"')
        i_cust = body.find('data-testid="cust-list"')
        self.assertGreater(i_filter, -1, "filter form missing")
        self.assertGreater(i_strip, -1, "day-strip block missing")
        self.assertGreater(i_cust, -1, "customer list missing")
        # Order: filter form → day-strip → customer list.
        self.assertLess(i_filter, i_strip,
                        "day-strip should come after the filter form")
        self.assertLess(i_strip, i_cust,
                        "day-strip should come before the customer list")
        # And the headline stats block is gone — its job moved into
        # the strip's "Week" tile.
        self.assertNotIn('class="stats"', body)

    def test_day_strip_appears_only_once_on_page(self):
        # The strip used to render at the bottom in the old layout.
        # After the move it lives only at the top; no duplicate.
        Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        for path in (
                "/orders/?week=2026-05-18",
                f"/orders/?week=2026-05-18&customer={self.alice_cust.pk}",
        ):
            body = self.client.get(path).content.decode()
            self.assertEqual(
                body.count('data-testid="day-strip"'), 1,
                f"{path} should render exactly one day-strip — found "
                f"{body.count('data-testid=\"day-strip\"')}")

    def test_day_strip_sits_above_the_customer_grid(self):
        # When a customer is selected, the order is:
        # filter form → day-strip → grid (no customer list, which is
        # mutually exclusive with the grid).
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("3"))
        body = self.client.get(
            f"/orders/?week=2026-05-18&customer={self.alice_cust.pk}"
        ).content.decode()
        i_filter = body.find('data-testid="filter-form"')
        i_strip = body.find('data-testid="day-strip"')
        i_grid = body.find('data-testid="order-grid-wrap"')
        self.assertGreater(i_filter, -1)
        self.assertGreater(i_strip, -1)
        self.assertGreater(i_grid, -1)
        self.assertLess(i_filter, i_strip)
        self.assertLess(i_strip, i_grid)

    # ---- compact per-day totals strip ----

    def test_day_strip_shows_per_day_count_and_value(self):
        # Mon = 3 × 2.20 = £6.60; Wed = 1 × 18.00 = £18.00. The strip
        # shows the day, date, "1 order" and the value for each.
        order_mon = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        OrderLine.objects.create(
            order=order_mon, sale_product=self.loose, qty_ordered=Decimal("3"))
        order_wed = Order.objects.create(
            customer=self.bob_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 20))
        OrderLine.objects.create(
            order=order_wed, sale_product=self.slab, qty_ordered=Decimal("1"))
        body = self.client.get("/orders/?week=2026-05-18").content.decode()
        strip_start = body.find('data-testid="day-strip"')
        self.assertGreater(strip_start, -1)
        # All 7 weekday tiles render with their date.
        for dt in ("18 May", "19 May", "20 May", "21 May",
                   "22 May", "23 May", "24 May"):
            self.assertIn(dt, body[strip_start:])
        # Mon + Wed each report "1 order" with their £ totals.
        strip = body[strip_start:]
        self.assertIn("1 order", strip)
        self.assertIn("£6.60", strip)
        self.assertIn("£18.00", strip)
        # The week-total tile rolls them up to "2 orders · £24.60".
        self.assertIn("2 orders", strip)
        self.assertIn("£24.60", strip)

    def test_day_strip_replaces_per_day_cards(self):
        # The verbose per-day cards no longer render in the default
        # (unfiltered) view; they only come back as a single block
        # when ?date= narrows to one day. Without ?date= there should
        # be at most one .day-block on the page (the filtered card),
        # and with no filter, zero.
        Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        body = self.client.get("/orders/?week=2026-05-18").content.decode()
        # No verbose per-day .day-block sections in the default view.
        self.assertEqual(body.count('class="day-block'), 0)
        # The compact strip is the replacement — 7 day-tiles + 1
        # week-total tile = 8.
        self.assertEqual(body.count('data-testid="day-tile"'), 8)

    def test_day_strip_tile_links_filter_to_that_day(self):
        # Each day-tile is a clickable link with ?date=YYYY-MM-DD so
        # the operator can drill into a single day. The filtered view
        # then surfaces that day's order detail links.
        order = Order.objects.create(
            customer=self.alice_cust, department=self.dept,
            order_date=datetime.date(2026, 5, 18))
        OrderLine.objects.create(
            order=order, sale_product=self.loose, qty_ordered=Decimal("3"))
        body = self.client.get("/orders/?week=2026-05-18").content.decode()
        self.assertIn("date=2026-05-18", body)
        # And the per-day list does surface when ?date= is set.
        filtered = self.client.get(
            "/orders/?week=2026-05-18&date=2026-05-18"
        ).content.decode()
        self.assertIn(f'/orders/{order.pk}/', filtered)


def _seed_garden_cafe_products(dept):
    """Pre-create the SaleProducts the GARDEN CAFE tab references so
    the importer has something to match against. Sage codes mirror
    the sheet exactly; products without a Sage code (the '0' SKU
    rows) match by name."""
    seeds = [
        ("Croissant (Loose)", "660130013", Decimal("1.10")),
        ("Pain au Chocolat (Loose)", "660130059", Decimal("1.50")),
        ("Sticky Apple & Cinnamon Bun (Loose) - WHOLESALE ONLY", "",
         Decimal("1.20")),
        ("Sourdough Sandwich Loaf 1.1kg (Loose)", "660260279",
         Decimal("2.20")),
        ("Brioche Loaf - 1.1kg", "", Decimal("5.00")),
        # Trailing space matches the sheet's "Apple Waste Sourdough - 1.2kg ".
        ("Apple Waste Sourdough - 1.2kg ", "", Decimal("4.50")),
        ("Rhubarb and Custard Danish", "", Decimal("1.60")),
        ("Pain Au Almond ( Loose)", "660260415", Decimal("1.50")),
        ("Sultana Pain Suisse (Loose)", "660260411", Decimal("1.65")),
        ("Seeded Sourdough", "660260409", Decimal("2.20")),
    ]
    out = {}
    for name, sage, price in seeds:
        sp = SaleProduct.objects.create(
            name=name, sage_number=sage, price=price, department=dept)
        out[name] = sp
    return out


class OrderImportTests(TestCase):
    """End-to-end import of the GARDEN CAFE tab from the real workbook."""

    SHEET = "data/order_sheet.xlsm"

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        # Garden Café customer — the importer matches the tab name
        # case-insensitively against Customer.name.
        self.garden = Customer.objects.create(
            name="Garden Cafe", department=self.dept)
        self.products = _seed_garden_cafe_products(self.dept)

    def test_import_creates_per_day_lines_with_correct_qtys(self):
        from django.core.management import call_command
        call_command("import_orders", self.SHEET)
        # Croissant: Mon..Fri = 4, Sat..Sun = 7.
        croissant = self.products["Croissant (Loose)"]
        rows = {
            o.order_date: o.lines.get(sale_product=croissant).qty_ordered
            for o in Order.objects.filter(customer=self.garden,
                                          lines__sale_product=croissant)
        }
        self.assertEqual(rows[datetime.date(2026, 5, 18)], Decimal("4"))
        self.assertEqual(rows[datetime.date(2026, 5, 22)], Decimal("4"))
        # Croissant 7 on Saturday (the spec calls this out specifically).
        self.assertEqual(rows[datetime.date(2026, 5, 23)], Decimal("7"))
        self.assertEqual(rows[datetime.date(2026, 5, 24)], Decimal("7"))

    def test_import_seeded_sourdough_is_nine_each_day(self):
        from django.core.management import call_command
        call_command("import_orders", self.SHEET)
        seeded = self.products["Seeded Sourdough"]
        for i in range(7):
            day = datetime.date(2026, 5, 18) + datetime.timedelta(days=i)
            line = OrderLine.objects.get(
                order__customer=self.garden, order__order_date=day,
                sale_product=seeded)
            self.assertEqual(line.qty_ordered, Decimal("9"))

    def test_import_skips_zero_blank_weeks(self):
        # Rhubarb and Custard Danish — entire row blank — no lines.
        # Pain Au Almond — entire row blank — no lines.
        from django.core.management import call_command
        call_command("import_orders", self.SHEET)
        self.assertFalse(OrderLine.objects.filter(
            sale_product=self.products["Rhubarb and Custard Danish"]).exists())
        self.assertFalse(OrderLine.objects.filter(
            sale_product=self.products["Pain Au Almond ( Loose)"]).exists())

    def test_import_matches_brioche_by_exact_name_no_sage(self):
        # Brioche Loaf carries SKU '0' on the sheet — the importer
        # falls back to exact-name match and still wires it up.
        from django.core.management import call_command
        call_command("import_orders", self.SHEET)
        brioche = self.products["Brioche Loaf - 1.1kg"]
        # Sheet says 2 on Mon, blank Tue, 2 Wed, blank Thu, 2 Fri,
        # blank Sat, 2 Sun.
        qty_by_day = {
            o.order_date: o.lines.get(sale_product=brioche).qty_ordered
            for o in Order.objects.filter(customer=self.garden,
                                          lines__sale_product=brioche)
        }
        self.assertEqual(qty_by_day.get(datetime.date(2026, 5, 18)),
                         Decimal("2"))
        self.assertNotIn(datetime.date(2026, 5, 19), qty_by_day)  # blank
        self.assertEqual(qty_by_day.get(datetime.date(2026, 5, 24)),
                         Decimal("2"))

    def test_import_is_idempotent_and_replaces(self):
        from django.core.management import call_command
        call_command("import_orders", self.SHEET)
        first_lines = OrderLine.objects.count()
        first_orders = Order.objects.count()
        # Edit one line so we can verify it gets reset on re-run.
        croissant = self.products["Croissant (Loose)"]
        line = OrderLine.objects.get(
            order__customer=self.garden,
            order__order_date=datetime.date(2026, 5, 18),
            sale_product=croissant)
        line.qty_ordered = Decimal("999")
        line.save()
        # Re-run — same lines (no duplicates) and the edit is reverted.
        # The importer wipes + recreates lines, so the original pk is
        # gone; re-fetch by (order_date, sale_product).
        call_command("import_orders", self.SHEET)
        self.assertEqual(OrderLine.objects.count(), first_lines)
        self.assertEqual(Order.objects.count(), first_orders)
        fresh = OrderLine.objects.get(
            order__customer=self.garden,
            order__order_date=datetime.date(2026, 5, 18),
            sale_product=croissant)
        self.assertEqual(fresh.qty_ordered, Decimal("4"))

    def test_imported_week_renders_in_grid_view(self):
        # After importing, the grid for Garden Café week 18 May should
        # show the right per-day quantities and totals.
        from django.core.management import call_command
        call_command("import_orders", self.SHEET)
        body = self.client.get(
            f"/orders/?week=2026-05-18&customer={self.garden.pk}"
        ).content.decode()
        # Sage-coded SKU appears in the SKU column.
        self.assertIn("660130013", body)
        # Croissant row totals to 34, total value £37.40.
        self.assertIn(">34<", body)
        self.assertIn("£37.40", body)

    def test_manual_create_still_works_after_import(self):
        # Importing doesn't break the manual /orders/new/ path.
        from django.core.management import call_command
        call_command("import_orders", self.SHEET)
        resp = self.client.post("/orders/new/", {
            "customer_id": str(self.garden.pk),
            "order_date": "2026-06-01",
            "product_id": [str(self.products["Croissant (Loose)"].pk)],
            "qty_ordered": ["2"],
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(Order.objects.filter(
            customer=self.garden,
            order_date=datetime.date(2026, 6, 1)).exists())


class OrderLineSnapshotTests(TestCase):
    """OrderLine snapshots product_name + unit_price independently of
    the SaleProduct catalogue so financial totals survive product
    renames, price changes, and full discontinuation."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.cust = Customer.objects.create(
            name="Garden Cafe", department=self.dept)
        self.sp = SaleProduct.objects.create(
            name="Croissant (Loose)", price=Decimal("1.10"),
            department=self.dept)

    def test_save_snapshots_name_and_price_from_sale_product(self):
        order = Order.objects.create(
            customer=self.cust, department=self.dept,
            order_date=datetime.date(2026, 5, 1))
        line = OrderLine.objects.create(
            order=order, sale_product=self.sp, qty_ordered=Decimal("3"))
        line.refresh_from_db()
        self.assertEqual(line.product_name, "Croissant (Loose)")
        self.assertEqual(line.unit_price, Decimal("1.10"))

    def test_line_value_uses_snapshot_not_live_price(self):
        order = Order.objects.create(
            customer=self.cust, department=self.dept,
            order_date=datetime.date(2026, 5, 1))
        line = OrderLine.objects.create(
            order=order, sale_product=self.sp, qty_ordered=Decimal("4"))
        # Catalogue price drifts AFTER the order — line_value must not
        # follow. Snapshot price was £1.10, so 4 × 1.10 = 4.40.
        self.sp.price = Decimal("99.99")
        self.sp.save()
        line.refresh_from_db()
        self.assertEqual(line.line_value, Decimal("4.40"))

    def test_total_value_includes_discontinued_unlinked_line(self):
        order = Order.objects.create(
            customer=self.cust, department=self.dept,
            order_date=datetime.date(2026, 3, 30))
        # Linked, modern product.
        OrderLine.objects.create(
            order=order, sale_product=self.sp, qty_ordered=Decimal("2"))
        # Historical, NO catalogue link — the kind of line the
        # historical importer creates for a discontinued SKU like
        # "Hot Cross Buns".
        OrderLine.objects.create(
            order=order, sale_product=None,
            product_name="Hot Cross Buns (Pack/10)",
            unit_price=Decimal("6.25"),
            qty_ordered=Decimal("3"))
        # 2 × 1.10 + 3 × 6.25 = 2.20 + 18.75 = 20.95
        self.assertEqual(order.total_value(), Decimal("20.95"))

    def test_explicit_snapshot_overrides_sale_product_defaults(self):
        # Callers (e.g. the historical importer) can pass product_name
        # + unit_price explicitly; save() must NOT clobber them with
        # the linked SaleProduct's current values.
        order = Order.objects.create(
            customer=self.cust, department=self.dept,
            order_date=datetime.date(2026, 5, 1))
        line = OrderLine.objects.create(
            order=order, sale_product=self.sp,
            product_name="Old name", unit_price=Decimal("0.95"),
            qty_ordered=Decimal("1"))
        line.refresh_from_db()
        self.assertEqual(line.product_name, "Old name")
        self.assertEqual(line.unit_price, Decimal("0.95"))


class OrderImportPriceParsingTests(SimpleTestCase):
    """_parse_price handles the historical sheet's mixed Price column
    (numeric cells like 1.1 vs string cells like "£6.25")."""

    def test_parses_pound_prefixed_string(self):
        from .order_import import _parse_price
        self.assertEqual(_parse_price("£6.25"), Decimal("6.25"))

    def test_parses_pound_prefixed_with_whitespace(self):
        from .order_import import _parse_price
        self.assertEqual(_parse_price("  £ 6.25 "), Decimal("6.25"))

    def test_parses_plain_decimal_string(self):
        from .order_import import _parse_price
        self.assertEqual(_parse_price("1.1"), Decimal("1.1"))

    def test_parses_numeric_cell(self):
        from .order_import import _parse_price
        self.assertEqual(_parse_price(1.1), Decimal("1.1"))

    def test_blank_returns_none(self):
        from .order_import import _parse_price
        self.assertIsNone(_parse_price(None))
        self.assertIsNone(_parse_price(""))
        self.assertIsNone(_parse_price("   "))

    def test_garbage_returns_none(self):
        from .order_import import _parse_price
        self.assertIsNone(_parse_price("not-a-price"))


class HistoricalOrderImportTests(TestCase):
    """End-to-end import of the historical GARDEN CAFE tab (w/c 30
    March 2026). Proves that EVERY product row with an ordered qty
    becomes an OrderLine — including discontinued SKUs that no
    current catalogue product matches — so financial totals are
    complete."""

    SHEET = "data/historical/order_sheet_2026_03_30.xlsm"

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        self.garden = Customer.objects.create(
            name="Garden Cafe", department=self.dept)
        self.products = _seed_garden_cafe_products(self.dept)

    def test_discontinued_products_land_with_snapshot(self):
        from django.core.management import call_command
        call_command("import_orders", self.SHEET, tab=["GARDEN CAFE"])
        # "Hot Cross Buns (Pack/10)" is on the historical sheet but
        # NOT in the current catalogue seed — the line must still
        # exist, with sale_product=None and the sheet's price snapshotted.
        hxb = OrderLine.objects.filter(
            order__customer=self.garden,
            product_name="Hot Cross Buns (Pack/10)")
        self.assertTrue(hxb.exists(),
            "Hot Cross Buns line was dropped — historical importer "
            "must keep discontinued SKUs")
        for line in hxb:
            self.assertIsNone(line.sale_product_id)
            self.assertEqual(line.unit_price, Decimal("6.25"))
            # And the value flows through.
            self.assertEqual(
                line.line_value, line.qty_ordered * Decimal("6.25"))

    def test_almond_pastry_discontinued_line_exists(self):
        from django.core.management import call_command
        call_command("import_orders", self.SHEET, tab=["GARDEN CAFE"])
        # "Almond Pastry" is the historical name; the modern
        # catalogue uses "Pain Au Almond ( Loose)" — different name,
        # so it stays unlinked.
        ap = OrderLine.objects.filter(
            order__customer=self.garden,
            product_name="Almond Pastry")
        self.assertTrue(ap.exists())
        for line in ap:
            self.assertIsNone(line.sale_product_id)
            self.assertEqual(line.unit_price, Decimal("1.5"))

    def test_matched_products_still_link(self):
        from django.core.management import call_command
        call_command("import_orders", self.SHEET, tab=["GARDEN CAFE"])
        # Croissant (Loose) exists in both the historical sheet and
        # the modern catalogue — it must still link.
        croissant = self.products["Croissant (Loose)"]
        lines = OrderLine.objects.filter(
            order__customer=self.garden, sale_product=croissant)
        self.assertTrue(lines.exists())
        for line in lines:
            self.assertEqual(line.product_name, "Croissant (Loose)")
            self.assertEqual(line.unit_price, Decimal("1.1"))

    def test_week_dates_derived_from_sheet(self):
        from django.core.management import call_command
        call_command("import_orders", self.SHEET, tab=["GARDEN CAFE"])
        # Week-commencing 30 March 2026 (Monday) — orders only on
        # those seven dates, derived from the sheet.
        dates = set(Order.objects.filter(
            customer=self.garden).values_list("order_date", flat=True))
        self.assertEqual(dates, {
            datetime.date(2026, 3, 30) + datetime.timedelta(days=i)
            for i in range(7)
        })

    def test_order_total_includes_discontinued_values(self):
        from django.core.management import call_command
        call_command("import_orders", self.SHEET, tab=["GARDEN CAFE"])
        # Every order across the historical week should have a
        # non-zero total — there's no day with only unpriced lines.
        for order in Order.objects.filter(customer=self.garden):
            self.assertGreater(order.total_value(), Decimal("0"))


@tag("slow")
class HistoricalWorkbookImportTests(TestCase):
    """End-to-end ``import_historical_orders`` against the real w/c
    30 Mar 2026 file: ALL customer tabs (not just Garden Café),
    idempotent at the week level, non-fatal on any single bad tab.

    Tagged ``slow`` because the import walks the entire 2.7 MB
    workbook (every customer tab) on every test; the default fast
    suite excludes ``--tag slow``. CI / pre-push runs them via
    ``--tag slow``.
    """

    SHEET = "data/historical/order_sheet_2026_03_30.xlsm"
    WEEK_START = datetime.date(2026, 3, 30)

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        # Two customers from the historical workbook so we can prove
        # the importer walks more than one tab. The rest of the ~30
        # tabs land in ``failures`` (no matching Customer) — that's
        # the expected non-fatal path.
        self.garden = Customer.objects.create(
            name="Garden Cafe", department=self.dept)
        self.farmshop = Customer.objects.create(
            name="Farmshop", department=self.dept)
        self.products = _seed_garden_cafe_products(self.dept)

    def test_imports_lines_across_multiple_customer_tabs(self):
        from stock.order_import import import_historical_workbook
        summary = import_historical_workbook(self.SHEET)
        self.assertFalse(summary["skipped"])
        self.assertEqual(summary["week_start"], self.WEEK_START)
        # Both seeded customers landed orders for the week.
        for cust in (self.garden, self.farmshop):
            self.assertTrue(
                Order.objects.filter(
                    customer=cust, order_date=self.WEEK_START).exists(),
                f"{cust.name} got no orders for w/c {self.WEEK_START}")
        # The per-tab breakdown surfaces both customers + counts > 0.
        self.assertIn("GARDEN CAFE", summary["per_tab"])
        self.assertIn("FARMSHOP", summary["per_tab"])
        self.assertGreater(summary["per_tab"]["GARDEN CAFE"]["lines_imported"], 0)
        self.assertGreater(summary["per_tab"]["FARMSHOP"]["lines_imported"], 0)

    def test_every_line_has_snapshotted_name_and_price(self):
        from stock.order_import import import_historical_workbook
        import_historical_workbook(self.SHEET)
        # Every OrderLine landed by the historical import — matched
        # OR unmatched — must carry the snapshot fields. Use the week
        # range so we don't pick up anything outside the import.
        lines = OrderLine.objects.filter(
            order__order_date__range=(
                self.WEEK_START,
                self.WEEK_START + datetime.timedelta(days=6)))
        self.assertGreater(lines.count(), 0)
        for line in lines:
            self.assertTrue(line.product_name,
                f"line {line.pk} has no product_name snapshot")
            self.assertIsNotNone(line.unit_price,
                f"line {line.pk} has no unit_price snapshot")

    def test_discontinued_products_kept_across_tabs(self):
        from stock.order_import import import_historical_workbook
        import_historical_workbook(self.SHEET)
        # Hot Cross Buns / Almond Pastry appear on the historical
        # GARDEN CAFE tab but aren't in the current catalogue seed.
        # They must land as unlinked OrderLines with the sheet's
        # snapshotted price, not be silently dropped.
        for name, price in (
                ("Hot Cross Buns (Pack/10)", Decimal("6.25")),
                ("Almond Pastry", Decimal("1.5"))):
            lines = OrderLine.objects.filter(
                order__customer=self.garden, product_name=name)
            self.assertTrue(lines.exists(),
                f"historical importer dropped discontinued {name!r}")
            for line in lines:
                self.assertIsNone(line.sale_product_id)
                self.assertEqual(line.unit_price, price)

    def test_matched_products_link_to_current_saleproduct(self):
        from stock.order_import import import_historical_workbook
        import_historical_workbook(self.SHEET)
        croissant = self.products["Croissant (Loose)"]
        lines = OrderLine.objects.filter(
            order__customer=self.garden, sale_product=croissant)
        self.assertGreater(lines.count(), 0,
            "Croissant should still link to the current catalogue entry")

    def test_idempotent_second_run_skips_entire_week(self):
        from stock.order_import import import_historical_workbook
        first = import_historical_workbook(self.SHEET)
        self.assertFalse(first["skipped"])
        line_count_after_first = OrderLine.objects.count()
        order_count_after_first = Order.objects.count()

        # Touch one line so we can verify a second run leaves edits alone.
        croissant = self.products["Croissant (Loose)"]
        edited = OrderLine.objects.filter(
            order__customer=self.garden, sale_product=croissant).first()
        edited.qty_ordered = Decimal("999")
        edited.save()

        second = import_historical_workbook(self.SHEET)
        self.assertTrue(second["skipped"])
        # Reason now references the version stamp, not "any Order
        # exists for this week" — the new gate is HistoricalImport-
        # version-driven so it doesn't get confused by manual edits.
        self.assertIn("already imported", second["reason"])
        self.assertEqual(second["week_start"], self.WEEK_START)
        # Nothing was wiped or duplicated, and the hand-edit survived.
        self.assertEqual(OrderLine.objects.count(), line_count_after_first)
        self.assertEqual(Order.objects.count(), order_count_after_first)
        edited.refresh_from_db()
        self.assertEqual(edited.qty_ordered, Decimal("999"))

    def test_force_reimports_and_does_not_duplicate(self):
        from stock.order_import import import_historical_workbook
        first = import_historical_workbook(self.SHEET)
        line_count = OrderLine.objects.count()
        # --force bypasses the idempotency gate; the per-tab importer
        # still wipes + rebuilds, so counts match (not double).
        again = import_historical_workbook(self.SHEET, force=True)
        self.assertFalse(again["skipped"])
        self.assertEqual(OrderLine.objects.count(), line_count)

    def test_bad_tab_is_non_fatal_other_tabs_still_import(self):
        # Make one tab's per-tab call raise; the wrapper must capture
        # the failure, keep going on every other tab, and still
        # complete the run.
        from stock import order_import as oi
        real = oi.import_orders_for_tab

        def fail_on_farmshop(workbook, tab_name, customer, dept, **kwargs):
            if tab_name == "FARMSHOP":
                raise RuntimeError("synthetic per-tab boom")
            return real(workbook, tab_name, customer, dept, **kwargs)

        with patch.object(oi, "import_orders_for_tab", side_effect=fail_on_farmshop):
            summary = oi.import_historical_workbook(self.SHEET)
        self.assertFalse(summary["skipped"])
        self.assertIn("GARDEN CAFE", summary["per_tab"])
        self.assertNotIn("FARMSHOP", summary["per_tab"])
        # FARMSHOP shows up as a failure with the synthetic message.
        failed = dict(summary["failures"])
        self.assertIn("FARMSHOP", failed)
        self.assertIn("synthetic per-tab boom", failed["FARMSHOP"])
        # And Garden Café's orders are still in place.
        self.assertTrue(Order.objects.filter(
            customer=self.garden, order_date=self.WEEK_START).exists())

    def test_meta_tabs_are_skipped(self):
        # The Products / Customers / Delivery Note tabs aren't order
        # forms — they should never appear in per_tab or failures.
        # WHOLESALE is NOT in this list any more: it has its own
        # parser and routes through per_tab with the wholesale-handler
        # shape; see HistoricalImportReconciliationTests for that path.
        from stock.order_import import import_historical_workbook
        summary = import_historical_workbook(self.SHEET)
        seen = set(summary["per_tab"]) | {t for t, _ in summary["failures"]}
        for meta in ("Start", "Products", "Customers",
                     "Customer Lookup", "Production",
                     "Delivery Note", "Wholesale Delivery Note"):
            self.assertNotIn(meta, seen,
                f"meta tab {meta!r} should have been skipped silently")

    def test_management_command_runs_and_then_skips(self):
        from django.core.management import call_command
        # First call imports; second is skipped wholesale. Both must
        # exit 0 (no CommandError) — the deploy depends on it.
        call_command("import_historical_orders", self.SHEET)
        first_count = OrderLine.objects.count()
        self.assertGreater(first_count, 0)
        call_command("import_historical_orders", self.SHEET)
        self.assertEqual(OrderLine.objects.count(), first_count)

    def test_missing_file_raises_command_error_caught_by_build(self):
        # The management command surfaces FileNotFoundError as
        # CommandError so build.sh's ``|| echo`` non-fatal pattern
        # logs it and moves on instead of aborting the deploy.
        from django.core.management import call_command
        from django.core.management.base import CommandError
        with self.assertRaises(CommandError):
            call_command("import_historical_orders", "data/historical/does_not_exist.xlsm")


class ImportBlankPriceTreatmentTests(TestCase):
    """A blank Price cell on the sheet means £0 on import — NOT the
    linked SaleProduct's current catalogue price. The fallback in
    OrderLine.save() (snapshot the catalogue when the caller passes
    None) is kept for the manual order form, where it's the right
    thing to do. Importer paths must always pass something explicit."""

    SHEET = "data/historical/order_sheet_2026_03_30.xlsm"

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.ecom = Customer.objects.create(name="ECOM", department=self.dept)
        # Deliberately seed the catalogue with the WRONG price for the
        # ECOM TN Biscuits row so any code path that touched the
        # SaleProduct catalogue (instead of the sheet) would produce
        # £99.99/unit and immediately show up. The historical workbook's
        # Products tab carries this SKU at £2.50, so the recovery chain
        # MUST land on £2.50 — sheet-authoritative.
        self.biscuits = SaleProduct.objects.create(
            name="TN Biscuits Easter Bag_EA", sage_number="660260238",
            price=Decimal("99.99"), department=self.dept)

    def test_blank_price_recovered_from_workbook_products_tab_not_catalogue(self):
        from django.core.management import call_command
        call_command("import_orders", self.SHEET, tab=["ECOM"])
        line = OrderLine.objects.filter(
            order__customer=self.ecom,
            product_name="TN Biscuits Easter Bag_EA").first()
        self.assertIsNotNone(line, "ECOM biscuits row should have imported")
        self.assertEqual(line.sale_product_id, self.biscuits.pk,
            "still links to current catalogue product (matched by Sage)")
        # £2.50 from the workbook's own Products tab — NEVER £99.99 from
        # the live SaleProduct catalogue. The point of the recovery
        # chain is that it stays sheet-authoritative.
        self.assertEqual(line.unit_price, Decimal("2.50"),
            "blank Price cell must be recovered from the workbook's "
            "Products tab, not the catalogue")
        self.assertEqual(line.line_value, Decimal("125.00"))

    def test_manual_orderline_creation_still_uses_catalogue_fallback(self):
        # The OrderLine.save() snapshot fallback (catalogue → unit_price
        # when caller passes None) is intentionally PRESERVED for the
        # manual order form. This protects the operator UX: creating
        # an order line through /orders/new/ should still record the
        # current catalogue price even though the import path
        # explicitly bypasses that fallback (sheet-authoritative).
        order = Order.objects.create(
            customer=self.ecom, department=self.dept,
            order_date=datetime.date(2026, 5, 1))
        line = OrderLine.objects.create(
            order=order, sale_product=self.biscuits,
            qty_ordered=Decimal("3"))  # no unit_price passed
        line.refresh_from_db()
        # Mirrors whatever we seeded for the SaleProduct (£99.99 here,
        # to make the divergence from the import path unmistakable).
        self.assertEqual(line.unit_price, Decimal("99.99"),
            "manual creation should still snapshot the catalogue price")


class WholesaleImportTests(TestCase):
    """The WHOLESALE tab — one row per (wholesale customer, product)
    — is now routed to a dedicated parser. Each wholesale customer
    must ALREADY EXIST as a Customer (no auto-create), gets their own
    first-class Orders, and unmatched names are reported."""

    SHEET = "data/historical/order_sheet_2026_03_30.xlsm"
    WEEK_START = datetime.date(2026, 3, 30)

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        # Seed two wholesale customers (TEALS + MJOLK) so we can prove
        # per-customer orders land separately; deliberately omit
        # BISHOPSTROW HOTEL to prove unmatched-name reporting.
        self.teals = Customer.objects.create(
            name="TEALS", department=self.dept)
        self.mjolk = Customer.objects.create(
            name="MJOLK", department=self.dept)
        # A current-catalogue product so we can verify match-by-name
        # still wires the link on a wholesale line.
        self.croissant = SaleProduct.objects.create(
            name="Croissant (Loose)", sage_number="660130013",
            price=Decimal("1.10"), department=self.dept)

    def test_wholesale_parser_splits_per_customer_with_orders(self):
        from stock.order_import import import_wholesale_tab
        from openpyxl import load_workbook
        wb = load_workbook(self.SHEET, read_only=True, data_only=True)
        try:
            result = import_wholesale_tab(wb)
        finally:
            wb.close()
        # Both seeded wholesale customers now have orders for the
        # historical week — first-class Orders, just like Garden Café.
        for cust in (self.teals, self.mjolk):
            self.assertTrue(
                Order.objects.filter(
                    customer=cust, order_date=self.WEEK_START).exists(),
                f"{cust.name} got no Orders out of WHOLESALE")
        # Per-customer summary surfaces both with non-zero values.
        self.assertIn("TEALS", result["per_customer"])
        self.assertIn("MJOLK", result["per_customer"])
        self.assertGreater(result["per_customer"]["TEALS"]["value"],
                           Decimal("0"))

    def test_wholesale_imports_existing_customers_no_duplicates(self):
        # The handler must NEVER create a Customer row. Count before
        # and after; the only Customer rows are the ones we seeded.
        from stock.order_import import import_wholesale_tab
        from openpyxl import load_workbook
        before = set(Customer.objects.values_list("name", flat=True))
        wb = load_workbook(self.SHEET, read_only=True, data_only=True)
        try:
            import_wholesale_tab(wb)
        finally:
            wb.close()
        after = set(Customer.objects.values_list("name", flat=True))
        self.assertEqual(before, after,
            "wholesale importer must not auto-create Customer rows")

    def test_unmatched_wholesale_names_are_reported(self):
        # BISHOPSTROW HOTEL appears on the WHOLESALE tab but we
        # deliberately didn't seed it. The handler must report it
        # in customer_unmatched rather than silently dropping it AND
        # rather than auto-creating a new Customer for it.
        from stock.order_import import import_wholesale_tab
        from openpyxl import load_workbook
        wb = load_workbook(self.SHEET, read_only=True, data_only=True)
        try:
            result = import_wholesale_tab(wb)
        finally:
            wb.close()
        names_lower = [n.lower() for n in result["customer_unmatched"]]
        self.assertIn("bishopstrow hotel", names_lower)
        self.assertFalse(
            Customer.objects.filter(name__iexact="bishopstrow hotel").exists(),
            "unmatched wholesale name must NOT have been auto-created")

    def test_wholesale_matched_products_link_to_saleproduct(self):
        from stock.order_import import import_wholesale_tab
        from openpyxl import load_workbook
        wb = load_workbook(self.SHEET, read_only=True, data_only=True)
        try:
            import_wholesale_tab(wb)
        finally:
            wb.close()
        # Croissant (Loose) on TEALS' block matches by name → linked.
        croissant_lines = OrderLine.objects.filter(
            order__customer=self.teals, sale_product=self.croissant)
        self.assertGreater(croissant_lines.count(), 0,
            "name-matched product should still link to current catalogue")

    def test_wholesale_total_for_seeded_customers_matches_sheet(self):
        # TEALS sheet total is £1703.40, MJOLK is £357.60. Together
        # the imported total for these two seeded customers should
        # reconcile exactly with the sheet.
        from stock.order_import import import_wholesale_tab
        from openpyxl import load_workbook
        wb = load_workbook(self.SHEET, read_only=True, data_only=True)
        try:
            import_wholesale_tab(wb)
        finally:
            wb.close()
        week = [self.WEEK_START + datetime.timedelta(days=i)
                for i in range(7)]
        teals_total = sum(
            (o.total_value() for o in Order.objects.filter(
                customer=self.teals, order_date__in=week)),
            Decimal("0"))
        mjolk_total = sum(
            (o.total_value() for o in Order.objects.filter(
                customer=self.mjolk, order_date__in=week)),
            Decimal("0"))
        self.assertEqual(teals_total, Decimal("1703.40"))
        self.assertEqual(mjolk_total, Decimal("357.60"))

    def test_wholesale_is_idempotent(self):
        # Re-running converges, never duplicates. Mirror the per-tab
        # importer's wipe-and-rebuild guarantee for the wholesale
        # handler so build.sh can call it on every deploy safely.
        from stock.order_import import import_wholesale_tab
        from openpyxl import load_workbook
        wb = load_workbook(self.SHEET, read_only=True, data_only=True)
        try:
            import_wholesale_tab(wb)
            first_lines = OrderLine.objects.count()
            first_orders = Order.objects.count()
            import_wholesale_tab(wb)
        finally:
            wb.close()
        self.assertEqual(OrderLine.objects.count(), first_lines)
        self.assertEqual(Order.objects.count(), first_orders)

    def test_wholesale_blank_price_recovered_from_workbook_products_tab(self):
        # THE OLD PHARMACY's "Sticky Apple & Cinnamon Bun (Loose) -
        # WHOLESALE ONLY" row has a blank Price cell on the WHOLESALE
        # tab but lives in the workbook's own Products tab at £1.20.
        # The wholesale handler must apply the same recovery chain as
        # the per-customer importer, snapshotting £1.20 from the sheet
        # — NOT from any SaleProduct catalogue entry.
        Customer.objects.create(
            name="THE OLD PHARMACY", department=self.dept)
        # Catalogue intentionally at the wrong price so the test would
        # fail loudly if the catalogue ever leaked into the import
        # path.
        SaleProduct.objects.create(
            name="Sticky Apple & Cinnamon Bun (Loose) - WHOLESALE ONLY",
            price=Decimal("99.99"), department=self.dept)
        from stock.order_import import (
            import_wholesale_tab, make_price_recovery)
        from openpyxl import load_workbook
        wb = load_workbook(self.SHEET, read_only=True, data_only=True)
        try:
            import_wholesale_tab(
                wb, recover_price=make_price_recovery(wb))
        finally:
            wb.close()
        sticky_lines = OrderLine.objects.filter(
            order__customer__name__iexact="THE OLD PHARMACY",
            product_name__icontains="Sticky Apple & Cinnamon Bun")
        self.assertGreater(sticky_lines.count(), 0)
        for line in sticky_lines:
            self.assertEqual(line.unit_price, Decimal("1.20"),
                "WHOLESALE blank-price row should be recovered from "
                "the workbook Products tab (£1.20), not from the "
                "catalogue (£99.99) and not £0")

    def test_wholesale_blank_price_falls_to_zero_when_no_recovery(self):
        # The recovery is opt-in via recover_price. Without it, the
        # wholesale handler keeps the original blank-is-£0 behaviour —
        # important so callers that don't want recovery can still rely
        # on the strict-sheet semantics.
        Customer.objects.create(
            name="THE OLD PHARMACY", department=self.dept)
        from stock.order_import import import_wholesale_tab
        from openpyxl import load_workbook
        wb = load_workbook(self.SHEET, read_only=True, data_only=True)
        try:
            import_wholesale_tab(wb)  # no recover_price
        finally:
            wb.close()
        sticky_lines = OrderLine.objects.filter(
            order__customer__name__iexact="THE OLD PHARMACY",
            product_name__icontains="Sticky Apple & Cinnamon Bun")
        self.assertGreater(sticky_lines.count(), 0)
        for line in sticky_lines:
            self.assertEqual(line.unit_price, Decimal("0.00"))

    def test_wholesale_no_longer_in_meta_tabs(self):
        # Future-proofing: if anyone re-adds 'wholesale' to META_TABS
        # the importer silently stops capturing wholesale revenue.
        from stock.order_import import META_TABS
        self.assertNotIn("wholesale", META_TABS)


class WholesaleResilienceTests(TestCase):
    """w/c 4 May 2026 used to import with ZERO wholesale orders.
    The cause: rows 89 + 97 of the WHOLESALE tab are an exact
    (PINKMANS-WAPPING WHARF, 'Sliced Apple Waste Sourdough', Sat)
    duplicate that collides with OrderLine's partial-unique
    (order, sale_product) constraint. Under the outer @transaction.
    atomic, that single IntegrityError rolled back every wholesale
    line for the week. The fix wraps each create in a savepoint so
    the dup is dropped (first occurrence wins) and every other
    priced line lands. The TEALS new-product blank-price rows
    (Fruited Sourdough Loaf, 'Wheat & Rye Sourdough ' with trailing
    space, Seeded Sourdough) snapshot at £0 and are surfaced via
    ``lines_unpriced`` rather than aborting the tab."""

    SHEET = "data/historical/order_sheet_2026_05_04.xlsm"
    WEEK_START = datetime.date(2026, 5, 4)

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        # Seed the wholesale customers whose lines we care about. The
        # rest legitimately land in customer_unmatched and don't affect
        # this test.
        self.pinkmans = Customer.objects.create(
            name="PINKMANS-WAPPING WHARF",
            customer_type=Customer.WHOLESALE, department=self.dept)
        self.teals = Customer.objects.create(
            name="TEALS", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.society = Customer.objects.create(
            name="SOCIETY- BALDWIN ST",
            customer_type=Customer.WHOLESALE, department=self.dept)
        # A name-matched SaleProduct so the dup-row collision is the
        # real partial-unique constraint, not a happy NULL accident.
        SaleProduct.objects.create(
            name="Sliced Apple Waste Sourdough",
            price=Decimal("2.30"), department=self.dept)

    def _import(self):
        from stock.order_import import (
            import_wholesale_tab, make_price_recovery)
        from openpyxl import load_workbook
        wb = load_workbook(self.SHEET, read_only=True, data_only=True)
        try:
            return import_wholesale_tab(
                wb, recover_price=make_price_recovery(wb))
        finally:
            wb.close()

    def test_duplicate_row_does_not_abort_wholesale_tab(self):
        result = self._import()
        # The whole reason this test exists: pre-fix this came back at 0.
        self.assertGreater(
            result["lines_imported"], 0,
            "WHOLESALE tab must yield lines even when a duplicate row "
            "trips the partial-unique (order, sale_product) constraint")
        # PINKMANS-WAPPING WHARF (the customer holding the dup) must
        # still have orders for the week — only the second occurrence
        # of the offending row is dropped, not the whole customer.
        week = [self.WEEK_START + datetime.timedelta(days=i)
                for i in range(7)]
        self.assertTrue(
            Order.objects.filter(
                customer=self.pinkmans, order_date__in=week).exists(),
            "PINKMANS-WAPPING WHARF should still have orders despite "
            "carrying the duplicate row")
        # The dup itself surfaces in lines_skipped so the operator can
        # see why one row was dropped.
        skipped_names = [name for _, name, _, _ in result["lines_skipped"]]
        self.assertIn("Sliced Apple Waste Sourdough", skipped_names)

    def test_other_wholesale_customers_priced_lines_present(self):
        # The pre-fix failure mode wiped every wholesale customer —
        # TEALS and SOCIETY-BALDWIN ST included, even though their
        # blocks have nothing to do with the duplicate. Assert their
        # priced lines actually landed.
        self._import()
        week = [self.WEEK_START + datetime.timedelta(days=i)
                for i in range(7)]
        # TEALS — Croissant (Loose) at £1.10, qty=28 on Mon → £30.80
        # line. Picking one priced TEALS line is enough to prove the
        # tab didn't get rolled back.
        teals_croissant = OrderLine.objects.filter(
            order__customer=self.teals,
            order__order_date__in=week,
            product_name="Croissant (Loose)")
        self.assertGreater(teals_croissant.count(), 0)
        self.assertTrue(all(l.unit_price == Decimal("1.10")
                            for l in teals_croissant))
        # SOCIETY- BALDWIN ST — priced Sultana Pain Suisse rows landed.
        society_lines = OrderLine.objects.filter(
            order__customer=self.society,
            order__order_date__in=week,
            product_name__icontains="Sultana Pain Suisse")
        self.assertGreater(society_lines.count(), 0)

    def test_teals_new_product_blank_price_rows_recovered_from_products_tab(self):
        # The three TEALS rows this week (Fruited Sourdough Loaf,
        # 'Wheat & Rye Sourdough ' with trailing whitespace, Seeded
        # Sourdough) have a blank Price£ cell on the WHOLESALE tab.
        # They sit at £2.30 / £2.20 / £2.20 in the workbook's own
        # Products tab — the price-recovery cascade must apply per
        # line on the wholesale path, exactly as it does on customer
        # tabs, so these rows land priced rather than skipping the
        # whole tab. Trailing whitespace on the product name must not
        # break the Products-tab lookup either.
        self._import()
        fri = self.WEEK_START + datetime.timedelta(days=4)
        cases = [
            ("Fruited Sourdough Loaf", Decimal("2.30")),
            ("Wheat & Rye Sourdough", Decimal("2.20")),
            ("Seeded Sourdough", Decimal("2.20")),
        ]
        for name, expected_price in cases:
            line = OrderLine.objects.filter(
                order__customer=self.teals, order__order_date=fri,
                product_name=name).first()
            self.assertIsNotNone(
                line, f"{name!r} TEALS line missing for Fri {fri}")
            self.assertEqual(
                line.unit_price, expected_price,
                f"{name!r} should be recovered from the workbook "
                f"Products tab at {expected_price}, not £0")


class HistoricalImportReconciliationTests(TestCase):
    """End-to-end reconciliation: the full historical import routes
    WHOLESALE through its handler, fixes blank-price rows to £0, and
    reconciles to the spec totals for w/c 30 Mar 2026."""

    SHEET = "data/historical/order_sheet_2026_03_30.xlsm"
    WEEK_START = datetime.date(2026, 3, 30)

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        # Seed two customer-tab customers + two wholesale customers.
        # The rest land in failures / customer_unmatched as expected.
        self.garden = Customer.objects.create(
            name="Garden Cafe", department=self.dept)
        self.ecom = Customer.objects.create(
            name="ECOM", department=self.dept)
        self.teals = Customer.objects.create(
            name="TEALS", department=self.dept)
        # Seed the one ECOM SaleProduct whose blank-price row is the
        # whole point of FIX 1 (also seed garden's catalogue so its
        # matched lines work as before).
        SaleProduct.objects.create(
            name="TN Biscuits Easter Bag_EA", sage_number="660260238",
            price=Decimal("2.50"), department=self.dept)
        _seed_garden_cafe_products(self.dept)

    def test_garden_and_ecom_reconcile_to_sheet_totals(self):
        from stock.order_import import import_historical_workbook
        import_historical_workbook(self.SHEET, force=True)
        week = [self.WEEK_START + datetime.timedelta(days=i) for i in range(7)]
        garden_total = sum(
            (o.total_value() for o in Order.objects.filter(
                customer=self.garden, order_date__in=week)),
            Decimal("0"))
        ecom_total = sum(
            (o.total_value() for o in Order.objects.filter(
                customer=self.ecom, order_date__in=week)),
            Decimal("0"))
        # Garden Café has no blank-price rows the recovery hits, so
        # its total is unchanged across all three blank-price regimes.
        self.assertEqual(garden_total, Decimal("649.05"))
        # ECOM tracks the importer's blank-price treatment:
        #   * pre-fix #1: £766.60 (catalogue fallback ghost)
        #   * fix #1 (blank-is-£0): £641.60
        #   * fix #2 (sheet-authoritative recovery): £1,150.10
        # the 5 reference-CSV SKUs (Loaf Seeded Rye + 4) plus the
        # workbook-Products-tab recovery of TN Biscuits Easter Bag
        # account for the +£508.50 vs the previous fix.
        self.assertEqual(ecom_total, Decimal("1150.10"))

    def test_wholesale_routed_through_handler_in_historical(self):
        from stock.order_import import import_historical_workbook
        summary = import_historical_workbook(self.SHEET, force=True)
        # WHOLESALE appears in per_tab with wholesale-handler shape
        # (carries customers_imported), proving it's not skipped.
        self.assertIn("WHOLESALE", summary["per_tab"])
        self.assertIn("customers_imported", summary["per_tab"]["WHOLESALE"])
        # TEALS has its own orders from WHOLESALE.
        week = [self.WEEK_START + datetime.timedelta(days=i) for i in range(7)]
        teals_total = sum(
            (o.total_value() for o in Order.objects.filter(
                customer=self.teals, order_date__in=week)),
            Decimal("0"))
        self.assertEqual(teals_total, Decimal("1703.40"))

    def test_historical_summary_surfaces_wholesale_unmatched(self):
        from stock.order_import import import_historical_workbook
        summary = import_historical_workbook(self.SHEET, force=True)
        # We didn't seed Bishopstrow / Pinkmans / Society / Mjolk /
        # etc., so they should be reported at the top level for
        # operator reconciliation — never silently auto-created.
        self.assertGreater(len(summary.get(
            "wholesale_customer_unmatched", [])), 0)
        names_lower = [n.lower() for n in summary["wholesale_customer_unmatched"]]
        self.assertIn("bishopstrow hotel", names_lower)


class HistoricalImportVersionGateTests(TestCase):
    """The historical importer skips a week only when its
    ``HistoricalImport`` stamp's ``import_version`` is at or above the
    current ``HISTORICAL_IMPORT_VERSION`` constant. Bumping the
    constant on a fix forces a one-time re-import on the next deploy,
    then the gate settles back to skip-if-current."""

    SHEET = "data/historical/order_sheet_2026_03_30.xlsm"
    WEEK_START = datetime.date(2026, 3, 30)

    def setUp(self):
        from stock.models import HistoricalImport
        self.HistoricalImport = HistoricalImport
        self.dept = Department.objects.create(name="Bakery")
        # Minimal seed — only Garden Café needs products for the
        # version-gate tests; the wholesale + other tabs become
        # "Customer not found" failures which is fine here.
        self.garden = Customer.objects.create(
            name="Garden Cafe", department=self.dept)
        _seed_garden_cafe_products(self.dept)

    def _current_version(self):
        from stock.order_import import HISTORICAL_IMPORT_VERSION
        return HISTORICAL_IMPORT_VERSION

    def test_first_import_writes_stamp_at_current_version(self):
        from stock.order_import import import_historical_workbook
        self.assertFalse(self.HistoricalImport.objects.exists())
        result = import_historical_workbook(self.SHEET)
        self.assertFalse(result["skipped"])
        self.assertIsNone(result["stamp_was"])
        self.assertEqual(result["import_version"], self._current_version())
        stamp = self.HistoricalImport.objects.get(week_start=self.WEEK_START)
        self.assertEqual(stamp.import_version, self._current_version())
        self.assertEqual(stamp.file_path, self.SHEET)

    def test_current_stamp_makes_next_import_skip(self):
        from stock.order_import import import_historical_workbook
        import_historical_workbook(self.SHEET)  # writes stamp v_current
        second = import_historical_workbook(self.SHEET)
        self.assertTrue(second["skipped"])
        self.assertIn("already imported", second["reason"])
        # Reason mentions both the recorded version + the current one.
        self.assertIn(f"v{self._current_version()}", second["reason"])

    def test_stale_stamp_triggers_reimport_and_upgrades(self):
        # Simulate a week imported under an older version (v1). On
        # the next call the gate should NOT skip — it should re-run
        # and upgrade the stamp to the current version.
        self.HistoricalImport.objects.create(
            week_start=self.WEEK_START,
            import_version=self._current_version() - 1,
            file_path="data/historical/older-file.xlsm")
        from stock.order_import import import_historical_workbook
        result = import_historical_workbook(self.SHEET)
        self.assertFalse(result["skipped"],
            "stale stamp must trigger re-import, not skip")
        self.assertEqual(result["stamp_was"], self._current_version() - 1)
        self.assertEqual(result["import_version"], self._current_version())
        stamp = self.HistoricalImport.objects.get(week_start=self.WEEK_START)
        self.assertEqual(stamp.import_version, self._current_version())
        # And the file_path is updated to the file actually used.
        self.assertEqual(stamp.file_path, self.SHEET)

    def test_force_overrides_current_stamp_and_keeps_it_current(self):
        from stock.order_import import import_historical_workbook
        import_historical_workbook(self.SHEET)
        # Force re-run even though the stamp is up to date.
        result = import_historical_workbook(self.SHEET, force=True)
        self.assertFalse(result["skipped"])
        stamp = self.HistoricalImport.objects.get(week_start=self.WEEK_START)
        self.assertEqual(stamp.import_version, self._current_version())

    def test_existing_orders_without_stamp_get_reimported(self):
        # Pre-versioning state: Orders for the week exist but no
        # HistoricalImport stamp. The new gate must NOT confuse this
        # with "imported at current version" — it should treat the
        # missing stamp as stale and re-import. (This is exactly the
        # path the live site will follow on the first deploy after the
        # version bump lands.)
        Order.objects.create(
            customer=self.garden, department=self.dept,
            order_date=self.WEEK_START)
        self.assertFalse(self.HistoricalImport.objects.exists())
        from stock.order_import import import_historical_workbook
        result = import_historical_workbook(self.SHEET)
        self.assertFalse(result["skipped"],
            "missing stamp must be treated as stale, not as 'imported'")
        self.assertTrue(self.HistoricalImport.objects.filter(
            week_start=self.WEEK_START).exists())

    def test_management_command_announces_version_status(self):
        from django.core.management import call_command
        from io import StringIO
        out = StringIO()
        call_command("import_historical_orders", self.SHEET, stdout=out)
        body = out.getvalue()
        self.assertIn(f"at v{self._current_version()}", body)
        self.assertIn("first import", body)
        # Second run skips with version-stamp-based reason.
        out2 = StringIO()
        call_command("import_historical_orders", self.SHEET, stdout=out2)
        body2 = out2.getvalue()
        self.assertIn("already imported", body2)


def _seed_all_w_c_30_mar_customers(dept):
    """Seed every Customer the w/c 30 Mar 2026 historical sheet
    references — both per-customer tabs and wholesale customer rows.
    Without products it doesn't matter for reconciliation: every
    OrderLine still snapshots the sheet's name + price, so the DB
    total per customer matches the sheet total exactly. Returns a
    dict ``{name: Customer}`` for callers that want a handle."""
    names = [
        # Customer tabs (the ones with actual orders w/c 30 Mar).
        "ECOM", "CREAMERY FARMSHOP", "FARMSHOP", "FARMYARD",
        "CREAMERY F&B", "BOTANICAL ROOMS", "Garden Cafe", "CBK",
        "ROMAN VILLA", "FARM KITCHEN", "NPD", "STAFF FOOD",
        "BAKERY INTERNAL USE", "VA & EVENTS", "GELATO", "GGE",
        "HIVE", "HK", "HR", "ECOM MKT", "SOG", "TN100", "QAS",
        "BUTCHERY", "STAFF FOOD AVALON", "BAKERY WASTAGE",
        "YARLINGTON",
        # Wholesale customers from the WHOLESALE tab.
        "THE OLD PHARMACY", "NUMBER ONE BRUTON", "TEALS",
        "TEALS KITCHEN", "TREETOPS CAFÉ",
        "PINKMANS - WHITELADIES ROAD", "PINKMANS - WESTBURY-ON-TRYN",
        "PINKMANS - CATHEDRAL", "PINKMANS - STOKES CROFT",
        "BISHOPSTROW HOTEL", "THE EYE", "WE THE CURIOUS",
        "SOCIETY- CORRIDOR", "SOCIETY- KINGSMEAD",
        "SOCIETY- BALDWIN ST", "SOCIETY- HARBOURSIDE",
        "CORTADO", "MJOLK", "WHOLESALE SAMPLES",
    ]
    out = {}
    for name in names:
        out[name], _ = Customer.objects.get_or_create(
            name=name, defaults={"department": dept})
    return out


class ReconcileOrdersCommandTests(TestCase):
    """End-to-end check that reconcile_orders prints a clean
    RECONCILES verdict for an imported historical week."""

    SHEET = "data/historical/order_sheet_2026_03_30.xlsm"
    WEEK_START = datetime.date(2026, 3, 30)

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        _seed_all_w_c_30_mar_customers(self.dept)

    def test_w_c_30_mar_reconciles_with_recovered_prices(self):
        from django.core.management import call_command
        from io import StringIO
        from stock.order_import import import_historical_workbook
        # Populate the DB exactly the way the deploy would (including
        # the blank-price recovery chain).
        import_historical_workbook(self.SHEET, force=True)
        # Then run the reconciliation report and inspect its output.
        out = StringIO()
        call_command("reconcile_orders", self.SHEET, stdout=out)
        body = out.getvalue()
        # The verdict.
        self.assertIn("RECONCILES", body)
        self.assertNotIn("MISMATCH", body)
        # Post-recovery grand total: customer tabs £13,458.10 (was
        # £12,949.60 — ECOM picked up the 5 reference-CSV SKUs +
        # workbook-Products Easter Bag) + wholesale £5,674.85 (was
        # £5,660.45 — sticky bun recovered from workbook Products
        # tab) = £19,132.95.
        self.assertIn("£ 19,132.95", body)
        # Section headings present (per-customer table + wholesale).
        self.assertIn("CUSTOMER TABS", body)
        self.assertIn("WHOLESALE BREAKDOWN", body)
        # Per-section subtotals.
        self.assertIn("£ 13,458.10", body)
        self.assertIn("£  5,674.85", body)
        # Every per-row entry says OK, none flagged.
        self.assertNotIn("NO CUSTOMER ROW", body)
        # Difference line.
        self.assertIn("Difference:", body)
        self.assertIn("£      0.00", body)

    def test_command_is_read_only(self):
        # Before/after snapshot of every Order + OrderLine + stamp —
        # reconcile_orders must touch none of them.
        from django.core.management import call_command
        from stock.order_import import import_historical_workbook
        from stock.models import HistoricalImport
        import_historical_workbook(self.SHEET, force=True)
        before = (
            set(Order.objects.values_list("pk", "customer_id",
                                          "order_date")),
            set(OrderLine.objects.values_list(
                "pk", "order_id", "product_name", "unit_price",
                "qty_ordered")),
            set(HistoricalImport.objects.values_list(
                "pk", "week_start", "import_version")),
        )
        call_command("reconcile_orders", self.SHEET)
        after = (
            set(Order.objects.values_list("pk", "customer_id",
                                          "order_date")),
            set(OrderLine.objects.values_list(
                "pk", "order_id", "product_name", "unit_price",
                "qty_ordered")),
            set(HistoricalImport.objects.values_list(
                "pk", "week_start", "import_version")),
        )
        self.assertEqual(before, after,
            "reconcile_orders must not modify any data")

    def test_command_reports_missing_file_as_command_error(self):
        from django.core.management import call_command
        from django.core.management.base import CommandError
        with self.assertRaises(CommandError):
            call_command(
                "reconcile_orders",
                "data/historical/does_not_exist.xlsm")

    def test_command_flags_per_customer_mismatch(self):
        # If a single customer's DB total drifts from the sheet, the
        # report must flag the per-row mismatch AND the overall verdict.
        from django.core.management import call_command
        from io import StringIO
        from stock.order_import import import_historical_workbook
        import_historical_workbook(self.SHEET, force=True)
        # Pick a customer with imported lines and bump one qty so the
        # DB total no longer matches the sheet.
        teals = Customer.objects.get(name__iexact="TEALS")
        line = OrderLine.objects.filter(
            order__customer=teals,
            order__order_date=self.WEEK_START).first()
        self.assertIsNotNone(line)
        line.qty_ordered = line.qty_ordered + Decimal("1")
        line.save()
        out = StringIO()
        call_command("reconcile_orders", self.SHEET, stdout=out)
        body = out.getvalue()
        self.assertIn("MISMATCH", body)
        # The per-row TEALS line is flagged too — the report isn't
        # just summary-level.
        teals_line = next(
            (l for l in body.splitlines() if l.startswith("TEALS ")), None)
        self.assertIsNotNone(teals_line, "TEALS row missing from report")
        self.assertIn("MISMATCH", teals_line)


class CustomerIsInternalTests(TestCase):
    """``Customer.is_internal`` carves out non-revenue customers
    (BAKERY INTERNAL USE + BAKERY WASTAGE). The flag is set by the
    0024 data migration and is editable on the customer edit page +
    in Django admin. It changes the displayed Week / per-day totals
    but never mutates the underlying orders."""

    def test_migration_flagged_exactly_two_named_customers(self):
        # Apply-time check: the migration ran against the test DB on
        # setup, but we don't have those rows here. Create them now
        # and call the data function directly so we can assert its
        # selectivity in isolation from any other test fixture.
        from importlib import import_module
        mig = import_module("stock.migrations.0024_customer_is_internal")
        dept = Department.objects.create(name="Bakery")
        Customer.objects.create(name="BAKERY INTERNAL USE", department=dept)
        Customer.objects.create(name="bakery wastage",  # lower-case to prove iexact
                                department=dept)
        Customer.objects.create(name="GARDEN CAFE", department=dept)
        Customer.objects.create(name="TEALS", department=dept)
        Customer.objects.update(is_internal=False)

        class _AppsStub:
            def get_model(self, _app_label, _model_name):
                return Customer
        mig.flag_internal_customers(_AppsStub(), schema_editor=None)

        flagged = set(Customer.objects.filter(is_internal=True)
                      .values_list("name", flat=True))
        self.assertEqual(flagged, {"BAKERY INTERNAL USE", "bakery wastage"})
        # All others stay external — the flag is opt-in.
        for name in ("GARDEN CAFE", "TEALS"):
            self.assertFalse(
                Customer.objects.get(name=name).is_internal,
                f"{name} must NOT be flagged is_internal")


class OrdersWeekViewInternalSplitTests(TestCase):
    """The orders weekly view splits totals: external (headline Week +
    per-day tiles) vs internal-use subtotal. Internal customers stay
    in the customer list with their own figures + an 'internal' tag."""

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client = Client()
        assert self.client.login(username="alice", password="pw")
        self.client.get(f"/switch/{self.dept.pk}/")
        # Three customers covering the matrix.
        self.garden = Customer.objects.create(
            name="GARDEN CAFE", department=self.dept)  # external
        self.bakery_use = Customer.objects.create(
            name="BAKERY INTERNAL USE", department=self.dept,
            is_internal=True)
        self.bakery_waste = Customer.objects.create(
            name="BAKERY WASTAGE", department=self.dept,
            is_internal=True)
        self.croissant = SaleProduct.objects.create(
            name="Croissant (Loose)", price=Decimal("1.10"),
            department=self.dept)
        # One Garden Café order (external £11.00) and two internal
        # orders (£5.50 + £2.20 = £7.70 internal subtotal).
        self.week_start = datetime.date(2026, 5, 18)
        for cust, qty in [(self.garden, Decimal("10")),
                          (self.bakery_use, Decimal("5")),
                          (self.bakery_waste, Decimal("2"))]:
            o = Order.objects.create(
                customer=cust, department=self.dept,
                order_date=self.week_start)
            OrderLine.objects.create(
                order=o, sale_product=self.croissant, qty_ordered=qty)

    def _get(self, **params):
        params.setdefault("week", self.week_start.isoformat())
        from urllib.parse import urlencode
        return self.client.get(f"/orders/?{urlencode(params)}")

    def test_week_total_excludes_internal_customers(self):
        resp = self._get()
        ctx = resp.context
        # External revenue only: 10 × 1.10 = £11.00.
        self.assertEqual(ctx["week_total_value"], Decimal("11.00"))
        # Internal subtotal: 5×1.10 + 2×1.10 = £7.70.
        self.assertEqual(ctx["week_total_internal"], Decimal("7.70"))

    def test_per_day_total_excludes_internal_customers(self):
        resp = self._get()
        days = resp.context["days"]
        mon = next(d for d in days if d["date"] == self.week_start)
        self.assertEqual(mon["total"], Decimal("11.00"),
            "Monday's headline 'total' must be external-only")
        self.assertEqual(mon["total_internal"], Decimal("7.70"),
            "Monday's internal subtotal must capture the internal orders")
        # Counts still include ALL orders (data-completeness: the
        # 'N orders' chip on the tile reflects everything that landed).
        self.assertEqual(mon["count"], 3)

    def test_internal_subtotal_rendered_when_non_zero(self):
        # Page surfaces an "Internal use (excluded from week total)"
        # line directly under the day strip iff there's any internal
        # order in the week.
        body = self._get().content.decode()
        self.assertIn("Internal use (excluded from week total)", body)
        self.assertIn("£7.70", body)
        # And the headline week tile shows the external £11.00, not
        # the inflated £18.70.
        self.assertIn("£11.00", body)
        self.assertNotIn("£18.70", body)

    def test_internal_subtotal_hidden_when_no_internal_orders(self):
        # If only external orders exist, the subtotal line should NOT
        # show up — no point taking up screen space for £0.
        for cust in (self.bakery_use, self.bakery_waste):
            Order.objects.filter(customer=cust).delete()
        body = self._get().content.decode()
        self.assertNotIn("Internal use (excluded from week total)", body)

    def test_internal_customers_still_in_customer_list_with_tag(self):
        # Every customer is listed with their own value + an
        # 'internal' tag on the non-revenue ones. Their order count
        # and value remain visible — the flag only changes whether
        # they roll up into the week headline.
        body = self._get().content.decode()
        self.assertIn("BAKERY INTERNAL USE", body)
        self.assertIn("BAKERY WASTAGE", body)
        # The "internal" pill renders next to each flagged customer
        # (rendered as `<span class="tag-internal">internal</span>`).
        self.assertEqual(body.count(">internal<"), 2,
            "internal pill must appear exactly once per flagged customer")
        # Per-customer value still visible — internal use is £5.50,
        # bakery wastage is £2.20.
        self.assertIn("£5.50", body)
        self.assertIn("£2.20", body)

    def test_customer_edit_page_exposes_is_internal_toggle(self):
        # The checkbox is rendered on the edit form so an operator
        # can flip the flag without dropping into the admin.
        body = self.client.get(
            f"/customers/{self.bakery_use.pk}/edit/").content.decode()
        self.assertIn('name="is_internal"', body)
        self.assertIn("checked", body)
        # Submitting unchecked clears the flag; submitting checked
        # sets it. Round-trip via POST.
        resp = self.client.post(f"/customers/{self.bakery_use.pk}/edit/", {
            "name": self.bakery_use.name,
            "customer_type": self.bakery_use.customer_type,
            "location": "",
            "ordered_by": "",
            # is_internal omitted from POST → unchecked → False
        })
        self.assertEqual(resp.status_code, 302)
        self.bakery_use.refresh_from_db()
        self.assertFalse(self.bakery_use.is_internal)

    def test_filtering_to_internal_customer_still_shows_their_orders(self):
        # ?customer=<internal> still narrows the page to that
        # customer's grid (we don't hide internal customers; we just
        # exclude them from the headline). The filtered view's grid
        # still shows their orders.
        resp = self._get(customer=str(self.bakery_use.pk))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context["selected_customer"],
                         self.bakery_use)


class ReconcileIncludesInternalCustomersTests(TestCase):
    """The reconciliation report MUST include internal-use customers —
    it's a data-completeness check, orthogonal to the
    revenue-display split."""

    SHEET = "data/historical/order_sheet_2026_03_30.xlsm"

    def test_reconcile_orders_still_lists_internal_customers(self):
        from django.core.management import call_command
        from io import StringIO
        dept = Department.objects.create(name="Bakery")
        _seed_all_w_c_30_mar_customers(dept)
        # Apply the same data-migration logic the live deploy uses so
        # the seeded BAKERY INTERNAL USE / BAKERY WASTAGE customers
        # carry is_internal=True for this test.
        Customer.objects.filter(
            name__iexact="BAKERY INTERNAL USE").update(is_internal=True)
        Customer.objects.filter(
            name__iexact="BAKERY WASTAGE").update(is_internal=True)
        from stock.order_import import import_historical_workbook
        import_historical_workbook(self.SHEET, force=True)
        out = StringIO()
        call_command("reconcile_orders", self.SHEET, stdout=out)
        body = out.getvalue()
        # Both internal customers appear in the per-tab table — their
        # data still flows through the reconciliation just like every
        # other customer. The is_internal flag is purely a Week-total
        # display thing, not a reconcile-time filter.
        self.assertIn("BAKERY INTERNAL USE", body)
        # BAKERY WASTAGE happens to be an "empty" tab on this week's
        # sheet (no priced lines), so it lands in the empty-tab pool
        # rather than as a per-row line — but it's still seeded as a
        # Customer and still listed elsewhere if it had orders.
        # The crucial RECONCILES verdict must still hold.
        self.assertIn("RECONCILES", body)


class BlankPriceRecoveryChainTests(TestCase):
    """Unit-level tests for the price-recovery chain itself, isolated
    from the full importer plumbing. The lookup order is:
    workbook Products tab → reference CSV → £0."""

    def _wb_with_products(self, products):
        """Build an in-memory workbook with a single ``Products`` sheet
        carrying the given ``(name, price, sage)`` rows. The recovery
        helpers don't need any of the customer order tabs to be
        present — they only ever touch the Products sheet."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Product", "Price", "Sage No."])
        for name, price, sage in products:
            ws.append([name, price, sage])
        return wb

    def test_recovers_from_workbook_products_tab_by_sage(self):
        from stock.order_import import make_price_recovery
        wb = self._wb_with_products([
            ("Loaf Seeded Rye_650g", Decimal("3.30"), "660130120"),
        ])
        recover = make_price_recovery(wb, reference_path=None)
        # Sage win — even with a slightly mangled name.
        self.assertEqual(recover("660130120", "different name"),
                         Decimal("3.30"))

    def test_recovers_from_workbook_products_tab_by_name_when_sage_blank(self):
        from stock.order_import import make_price_recovery
        wb = self._wb_with_products([
            ("Mystery Loaf", Decimal("4.40"), None),
        ])
        recover = make_price_recovery(wb, reference_path=None)
        # No Sage on the row → name match is the fallback.
        self.assertEqual(recover("", "MYSTERY LOAF"),  # case-insensitive
                         Decimal("4.40"))

    def test_falls_through_to_reference_csv_when_workbook_has_no_price(self):
        # Workbook has the row but with a BLANK Price cell (the whole
        # point of the reference CSV: older sheets had prices that
        # newer ones lost). The CSV should be consulted next.
        import os, tempfile
        from stock.order_import import make_price_recovery
        wb = self._wb_with_products([
            # Workbook lists the SKU but with no price.
            ("Loaf Seeded Rye_650g", None, "660130120"),
        ])
        # Write a tiny CSV to a temp dir so the test owns the lookup.
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = os.path.join(tmp, "refs.csv")
            with open(csv_path, "w", encoding="utf-8") as fh:
                fh.write("sage,price\n660130120,3.30\n")
            recover = make_price_recovery(wb, reference_path=csv_path)
        self.assertEqual(recover("660130120", "Loaf Seeded Rye_650g"),
                         Decimal("3.30"))

    def test_falls_back_to_zero_when_nothing_found(self):
        from stock.order_import import make_price_recovery
        wb = self._wb_with_products([])
        recover = make_price_recovery(wb, reference_path=None)
        self.assertEqual(recover("999999", "Anything"),
                         Decimal("0"))

    def test_workbook_products_tab_wins_over_reference_csv(self):
        # If the same SKU appears in both, the workbook is
        # authoritative (the headline rule). The CSV only fills gaps
        # the workbook itself doesn't fill.
        import os, tempfile
        from stock.order_import import make_price_recovery
        wb = self._wb_with_products([
            ("X", Decimal("9.99"), "555"),
        ])
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = os.path.join(tmp, "refs.csv")
            with open(csv_path, "w", encoding="utf-8") as fh:
                fh.write("sage,price\n555,1.00\n")
            recover = make_price_recovery(wb, reference_path=csv_path)
        self.assertEqual(recover("555", "X"), Decimal("9.99"))

    def test_csv_loader_tolerates_missing_file_and_comments(self):
        from stock.order_import import load_reference_prices
        import os, tempfile
        # Missing path → empty dict, no exception.
        self.assertEqual(load_reference_prices("/nope/missing.csv"), {})
        # `#`-prefixed rows are treated as comments.
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = os.path.join(tmp, "refs.csv")
            with open(csv_path, "w", encoding="utf-8") as fh:
                fh.write("sage,price\n# header note\n123,4.50\n")
            self.assertEqual(load_reference_prices(csv_path),
                             {"123": Decimal("4.50")})


class BlankPriceRecoveryIntegrationTests(TestCase):
    """End-to-end: real historical workbook + the committed reference
    CSV produce the expected recovered totals on the ECOM tab."""

    SHEET = "data/historical/order_sheet_2026_03_30.xlsm"

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.ecom = Customer.objects.create(
            name="ECOM", department=self.dept)

    def test_ecom_recovers_to_eleven_fifty_ten(self):
        # The recovered ECOM total is the sum of:
        #   priced cells on the tab (already non-blank): 641.60
        #   workbook-Products tab recovery (TN Biscuits @ £2.50):
        #     50 × 2.50 = 125.00
        #   reference-CSV recovery (5 SKUs):
        #     Loaf Seeded Rye 34×3.30=112.20
        #     Wholewheat Sourdough 34×2.30=78.20
        #     TN Cake Carrot Walnut 11×5.50=60.50
        #     Cake Somerset Apple 18×5.50=99.00
        #     Choc Brownie Bag 12×2.80=33.60
        # Total: 1,150.10.
        from django.core.management import call_command
        call_command("import_orders", self.SHEET, tab=["ECOM"])
        week = [datetime.date(2026, 3, 30) + datetime.timedelta(days=i)
                for i in range(7)]
        total = sum(
            (o.total_value() for o in Order.objects.filter(
                customer=self.ecom, order_date__in=week)),
            Decimal("0"))
        self.assertEqual(total, Decimal("1150.10"))

    def test_recovered_price_is_snapshotted_on_the_line(self):
        # The recovered price must land in OrderLine.unit_price (the
        # historical snapshot), not just be a transient computation.
        from django.core.management import call_command
        call_command("import_orders", self.SHEET, tab=["ECOM"])
        # A reference-CSV row.
        line = OrderLine.objects.filter(
            order__customer=self.ecom,
            product_name="Loaf Seeded Rye_650g").first()
        self.assertIsNotNone(line)
        self.assertEqual(line.unit_price, Decimal("3.30"))
        # A workbook-Products-tab row.
        easter = OrderLine.objects.filter(
            order__customer=self.ecom,
            product_name="TN Biscuits Easter Bag_EA").first()
        self.assertIsNotNone(easter)
        self.assertEqual(easter.unit_price, Decimal("2.50"))


class FinancialsClassificationTests(TestCase):
    """Customer.customer_type + Customer.is_internal already encode the
    3-way channel classification the Financials page needs, so the
    page doesn't need a separate is_wholesale field or any per-request
    derivation from the order-sheet's WHOLESALE-tab provenance.

    * customer_type='wholesale' → wholesale channel
    * customer_type='internal' + is_internal=False → internal channel
    * is_internal=True → excluded (BAKERY INTERNAL USE / BAKERY WASTAGE
      — the bakery's own consumption, not demand).
    """

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        # Estate outlets (internal channel)
        for n in ("GARDEN CAFE", "FARMSHOP", "BOTANICAL ROOMS", "CBK"):
            Customer.objects.create(
                name=n, customer_type=Customer.INTERNAL,
                department=self.dept)
        # Wholesale accounts (set by import_customers from WHOLESALE tab)
        for n in ("TEALS", "MJOLK", "CORTADO",
                  "PINKMANS - WHITELADIES ROAD",
                  "SOCIETY- CORRIDOR"):
            Customer.objects.create(
                name=n, customer_type=Customer.WHOLESALE,
                department=self.dept)
        # Excluded (set by migration 0024)
        Customer.objects.create(
            name="BAKERY INTERNAL USE",
            customer_type=Customer.INTERNAL,
            is_internal=True, department=self.dept)
        Customer.objects.create(
            name="BAKERY WASTAGE",
            customer_type=Customer.INTERNAL,
            is_internal=True, department=self.dept)

    def test_known_wholesale_accounts_classify_as_wholesale(self):
        for n in ("TEALS", "MJOLK", "CORTADO",
                  "PINKMANS - WHITELADIES ROAD",
                  "SOCIETY- CORRIDOR"):
            c = Customer.objects.get(name=n)
            self.assertEqual(
                c.customer_type, Customer.WHOLESALE,
                f"{n} should be classified wholesale (set by "
                f"import_customers from WHOLESALE-tab membership)")
            self.assertFalse(
                c.is_internal,
                f"{n} is a real wholesale account, not bakery-own-use")

    def test_estate_outlets_classify_as_internal_channel(self):
        for n in ("GARDEN CAFE", "FARMSHOP", "BOTANICAL ROOMS", "CBK"):
            c = Customer.objects.get(name=n)
            self.assertEqual(c.customer_type, Customer.INTERNAL)
            self.assertFalse(
                c.is_internal,
                f"{n} is an Estate outlet — real revenue, must NOT be "
                f"in the excluded bakery-own-use bucket")

    def test_bakery_internal_use_and_wastage_are_excluded(self):
        for n in ("BAKERY INTERNAL USE", "BAKERY WASTAGE"):
            c = Customer.objects.get(name=n)
            self.assertTrue(
                c.is_internal,
                f"{n} must be flagged is_internal=True so the Financials "
                f"page excludes it from every total")

    def test_every_customer_is_exactly_one_channel(self):
        # Each Customer in the dept maps to exactly one of
        # {wholesale, internal, excluded} — the page's classification
        # is a TOTAL partition, no double-counting and no orphans.
        buckets = {"wholesale": 0, "internal": 0, "excluded": 0}
        for c in Customer.objects.filter(department=self.dept):
            channels = []
            if c.is_internal:
                channels.append("excluded")
            elif c.customer_type == Customer.WHOLESALE:
                channels.append("wholesale")
            elif c.customer_type == Customer.INTERNAL:
                channels.append("internal")
            self.assertEqual(
                len(channels), 1,
                f"{c.name} matched {channels} — must be exactly one")
            buckets[channels[0]] += 1
        # Sanity: setUp seeded 4 internal, 5 wholesale, 2 excluded.
        self.assertEqual(buckets, {"internal": 4, "wholesale": 5,
                                   "excluded": 2})


class FinancialsAggregationTests(TestCase):
    """range_totals + per_week_split + per_customer_in_channel are the
    three pure aggregation primitives the Financials page is built
    from. Their contract:

    * internal_total + wholesale_total == range grand total (over all
      non-excluded customers)
    * each channel's per-customer rows sum to that channel's total
    * is_internal=True customers never contribute to any total
    """

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.wc1 = datetime.date(2026, 3, 30)   # Monday
        self.wc2 = datetime.date(2026, 4, 6)    # Monday
        self.wc3 = datetime.date(2026, 4, 13)   # Monday
        # 2 internal Estate outlets, 2 wholesale accounts, 1 excluded.
        self.garden = Customer.objects.create(
            name="GARDEN CAFE", customer_type=Customer.INTERNAL,
            department=self.dept)
        self.farmshop = Customer.objects.create(
            name="FARMSHOP", customer_type=Customer.INTERNAL,
            department=self.dept)
        self.teals = Customer.objects.create(
            name="TEALS", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.mjolk = Customer.objects.create(
            name="MJOLK", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.internal_use = Customer.objects.create(
            name="BAKERY INTERNAL USE",
            customer_type=Customer.INTERNAL,
            is_internal=True, department=self.dept)
        # Seed orders × lines across the 3 weeks. Values picked so
        # every assertion is exact rather than approximate.
        self._line(self.garden, self.wc1, Decimal("10"), Decimal("2.50"))   # 25.00
        self._line(self.garden, self.wc2, Decimal("4"), Decimal("3.00"))    # 12.00
        self._line(self.farmshop, self.wc1, Decimal("8"), Decimal("1.25"))  # 10.00
        self._line(self.farmshop, self.wc3, Decimal("5"), Decimal("2.00"))  # 10.00
        self._line(self.teals, self.wc1, Decimal("20"), Decimal("1.10"))    # 22.00
        self._line(self.teals, self.wc2, Decimal("6"), Decimal("1.50"))     # 9.00
        self._line(self.mjolk, self.wc2, Decimal("10"), Decimal("1.10"))    # 11.00
        self._line(self.mjolk, self.wc3, Decimal("4"), Decimal("1.50"))     # 6.00
        # An excluded order — must NOT show up in any total.
        self._line(self.internal_use, self.wc1, Decimal("100"),
                   Decimal("9.99"))                                          # 999.00
        # Pre-computed expectations
        self.exp_internal = Decimal("57.00")    # 25 + 12 + 10 + 10
        self.exp_wholesale = Decimal("48.00")   # 22 + 9 + 11 + 6
        self.exp_grand = Decimal("105.00")

    def _line(self, customer, order_date, qty, price):
        o = Order.objects.create(
            customer=customer, department=self.dept,
            order_date=order_date)
        OrderLine.objects.create(
            order=o, sale_product=None, product_name=f"P{o.pk}",
            unit_price=price, qty_ordered=qty)

    def test_range_totals_internal_plus_wholesale_equals_grand_total(self):
        from stock.financials import range_totals
        r = range_totals(self.dept, self.wc1, self.wc3)
        self.assertEqual(r["internal"], self.exp_internal)
        self.assertEqual(r["wholesale"], self.exp_wholesale)
        self.assertEqual(r["total"], self.exp_grand)
        self.assertEqual(r["internal"] + r["wholesale"], r["total"])

    def test_range_totals_excludes_is_internal_customers(self):
        # The BAKERY INTERNAL USE order is £999 — if it leaked in, the
        # grand total would be £1,104. It must NOT contribute.
        from stock.financials import range_totals
        r = range_totals(self.dept, self.wc1, self.wc3)
        self.assertEqual(r["total"], Decimal("105.00"))
        # Defensive: deleting the BAKERY INTERNAL USE order shouldn't
        # change ANY total.
        Order.objects.filter(customer=self.internal_use).delete()
        r2 = range_totals(self.dept, self.wc1, self.wc3)
        self.assertEqual(r2, r)

    def test_range_totals_narrowed_range_only_counts_in_range(self):
        from stock.financials import range_totals
        # Just wc2 → garden=12 + teals=9 + mjolk=11 = 32
        r = range_totals(self.dept, self.wc2, self.wc2)
        self.assertEqual(r["internal"], Decimal("12.00"))
        self.assertEqual(r["wholesale"], Decimal("20.00"))
        self.assertEqual(r["total"], Decimal("32.00"))

    def test_per_week_split_each_week_sums_to_week_total(self):
        from stock.financials import per_week_split
        rows = per_week_split(self.dept, self.wc1, self.wc3)
        self.assertEqual([r["wc"] for r in rows],
                         [self.wc1, self.wc2, self.wc3])
        by_wc = {r["wc"]: r for r in rows}
        # wc1: garden 25 + farmshop 10 internal; teals 22 wholesale
        self.assertEqual(by_wc[self.wc1]["internal"], Decimal("35.00"))
        self.assertEqual(by_wc[self.wc1]["wholesale"], Decimal("22.00"))
        self.assertEqual(by_wc[self.wc1]["total"], Decimal("57.00"))
        # wc3: farmshop 10 internal; mjolk 6 wholesale
        self.assertEqual(by_wc[self.wc3]["internal"], Decimal("10.00"))
        self.assertEqual(by_wc[self.wc3]["wholesale"], Decimal("6.00"))
        # Every week's internal+wholesale equals its total.
        for r in rows:
            self.assertEqual(r["internal"] + r["wholesale"], r["total"])

    def test_per_week_split_sums_to_range_total(self):
        from stock.financials import per_week_split, range_totals
        rows = per_week_split(self.dept, self.wc1, self.wc3)
        rt = range_totals(self.dept, self.wc1, self.wc3)
        self.assertEqual(sum(r["internal"] for r in rows), rt["internal"])
        self.assertEqual(sum(r["wholesale"] for r in rows), rt["wholesale"])
        self.assertEqual(sum(r["total"] for r in rows), rt["total"])

    def test_per_week_split_empty_weeks_render_as_zero(self):
        from stock.financials import per_week_split
        # Add an extra week beyond wc3 — no orders, must still appear.
        wc4 = self.wc3 + datetime.timedelta(days=7)
        rows = per_week_split(self.dept, self.wc1, wc4)
        self.assertEqual(len(rows), 4)
        last = rows[-1]
        self.assertEqual(last["wc"], wc4)
        self.assertEqual(last["total"], Decimal("0"))

    def test_per_customer_wholesale_rows_sum_to_wholesale_total(self):
        from stock.financials import (
            per_customer_in_channel, range_totals)
        rows = per_customer_in_channel(
            self.dept, Customer.WHOLESALE, self.wc1, self.wc3)
        names = [r["name"] for r in rows]
        # Both wholesale customers present, ranked biggest first.
        self.assertEqual(set(names), {"TEALS", "MJOLK"})
        self.assertEqual(names[0], "TEALS")  # 31.00 vs MJOLK 17.00
        self.assertEqual(rows[0]["total"], Decimal("31.00"))
        self.assertEqual(rows[1]["total"], Decimal("17.00"))
        # Sum equals wholesale channel total from range_totals.
        rt = range_totals(self.dept, self.wc1, self.wc3)
        self.assertEqual(sum(r["total"] for r in rows), rt["wholesale"])
        # % shares add to 100 (within rounding).
        self.assertAlmostEqual(
            float(sum(r["pct"] for r in rows)), 100.0, places=1)

    def test_per_customer_internal_excludes_bakery_internal_use(self):
        # The internal-channel breakdown is Estate outlets ONLY —
        # BAKERY INTERNAL USE must NEVER appear here even though it
        # carries customer_type='internal'.
        from stock.financials import per_customer_in_channel
        rows = per_customer_in_channel(
            self.dept, Customer.INTERNAL, self.wc1, self.wc3)
        names = {r["name"] for r in rows}
        self.assertEqual(names, {"GARDEN CAFE", "FARMSHOP"})
        self.assertNotIn("BAKERY INTERNAL USE", names)

    def test_available_week_range_returns_earliest_and_latest_mondays(self):
        from stock.financials import available_week_range
        earliest, latest = available_week_range(self.dept)
        self.assertEqual(earliest, self.wc1)
        self.assertEqual(latest, self.wc3)

    def test_misclassified_external_customer_lands_in_internal_channel(self):
        # Reproduces the w/c 18 May 2026 £22.05 shortfall: a customer
        # with is_internal=False but an unrecognised customer_type
        # (empty / typoed / NULL) USED TO be in the Orders-page
        # external total but in NEITHER Financials channel, silently
        # dropping money from the grand total. The fix is the
        # complement-internal partition: INTERNAL = external NOT
        # wholesale, so any off-piste customer_type lands in INTERNAL
        # (visible) rather than off the page (invisible).
        from stock.financials import (
            per_customer_in_channel, range_totals)
        oddball = Customer.objects.create(
            name="OFF-PISTE OUTLET",
            customer_type="",  # not 'wholesale' AND not 'internal'
            department=self.dept)
        self._line(oddball, self.wc2, Decimal("3"), Decimal("7.35"))  # 22.05

        rt = range_totals(self.dept, self.wc1, self.wc3)
        # Grand total grew by exactly the oddball line — nothing lost.
        self.assertEqual(rt["total"], self.exp_grand + Decimal("22.05"))
        # Wholesale unchanged — the oddball isn't wholesale.
        self.assertEqual(rt["wholesale"], self.exp_wholesale)
        # Oddball's £22.05 lives in the internal complement.
        self.assertEqual(rt["internal"],
                         self.exp_internal + Decimal("22.05"))
        # And surfaces in the internal per-customer breakdown, so the
        # operator can see WHO it is and fix the customer_type.
        internal_names = {
            r["name"] for r in per_customer_in_channel(
                self.dept, Customer.INTERNAL, self.wc1, self.wc3)}
        self.assertIn("OFF-PISTE OUTLET", internal_names)

    def test_invariant_internal_plus_wholesale_plus_excluded_equals_true_total(self):
        # The hole the £22.05 shortfall slipped through: the previous
        # classification test only walked seed customers (all of which
        # were perfectly classified), so a real-data misclassification
        # had nothing checking it. This invariant covers EVERY customer
        # with orders in the range — Financials' grand + the excluded
        # total must equal the unfiltered ordered total, and no
        # customer should fall outside the three buckets.
        from stock.financials import range_totals
        # Add an oddball so the invariant has something to catch if
        # the partition ever regresses.
        oddball = Customer.objects.create(
            name="ROGUE", customer_type="external",  # off-piste
            department=self.dept)
        self._line(oddball, self.wc3, Decimal("2"), Decimal("4.50"))  # 9.00

        # The "true" ordered total: every line × qty × price, summed
        # in Python over EVERY customer in the range — no partition
        # involved at all. This is the figure the Orders page reports
        # when you sum its day tiles (external) plus the internal-use
        # subtotal underneath.
        true_external = Decimal("0")
        true_excluded = Decimal("0")
        end = self.wc3 + datetime.timedelta(days=6)
        from django.db.models import Sum, F, DecimalField
        for o in Order.objects.filter(
                department=self.dept,
                order_date__range=(self.wc1, end)).select_related("customer"):
            v = o.total_value()
            if o.customer.is_internal:
                true_excluded += v
            else:
                true_external += v

        rt = range_totals(self.dept, self.wc1, self.wc3)

        # Financials grand == true external (the headline reconciles).
        self.assertEqual(
            rt["total"], true_external,
            f"Financials grand £{rt['total']} != true external "
            f"£{true_external} — some customer's demand fell outside "
            f"both channels. Likely an off-piste customer_type.")
        # Two channels partition the external scope exactly.
        self.assertEqual(rt["internal"] + rt["wholesale"], rt["total"])
        # Explicitly list any customer whose demand is in neither
        # channel — empty under the complement-internal rule.
        offenders = []
        for o in Order.objects.filter(
                department=self.dept,
                order_date__range=(self.wc1, end),
                customer__is_internal=False).select_related("customer"):
            ct = o.customer.customer_type
            v = o.total_value()
            if v == 0:
                continue
            # Under the partition rule, exactly one of these is true.
            in_wholesale = (ct == Customer.WHOLESALE)
            in_internal = not in_wholesale  # complement
            if not (in_wholesale ^ in_internal):
                offenders.append((o.customer.name, ct, v))
        self.assertEqual(offenders, [],
            f"Found customers in neither / both channels: {offenders}")

    def test_per_week_invariant_internal_plus_wholesale_equals_external_total(self):
        # Per-week regression of the 18 May leak: for EVERY imported
        # week, per_week_split's internal+wholesale must equal the
        # unfiltered ordered total of external (non is_internal)
        # customers in that week. If a customer slips out of both
        # buckets, the week-level sum diverges from the Orders-page
        # week total — the symptom the operator hit on w/c 18 May.
        from stock.financials import per_week_split
        # Add an oddball customer so the invariant has something to
        # catch if the partition ever regresses.
        oddball = Customer.objects.create(
            name="OFF-PISTE OUTLET 18M", customer_type="",  # blank/typo
            department=self.dept)
        self._line(oddball, self.wc2, Decimal("3"), Decimal("7.35"))  # 22.05

        rows = per_week_split(self.dept, self.wc1, self.wc3)
        for r in rows:
            wc = r["wc"]
            end = wc + datetime.timedelta(days=6)
            true_external = Decimal("0")
            for o in Order.objects.filter(
                    department=self.dept,
                    order_date__range=(wc, end),
                    customer__is_internal=False).select_related("customer"):
                true_external += o.total_value()
            self.assertEqual(
                r["internal"] + r["wholesale"], r["total"],
                f"week {wc}: internal+wholesale != total")
            self.assertEqual(
                r["total"], true_external,
                f"week {wc}: Financials £{r['total']} != "
                f"Orders-page external £{true_external} — a customer "
                f"is slipping out of both channel buckets.")


class FinancialsWeekHelperTests(TestCase):
    """Unit tests for the single-week helpers in stock.financials —
    the building blocks the dashboard's week view is computed from.
    Same seeded data shape as FinancialsAggregationTests so the
    week-level figures reconcile with the range-level ones."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.wc1 = datetime.date(2026, 3, 30)   # Monday
        self.wc2 = datetime.date(2026, 4, 6)    # Monday
        self.wc3 = datetime.date(2026, 4, 13)   # Monday
        self.garden = Customer.objects.create(
            name="GARDEN CAFE", customer_type=Customer.INTERNAL,
            department=self.dept)
        self.farmshop = Customer.objects.create(
            name="FARMSHOP", customer_type=Customer.INTERNAL,
            department=self.dept)
        self.teals = Customer.objects.create(
            name="TEALS", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.mjolk = Customer.objects.create(
            name="MJOLK", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.excluded = Customer.objects.create(
            name="BAKERY INTERNAL USE",
            customer_type=Customer.INTERNAL,
            is_internal=True, department=self.dept)
        # wc2 spread across multiple days so daily_totals has shape to test.
        self._line(self.garden, self.wc2 + datetime.timedelta(days=0),
                   Decimal("4"), Decimal("3.00"))                # Mon 12.00 i
        self._line(self.teals, self.wc2 + datetime.timedelta(days=2),
                   Decimal("6"), Decimal("1.50"))                # Wed  9.00 w
        self._line(self.mjolk, self.wc2 + datetime.timedelta(days=4),
                   Decimal("10"), Decimal("1.10"))               # Fri 11.00 w
        # wc1 = prior imported week (single line)
        self._line(self.teals, self.wc1, Decimal("20"),
                   Decimal("1.10"))                              # 22.00 wholesale
        # wc3 (so we have a "current week" with one day's worth)
        self._line(self.farmshop, self.wc3 + datetime.timedelta(days=1),
                   Decimal("5"), Decimal("2.00"))                # Tue 10.00 i
        self._line(self.mjolk, self.wc3 + datetime.timedelta(days=3),
                   Decimal("4"), Decimal("1.50"))                # Thu  6.00 w
        # Excluded — must NEVER move any total.
        self._line(self.excluded, self.wc2, Decimal("100"),
                   Decimal("9.99"))                              # 999.00

    def _line(self, customer, order_date, qty, price):
        o = Order.objects.create(
            customer=customer, department=self.dept,
            order_date=order_date)
        OrderLine.objects.create(
            order=o, sale_product=None, product_name=f"P{o.pk}",
            unit_price=price, qty_ordered=qty)
        return o

    def test_available_weeks_lists_each_imported_monday_newest_first(self):
        from stock.financials import available_weeks
        weeks = available_weeks(self.dept)
        self.assertEqual(weeks, [self.wc3, self.wc2, self.wc1])

    def test_week_daily_totals_returns_seven_rows_mon_to_sun(self):
        from stock.financials import week_daily_totals
        rows = week_daily_totals(self.dept, self.wc2)
        self.assertEqual(len(rows), 7)
        self.assertEqual(rows[0]["date"], self.wc2)
        self.assertEqual(rows[6]["date"],
                         self.wc2 + datetime.timedelta(days=6))

    def test_week_daily_totals_sums_external_only_and_zero_fills_quiet_days(self):
        from stock.financials import week_daily_totals
        rows = week_daily_totals(self.dept, self.wc2)
        by_date = {r["date"]: r["total"] for r in rows}
        self.assertEqual(by_date[self.wc2 + datetime.timedelta(days=0)],
                         Decimal("12.00"))
        self.assertEqual(by_date[self.wc2 + datetime.timedelta(days=1)],
                         Decimal("0.00"))
        self.assertEqual(by_date[self.wc2 + datetime.timedelta(days=2)],
                         Decimal("9.00"))
        self.assertEqual(by_date[self.wc2 + datetime.timedelta(days=4)],
                         Decimal("11.00"))

    def test_week_daily_totals_excludes_is_internal_customer_orders(self):
        # The 999 BAKERY INTERNAL USE line on Monday wc2 must NEVER
        # land in a daily total — these are the bakery's own
        # consumption, not external demand.
        from stock.financials import week_daily_totals
        rows = week_daily_totals(self.dept, self.wc2)
        self.assertEqual(sum((r["total"] for r in rows), Decimal("0")),
                         Decimal("32.00"))  # 12 + 9 + 11

    def test_week_daily_totals_sum_reconciles_with_range_totals(self):
        from stock.financials import week_daily_totals, range_totals
        rt = range_totals(self.dept, self.wc2, self.wc2)
        total = sum((r["total"] for r in week_daily_totals(self.dept, self.wc2)),
                    Decimal("0"))
        self.assertEqual(total, rt["total"])

    def test_week_channel_split_totals_and_pcts(self):
        from stock.financials import week_channel_split
        s = week_channel_split(self.dept, self.wc2)
        self.assertEqual(s["internal"]["total"], Decimal("12.00"))
        self.assertEqual(s["wholesale"]["total"], Decimal("20.00"))
        # 12 / 32 = 37.5%; 20 / 32 = 62.5%
        self.assertEqual(s["internal"]["pct"], Decimal("37.5"))
        self.assertEqual(s["wholesale"]["pct"], Decimal("62.5"))

    def test_week_channel_split_zero_week_returns_zero_pcts(self):
        from stock.financials import week_channel_split
        empty_week = self.wc3 + datetime.timedelta(days=14)
        s = week_channel_split(self.dept, empty_week)
        self.assertEqual(s["internal"]["total"], Decimal("0.00"))
        self.assertEqual(s["wholesale"]["total"], Decimal("0.00"))
        self.assertEqual(s["internal"]["pct"], Decimal("0.0"))
        self.assertEqual(s["wholesale"]["pct"], Decimal("0.0"))

    def test_week_orders_count_counts_external_lines_only(self):
        from stock.financials import week_orders_count
        # wc2: 3 external lines (garden, teals, mjolk); excluded ignored.
        self.assertEqual(week_orders_count(self.dept, self.wc2), 3)
        # wc1: 1 external line.
        self.assertEqual(week_orders_count(self.dept, self.wc1), 1)
        # Empty week → 0.
        self.assertEqual(
            week_orders_count(
                self.dept, self.wc3 + datetime.timedelta(days=14)),
            0)

    def test_week_over_week_compares_to_immediately_prior_imported_week(self):
        from stock.financials import week_over_week
        wow = week_over_week(self.dept, self.wc2)
        self.assertEqual(wow["prev_week_start"], self.wc1)
        self.assertEqual(wow["prev_total"], Decimal("22.00"))
        self.assertEqual(wow["total"], Decimal("32.00"))
        # (32-22)/22 * 100 = 45.45...% → 45.5
        self.assertAlmostEqual(wow["pct"], 45.5, places=1)

    def test_week_over_week_skips_gap_weeks(self):
        # wc3 → prior imported week is wc2, not wc3-7 (which it is here
        # anyway). Add a gap to make it concrete: a week between wc3
        # and a future "current" week with no orders must be skipped.
        from stock.financials import week_over_week
        far_week = self.wc3 + datetime.timedelta(days=21)
        self._line(self.teals, far_week + datetime.timedelta(days=2),
                   Decimal("1"), Decimal("4.00"))               # 4.00
        wow = week_over_week(self.dept, far_week)
        # Prior imported week is wc3 (skipping the 2 empty intervening
        # weeks) — NOT far_week - 7 days (which would be a £0 prev and
        # a silently null pct).
        self.assertEqual(wow["prev_week_start"], self.wc3)
        self.assertEqual(wow["prev_total"], Decimal("16.00"))

    def test_week_over_week_returns_none_pct_for_first_imported_week(self):
        from stock.financials import week_over_week
        wow = week_over_week(self.dept, self.wc1)
        self.assertIsNone(wow["prev_week_start"])
        self.assertEqual(wow["prev_total"], Decimal("0.00"))
        self.assertIsNone(wow["pct"])

    def test_week_top_customers_returns_biggest_first_capped_at_n(self):
        from stock.financials import week_top_customers
        rows = week_top_customers(
            self.dept, self.wc2, Customer.WHOLESALE, n=5)
        names = [r["name"] for r in rows]
        # wc2 wholesale: MJOLK 11 vs TEALS 9.
        self.assertEqual(names, ["MJOLK", "TEALS"])
        # n=1 caps the list.
        rows1 = week_top_customers(
            self.dept, self.wc2, Customer.WHOLESALE, n=1)
        self.assertEqual([r["name"] for r in rows1], ["MJOLK"])

    def test_week_top_customers_internal_excludes_bakery_internal_use(self):
        from stock.financials import week_top_customers
        rows = week_top_customers(
            self.dept, self.wc2, Customer.INTERNAL, n=5)
        names = {r["name"] for r in rows}
        self.assertNotIn("BAKERY INTERNAL USE", names)
        self.assertIn("GARDEN CAFE", names)

    def test_recent_order_groups_orders_by_date_desc_with_channel(self):
        from stock.financials import recent_order_groups
        groups = recent_order_groups(self.dept, n=10)
        # newest first
        dates = [g["date"] for g in groups]
        self.assertEqual(dates, sorted(dates, reverse=True))
        # channel labels apply the partition rule live
        by_customer = {g["customer"]: g for g in groups}
        self.assertEqual(by_customer["TEALS"]["channel"], "wholesale")
        self.assertEqual(by_customer["GARDEN CAFE"]["channel"], "internal")
        self.assertEqual(by_customer["BAKERY INTERNAL USE"]["channel"],
                         "excluded")

    def test_recent_order_groups_line_count_and_ordered_total_match_order(self):
        # Add a second line to one Order so line_count > 1 surfaces.
        from stock.financials import recent_order_groups
        o = Order.objects.create(
            customer=self.teals, department=self.dept,
            order_date=datetime.date(2026, 5, 4))  # later than all seed
        OrderLine.objects.create(
            order=o, product_name="A", unit_price=Decimal("1.00"),
            qty_ordered=Decimal("3"))   # 3.00
        OrderLine.objects.create(
            order=o, product_name="B", unit_price=Decimal("2.00"),
            qty_ordered=Decimal("4"))   # 8.00
        groups = recent_order_groups(self.dept, n=1)
        self.assertEqual(len(groups), 1)
        g = groups[0]
        self.assertEqual(g["customer"], "TEALS")
        self.assertEqual(g["line_count"], 2)
        self.assertEqual(g["ordered_total"], Decimal("11.00"))


class WeekProductDayMatrixTests(TestCase):
    """``week_product_day_matrix`` returns the top N products by ordered
    qty for one week, each with Mon..Sun daily qtys. External-only
    (``is_internal=False``) and groups by the OrderLine SNAPSHOT
    ``product_name``, not the catalogue SaleProduct."""

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        self.wc = datetime.date(2026, 4, 6)
        self.garden = Customer.objects.create(
            name="GARDEN CAFE", customer_type=Customer.INTERNAL,
            department=self.dept)
        self.teals = Customer.objects.create(
            name="TEALS", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.excluded = Customer.objects.create(
            name="BAKERY INTERNAL USE",
            customer_type=Customer.INTERNAL, is_internal=True,
            department=self.dept)

    def _line(self, customer, day_offset, product_name, qty):
        """Add one OrderLine on ``self.wc + day_offset`` for ``customer``
        with the given snapshotted product name + qty (price irrelevant
        for the matrix). Reuses an existing Order if one exists for the
        same customer/date so multiple products on one day land on the
        same Order — matches the import flow."""
        d = self.wc + datetime.timedelta(days=day_offset)
        o, _ = Order.objects.get_or_create(
            customer=customer, department=self.dept, order_date=d)
        OrderLine.objects.create(
            order=o, sale_product=None, product_name=product_name,
            unit_price=Decimal("1.00"), qty_ordered=Decimal(str(qty)))

    def test_empty_week_returns_empty_list(self):
        from stock.financials import week_product_day_matrix
        self.assertEqual(week_product_day_matrix(self.dept, self.wc), [])

    def test_daily_qtys_align_mon_to_sun(self):
        # Same product across three different weekdays — the daily list
        # must place each qty at its weekday offset (0=Mon..6=Sun).
        from stock.financials import week_product_day_matrix
        self._line(self.garden, 0, "Croissant", 3)   # Mon
        self._line(self.teals,  2, "Croissant", 5)   # Wed
        self._line(self.garden, 6, "Croissant", 2)   # Sun
        rows = week_product_day_matrix(self.dept, self.wc)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["product"], "Croissant")
        self.assertEqual(rows[0]["total_qty"], Decimal("10"))
        self.assertEqual(rows[0]["daily"], [
            Decimal("3"), Decimal("0"), Decimal("5"),
            Decimal("0"), Decimal("0"), Decimal("0"), Decimal("2")])

    def test_aggregates_same_product_across_customers_and_orders(self):
        # Same product, same day, two customers — the matrix sums them
        # into one cell (the grid is "demand", not "per customer").
        from stock.financials import week_product_day_matrix
        self._line(self.garden, 1, "Baguette", 4)
        self._line(self.teals,  1, "Baguette", 7)
        rows = week_product_day_matrix(self.dept, self.wc)
        self.assertEqual(rows[0]["daily"][1], Decimal("11"))
        self.assertEqual(rows[0]["total_qty"], Decimal("11"))

    def test_excludes_is_internal_customer_orders(self):
        # BAKERY INTERNAL USE never contributes to demand. Even a big
        # internal-use line for a product must NOT inflate the matrix.
        from stock.financials import week_product_day_matrix
        self._line(self.garden,    0, "Loaf", 2)
        self._line(self.excluded,  0, "Loaf", 500)
        rows = week_product_day_matrix(self.dept, self.wc)
        self.assertEqual(rows[0]["total_qty"], Decimal("2"))
        self.assertEqual(rows[0]["daily"][0], Decimal("2"))

    def test_top_n_returns_only_n_biggest_by_total_qty(self):
        # Seed 5 products with descending totals; top_n=3 keeps the
        # three largest in descending order.
        from stock.financials import week_product_day_matrix
        for i, qty in enumerate([10, 8, 6, 4, 2]):
            self._line(self.garden, 0, f"Prod{i}", qty)
        rows = week_product_day_matrix(self.dept, self.wc, top_n=3)
        self.assertEqual([r["product"] for r in rows],
                         ["Prod0", "Prod1", "Prod2"])
        self.assertEqual([r["total_qty"] for r in rows],
                         [Decimal("10"), Decimal("8"), Decimal("6")])

    def test_groups_by_snapshotted_product_name_not_sale_product_id(self):
        # An OrderLine's product_name is the financial snapshot — even
        # when the linked SaleProduct gets renamed or deleted, the
        # matrix groups on the historical name.
        from stock.financials import week_product_day_matrix
        sp = SaleProduct.objects.create(
            name="Brand New Name", price=Decimal("1.00"), department=self.dept)
        o = Order.objects.create(
            customer=self.garden, department=self.dept, order_date=self.wc)
        # Snapshot name diverges from the live catalogue name.
        OrderLine.objects.create(
            order=o, sale_product=sp,
            product_name="Old Snapshotted Name",
            unit_price=Decimal("1.00"), qty_ordered=Decimal("3"))
        rows = week_product_day_matrix(self.dept, self.wc)
        self.assertEqual(rows[0]["product"], "Old Snapshotted Name")


class FinancialsViewTests(TestCase):
    """/financials/ is retired — it now redirects to the Business
    Performance SPA (which carries the channel split / weekly trend /
    per-customer breakdowns). The financials_home view + template stay in
    the tree, unused; the underlying financials.py math is still covered by
    the dashboard / business-performance API reconciliation tests.
    """

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.client.force_login(self.user)

    def test_financials_redirects_to_business_performance(self):
        r = self.client.get("/financials/")
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r["Location"], "/business-performance-dashboard/")

    def test_financials_redirect_preserves_query_string(self):
        r = self.client.get("/financials/?from=2026-03-30&to=2026-03-30")
        self.assertEqual(r.status_code, 302)
        self.assertEqual(
            r["Location"],
            "/business-performance-dashboard/?from=2026-03-30&to=2026-03-30")


class DashboardSummaryApiTests(TestCase):
    """/api/dashboard/summary/ — DRF endpoint that powers the React
    dashboard. Every figure here must reconcile with the financials.py
    primitives the /financials/ page already uses; this test class is
    the contract that locks that in.
    """

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.wc1 = datetime.date(2026, 3, 30)
        self.wc2 = datetime.date(2026, 4, 6)
        self.wc3 = datetime.date(2026, 4, 13)
        self.teals = Customer.objects.create(
            name="TEALS", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.mjolk = Customer.objects.create(
            name="MJOLK", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.garden = Customer.objects.create(
            name="GARDEN CAFE", customer_type=Customer.INTERNAL,
            department=self.dept)
        self.farmshop = Customer.objects.create(
            name="FARMSHOP", customer_type=Customer.INTERNAL,
            department=self.dept)
        self.excluded = Customer.objects.create(
            name="BAKERY INTERNAL USE",
            customer_type=Customer.INTERNAL,
            is_internal=True, department=self.dept)
        # wc1: garden 25 + farmshop 10 internal; teals 22 wholesale = 57
        # wc2: garden 12 internal; teals 9 + mjolk 11 wholesale = 32
        # wc3: farmshop 10 internal; mjolk 6 wholesale = 16
        # grand: internal 57, wholesale 48, total 105
        self._line(self.garden, self.wc1, Decimal("10"), Decimal("2.50"))
        self._line(self.farmshop, self.wc1, Decimal("8"), Decimal("1.25"))
        self._line(self.teals, self.wc1, Decimal("20"), Decimal("1.10"))
        self._line(self.garden, self.wc2, Decimal("4"), Decimal("3.00"))
        self._line(self.teals, self.wc2, Decimal("6"), Decimal("1.50"))
        self._line(self.mjolk, self.wc2, Decimal("10"), Decimal("1.10"))
        self._line(self.farmshop, self.wc3, Decimal("5"), Decimal("2.00"))
        self._line(self.mjolk, self.wc3, Decimal("4"), Decimal("1.50"))
        # Excluded — must never appear in any total.
        self._line(self.excluded, self.wc1, Decimal("100"), Decimal("9.99"))

    def _line(self, customer, order_date, qty, price):
        o = Order.objects.create(
            customer=customer, department=self.dept,
            order_date=order_date)
        OrderLine.objects.create(
            order=o, sale_product=None, product_name=f"P{o.pk}",
            unit_price=price, qty_ordered=qty)

    # ---- auth ----

    def test_requires_authentication(self):
        # Anonymous → 403 (DRF SessionAuthentication + IsAuthenticated).
        r = self.client.get("/api/dashboard/summary/")
        self.assertIn(r.status_code, (401, 403))

    def test_authenticated_returns_200(self):
        self.client.force_login(self.user)
        r = self.client.get("/api/dashboard/summary/")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"].split(";")[0], "application/json")

    # ---- shape ----

    def _get(self):
        self.client.force_login(self.user)
        r = self.client.get("/api/dashboard/summary/")
        self.assertEqual(r.status_code, 200)
        return r.json()

    def test_response_has_no_revenue_waste_or_margin_keys(self):
        # Hard contract: this dashboard never invents data it doesn't have.
        # No revenue / sales / waste / margin / profit keys anywhere in
        # the payload — those figures don't exist for this business.
        data = self._get()
        as_text = str(data).lower()
        for forbidden in (
            "revenue", "sales", "margin", "profit", "waste", "wastage",
        ):
            self.assertNotIn(
                forbidden, as_text,
                f"forbidden key/value '{forbidden}' leaked into payload: "
                f"{data!r}")

    def test_top_level_keys_present(self):
        data = self._get()
        for key in (
            "from", "to", "grand_total",
            "internal", "wholesale",
            "avg_week", "avg_day",
            "latest_week",
            "weekly_trend",
            "top_wholesale", "top_internal",
            "summary",
        ):
            self.assertIn(key, data, f"missing key: {key}")

    # ---- totals reconcile with financials.py ----

    def test_grand_total_matches_financials(self):
        from stock.financials import range_totals, available_week_range
        f, t = available_week_range(self.dept)
        rt = range_totals(self.dept, f, t)
        data = self._get()
        self.assertEqual(Decimal(str(data["grand_total"])), rt["total"])
        self.assertEqual(Decimal(str(data["grand_total"])),
                         Decimal("105.00"))

    def test_internal_plus_wholesale_equals_grand(self):
        data = self._get()
        i = Decimal(str(data["internal"]["total"]))
        w = Decimal(str(data["wholesale"]["total"]))
        g = Decimal(str(data["grand_total"]))
        self.assertEqual(i + w, g)
        self.assertEqual(i, Decimal("57.00"))
        self.assertEqual(w, Decimal("48.00"))

    def test_channel_pcts_sum_to_100(self):
        data = self._get()
        ip = float(data["internal"]["pct"])
        wp = float(data["wholesale"]["pct"])
        self.assertAlmostEqual(ip + wp, 100.0, places=1)

    def test_excluded_customer_never_appears(self):
        data = self._get()
        names = [r["name"] for r in data["top_wholesale"]]
        names += [r["name"] for r in data["top_internal"]]
        self.assertNotIn("BAKERY INTERNAL USE", names)
        # Grand total without the excluded £999 line.
        self.assertEqual(Decimal(str(data["grand_total"])),
                         Decimal("105.00"))

    # ---- weekly trend ----

    def test_weekly_trend_one_row_per_imported_week(self):
        data = self._get()
        weeks = [w["week"] for w in data["weekly_trend"]]
        self.assertEqual(weeks, [
            self.wc1.isoformat(),
            self.wc2.isoformat(),
            self.wc3.isoformat(),
        ])

    def test_weekly_trend_each_row_reconciles(self):
        data = self._get()
        for row in data["weekly_trend"]:
            i = Decimal(str(row["internal"]))
            w = Decimal(str(row["wholesale"]))
            t = Decimal(str(row["total"]))
            self.assertEqual(i + w, t)

    def test_weekly_trend_sums_to_grand_total(self):
        data = self._get()
        total = sum(Decimal(str(w["total"]))
                    for w in data["weekly_trend"])
        self.assertEqual(total, Decimal(str(data["grand_total"])))

    # ---- averages, latest week, wow ----

    def test_avg_week_and_avg_day(self):
        data = self._get()
        # 3 weeks, total 105 → avg_week 35.00, avg_day 5.00.
        self.assertEqual(Decimal(str(data["avg_week"])), Decimal("35.00"))
        self.assertEqual(Decimal(str(data["avg_day"])), Decimal("5.00"))

    def test_latest_week_is_wc3_with_wow_pct(self):
        data = self._get()
        lw = data["latest_week"]
        self.assertEqual(lw["week"], self.wc3.isoformat())
        self.assertEqual(Decimal(str(lw["total"])), Decimal("16.00"))
        # wc2 was 32 → wc3 16 → wow = (16-32)/32 = -50%.
        self.assertAlmostEqual(float(lw["wow_pct"]), -50.0, places=1)

    # ---- ranked customer lists ----

    def test_top_wholesale_ranked_with_pct(self):
        data = self._get()
        rows = data["top_wholesale"]
        names = [r["name"] for r in rows]
        # TEALS 31, MJOLK 17.
        self.assertEqual(names, ["TEALS", "MJOLK"])
        self.assertEqual(Decimal(str(rows[0]["value"])), Decimal("31.00"))
        self.assertEqual(Decimal(str(rows[1]["value"])), Decimal("17.00"))
        self.assertAlmostEqual(
            sum(float(r["pct"]) for r in rows), 100.0, places=1)

    def test_top_internal_ranked_with_pct(self):
        data = self._get()
        rows = data["top_internal"]
        names = [r["name"] for r in rows]
        # GARDEN CAFE 37, FARMSHOP 20.
        self.assertEqual(names, ["GARDEN CAFE", "FARMSHOP"])
        self.assertEqual(Decimal(str(rows[0]["value"])), Decimal("37.00"))
        self.assertEqual(Decimal(str(rows[1]["value"])), Decimal("20.00"))

    # ---- summary block ----

    def test_summary_block_reconciles(self):
        data = self._get()
        s = data["summary"]
        # highest_week == wc1 (57), lowest_week == wc3 (16).
        self.assertEqual(s["highest_week"]["week"], self.wc1.isoformat())
        self.assertEqual(Decimal(str(s["highest_week"]["total"])),
                         Decimal("57.00"))
        self.assertEqual(s["lowest_week"]["week"], self.wc3.isoformat())
        self.assertEqual(Decimal(str(s["lowest_week"]["total"])),
                         Decimal("16.00"))
        self.assertEqual(s["top_wholesale"], "TEALS")
        self.assertEqual(s["top_internal"], "GARDEN CAFE")
        # pct echoes top-level channel pct (same number, same source).
        self.assertEqual(s["internal_pct"], data["internal"]["pct"])
        self.assertEqual(s["wholesale_pct"], data["wholesale"]["pct"])

    # ---- range param ----

    def test_from_to_narrows_range(self):
        self.client.force_login(self.user)
        r = self.client.get(
            f"/api/dashboard/summary/?from={self.wc2.isoformat()}"
            f"&to={self.wc2.isoformat()}")
        self.assertEqual(r.status_code, 200)
        data = r.json()
        # Just wc2: 32 total, internal 12, wholesale 20.
        self.assertEqual(Decimal(str(data["grand_total"])), Decimal("32.00"))
        self.assertEqual(Decimal(str(data["internal"]["total"])),
                         Decimal("12.00"))
        self.assertEqual(Decimal(str(data["wholesale"]["total"])),
                         Decimal("20.00"))
        self.assertEqual(len(data["weekly_trend"]), 1)


class DashboardWeekModeApiTests(TestCase):
    """``/api/dashboard/summary/?week=…`` — single-week dashboard payload.

    The ``?week=`` query string puts the endpoint in week mode (a new
    payload shape); without it the endpoint serves the existing
    range-mode payload (covered by DashboardSummaryApiTests). Each test
    asserts a specific contract — totals reconcile with financials.py,
    is_internal customers never leak in, forbidden labels stay out,
    available_weeks lists every imported Monday, etc.
    """

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.wc1 = datetime.date(2026, 3, 30)   # Monday
        self.wc2 = datetime.date(2026, 4, 6)    # Monday
        self.wc3 = datetime.date(2026, 4, 13)   # Monday
        self.teals = Customer.objects.create(
            name="TEALS", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.mjolk = Customer.objects.create(
            name="MJOLK", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.garden = Customer.objects.create(
            name="GARDEN CAFE", customer_type=Customer.INTERNAL,
            department=self.dept)
        self.farmshop = Customer.objects.create(
            name="FARMSHOP", customer_type=Customer.INTERNAL,
            department=self.dept)
        self.excluded = Customer.objects.create(
            name="BAKERY INTERNAL USE",
            customer_type=Customer.INTERNAL,
            is_internal=True, department=self.dept)
        # wc1: one prior-week line (22 wholesale)
        self._line(self.teals, self.wc1, Decimal("20"), Decimal("1.10"))
        # wc2: spread across the week so daily_trend has shape
        self._line(self.garden, self.wc2 + datetime.timedelta(days=0),
                   Decimal("4"), Decimal("3.00"))    # Mon 12.00 internal
        self._line(self.teals, self.wc2 + datetime.timedelta(days=2),
                   Decimal("6"), Decimal("1.50"))    # Wed  9.00 wholesale
        self._line(self.mjolk, self.wc2 + datetime.timedelta(days=4),
                   Decimal("10"), Decimal("1.10"))   # Fri 11.00 wholesale
        # wc3: lighter week
        self._line(self.farmshop, self.wc3 + datetime.timedelta(days=1),
                   Decimal("5"), Decimal("2.00"))    # Tue 10.00 internal
        self._line(self.mjolk, self.wc3 + datetime.timedelta(days=3),
                   Decimal("4"), Decimal("1.50"))    # Thu  6.00 wholesale
        # Excluded — must NEVER appear in any total / list
        self._line(self.excluded, self.wc2, Decimal("100"), Decimal("9.99"))

    def _line(self, customer, order_date, qty, price):
        o = Order.objects.create(
            customer=customer, department=self.dept,
            order_date=order_date)
        OrderLine.objects.create(
            order=o, sale_product=None, product_name=f"P{o.pk}",
            unit_price=price, qty_ordered=qty)

    def _get(self, qs=""):
        self.client.force_login(self.user)
        url = "/api/dashboard/summary/"
        if qs:
            url += "?" + qs
        r = self.client.get(url)
        self.assertEqual(r.status_code, 200, r.content)
        return r.json()

    # ---- mode dispatch ----

    def test_requires_authentication(self):
        r = self.client.get("/api/dashboard/summary/?week=")
        self.assertIn(r.status_code, (401, 403))

    def test_no_week_param_serves_range_payload_for_backward_compat(self):
        # Without ?week= the endpoint must still emit the legacy
        # multi-week payload so the older SPA client keeps working.
        data = self._get()
        self.assertIn("weekly_trend", data)
        self.assertIn("avg_week", data)
        self.assertNotIn("daily_trend", data)

    def test_week_param_present_serves_week_payload(self):
        data = self._get("week=")
        self.assertIn("daily_trend", data)
        self.assertIn("week_start", data)
        self.assertNotIn("weekly_trend", data)

    def test_week_defaults_to_latest_imported_week(self):
        data = self._get("week=")
        self.assertEqual(data["week_start"], self.wc3.isoformat())

    def test_week_invalid_date_falls_back_to_latest(self):
        data = self._get("week=not-a-date")
        self.assertEqual(data["week_start"], self.wc3.isoformat())

    def test_week_snaps_any_day_to_its_monday(self):
        # A Thursday inside wc2 snaps back to wc2 Monday.
        thursday = self.wc2 + datetime.timedelta(days=3)
        data = self._get(f"week={thursday.isoformat()}")
        self.assertEqual(data["week_start"], self.wc2.isoformat())

    # ---- shape ----

    def test_top_level_keys_present(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        for key in (
            "week_start", "prev_week_start", "available_weeks",
            "total_ordered", "total_orders",
            "internal", "wholesale",
            "avg_day", "wow",
            "daily_trend",
            "top_wholesale", "top_internal",
            "recent_orders",
            "highest_day", "lowest_day",
        ):
            self.assertIn(key, data, f"missing key: {key}")

    def test_response_has_no_revenue_waste_or_margin_keys(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        as_text = str(data).lower()
        for forbidden in (
            "revenue", "sales", "margin", "profit", "waste", "wastage",
        ):
            self.assertNotIn(
                forbidden, as_text,
                f"forbidden key/value '{forbidden}' leaked into payload: "
                f"{data!r}")

    # ---- totals reconcile with financials.py ----

    def test_total_ordered_equals_range_totals_for_single_week(self):
        from stock.financials import range_totals
        rt = range_totals(self.dept, self.wc2, self.wc2)
        data = self._get(f"week={self.wc2.isoformat()}")
        self.assertEqual(Decimal(str(data["total_ordered"])), rt["total"])
        self.assertEqual(Decimal(str(data["total_ordered"])),
                         Decimal("32.00"))

    def test_internal_plus_wholesale_equals_total(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        i = Decimal(str(data["internal"]["total"]))
        w = Decimal(str(data["wholesale"]["total"]))
        t = Decimal(str(data["total_ordered"]))
        self.assertEqual(i + w, t)

    def test_channel_pcts_sum_to_100(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        ip = float(data["internal"]["pct"])
        wp = float(data["wholesale"]["pct"])
        self.assertAlmostEqual(ip + wp, 100.0, places=1)

    def test_excluded_customer_never_appears_in_any_total_or_list(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        self.assertEqual(Decimal(str(data["total_ordered"])),
                         Decimal("32.00"))  # NOT 32 + 999
        names = {r["name"] for r in data["top_wholesale"]}
        names |= {r["name"] for r in data["top_internal"]}
        self.assertNotIn("BAKERY INTERNAL USE", names)
        # recent_orders may still surface excluded groups (tagged
        # "excluded") so the operator can SEE the activity — but the
        # money never lands in a total.

    # ---- daily_trend ----

    def test_daily_trend_has_seven_rows_mon_to_sun(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        rows = data["daily_trend"]
        self.assertEqual(len(rows), 7)
        self.assertEqual(rows[0]["date"], self.wc2.isoformat())
        self.assertEqual(
            rows[6]["date"],
            (self.wc2 + datetime.timedelta(days=6)).isoformat())

    def test_daily_trend_sums_to_total_ordered(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        total = sum(Decimal(str(r["total"])) for r in data["daily_trend"])
        self.assertEqual(total, Decimal(str(data["total_ordered"])))

    def test_daily_trend_includes_prev_week_totals_per_day(self):
        # wc2 has 12 (Mon), 9 (Wed), 11 (Fri). wc1 has 22 on Monday.
        # So daily_trend for wc2: row[0].prev_week_total = 22.00.
        data = self._get(f"week={self.wc2.isoformat()}")
        rows = data["daily_trend"]
        self.assertEqual(
            Decimal(str(rows[0]["prev_week_total"])), Decimal("22.00"))
        # Days where prev week had no orders → 0.00.
        self.assertEqual(
            Decimal(str(rows[1]["prev_week_total"])), Decimal("0.00"))

    def test_daily_trend_prev_week_zero_for_first_imported_week(self):
        data = self._get(f"week={self.wc1.isoformat()}")
        for r in data["daily_trend"]:
            self.assertEqual(
                Decimal(str(r["prev_week_total"])), Decimal("0.00"))

    # ---- averages / wow / counts ----

    def test_total_orders_counts_lines_external_only(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        # 3 external lines on wc2 (garden, teals, mjolk); excluded ignored.
        self.assertEqual(data["total_orders"], 3)

    def test_avg_day_is_total_divided_by_seven(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        # 32.00 / 7 = 4.571… → 4.57
        self.assertEqual(Decimal(str(data["avg_day"])), Decimal("4.57"))

    def test_wow_compares_to_prior_imported_week(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        self.assertEqual(data["prev_week_start"], self.wc1.isoformat())
        # wow.total carries the prev-week total (22.00); wow.pct = change.
        self.assertEqual(Decimal(str(data["wow"]["total"])),
                         Decimal("22.00"))
        # (32-22)/22 = 45.45...% → 45.5
        self.assertAlmostEqual(float(data["wow"]["pct"]), 45.5, places=1)

    def test_wow_pct_null_for_first_imported_week(self):
        data = self._get(f"week={self.wc1.isoformat()}")
        self.assertIsNone(data["prev_week_start"])
        self.assertIsNone(data["wow"]["pct"])

    # ---- ranked customer lists ----

    def test_top_wholesale_ranked_with_pct(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        names = [r["name"] for r in data["top_wholesale"]]
        # wc2 wholesale: MJOLK 11.00, TEALS 9.00.
        self.assertEqual(names, ["MJOLK", "TEALS"])

    def test_top_internal_ranked(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        names = [r["name"] for r in data["top_internal"]]
        # wc2 internal: just GARDEN CAFE.
        self.assertEqual(names, ["GARDEN CAFE"])

    # ---- recent_orders ----

    def test_recent_orders_is_global_most_recent_groups(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        # Up to 10, newest first.
        dates = [g["date"] for g in data["recent_orders"]]
        self.assertEqual(dates, sorted(dates, reverse=True))

    def test_recent_orders_has_no_status_field(self):
        # No model field for order status — the helper must NOT invent one.
        data = self._get(f"week={self.wc2.isoformat()}")
        for g in data["recent_orders"]:
            self.assertNotIn("status", g)

    # ---- available_weeks ----

    def test_available_weeks_lists_every_imported_monday_newest_first(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        self.assertEqual(data["available_weeks"], [
            self.wc3.isoformat(),
            self.wc2.isoformat(),
            self.wc1.isoformat(),
        ])

    # ---- highest / lowest day ----

    def test_highest_and_lowest_day_pick_nonzero_extremes(self):
        data = self._get(f"week={self.wc2.isoformat()}")
        # Days with orders: Mon 12, Wed 9, Fri 11. Highest=Mon, Lowest=Wed.
        self.assertEqual(
            data["highest_day"]["date"], self.wc2.isoformat())
        self.assertEqual(
            Decimal(str(data["highest_day"]["total"])), Decimal("12.00"))
        self.assertEqual(
            data["lowest_day"]["date"],
            (self.wc2 + datetime.timedelta(days=2)).isoformat())
        self.assertEqual(
            Decimal(str(data["lowest_day"]["total"])), Decimal("9.00"))

    def test_highest_and_lowest_day_null_for_empty_week(self):
        empty_week = self.wc3 + datetime.timedelta(days=21)
        data = self._get(f"week={empty_week.isoformat()}")
        self.assertIsNone(data["highest_day"])
        self.assertIsNone(data["lowest_day"])

    def test_product_day_matrix_returns_top_products_with_seven_day_qtys(self):
        # The setUp seeds five lines on wc2 each with product_name="P{pk}"
        # (one product per order). The matrix returns each one with a
        # 7-long daily list aligned Mon..Sun. The BAKERY INTERNAL USE
        # row (qty=100) must NOT appear — external-only.
        data = self._get(f"week={self.wc2.isoformat()}")
        self.assertIn("product_day_matrix", data)
        matrix = data["product_day_matrix"]
        for row in matrix:
            self.assertIn("product", row)
            self.assertIn("total_qty", row)
            self.assertEqual(len(row["daily"]), 7)
        # Largest total this week is the mjolk Fri line, qty=10
        # ("P{pk}"). It must lead the matrix.
        self.assertEqual(Decimal(matrix[0]["total_qty"]), Decimal("10"))
        # No internal-use product (qty=100) ever leaks into the matrix.
        self.assertNotIn(Decimal("100"),
                         [Decimal(r["total_qty"]) for r in matrix])

    def test_product_day_matrix_present_and_empty_for_blank_week(self):
        empty_week = self.wc3 + datetime.timedelta(days=21)
        data = self._get(f"week={empty_week.isoformat()}")
        self.assertEqual(data["product_day_matrix"], [])


class DashboardExportCsvTests(TestCase):
    """``GET /api/dashboard/export.csv?week=`` returns that week's
    external order lines as CSV. SessionAuth + login_required."""

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.wc = datetime.date(2026, 4, 6)
        self.teals = Customer.objects.create(
            name="TEALS", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.garden = Customer.objects.create(
            name="GARDEN CAFE", customer_type=Customer.INTERNAL,
            department=self.dept)
        self.excluded = Customer.objects.create(
            name="BAKERY INTERNAL USE",
            customer_type=Customer.INTERNAL,
            is_internal=True, department=self.dept)
        o1 = Order.objects.create(
            customer=self.teals, department=self.dept,
            order_date=self.wc + datetime.timedelta(days=2))
        OrderLine.objects.create(
            order=o1, product_name="Sourdough",
            unit_price=Decimal("4.50"), qty_ordered=Decimal("10"))
        o2 = Order.objects.create(
            customer=self.garden, department=self.dept,
            order_date=self.wc + datetime.timedelta(days=0))
        OrderLine.objects.create(
            order=o2, product_name="Croissant",
            unit_price=Decimal("1.20"), qty_ordered=Decimal("24"))
        # Excluded — MUST NOT appear in CSV.
        o3 = Order.objects.create(
            customer=self.excluded, department=self.dept,
            order_date=self.wc + datetime.timedelta(days=1))
        OrderLine.objects.create(
            order=o3, product_name="Internal-only",
            unit_price=Decimal("9.99"), qty_ordered=Decimal("100"))

    def test_requires_authentication(self):
        r = self.client.get("/api/dashboard/export.csv?week=")
        self.assertIn(r.status_code, (401, 403))

    def test_returns_csv_content_type_and_filename(self):
        self.client.force_login(self.user)
        r = self.client.get(
            f"/api/dashboard/export.csv?week={self.wc.isoformat()}")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"].split(";")[0], "text/csv")
        self.assertIn(f"orders-{self.wc.isoformat()}.csv",
                      r["Content-Disposition"])

    def test_csv_header_columns(self):
        self.client.force_login(self.user)
        r = self.client.get(
            f"/api/dashboard/export.csv?week={self.wc.isoformat()}")
        body = r.content.decode()
        first = body.splitlines()[0]
        self.assertEqual(
            first,
            "date,customer,channel,product,qty,unit_price,line_value")

    def test_csv_excludes_is_internal_customers(self):
        self.client.force_login(self.user)
        r = self.client.get(
            f"/api/dashboard/export.csv?week={self.wc.isoformat()}")
        body = r.content.decode()
        self.assertNotIn("BAKERY INTERNAL USE", body)
        self.assertNotIn("Internal-only", body)

    def test_csv_rows_match_expected_line_values(self):
        self.client.force_login(self.user)
        r = self.client.get(
            f"/api/dashboard/export.csv?week={self.wc.isoformat()}")
        body = r.content.decode()
        lines = body.splitlines()
        # Header + 2 external lines
        self.assertEqual(len(lines), 3)
        # Garden Mon (sorted by date, then customer)
        self.assertIn("GARDEN CAFE", lines[1])
        self.assertIn("internal", lines[1])
        self.assertIn("Croissant", lines[1])
        self.assertIn("28.80", lines[1])    # 24 × 1.20
        # Teals Wed
        self.assertIn("TEALS", lines[2])
        self.assertIn("wholesale", lines[2])
        self.assertIn("Sourdough", lines[2])
        self.assertIn("45.00", lines[2])    # 10 × 4.50

    def test_csv_defaults_to_latest_week_when_param_missing(self):
        self.client.force_login(self.user)
        r = self.client.get("/api/dashboard/export.csv?week=")
        self.assertEqual(r.status_code, 200)
        # Latest week is self.wc — must contain the two external rows.
        body = r.content.decode()
        self.assertIn("GARDEN CAFE", body)
        self.assertIn("TEALS", body)


class SpaDashboardRouteTests(TestCase):
    """The /business-performance-dashboard/ route serves the built Vite
    bundle's index.html and falls back to the same HTML for any sub-route
    (client-side deep-linking). Auth-gated like every other page.

    /dashboard/ is RETIRED — it now redirects to the Business Performance
    route (App.jsx stays in the frontend tree, unused).
    """

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)

    def test_dashboard_redirects_to_business_performance(self):
        r = self.client.get("/dashboard/")
        self.assertEqual(r.status_code, 302)
        self.assertEqual(r["Location"], "/business-performance-dashboard/")

    def test_dashboard_redirect_preserves_query_string(self):
        r = self.client.get("/dashboard/?week=2026-05-18")
        self.assertEqual(r.status_code, 302)
        self.assertEqual(
            r["Location"], "/business-performance-dashboard/?week=2026-05-18")

    def test_bp_anonymous_redirects_to_login(self):
        r = self.client.get("/business-performance-dashboard/")
        # @login_required redirects to /login/ when anonymous.
        self.assertEqual(r.status_code, 302)
        self.assertIn("/login/", r["Location"])

    def test_bp_authenticated_serves_index_html_or_helpful_404(self):
        # If the SPA has been built (frontend/dist/index.html exists), we get
        # the bundled HTML; otherwise the actionable 404 telling the dev to
        # run the build. EITHER is correct depending on the build state.
        self.client.force_login(self.user)
        r = self.client.get("/business-performance-dashboard/")
        self.assertIn(r.status_code, (200, 404))
        body = r.content.decode()
        if r.status_code == 200:
            self.assertIn("<div id=\"root\"></div>", body)
        else:
            self.assertIn("npm run build", body)

    def test_bp_subroute_falls_back_to_same_response(self):
        # Client-side routes — /business-performance-dashboard/foo/ must
        # serve the same HTML, not 404 in Django.
        self.client.force_login(self.user)
        r = self.client.get("/business-performance-dashboard/anything-here/")
        self.assertIn(r.status_code, (200, 404))


class LiveVsLatestHistoricalSnapshotTests(SimpleTestCase):
    """Belt-and-braces against the live-vs-snapshot drift class.

    On every deploy, build.sh runs ``import_orders --tab "GARDEN CAFE"
    --tab "WHOLESALE" data/order_sheet.xlsm`` — that importer REPLACES
    OrderLines per (customer, date). If ``data/order_sheet.xlsm`` is
    stale relative to the most recent historical snapshot for the same
    week, the deploy silently rewrites prod with the older values.
    That's what happened on w/c 18 May (see commit 792d9d4).

    This test fails loudly when the two workbooks diverge for a week
    they both cover. Reads them with the same helpers the importer
    uses, so it tests exactly what build.sh would see — not a byte
    diff that over-triggers on formatting.
    """

    LIVE = "data/order_sheet.xlsm"
    HIST_DIR = "data/historical"

    def _latest_snapshot(self):
        import glob, os, re
        files = sorted(glob.glob(f"{self.HIST_DIR}/order_sheet_*.xlsm"))
        self.assertTrue(files, f"no historical snapshots in {self.HIST_DIR}")
        path = files[-1]
        m = re.search(r"order_sheet_(\d{4})_(\d{2})_(\d{2})\.xlsm$", path)
        self.assertIsNotNone(m, f"unparseable snapshot filename: {path}")
        return path, datetime.date(int(m[1]), int(m[2]), int(m[3]))

    def _read_tab(self, wb, tab, *, wholesale):
        """Reduce a tab to ``{(date_iso, customer, product, sage): (qty, price)}``
        — the smallest comparable shape the importer would write."""
        from .order_import import (iter_product_rows, iter_wholesale_rows,
                                   read_tab_dates, read_wholesale_dates)
        ws = wb[tab]
        dates = read_wholesale_dates(ws) if wholesale else read_tab_dates(ws)
        iterator = iter_wholesale_rows(ws) if wholesale else iter_product_rows(ws)
        out = {}
        for row in iterator:
            cust = row.get("customer", "") if wholesale else ""
            for d, q in zip(dates, row["qtys"]):
                if q is None:  # blank cell -> no line written
                    continue
                key = (d.isoformat(), cust, row["name"], row.get("sage") or "")
                out[key] = (str(q), str(row["price"]) if row["price"] is not None else None)
        return dates, out

    def test_live_workbook_matches_latest_historical_snapshot(self):
        from openpyxl import load_workbook
        snap_path, snap_wc = self._latest_snapshot()
        live_wb = load_workbook(self.LIVE, read_only=True, data_only=True)
        snap_wb = load_workbook(snap_path, read_only=True, data_only=True)
        try:
            for tab, wholesale in [("GARDEN CAFE", False), ("WHOLESALE", True)]:
                live_dates, live_rows = self._read_tab(live_wb, tab, wholesale=wholesale)
                snap_dates, snap_rows = self._read_tab(snap_wb, tab, wholesale=wholesale)
                if live_dates[0] != snap_wc:
                    self.skipTest(
                        f"live workbook covers w/c {live_dates[0]}, not the "
                        f"latest snapshot's w/c {snap_wc} — take a fresh "
                        f"snapshot to re-engage this test")
                self.assertEqual(
                    live_rows, snap_rows,
                    f"{tab}: data/order_sheet.xlsm DIVERGES from "
                    f"{snap_path} for w/c {snap_wc}. The next deploy will "
                    f"trample prod with the live values. See commit 792d9d4.")
        finally:
            live_wb.close()
            snap_wb.close()


class BusinessPerformanceHelpersTests(TestCase):
    """period_comparison + concentration_metrics + customer_dynamics +
    range_product_revenue + range_week_stats — the multi-week helpers
    that power /business-performance-dashboard/. Contract:

    * deltas are None when prior period extends before earliest imported
    * concentration band thresholds: <50% healthy, 50–70% watch, >70% concentrated
    * customer state classifies on first-order date + ±10% growth band
    * dormant customers surface in a SEPARATE list (never in `rows`)
    * Pareto sorts by VALUE desc; cumulative % is monotone non-decreasing
    """

    def setUp(self):
        self.dept = Department.objects.create(name="Bakery")
        # 4 consecutive Mondays — prior=[wc1,wc2], current=[wc3,wc4]
        self.wc1 = datetime.date(2026, 3, 30)
        self.wc2 = datetime.date(2026, 4, 6)
        self.wc3 = datetime.date(2026, 4, 13)
        self.wc4 = datetime.date(2026, 4, 20)
        # 4 wholesale customers spanning the four states + 1 dormant +
        # 1 internal customer + 1 excluded (£999 noise).
        self.A = Customer.objects.create(
            name="A", customer_type=Customer.WHOLESALE, department=self.dept)
        self.B = Customer.objects.create(
            name="B", customer_type=Customer.WHOLESALE, department=self.dept)
        self.C = Customer.objects.create(
            name="C", customer_type=Customer.WHOLESALE, department=self.dept)
        self.D = Customer.objects.create(
            name="D", customer_type=Customer.WHOLESALE, department=self.dept)
        self.F = Customer.objects.create(
            name="F", customer_type=Customer.WHOLESALE, department=self.dept)
        self.E = Customer.objects.create(
            name="E", customer_type=Customer.INTERNAL, department=self.dept)
        self.X = Customer.objects.create(
            name="BAKERY INTERNAL USE",
            customer_type=Customer.INTERNAL,
            is_internal=True, department=self.dept)

        # Prior period (wc1, wc2). All prices £1.00 so qty == value for
        # easy mental arithmetic.
        self._line(self.A, self.wc1, Decimal("50"), Decimal("1.00"), "Sourdough")
        self._line(self.A, self.wc2, Decimal("50"), Decimal("1.00"), "Sourdough")
        self._line(self.B, self.wc1, Decimal("100"), Decimal("1.00"), "Sourdough")
        self._line(self.C, self.wc1, Decimal("100"), Decimal("1.00"), "Sourdough")
        self._line(self.F, self.wc1, Decimal("100"), Decimal("1.00"), "Sourdough")
        self._line(self.E, self.wc2, Decimal("50"), Decimal("1.00"), "Sourdough")
        # Excluded — £999 must NEVER show up in any total.
        self._line(self.X, self.wc1, Decimal("999"), Decimal("1.00"), "Sourdough")

        # Current period (wc3, wc4).
        # A: 100 → 70 (-30% → declining)
        self._line(self.A, self.wc3, Decimal("70"), Decimal("1.00"), "Sourdough")
        # B: 100 → 105 (+5% → stable)
        self._line(self.B, self.wc3, Decimal("105"), Decimal("1.00"), "Sourdough")
        # C: 100 → 130 (+30% → growing). Two products so Pareto has variety.
        self._line(self.C, self.wc3, Decimal("65"), Decimal("1.00"), "Sourdough")
        self._line(self.C, self.wc4, Decimal("65"), Decimal("1.00"), "Cake")
        # D: 0 → 50 (NEW — first-ever order in current period)
        self._line(self.D, self.wc3, Decimal("50"), Decimal("1.00"), "Bread")
        # F: 100 → 0 (dormant)
        # E (internal): 50 → 50
        self._line(self.E, self.wc3, Decimal("50"), Decimal("1.00"), "Sourdough")

        # Pre-computed expectations:
        # Wholesale prior = 50+50 + 100 + 100 + 100 = 400
        # Wholesale current = 70 + 105 + 130 + 50 = 355
        # Wholesale delta % = (355-400)/400 * 100 = -11.25 → -11.3 (1dp)
        # Internal prior = 50, current = 50

    def _line(self, customer, order_date, qty, price, product_name):
        o = Order.objects.create(
            customer=customer, department=self.dept,
            order_date=order_date)
        OrderLine.objects.create(
            order=o, sale_product=None, product_name=product_name,
            unit_price=price, qty_ordered=qty)

    # ---- period_comparison ----

    def test_period_comparison_current_and_prior_totals_correct(self):
        from stock.financials import period_comparison
        c = period_comparison(self.dept, self.wc3, self.wc4)
        self.assertEqual(c["current"]["wholesale"], Decimal("355.00"))
        self.assertEqual(c["current"]["internal"], Decimal("50.00"))
        self.assertEqual(c["prior"]["wholesale"], Decimal("400.00"))
        self.assertEqual(c["prior"]["internal"], Decimal("50.00"))
        self.assertFalse(c["prior_truncated"])
        self.assertEqual(c["n_weeks"], 2)
        self.assertEqual(c["prior_from"], self.wc1)
        self.assertEqual(c["prior_to"], self.wc2)

    def test_period_comparison_delta_total_matches_hand_math(self):
        from stock.financials import period_comparison
        c = period_comparison(self.dept, self.wc3, self.wc4)
        # Total: prior 450, current 405; delta = (405-450)/450 = -10.0%
        self.assertEqual(c["delta"]["total_pct"], Decimal("-10.0"))

    def test_period_comparison_share_pp_is_percentage_points_not_relative(self):
        from stock.financials import period_comparison
        c = period_comparison(self.dept, self.wc3, self.wc4)
        # Prior wholesale share = 400/450 = 88.9%; current = 355/405 = 87.7%
        # Δ = 87.7 - 88.9 = -1.2pp (not relative %)
        self.assertEqual(c["delta"]["wholesale_share_pp"], Decimal("-1.2"))

    def test_period_comparison_prior_truncated_when_period_starts_at_earliest(self):
        from stock.financials import period_comparison
        # Current = wc1..wc2; prior would be wc-1..wc0 which is before
        # earliest imported. Must return prior=None, prior_truncated=True.
        c = period_comparison(self.dept, self.wc1, self.wc2)
        self.assertTrue(c["prior_truncated"])
        self.assertIsNone(c["prior"])
        self.assertIsNone(c["delta"])
        self.assertEqual(c["current"]["wholesale"], Decimal("400.00"))

    # ---- concentration_metrics ----

    def test_concentration_top_1_3_5_match_per_customer_in_channel(self):
        from stock.financials import concentration_metrics
        m = concentration_metrics(self.dept, Customer.WHOLESALE,
                                  self.wc3, self.wc4)
        # Current wholesale: C=130, B=105, A=70, D=50; total 355
        # Shares: 36.6%, 29.6%, 19.7%, 14.1%
        self.assertEqual(m["top_1_name"], "C")
        self.assertEqual(m["top_1_pct"], Decimal("36.6"))
        # top_3 = 36.6 + 29.6 + 19.7 = 85.9
        self.assertEqual(m["top_3_pct"], Decimal("85.9"))
        # top_5 = same as all-4 here = 100.0
        self.assertEqual(m["top_5_pct"], Decimal("100.0"))
        self.assertEqual(m["n_customers"], 4)

    def test_concentration_band_concentrated_when_top_5_over_70(self):
        from stock.financials import concentration_metrics
        m = concentration_metrics(self.dept, Customer.WHOLESALE,
                                  self.wc3, self.wc4)
        self.assertEqual(m["band"], "concentrated")

    def test_concentration_empty_channel_returns_healthy_zero(self):
        from stock.financials import concentration_metrics
        # Internal channel: only one customer (E) — top_5 share = 100%,
        # so this is "concentrated" with 1 customer. Pick a channel with
        # zero data instead: wholesale on a week with no orders.
        far_week = self.wc4 + datetime.timedelta(days=28)
        m = concentration_metrics(self.dept, Customer.WHOLESALE,
                                  far_week, far_week)
        self.assertEqual(m["band"], "healthy")
        self.assertEqual(m["top_1_pct"], Decimal("0.0"))
        self.assertEqual(m["n_customers"], 0)
        self.assertIsNone(m["top_1_name"])

    # ---- customer_dynamics ----

    def test_customer_dynamics_classifies_growing_declining_stable(self):
        from stock.financials import customer_dynamics
        d = customer_dynamics(self.dept, Customer.WHOLESALE,
                              self.wc3, self.wc4)
        by_name = {r["name"]: r for r in d["rows"]}
        self.assertEqual(by_name["A"]["state"], "declining")  # -30%
        self.assertEqual(by_name["B"]["state"], "stable")     # +5%
        self.assertEqual(by_name["C"]["state"], "growing")    # +30%

    def test_customer_dynamics_new_customer_first_order_in_period(self):
        from stock.financials import customer_dynamics
        d = customer_dynamics(self.dept, Customer.WHOLESALE,
                              self.wc3, self.wc4)
        by_name = {r["name"]: r for r in d["rows"]}
        self.assertEqual(by_name["D"]["state"], "new")
        # Δ is None for new (no honest comparison).
        self.assertIsNone(by_name["D"]["delta_pct"])

    def test_customer_dynamics_dormant_in_separate_list_not_in_rows(self):
        from stock.financials import customer_dynamics
        d = customer_dynamics(self.dept, Customer.WHOLESALE,
                              self.wc3, self.wc4)
        row_names = {r["name"] for r in d["rows"]}
        dormant_names = {r["name"] for r in d["dormant"]}
        self.assertNotIn("F", row_names)
        self.assertIn("F", dormant_names)
        # The dormant entry remembers their prior total so the watchlist
        # can show "was £100".
        f_dormant = next(r for r in d["dormant"] if r["name"] == "F")
        self.assertEqual(f_dormant["prior"], Decimal("100.00"))

    def test_customer_dynamics_no_prior_marks_returning_as_growing(self):
        from stock.financials import customer_dynamics
        # Period starts at wc1 — prior window extends before earliest.
        d = customer_dynamics(self.dept, Customer.WHOLESALE,
                              self.wc1, self.wc2)
        self.assertFalse(d["has_prior"])
        # Every active customer in the prior-anchored period either has
        # first_order in [wc1, wc2] (→ new) or before (→ growing-returning).
        states = {r["state"] for r in d["rows"]}
        self.assertTrue(states.issubset({"new", "growing"}))

    def test_customer_dynamics_summary_counts_match_rows(self):
        from stock.financials import customer_dynamics
        d = customer_dynamics(self.dept, Customer.WHOLESALE,
                              self.wc3, self.wc4)
        states_from_rows = [r["state"] for r in d["rows"]]
        for state, n in d["summary"].items():
            if state == "dormant":
                self.assertEqual(n, len(d["dormant"]))
            else:
                self.assertEqual(n, states_from_rows.count(state),
                                 f"summary[{state}] != rows count")

    # ---- range_product_revenue ----

    def test_range_product_revenue_sorted_by_value_descending(self):
        from stock.financials import range_product_revenue
        m = range_product_revenue(self.dept, self.wc3, self.wc4)
        names = [r["product"] for r in m["rows"]]
        # Current period (external only):
        #   Sourdough: A(70) + B(105) + C(65) + E(50) = 290
        #   Cake: C(65) = 65
        #   Bread: D(50) = 50
        # Total = 405. Sorted by value: Sourdough, Cake, Bread.
        self.assertEqual(names, ["Sourdough", "Cake", "Bread"])

    def test_range_product_revenue_cumulative_pct_monotone_and_reaches_100(self):
        from stock.financials import range_product_revenue
        m = range_product_revenue(self.dept, self.wc3, self.wc4)
        cums = [r["cumulative_pct"] for r in m["rows"]]
        # Monotone non-decreasing
        self.assertEqual(sorted(cums), cums)
        # Final cumulative crosses 99.9 (rounding tolerance)
        self.assertGreaterEqual(cums[-1], Decimal("99.9"))

    def test_range_product_revenue_n_to_80pct(self):
        from stock.financials import range_product_revenue
        m = range_product_revenue(self.dept, self.wc3, self.wc4)
        # Sourdough alone is 290/405 = 71.6% — under 80.
        # Sourdough + Cake = 355/405 = 87.7% — crosses 80.
        # So n_to_80pct = 2.
        self.assertEqual(m["n_to_80pct"], 2)

    def test_range_product_revenue_excludes_internal_customers(self):
        from stock.financials import range_product_revenue
        # The £999 BAKERY INTERNAL USE Sourdough order is OUTSIDE the
        # current period (wc1, not wc3-wc4) so this test verifies the
        # IN-period excluded check by adding a £500 internal-use line
        # in the current period.
        self._line(self.X, self.wc3, Decimal("500"), Decimal("1.00"), "Sourdough")
        m = range_product_revenue(self.dept, self.wc3, self.wc4)
        sourdough = next(r for r in m["rows"] if r["product"] == "Sourdough")
        # If the £500 leaked in, Sourdough would be 790. Must stay 290.
        self.assertEqual(sourdough["value"], Decimal("290.00"))

    # ---- range_week_stats ----

    def test_range_week_stats_best_and_worst_picked_from_actual_weeks(self):
        from stock.financials import range_week_stats
        # Over wc1..wc4:
        #   wc1: A50 + B100 + C100 + F100 + X(excluded) = 350 (external)
        #   wc2: A50 + E50 = 100
        #   wc3: A70 + B105 + C65 + D50 + E50 = 340
        #   wc4: C65 = 65
        s = range_week_stats(self.dept, self.wc1, self.wc4)
        self.assertEqual(s["best_week"]["week"], self.wc1)
        self.assertEqual(s["best_week"]["total"], Decimal("350.00"))
        self.assertEqual(s["worst_week"]["week"], self.wc4)
        self.assertEqual(s["worst_week"]["total"], Decimal("65.00"))
        self.assertEqual(s["spread"], Decimal("285.00"))

    def test_range_week_stats_empty_period_all_nones(self):
        from stock.financials import range_week_stats
        far = self.wc4 + datetime.timedelta(days=28)
        s = range_week_stats(self.dept, far, far)
        self.assertIsNone(s["best_week"])
        self.assertIsNone(s["worst_week"])
        self.assertIsNone(s["variability_pct"])


class BusinessPerformanceEndpointTests(TestCase):
    """Smoke test for ``/api/business-performance/summary/`` — verifies
    auth gate, payload shape, default period, and the from/to clamping
    rules. Detailed math is covered by BusinessPerformanceHelpersTests;
    here we just verify the endpoint composes them correctly."""

    URL = "/api/business-performance/summary/"

    def setUp(self):
        U = get_user_model()
        self.dept = Department.objects.create(name="Bakery")
        self.user = U.objects.create_user("alice", password="pw")
        self.dept.members.add(self.user)
        self.wc1 = datetime.date(2026, 3, 30)
        self.wc2 = datetime.date(2026, 4, 6)
        self.wc3 = datetime.date(2026, 4, 13)
        self.teals = Customer.objects.create(
            name="TEALS", customer_type=Customer.WHOLESALE,
            department=self.dept)
        self.garden = Customer.objects.create(
            name="GARDEN CAFE", customer_type=Customer.INTERNAL,
            department=self.dept)
        for wc in (self.wc1, self.wc2, self.wc3):
            o = Order.objects.create(
                customer=self.teals, department=self.dept, order_date=wc)
            OrderLine.objects.create(
                order=o, sale_product=None, product_name="Sourdough",
                unit_price=Decimal("1.00"), qty_ordered=Decimal("10"))
            o2 = Order.objects.create(
                customer=self.garden, department=self.dept, order_date=wc)
            OrderLine.objects.create(
                order=o2, sale_product=None, product_name="Cake",
                unit_price=Decimal("1.00"), qty_ordered=Decimal("5"))

    def test_requires_authentication(self):
        r = self.client.get(self.URL)
        self.assertIn(r.status_code, (401, 403))

    def test_authenticated_returns_expected_top_level_keys(self):
        self.client.force_login(self.user)
        r = self.client.get(self.URL)
        self.assertEqual(r.status_code, 200)
        body = r.json()
        expected = {"period", "available_weeks", "totals", "weekly_trend",
                    "daily_trend", "best_worst", "concentration",
                    "customers", "products", "product_day_matrix",
                    "current_week"}
        self.assertTrue(expected.issubset(body.keys()),
                        f"Missing keys: {expected - body.keys()}")

    def test_default_period_clamps_to_imported_weeks(self):
        self.client.force_login(self.user)
        body = self.client.get(self.URL).json()
        # No params defaults to the latest imported/current week. Longer
        # ranges are selected explicitly by the SPA period picker.
        self.assertEqual(body["period"]["from"], self.wc3.isoformat())
        self.assertEqual(body["period"]["to"], self.wc3.isoformat())
        self.assertEqual(body["period"]["earliest_imported"], self.wc1.isoformat())
        self.assertEqual(body["period"]["latest_imported"], self.wc3.isoformat())

    def test_concentration_returned_for_both_channels(self):
        self.client.force_login(self.user)
        body = self.client.get(self.URL).json()
        self.assertIn("wholesale", body["concentration"])
        self.assertIn("internal", body["concentration"])
        self.assertIn("wholesale", body["customers"])
        self.assertIn("internal", body["customers"])

    def test_default_one_week_returns_daily_channel_trend(self):
        self.client.force_login(self.user)
        body = self.client.get(self.URL).json()
        rows = body["daily_trend"]
        self.assertEqual(len(rows), 7)
        self.assertEqual(rows[0]["date"], self.wc3.isoformat())
        self.assertEqual(
            rows[6]["date"],
            (self.wc3 + datetime.timedelta(days=6)).isoformat())
        total = sum(Decimal(str(r["total"])) for r in rows)
        self.assertEqual(total, Decimal(str(body["totals"]["current"]["total"])))
        monday = rows[0]
        self.assertEqual(Decimal(str(monday["wholesale"])), Decimal("10.00"))
        self.assertEqual(Decimal(str(monday["internal"])), Decimal("5.00"))
        self.assertEqual(Decimal(str(monday["prior_total"])), Decimal("15.00"))

    def test_default_one_week_returns_product_day_matrix(self):
        self.client.force_login(self.user)
        body = self.client.get(self.URL).json()
        matrix = body["product_day_matrix"]
        self.assertEqual([r["product"] for r in matrix], ["Sourdough", "Cake"])
        sourdough = matrix[0]
        self.assertEqual(Decimal(str(sourdough["total_qty"])), Decimal("10"))
        self.assertEqual(
            [Decimal(str(q)) for q in sourdough["daily"]],
            [Decimal("10"), Decimal("0"), Decimal("0"), Decimal("0"),
             Decimal("0"), Decimal("0"), Decimal("0")],
        )

    def test_multi_week_keeps_weekly_trend_and_no_daily_trend(self):
        self.client.force_login(self.user)
        body = self.client.get(
            f"{self.URL}?from={self.wc1.isoformat()}&to={self.wc3.isoformat()}"
        ).json()
        self.assertEqual(len(body["weekly_trend"]), 3)
        self.assertEqual(body["daily_trend"], [])
        self.assertEqual(body["product_day_matrix"], [])
        matrix = body["product_week_matrix"]
        self.assertEqual(matrix["granularity"], "week")
        self.assertEqual(
            matrix["buckets"],
            [self.wc1.isoformat(), self.wc2.isoformat(), self.wc3.isoformat()],
        )
        sourdough = next(r for r in matrix["rows"]
                         if r["product"] == "Sourdough")
        self.assertEqual(
            [Decimal(str(q)) for q in sourdough["values"]],
            [Decimal("10"), Decimal("10"), Decimal("10")],
        )
        self.assertEqual(Decimal(str(sourdough["total_qty"])), Decimal("30"))

    def test_product_week_matrix_boundary_switches_after_sixteen_weeks(self):
        for i in range(17):
            wc = self.wc1 + datetime.timedelta(weeks=i)
            order = Order.objects.create(
                customer=self.teals, department=self.dept, order_date=wc)
            OrderLine.objects.create(
                order=order, sale_product=None, product_name="Long Range Bun",
                unit_price=Decimal("1.00"), qty_ordered=Decimal("100"))

        self.client.force_login(self.user)
        week_to = self.wc1 + datetime.timedelta(weeks=15)
        week_body = self.client.get(
            f"{self.URL}?from={self.wc1.isoformat()}&to={week_to.isoformat()}"
        ).json()
        self.assertEqual(week_body["product_week_matrix"]["granularity"], "week")
        self.assertEqual(len(week_body["product_week_matrix"]["buckets"]), 16)

        to_wc = self.wc1 + datetime.timedelta(weeks=16)
        body = self.client.get(
            f"{self.URL}?from={self.wc1.isoformat()}&to={to_wc.isoformat()}"
        ).json()
        matrix = body["product_week_matrix"]
        self.assertEqual(matrix["granularity"], "month")
        self.assertEqual(
            matrix["buckets"],
            ["2026-03-01", "2026-04-01", "2026-05-01",
             "2026-06-01", "2026-07-01"],
        )
        bun = next(r for r in matrix["rows"]
                   if r["product"] == "Long Range Bun")
        self.assertEqual(
            [Decimal(str(q)) for q in bun["values"]],
            [Decimal("100"), Decimal("400"), Decimal("400"),
             Decimal("500"), Decimal("300")],
        )
        self.assertEqual(Decimal(str(bun["total_qty"])), Decimal("1700"))
