"""Parse and save a Recipe Breakdown Excel export.

The workbook is shaped like the WildFarmed recipe report: a main recipe at
the top (Recipe Code / Description / a single ingredient table with
totals), followed by repeated "Sub Recipe Breakdown" sections, one per
nested recipe. Each section has its own code-headed ingredient table, a
Total / Deposit Weight / Cook Loss / Finished Weight block, and a Method
section split into named stages. Codes prefixed NPD-I are raw ingredients
(Product); NPD-R are other recipes (sub-recipes), nestable arbitrarily.

`parse_recipe_workbook` returns a list of plain dicts — used by both the
management command and the upload view (so the view can show a preview
before committing). `save_recipes` is the idempotent write step.
"""
import re
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from .models import Product, Recipe, RecipeCycleError, RecipeLine


HEADING_RE = re.compile(r"^(NPD-R\d+)\s*-\s*(.+)$", re.IGNORECASE)
INGREDIENT_CODE_RE = re.compile(r"^NPD-[IR]\d+$", re.IGNORECASE)
WEIGHT_RE = re.compile(r"([\d.]+)\s*g?$", re.IGNORECASE)
PERCENT_RE = re.compile(r"([\d.]+)\s*%")

# Cells that appear next to ingredient lines but aren't the description.
_STATE_WORDS = {"live", "approved", "draft", "obsolete"}


class RecipeParseError(Exception):
    """The workbook doesn't look like a Recipe Breakdown export."""


def _dec(v):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _parse_weight(v):
    """'660g' / '15.8 g' / 600 → Decimal, else None."""
    if v is None or v == "":
        return None
    s = str(v).strip()
    m = WEIGHT_RE.search(s)
    if m:
        return _dec(m.group(1))
    return _dec(s)


def _parse_percent(v):
    if v is None or v == "":
        return None
    s = str(v).strip()
    m = PERCENT_RE.search(s)
    if m:
        return _dec(m.group(1))
    return _dec(s)


def _norm_label(s):
    return str(s).strip().lower().rstrip(":")


def _value_right_of(row, label):
    """First non-empty cell to the right of the cell whose text matches `label`.

    Case-insensitive; trailing colons ignored. Returns None if not found.
    """
    target = _norm_label(label)
    for i, c in enumerate(row):
        if c is None:
            continue
        if _norm_label(c) == target:
            for k in range(i + 1, len(row)):
                if row[k] is not None and str(row[k]).strip() != "":
                    return row[k]
            return None
    return None


def _row_has(row, label):
    target = _norm_label(label)
    for c in row:
        if c is None:
            continue
        if _norm_label(c) == target:
            return True
    return False


def _non_empty_cells(row):
    out = []
    for c in row:
        if c is None:
            continue
        s = str(c).strip()
        if s == "":
            continue
        out.append(s)
    return out


def _rightmost_text(row):
    for c in reversed(row):
        if c is None:
            continue
        s = str(c).strip()
        if s == "":
            continue
        return s
    return None


def _find_col(row, *labels):
    """Index of the first cell whose normalised text matches any of `labels`.

    Case-insensitive; trailing colons ignored. Returns None when not found.
    """
    targets = {_norm_label(n) for n in labels}
    for i, c in enumerate(row):
        if c is None:
            continue
        if _norm_label(c) in targets:
            return i
    return None


# Labels used by the two known layouts for the *weight* column of an
# ingredient/component table:
#   - "Weight (g)"  — top-level recipe tables in both files
#   - "g"           — sub-recipe tables in both files
# Packaging tables sit alongside ingredient tables and share Code/Description
# headers, but use "Quantity" / "UOM" instead — never "g" or "Weight (g)".
# Distinguishing on the weight-column header is what stops the packaging
# rows from clobbering the parent recipe's actual ingredient lines.
_WEIGHT_HEADER_LABELS = ("g", "weight (g)", "weight")


def _classify_ingredient_header(row):
    """If this row is a Code/Description/.../weight ingredient-table header,
    return ``(code_col, desc_col, weight_col)``. Otherwise return ``None``.

    Recognises both column layouts seen in production exports:
        Code | Description | (...) | Weight (g)   (top-level tables)
        Code | Description | State | g            (sub-recipe tables, newer)
    A header with Code+Description but no weight column (e.g. a Packaging
    table with Quantity/UOM) deliberately returns None so we don't enter
    ingredient-reading mode for it.
    """
    code_col = _find_col(row, "Code")
    desc_col = _find_col(row, "Description")
    if code_col is None or desc_col is None:
        return None
    weight_col = _find_col(row, *_WEIGHT_HEADER_LABELS)
    if weight_col is None:
        return None
    return code_col, desc_col, weight_col


def parse_recipe_workbook(file_or_path):
    """Walk the workbook and return a list of recipe dicts.

    First entry is the main recipe; the rest are sub-recipes in workbook
    order. Each dict carries `code`, `name`, `finished_weight_g`,
    `deposit_weight_g`, `cook_loss_pct`, `method_text`, `units_requested`,
    and `lines` (each `{code, name, weight_g, is_subrecipe}`).
    """
    wb = load_workbook(file_or_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    recipes = []
    by_code = {}
    current = {"r": None}
    method_state = {"buffer": [], "stage_name": None}

    main_code = None
    main_name = None
    main_units = None

    # Find header values in the first few rows.
    for row in rows[:25]:
        if not main_code:
            v = _value_right_of(row, "Recipe Code")
            if v:
                main_code = str(v).strip()
        if not main_name:
            v = _value_right_of(row, "Recipe Description")
            if v:
                main_name = str(v).strip()
        if main_units is None:
            v = _value_right_of(row, "Units Requested")
            if v is not None:
                main_units = _dec(v)

    if not main_code:
        raise RecipeParseError(
            "Could not find 'Recipe Code:' header — not a Recipe Breakdown export?")

    def _finish():
        r = current["r"]
        if r is not None and method_state["buffer"]:
            r["method_text"] = "\n\n".join(method_state["buffer"]).strip()
        method_state["buffer"] = []
        method_state["stage_name"] = None

    def _start(code, name, units=None):
        _finish()
        code = code.upper()
        if code in by_code:
            r = by_code[code]
            r["lines"] = []  # re-parsing this section: lines repopulate below
            if name and (not r.get("name") or r.get("name") == code):
                r["name"] = name
            current["r"] = r
            return r
        r = {
            "code": code, "name": (name or code).strip(),
            "units_requested": units,
            "finished_weight_g": None,
            "deposit_weight_g": None,
            "cook_loss_pct": None,
            "method_text": "",
            "lines": [],
            "_total": None,
        }
        recipes.append(r)
        by_code[code] = r
        current["r"] = r
        return r

    state = "idle"  # idle | ingredients | method
    # Column indices recorded the last time we saw an ingredient-table
    # header — used so subsequent line rows pull from the right cells
    # regardless of how the exporter spaced its columns. Reset on every
    # new header (which may have a different layout for sub-recipes).
    ing_cols = {"code": None, "desc": None, "weight": None}

    for row in rows:
        cells_non_empty = _non_empty_cells(row)
        if not cells_non_empty:
            # blank row closes an ingredient block (the Total row that follows
            # will land in idle and still pick up its value via _value_right_of).
            if state == "ingredients":
                state = "idle"
            continue

        # New section heading "NPD-R\d+ - Name". Anchored to single-cell rows
        # so a method "Materials" blob that lists "NPD-R123 - ..." as plain
        # text (with sibling lines in the same cell) can't trigger a section
        # change. Real headings are always alone in their row.
        heading = None
        if len(cells_non_empty) == 1:
            m = HEADING_RE.match(cells_non_empty[0])
            if m:
                heading = m
        if heading:
            _start(heading.group(1), heading.group(2))
            state = "idle"
            ing_cols = {"code": None, "desc": None, "weight": None}
            continue

        # "Method" section start (only when "Method" is the lone label in B,
        # not the "Materials/Method/Image" stage sub-header).
        second = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if state != "method" and second.lower() == "method" and len(cells_non_empty) == 1:
            state = "method"
            method_state["stage_name"] = None
            continue

        # Start of an ingredient block. Only count rows that ALSO carry a
        # weight column header ("g" or "Weight (g)") — packaging/quantity
        # tables share the Code/Description header but use Quantity/UOM,
        # so they must not flip us into ingredient-reading mode.
        cls = _classify_ingredient_header(row)
        if cls is not None:
            if current["r"] is None and main_code:
                _start(main_code, main_name, units=main_units)
            if current["r"]:
                current["r"]["lines"] = []
            ing_cols["code"], ing_cols["desc"], ing_cols["weight"] = cls
            state = "ingredients"
            continue

        # Ingredient row (in or near an ingredients block).
        if state == "ingredients":
            code_col = ing_cols["code"]
            desc_col = ing_cols["desc"]
            weight_col = ing_cols["weight"]
            code_cell = None
            if code_col is not None and code_col < len(row) and row[code_col] is not None:
                cand = str(row[code_col]).strip()
                if INGREDIENT_CODE_RE.match(cand):
                    code_cell = cand.upper()
            if code_cell and current["r"] is not None:
                # Description from the recorded column (fall back to blank
                # if the cell is empty or out of range).
                desc = ""
                if desc_col is not None and desc_col < len(row) and row[desc_col] is not None:
                    desc = str(row[desc_col]).strip()
                # Weight from the recorded column. Tolerate both numeric
                # cells and "660g"-style strings via _parse_weight.
                weight = None
                if weight_col is not None and weight_col < len(row):
                    weight = _parse_weight(row[weight_col])
                if weight is not None:
                    current["r"]["lines"].append({
                        "code": code_cell,
                        "name": desc,
                        "weight_g": weight,
                        "is_subrecipe": code_cell.upper().startswith("NPD-R"),
                    })
                continue

        # Totals (these can appear in any state and don't depend on order).
        r = current["r"]
        if r is not None:
            v = _value_right_of(row, "Finished Weight")
            if v is not None and r.get("finished_weight_g") is None:
                r["finished_weight_g"] = _parse_weight(v)
            v = _value_right_of(row, "Deposit Weight")
            if v is not None and r.get("deposit_weight_g") is None:
                r["deposit_weight_g"] = _parse_weight(v)
            v = _value_right_of(row, "Cook loss")
            if v is None:
                v = _value_right_of(row, "Cook Loss")
            if v is not None and r.get("cook_loss_pct") is None:
                r["cook_loss_pct"] = _parse_percent(v)
            v = _value_right_of(row, "Total")
            if v is not None and r.get("_total") is None:
                r["_total"] = _parse_weight(v)

        # Method content.
        if state == "method":
            # The "Materials/Method/Image" sub-header inside a method stage —
            # skip and wait for the content row that follows.
            if second.lower() == "materials" and _row_has(row, "Method"):
                continue
            # A bare stage name (single non-empty cell in column B).
            if len(cells_non_empty) == 1 and second:
                method_state["stage_name"] = second
                continue
            # Content row: take the rightmost substantive text (the leftmost
            # cell sometimes carries a "materials" listing for that stage,
            # which is recipe-builder metadata rather than user-facing method).
            content = _rightmost_text(row)
            if content:
                if method_state["stage_name"]:
                    method_state["buffer"].append(
                        f"{method_state['stage_name']}:\n{content}")
                else:
                    method_state["buffer"].append(content)
                method_state["stage_name"] = None

    _finish()

    # Main recipe may not have a separate "Deposit Weight" row — fall back
    # to the Total of its single (top-level) ingredient table.
    for r in recipes:
        if r.get("deposit_weight_g") is None and r.get("_total") is not None:
            r["deposit_weight_g"] = r["_total"]
        r.pop("_total", None)

    return recipes


def summarize_parse(parsed):
    """Cheap pre-save analysis for the preview screen.

    Returns the main recipe dict, the flat list of sub-recipes, the set of
    NPD-I codes that don't currently exist as Products (the parts the user
    will need to fill in), and the list of NPD-R codes referenced as
    sub-recipes but not present as their own section (auto-stubbed on save).
    """
    if not parsed:
        return {"main": None, "subs": [], "unknown_ingredients": [],
                "stub_subrecipes": []}
    main = parsed[0]
    subs = parsed[1:]
    have_codes = {r["code"] for r in parsed}
    referenced_ing_codes = {ln["code"] for r in parsed for ln in r["lines"]
                            if not ln["is_subrecipe"]}
    referenced_sub_codes = {ln["code"] for r in parsed for ln in r["lines"]
                            if ln["is_subrecipe"]}
    known_products = set(
        Product.objects.filter(code__in=referenced_ing_codes)
        .values_list("code", flat=True))
    unknown_ingredients = sorted(referenced_ing_codes - known_products)
    stub_subrecipes = sorted(referenced_sub_codes - have_codes)
    return {
        "main": main,
        "subs": subs,
        "unknown_ingredients": unknown_ingredients,
        "stub_subrecipes": stub_subrecipes,
    }


@transaction.atomic
def save_recipes(parsed, department):
    """Persist parsed recipes idempotently. Returns counts + diagnostics.

    Two passes: first upserts the Recipe rows so cross-references resolve,
    then rebuilds each recipe's RecipeLine rows from scratch (so re-running
    an updated workbook doesn't leave stale lines lying around).

    Unknown NPD-I codes get a stub Product (code + name + dept + g, no
    supplier price); they're returned in `unknown_ingredients` so the
    operator knows to fill them in. NPD-R codes that are referenced as
    sub-recipes but never sectioned (the workbook is incomplete) get a
    stub Recipe with no lines.
    """
    created = []
    updated = []
    stub_subrecipes = []
    stub_products = []
    unknown_ingredients = []

    recipes_by_code = {}
    for spec in parsed:
        recipe, was_created = Recipe.objects.update_or_create(
            code=spec["code"],
            defaults={
                "name": (spec["name"] or spec["code"]).strip(),
                "finished_weight_g": spec.get("finished_weight_g"),
                "deposit_weight_g": spec.get("deposit_weight_g"),
                "cook_loss_pct": spec.get("cook_loss_pct"),
                "method_text": spec.get("method_text") or "",
                "department": department,
            },
        )
        recipes_by_code[spec["code"]] = recipe
        (created if was_created else updated).append(recipe)

    product_by_code = {
        p.code: p for p in Product.objects.filter(code__isnull=False)
    }

    for spec in parsed:
        recipe = recipes_by_code[spec["code"]]
        recipe.lines.all().delete()
        for i, line in enumerate(spec["lines"]):
            code = line["code"]
            weight = line["weight_g"]
            if line["is_subrecipe"]:
                sub = recipes_by_code.get(code)
                if sub is None:
                    sub, was_new = Recipe.objects.get_or_create(
                        code=code,
                        defaults={
                            "name": (line.get("name") or code).strip(),
                            "department": department,
                        },
                    )
                    recipes_by_code[code] = sub
                    if was_new:
                        stub_subrecipes.append(sub)
                if recipe.contains_cycle(sub.pk):
                    raise RecipeCycleError(
                        f"{recipe.code} would contain itself via {sub.code}")
                RecipeLine.objects.create(
                    recipe=recipe, sub_recipe=sub,
                    weight_g=weight, ordering=i,
                )
            else:
                product = product_by_code.get(code)
                if product is None:
                    product = Product.objects.create(
                        code=code,
                        name=(line.get("name") or code).strip(),
                        department=department,
                        unit="g",
                        category="unassigned",
                        minimum=0,
                    )
                    product_by_code[code] = product
                    stub_products.append(product)
                    unknown_ingredients.append((recipe.code, code, line.get("name") or ""))
                RecipeLine.objects.create(
                    recipe=recipe, ingredient=product,
                    weight_g=weight, ordering=i,
                )

    # After the lines are written, refresh sold_as_product defaults based
    # on the new reference graph (not referenced = sold, referenced = not
    # sold). Manual overrides (is_sold_manual=True) are preserved.
    Recipe.recompute_all_sold_defaults()

    return {
        "created": created,
        "updated": updated,
        "stub_subrecipes": stub_subrecipes,
        "stub_products": stub_products,
        "unknown_ingredients": unknown_ingredients,
    }


# ---- session serialization for the upload preview screen -------------

def serialize_parsed(parsed):
    """Round-trippable JSON form (Decimals become strings)."""
    out = []
    for r in parsed:
        out.append({
            "code": r["code"],
            "name": r["name"],
            "units_requested": str(r["units_requested"]) if r.get("units_requested") is not None else None,
            "finished_weight_g": str(r["finished_weight_g"]) if r.get("finished_weight_g") is not None else None,
            "deposit_weight_g": str(r["deposit_weight_g"]) if r.get("deposit_weight_g") is not None else None,
            "cook_loss_pct": str(r["cook_loss_pct"]) if r.get("cook_loss_pct") is not None else None,
            "method_text": r.get("method_text") or "",
            "lines": [{
                "code": ln["code"],
                "name": ln["name"],
                "weight_g": str(ln["weight_g"]),
                "is_subrecipe": ln["is_subrecipe"],
            } for ln in r["lines"]],
        })
    return out


def deserialize_parsed(raw):
    out = []
    for r in raw:
        out.append({
            "code": r["code"],
            "name": r["name"],
            "units_requested": _dec(r.get("units_requested")),
            "finished_weight_g": _dec(r.get("finished_weight_g")),
            "deposit_weight_g": _dec(r.get("deposit_weight_g")),
            "cook_loss_pct": _dec(r.get("cook_loss_pct")),
            "method_text": r.get("method_text") or "",
            "lines": [{
                "code": ln["code"],
                "name": ln["name"],
                "weight_g": _dec(ln["weight_g"]),
                "is_subrecipe": ln["is_subrecipe"],
            } for ln in r["lines"]],
        })
    return out
