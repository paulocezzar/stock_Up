"""Read the order-sheet "Products" tab and upsert SaleProduct rows.

Shared by the ``import_sale_products`` management command and the
matching tests. Idempotent on ``SaleProduct.name`` (case-insensitive
lookup, case-preserved on write — the order sheet is authoritative for
display). Names are NEVER overwritten by the linked recipe's name.

Auto-link, in priority order:

  1. Sage No.: ``sage_number`` == any ``Recipe.code`` (exact, case-
     insensitive). The Sage column is reliable so the resulting link
     is ``link_source=sage`` and auto-``link_confirmed=True``.
  2. Exact name fallback (only when no Sage match): ``name`` ==
     ``Recipe.name`` (case-insensitive, whitespace-trimmed). Same
     auto-confirmed treatment.
  3. Otherwise the SaleProduct is left unlinked
     (``link_source=none``, ``link_confirmed=False``); the UI's
     link-review page surfaces fuzzy suggestions for the operator
     to confirm by hand.

Manual links — ``link_source=manual`` or ``is_manual_entry=True`` —
are preserved across re-imports: the importer never touches them.
"""
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook


# Header labels on the Products tab. The column order in the workbook
# is `Product, Price, Sage No., (blank), Stock Managed, Pack Size` —
# header text is the source of truth so a future reorder doesn't break
# the importer. "Stock Managed" is deliberately ignored per spec.
PRODUCT_HEADERS = {
    "name": ("product", "product name"),
    "price": ("price",),
    "sage": ("sage no.", "sage no", "sage number"),
    "pack_size": ("pack size", "pack"),
}


def _norm_header(s):
    return (str(s) if s is not None else "").strip().lower()


def _find_columns(header_row):
    """Map our logical fields to column indices on the Products header row.

    Returns ``{"name": int, "price": int|None, "sage": int|None,
    "pack_size": int|None}``. Name is required; if missing we treat
    the sheet as malformed. The others are optional — a missing
    Pack Size column just leaves pack_size blank on every row.
    """
    out = {"name": None, "price": None, "sage": None, "pack_size": None}
    for i, cell in enumerate(header_row):
        label = _norm_header(cell)
        if not label:
            continue
        for key, choices in PRODUCT_HEADERS.items():
            if out[key] is None and label in choices:
                out[key] = i
    return out


def _as_text(v):
    if v is None:
        return ""
    s = str(v).strip()
    # openpyxl returns Excel integers as e.g. 660130086 (int) — render
    # without ".0" so it round-trips to the Sage code on the Recipe.
    if isinstance(v, float) and v.is_integer():
        s = str(int(v))
    return s


def _as_price(v):
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def read_products_tab(file_or_path):
    """Yield ``{name, price, sage, pack_size}`` dicts from the workbook.

    Reads the sheet named "Products" exactly. Rows whose name column is
    empty are skipped (the tab has a few right-aligned helper columns
    that don't carry a Product Name and aren't sale-product rows).
    """
    wb = load_workbook(file_or_path, data_only=True)
    if "Products" not in wb.sheetnames:
        raise ValueError("workbook has no 'Products' sheet")
    ws = wb["Products"]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return
    cols = _find_columns(header_row)
    if cols["name"] is None:
        raise ValueError("'Products' sheet header has no Product/Name column")

    def _col(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in rows_iter:
        if not row:
            continue
        name = _as_text(_col(row, cols["name"]))
        if not name:
            continue
        yield {
            "name": name,
            "price": _as_price(_col(row, cols["price"])),
            "sage": _as_text(_col(row, cols["sage"])),
            "pack_size": _as_text(_col(row, cols["pack_size"])),
        }


def auto_link(product, recipes_by_sage, recipes_by_name):
    """Compute the recipe link for one SaleProduct.

    Returns ``(recipe_or_none, link_source)`` where link_source is one
    of the SaleProduct.SAGE / NAME / NONE constants. Caller decides
    whether to apply (manual links are preserved untouched).
    """
    from .models import SaleProduct  # avoid circular at module import
    sage = (product.get("sage") or "").strip()
    if sage:
        r = recipes_by_sage.get(sage.lower())
        if r is not None:
            return r, SaleProduct.SAGE
    name = (product.get("name") or "").strip().lower()
    if name:
        r = recipes_by_name.get(name)
        if r is not None:
            return r, SaleProduct.NAME
    return None, SaleProduct.NONE


def import_sale_products(file_or_path, department):
    """Upsert every row from the Products tab. Returns a stats dict.

    Idempotent on name (case-insensitive). Names are kept verbatim from
    the workbook — never overwritten by the recipe's name on re-import.
    Manual links (``link_source=manual``) and hand-created rows
    (``is_manual_entry=True``) are preserved untouched.
    """
    from .models import Recipe, SaleProduct
    rows = list(read_products_tab(file_or_path))

    # Index recipes by code (lowercase) and name (lowercase, trimmed)
    # so the auto-link pass is O(N) instead of O(N*M).
    recipes_by_sage = {}
    recipes_by_name = {}
    for r in Recipe.objects.all():
        code_key = (r.code or "").strip().lower()
        if code_key:
            recipes_by_sage[code_key] = r
        name_key = (r.name or "").strip().lower()
        if name_key:
            # Several recipes can share a name in theory; first wins so
            # the lookup is deterministic. If you hit a real ambiguity
            # the operator can break it with a manual link in the UI.
            recipes_by_name.setdefault(name_key, r)

    existing_by_key = {sp.name.lower(): sp for sp in SaleProduct.objects.all()}

    created = updated = skipped_manual = 0
    n_sage = n_name = n_unlinked = 0

    for row in rows:
        key = row["name"].lower()
        existing = existing_by_key.get(key)
        # Compute the auto-link from the row, regardless of whether we
        # apply it (we still want the stats).
        recipe, source = auto_link(row, recipes_by_sage, recipes_by_name)

        if existing is None:
            sp = SaleProduct(
                name=row["name"],
                price=row["price"],
                sage_number=row["sage"],
                pack_size=row["pack_size"],
                department=department,
                is_manual_entry=False,
            )
            sp.recipe = recipe
            sp.link_source = source
            sp.link_confirmed = source in (SaleProduct.SAGE, SaleProduct.NAME)
            sp.save()
            existing_by_key[key] = sp
            created += 1
        else:
            if existing.is_manual_entry:
                # Operator-owned row — leave entirely alone.
                skipped_manual += 1
                continue
            # Refresh basic fields from the workbook on every run.
            existing.name = row["name"]  # whitespace-normalised re-write
            existing.price = row["price"]
            existing.sage_number = row["sage"]
            existing.pack_size = row["pack_size"]
            existing.department = department
            # Manual links: NEVER override. Otherwise refresh the auto-
            # link from the current Recipe set so a renamed Sage code or
            # a newly-imported recipe wires up on the next deploy.
            if existing.link_source != SaleProduct.MANUAL:
                existing.recipe = recipe
                existing.link_source = source
                existing.link_confirmed = source in (
                    SaleProduct.SAGE, SaleProduct.NAME)
            existing.save()
            updated += 1

        # Stats reflect what the importer would (or did) wire up — manual
        # links are counted as "linked by sage/name" too if that's what
        # the workbook says they were, so the deploy log stays honest
        # without exposing the operator's overrides.
        if source == SaleProduct.SAGE:
            n_sage += 1
        elif source == SaleProduct.NAME:
            n_name += 1
        else:
            n_unlinked += 1

    return {
        "rows": len(rows),
        "created": created,
        "updated": updated,
        "skipped_manual": skipped_manual,
        "linked_via_sage": n_sage,
        "linked_via_name": n_name,
        "unlinked": n_unlinked,
    }
