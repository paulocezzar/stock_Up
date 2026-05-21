"""
Import products, supplier prices and an initial stocktake from the two
spreadsheets. Usage:
    python manage.py import_data --stocktake Stocktake_18_05_2026.xlsx --prices Price_Comparison.xlsx
Safe to re-run: products are matched on code, prices on (product, supplier).
"""
import datetime
from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook
from stock.models import Supplier, Product, SupplierPrice, Stocktake, StockLine


def dec(v):
    if v in (None, "", "#REF!"):
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = "Import products, prices and a stocktake from the bakery spreadsheets."

    def add_arguments(self, parser):
        parser.add_argument("--stocktake", required=True)
        parser.add_argument("--prices")

    @transaction.atomic
    def handle(self, *args, **opts):
        st_products = self._import_stocktake(opts["stocktake"])
        if opts.get("prices"):
            self._import_prices(opts["prices"])
        self.stdout.write(self.style.SUCCESS(
            f"Done. {Product.objects.count()} products, "
            f"{Supplier.objects.count()} suppliers, "
            f"{SupplierPrice.objects.count()} prices, "
            f"{StockLine.objects.count()} stock lines."
        ))

    def _import_stocktake(self, path):
        wb = load_workbook(path, data_only=True)
        ws = wb["Current"] if "Current" in wb.sheetnames else wb.worksheets[0]
        # header is on row 2; data from row 3
        st, _ = Stocktake.objects.get_or_create(
            date=datetime.date(2026, 5, 18),
            defaults={"completed_by": "Paulo", "note": "Imported from spreadsheet"},
        )
        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            code = row[0]
            if not code or str(code).startswith("#REF"):
                continue
            name = (row[2] or "").strip()
            supplier_name = (row[1] or "").strip()
            unit = (row[4] or "g").strip()
            price = dec(row[5])
            weight = dec(row[3])
            weekly = dec(row[8])
            minimum = dec(row[9]) or Decimal("0")
            current = dec(row[10])

            product, _ = Product.objects.update_or_create(
                code=str(code).strip(),
                defaults={"name": name, "unit": unit,
                          "weekly_usage": weekly, "minimum": minimum},
            )
            if supplier_name and price is not None and weight:
                supplier, _ = Supplier.objects.get_or_create(name=supplier_name)
                SupplierPrice.objects.update_or_create(
                    product=product, supplier=supplier,
                    defaults={"pack_weight": weight, "pack_price": price},
                )
            StockLine.objects.update_or_create(
                stocktake=st, product=product, defaults={"current": current},
            )
            count += 1
        self.stdout.write(f"  Stocktake: {count} product rows")
        return count

    def _import_prices(self, path):
        wb = load_workbook(path, data_only=True)
        if "Cheapest Suppliers" not in wb.sheetnames:
            return
        ws = wb["Cheapest Suppliers"]
        added = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[0]
            if not name or str(name).startswith("#REF"):
                continue
            supplier_name = (row[1] or "").strip()
            weight = dec(row[2])
            price = dec(row[4])
            if not (supplier_name and weight and price is not None):
                continue
            product = Product.objects.filter(name__iexact=str(name).strip()).first()
            if not product:
                continue
            supplier, _ = Supplier.objects.get_or_create(name=supplier_name)
            _, created = SupplierPrice.objects.update_or_create(
                product=product, supplier=supplier,
                defaults={"pack_weight": weight, "pack_price": price},
            )
            added += created
        self.stdout.write(f"  Prices: {added} extra supplier prices added")
