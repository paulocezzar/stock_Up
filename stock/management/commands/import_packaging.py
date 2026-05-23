"""Import the packaging master from the supplier-spec spreadsheet.

Usage:
    python manage.py import_packaging data/packaging.xlsx
    python manage.py import_packaging data/packaging.xlsx --department Bakery

Packaging items are stored as ordinary Products in the "packaging" category
so the existing stock machinery (stocktakes, deliveries, adjustments,
reorder, value, alerts) applies unchanged — no separate model.

Reads four tabs (mirrors import_ingredients):
  Packaging         master row per item (NPD-P code, name, cost, supply unit)
  Units Of Measure  rows of "N base-units per 1 purchase-unit" — e.g. for
                    NPD-P25 the row is "Each | 1000 | Box | 1" meaning a
                    Box contains 1000 Each. Note the direction is the
                    REVERSE of the ingredients UOM tab (which says
                    "1 Trade Paper Sack = 16 Kilograms").
  Suppliers         per-item supplier rows; one is flagged "Is Primary
                    Supplier = Yes" (column 6 here, not 5 as in ingredients)
  Reference         lookup whose Suppliers section maps Supplier Code → name

For each row a Product is created/updated keyed on the NPD-P code. The
unit is always "ea" (packaging is counted, not weighed). pack_weight on
the SupplierPrice is the number of "Each" per purchase unit (e.g. 1000
for a Box of 1000 labels); pack_price is the per-purchase-unit cost.
This keeps the existing £/1000 maths consistent: £110.12 / 1000 × 1000
= £110.12 per 1000 each.
"""
from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from stock.models import Department, Supplier, Product, SupplierPrice


def norm(s):
    if s is None:
        return ""
    return str(s).strip().lower()


def dec(v):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def resolve_pack(code, supply_unit, uom_by_code):
    """Return (pack_weight_in_each, base_unit='ea') for a packaging item.

    A UOM row here reads ``UoM=Each, Quantity=N, RUoM=<supply_unit>,
    Ref=M`` and means "M of the supply unit = N Each". So when the
    Packaging tab says supply_unit=Box for an item, we find the row
    whose RUoM is "Box" and compute pack = Quantity / Ref (usually
    1) base units per supply unit.

    If supply_unit is "Each" itself, treat as 1 base per purchase unit
    (no UOM row needed). If no UOM row matches, return None — the
    item still imports as a Product, just without a price.
    """
    key = norm(supply_unit)
    if not key:
        return None
    if key in ("each", "ea"):
        return (Decimal("1"), "ea")
    for row in uom_by_code.get(code, []):
        if norm(row["ruom"]) != key:
            continue
        qty = row["quantity"] or Decimal("1")
        ref = row["ref_quantity"]
        if ref is None or ref == 0:
            continue
        return (qty / ref, "ea")
    return None


class Command(BaseCommand):
    help = "Import packaging items (Products in the Packaging category) from Excel."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to packaging.xlsx")
        parser.add_argument("--department", default="Bakery",
                            help="Department to assign packaging to (default: Bakery)")

    @transaction.atomic
    def handle(self, *args, **opts):
        wb = load_workbook(opts["path"], data_only=True)
        for required in ("Packaging", "Units Of Measure", "Suppliers", "Reference"):
            if required not in wb.sheetnames:
                raise CommandError(f"workbook has no '{required}' sheet")

        dept, _ = Department.objects.get_or_create(name=opts["department"])

        uom_by_code = self._read_uom(wb["Units Of Measure"])
        supplier_name_by_code = self._read_supplier_lookup(wb["Reference"])
        primary_by_item = self._read_item_suppliers(wb["Suppliers"])

        supplier_cache = {}

        def get_supplier(name):
            if name not in supplier_cache:
                sup, _ = Supplier.objects.get_or_create(name=name)
                supplier_cache[name] = sup
            return supplier_cache[name]

        created = updated = priced = 0
        flagged = []
        unresolved_codes = set()

        for row in wb["Packaging"].iter_rows(min_row=2, values_only=True):
            code = row[0]
            if not code or str(code).strip() == "":
                continue
            code = str(code).strip()
            name = (row[1] or "").strip()
            cost = dec(row[24])
            supply_unit = row[25]

            pack_info = resolve_pack(code, supply_unit, uom_by_code)
            base_unit = pack_info[1] if pack_info else "ea"

            product, was_created = Product.objects.update_or_create(
                code=code,
                defaults={
                    "name": name,
                    "category": "packaging",
                    "unit": base_unit,
                    "department": dept,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

            supplier_code = primary_by_item.get(code)
            if supplier_code:
                supplier_name = supplier_name_by_code.get(supplier_code)
                if not supplier_name:
                    supplier_name = supplier_code
                    unresolved_codes.add(supplier_code)
            else:
                supplier_name = None

            if pack_info and cost is not None and supplier_name:
                pack_weight, _ = pack_info
                supplier = get_supplier(supplier_name)
                existing = (SupplierPrice.objects
                            .filter(product=product, supplier=supplier)
                            .order_by("-effective_date", "-id").first())
                if existing:
                    existing.pack_weight = pack_weight
                    existing.pack_price = cost
                    existing.save()
                else:
                    SupplierPrice.objects.create(
                        product=product, supplier=supplier,
                        pack_weight=pack_weight, pack_price=cost,
                    )
                priced += 1
            else:
                reasons = []
                if not pack_info:
                    reasons.append("no UOM")
                if cost is None:
                    reasons.append("no cost")
                if not supplier_name:
                    reasons.append("no supplier")
                flagged.append((code, name, supply_unit, ", ".join(reasons) or "skipped"))

        self.stdout.write(self.style.SUCCESS(
            f"Imported packaging into '{dept.name}':"))
        self.stdout.write(f"  {created} created, {updated} updated")
        self.stdout.write(
            f"  {priced} with pack size + price, {len(flagged)} flagged")
        if unresolved_codes:
            self.stdout.write(self.style.WARNING(
                f"  Unresolved supplier codes (used the code as the name): "
                f"{', '.join(sorted(unresolved_codes))}"))
        if flagged:
            self.stdout.write(self.style.WARNING("  Flagged items (needs attention):"))
            for code, name, supply_unit, reason in flagged:
                self.stdout.write(f"    {code}  {name}  [{supply_unit}]  - {reason}")

    def _read_uom(self, ws):
        out = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[0]
            if not code:
                continue
            code = str(code).strip()
            out.setdefault(code, []).append({
                "uom": row[2],
                "quantity": dec(row[3]) or Decimal("1"),
                "ruom": row[4],
                "ref_quantity": dec(row[5]),
            })
        return out

    def _read_supplier_lookup(self, ws):
        """Mirror of import_ingredients._read_supplier_lookup.

        The Reference tab has the same wide-section layout in both
        workbooks: row 1 is section labels, row 3 is sub-headers, then
        rows of (Supplier Code, Description) pairs under the Suppliers
        section.
        """
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        header = rows[0]
        section_col = None
        for i, label in enumerate(header):
            if label and str(label).strip().lower() == "suppliers":
                section_col = i
                break
        if section_col is None:
            return {}
        if len(rows) < 3:
            return {}
        sub_header = rows[2]
        code_col = name_col = None
        for j in range(section_col, min(section_col + 6, len(sub_header))):
            label = sub_header[j]
            if label is None:
                continue
            text = str(label).strip().lower()
            if text == "supplier code":
                code_col = j
            elif text == "description" and code_col is not None and name_col is None:
                name_col = j
        if code_col is None or name_col is None:
            return {}
        lookup = {}
        for row in rows[3:]:
            if code_col >= len(row) or name_col >= len(row):
                continue
            sc = row[code_col]
            nm = row[name_col]
            if not sc:
                continue
            sc = str(sc).strip()
            if not sc or not sc.upper().startswith("S"):
                continue
            if nm is None or str(nm).strip() == "":
                continue
            lookup[sc] = str(nm).strip()
        return lookup

    def _read_item_suppliers(self, ws):
        """Return {packaging code → primary supplier code}.

        Packaging Suppliers columns: 0=Code, 2=Supplier Code, 6=Is Primary
        (the Is-Primary column sits at index 6 here, not 5 as in the
        ingredients spreadsheet, because there's an extra "Ext. Code"
        column in between).
        """
        all_by_item = {}
        primary_by_item = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[0]
            if not code:
                continue
            code = str(code).strip()
            supplier_code = row[2]
            if not supplier_code:
                continue
            supplier_code = str(supplier_code).strip()
            is_primary = str(row[6] or "").strip().lower() == "yes"
            all_by_item.setdefault(code, []).append(supplier_code)
            if is_primary and code not in primary_by_item:
                primary_by_item[code] = supplier_code
        for code, supplier_codes in all_by_item.items():
            if code not in primary_by_item:
                primary_by_item[code] = supplier_codes[0]
        return primary_by_item
