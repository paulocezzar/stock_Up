"""Import the ingredient master from the supplier spec spreadsheet.

Usage:
    python manage.py import_ingredients data/ingredients.xlsx
    python manage.py import_ingredients data/ingredients.xlsx --department Bakery

Reads up to five tabs:
  Ingredients       master row per ingredient (NPD-I code, name, cost, supply unit, category)
  Units Of Measure  converts a supply unit to a reference weight (e.g. Trade Paper Sack = 16 Kilograms)
  Suppliers         per-ingredient supplier rows; one is flagged "Is Primary Supplier = Yes"
  Reference         lookup table whose "Suppliers" section maps Supplier Code → real name
  Allergens         per-ingredient allergen rows (Contains Yes/No, May Contain Yes/No); optional

For each ingredient row a Product is created/updated keyed on code. The pack weight comes
from the matching UOM row, normalised to base units (g for kg/grams, ml for litres). The
ingredient's primary supplier (Is Primary = Yes; first row if none flagged primary) drives
which Supplier the SupplierPrice is attached to; the supplier name is resolved via the
Reference tab. Allergen rows are upserted per (product, allergen) so re-runs don't duplicate.
Ingredients with no usable UOM row or no supplier row are still imported, but with no price -
the command prints them at the end so they can be addressed manually.
"""
from decimal import Decimal, InvalidOperation
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from stock.models import Department, Supplier, Product, SupplierPrice, IngredientAllergen


CATEGORY_MAP = {
    "dry goods": "dry_goods",
    "dairy & eggs": "dairy_eggs",
    "dairy and eggs": "dairy_eggs",
    "frozen goods": "frozen_goods",
    "fruit & veg": "fruit_veg",
    "fruit and veg": "fruit_veg",
    "unassigned": "unassigned",
}

# RUOM (reference unit) → (multiplier to base unit, base unit). Anything not listed
# here (e.g. another supply-unit name) is resolved indirectly via the same code's
# other UOM rows.
BASE_UNITS = {
    "kilograms": (Decimal("1000"), "g"),
    "kilogram": (Decimal("1000"), "g"),
    "kilos": (Decimal("1000"), "g"),
    "kilo": (Decimal("1000"), "g"),
    "kg": (Decimal("1000"), "g"),
    "grams": (Decimal("1"), "g"),
    "gram": (Decimal("1"), "g"),
    "g": (Decimal("1"), "g"),
    "litres": (Decimal("1000"), "ml"),
    "litre": (Decimal("1000"), "ml"),
    "liters": (Decimal("1000"), "ml"),
    "l": (Decimal("1000"), "ml"),
    "millilitres": (Decimal("1"), "ml"),
    "ml": (Decimal("1"), "ml"),
}


def norm(s):
    if s is None:
        return ""
    return str(s).strip().lower()


def dec(v):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def resolve_weight(code, supply_unit, uom_by_code, _seen=None):
    """Walk the UOM rows for this code to find (pack_weight, base_unit).

    UOM rows can chain (e.g. Box → 40 Pack → 0.25 Kilograms). We follow the
    chain until we hit a recognised base unit, with a guard against cycles.
    If the supply unit is itself a base unit name (e.g. "Kilograms"), we treat
    the pack as 1 × that unit — no UOM row needed.
    Returns (Decimal pack_weight in base unit, base unit code) or None.
    """
    _seen = _seen or set()
    key = norm(supply_unit)
    if not key or key in _seen:
        return None
    if key in BASE_UNITS:
        mult, base = BASE_UNITS[key]
        return (mult, base)
    _seen.add(key)
    rows = uom_by_code.get(code, [])
    for row in rows:
        if norm(row["uom"]) != key:
            continue
        qty = row["quantity"] or Decimal("1")
        ref_qty = row["ref_quantity"]
        ruom = row["ruom"]
        if ref_qty is None:
            continue
        ruom_key = norm(ruom)
        if ruom_key in BASE_UNITS:
            mult, base = BASE_UNITS[ruom_key]
            return (qty * ref_qty * mult, base)
        # RUOM is another supply unit on the same code - resolve recursively.
        nested = resolve_weight(code, ruom, uom_by_code, _seen)
        if nested:
            inner_weight, base = nested
            return (qty * ref_qty * inner_weight, base)
    return None


class Command(BaseCommand):
    help = "Import ingredients (Products + SupplierPrices) from the Excel master."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to ingredients.xlsx")
        parser.add_argument("--department", default="Bakery",
                            help="Department to assign ingredients to (default: Bakery)")

    @transaction.atomic
    def handle(self, *args, **opts):
        wb = load_workbook(opts["path"], data_only=True)
        for required in ("Ingredients", "Units Of Measure", "Suppliers", "Reference"):
            if required not in wb.sheetnames:
                raise CommandError(f"workbook has no '{required}' sheet")

        dept, _ = Department.objects.get_or_create(name=opts["department"])

        uom_by_code = self._read_uom(wb["Units Of Measure"])
        supplier_name_by_code = self._read_supplier_lookup(wb["Reference"])
        primary_by_ing = self._read_ingredient_suppliers(wb["Suppliers"])

        # Cache Supplier objects within this run so we touch the DB once per name.
        supplier_cache = {}

        def get_supplier(name):
            if name not in supplier_cache:
                sup, _ = Supplier.objects.get_or_create(name=name)
                supplier_cache[name] = sup
            return supplier_cache[name]

        created = updated = priced = 0
        flagged = []                # (code, name, supply_unit, reason)
        unresolved_codes = set()    # S-codes we couldn't translate to a name

        for row in wb["Ingredients"].iter_rows(min_row=2, values_only=True):
            code = row[0]
            if not code or str(code).strip() == "":
                continue
            code = str(code).strip()
            name = (row[1] or "").strip()
            cost = dec(row[33])
            supply_unit = row[34]
            category_raw = norm(row[35])
            category = CATEGORY_MAP.get(category_raw, "unassigned")

            weight_info = resolve_weight(code, supply_unit, uom_by_code)
            base_unit = weight_info[1] if weight_info else "g"

            product, was_created = Product.objects.update_or_create(
                code=code,
                defaults={
                    "name": name,
                    "category": category,
                    "unit": base_unit,
                    "department": dept,
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

            # Resolve the supplier for this ingredient (primary, else first row).
            supplier_code = primary_by_ing.get(code)
            if supplier_code:
                supplier_name = supplier_name_by_code.get(supplier_code)
                if not supplier_name:
                    # Unresolvable S-code: keep the code as the name and flag.
                    supplier_name = supplier_code
                    unresolved_codes.add(supplier_code)
            else:
                supplier_name = None

            # Write a price only if we have pack weight + cost + supplier.
            if weight_info and cost is not None and supplier_name:
                pack_weight, _ = weight_info
                supplier = get_supplier(supplier_name)
                existing = (SupplierPrice.objects
                            .filter(product=product, supplier=supplier)
                            .order_by("-effective_date", "-id").first())
                if existing:
                    existing.pack_weight = pack_weight
                    existing.pack_price = cost
                    existing.save()
                else:
                    SupplierPrice.objects.create(
                        product=product, supplier=supplier,
                        pack_weight=pack_weight, pack_price=cost,
                    )
                priced += 1
            else:
                reasons = []
                if not weight_info:
                    reasons.append("no UOM")
                if cost is None:
                    reasons.append("no cost")
                if not supplier_name:
                    reasons.append("no supplier")
                flagged.append((code, name, supply_unit, ", ".join(reasons) or "skipped"))

        allergens_touched = 0
        if "Allergens" in wb.sheetnames:
            allergens_touched = self._import_allergens(wb["Allergens"])

        self.stdout.write(self.style.SUCCESS(
            f"Imported ingredients into '{dept.name}':"))
        self.stdout.write(f"  {created} created, {updated} updated")
        self.stdout.write(
            f"  {priced} with pack weight + price, {len(flagged)} flagged (no pack info)")
        if allergens_touched:
            self.stdout.write(f"  {allergens_touched} allergen declarations imported")
        if unresolved_codes:
            self.stdout.write(self.style.WARNING(
                f"  Unresolved supplier codes (used the code as the name): "
                f"{', '.join(sorted(unresolved_codes))}"))
        if flagged:
            self.stdout.write(self.style.WARNING("  Flagged ingredients (needs attention):"))
            for code, name, supply_unit, reason in flagged:
                self.stdout.write(f"    {code}  {name}  [{supply_unit}]  - {reason}")

    def _read_uom(self, ws):
        """Index UOM rows by code so we can resolve indirect chains."""
        out = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[0]
            if not code:
                continue
            code = str(code).strip()
            out.setdefault(code, []).append({
                "uom": row[2],
                "quantity": dec(row[3]) or Decimal("1"),
                "ruom": row[4],
                "ref_quantity": dec(row[5]),
            })
        return out

    def _read_supplier_lookup(self, ws):
        """Build {Supplier Code → name} from the Reference tab's Suppliers section.

        Reference is a wide sheet whose first row is a list of section labels
        ("Status", "Formats", ..., "Suppliers", ...). We find the column where
        "Suppliers" appears, then read the "Supplier Code" / "Description"
        sub-columns from the third row down.
        """
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        header = rows[0]
        section_col = None
        for i, label in enumerate(header):
            if label and str(label).strip().lower() == "suppliers":
                section_col = i
                break
        if section_col is None:
            return {}
        # Sub-headers live two rows below the section label. Find which
        # offsets are "Supplier Code" and "Description".
        if len(rows) < 3:
            return {}
        sub_header = rows[2]
        code_col = name_col = None
        # Look within a reasonable window after the section column.
        for j in range(section_col, min(section_col + 6, len(sub_header))):
            label = sub_header[j]
            if label is None:
                continue
            text = str(label).strip().lower()
            if text == "supplier code":
                code_col = j
            elif text == "description" and code_col is not None and name_col is None:
                name_col = j
        if code_col is None or name_col is None:
            return {}
        lookup = {}
        for row in rows[3:]:
            if code_col >= len(row) or name_col >= len(row):
                continue
            sc = row[code_col]
            nm = row[name_col]
            if not sc:
                continue
            sc = str(sc).strip()
            if not sc or not sc.upper().startswith("S"):
                continue
            if nm is None or str(nm).strip() == "":
                continue
            lookup[sc] = str(nm).strip()
        return lookup

    def _import_allergens(self, ws):
        """Upsert IngredientAllergen rows from the Allergens tab.

        Columns: 0=Code, 2=Allergen, 4=Contains (Yes/No), 5=May Contain (Yes/No).
        Rows whose ingredient code isn't a Product we just imported are ignored.
        Returns the number of allergen rows touched.
        """
        # Index products by code once - cheaper than a query per row.
        product_by_code = {p.code: p for p in Product.objects.exclude(code__isnull=True)}
        touched = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[0]
            if not code:
                continue
            code = str(code).strip()
            product = product_by_code.get(code)
            if not product:
                continue
            allergen = (row[2] or "").strip() if row[2] is not None else ""
            if not allergen:
                continue
            contains = norm(row[4]) == "yes"
            may_contain = norm(row[5]) == "yes"
            IngredientAllergen.objects.update_or_create(
                product=product, name=allergen,
                defaults={"contains": contains, "may_contain": may_contain},
            )
            touched += 1
        return touched

    def _read_ingredient_suppliers(self, ws):
        """Return {ingredient code → supplier code} preferring the primary row.

        Ingredients can have several supplier rows. We collect them in order,
        then pick the first one marked Is Primary = Yes; if none, the first
        row wins.
        """
        all_by_ing = {}
        primary_by_ing = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[0]
            if not code:
                continue
            code = str(code).strip()
            supplier_code = row[2]
            if not supplier_code:
                continue
            supplier_code = str(supplier_code).strip()
            is_primary = str(row[5] or "").strip().lower() == "yes"
            all_by_ing.setdefault(code, []).append(supplier_code)
            if is_primary and code not in primary_by_ing:
                primary_by_ing[code] = supplier_code
        # Fallback: ingredients with no primary row get their first listed supplier.
        for code, supplier_codes in all_by_ing.items():
            if code not in primary_by_ing:
                primary_by_ing[code] = supplier_codes[0]
        return primary_by_ing
