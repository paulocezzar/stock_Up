"""Import customers from the bakery order-sheet workbook.

Usage:
    python manage.py import_customers data/order_sheet.xlsm
    python manage.py import_customers data/order_sheet.xlsm --department Bakery

Two source sheets — plus one piece of data pulled from each customer's
own order-form tab:

  Customers    columns "Customer Name", "Location" — one row per Estate
               outlet or named contact. Blank-name rows (the trailing
               staff-only "Ordered by" entries) are skipped. The
               "Ordered by" column on this sheet is a misaligned
               alphabetical list and is deliberately IGNORED.
  WHOLESALE    column 1 ("Wholesale Customer") from row 11 down — the
               authoritative list of wholesale accounts. A customer is
               classified ``wholesale`` iff their name appears here
               (case-insensitive). Everything else is ``internal``.
  per-customer The workbook has one tab per customer named to match
  tabs         the customer (e.g. "GARDEN CAFE"). Each carries an
               "Ordered by" label in its header with the real contact a
               few cells to the right (offset +3 in practice). This is
               the authoritative contact source. Meta-tabs (Start,
               Products, Customers, Production, …) are skipped.

Wholesale accounts that exist ONLY in the WHOLESALE tab (PINKMANS,
SOCIETY branches) are imported with empty location/contact.

Idempotent on name (case-insensitive lookup). Two flags protect operator
work from being clobbered by re-import:

- ``is_type_manual=True`` — set whenever the operator edits an existing
  row via the UI. Protects ALL editable fields (customer_type, ordered_by,
  location) from being overwritten. Re-import touches only ``department``.
- ``is_manual_entry=True`` — set when the operator creates a customer by
  hand from the UI. The importer SKIPS these rows entirely: no field is
  ever touched, and they're never deleted, even if their name happens to
  later appear in the order sheet.
"""
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from stock.models import Customer, Department


# Workbook tabs that aren't real customer order forms — must be skipped
# when matching sheet titles to customer names.
META_TABS = {
    "start", "products", "customers", "customer lookup",
    "production", "starter", "daily production", "weekly production",
    "delivery note", "wholesale delivery note", "wholesale",
}


def _norm(s):
    if s is None:
        return ""
    return str(s).strip()


def _read_customers_tab(ws):
    """Yield ``{name, location}`` dicts from the Customers sheet.

    Header row (row 1) is skipped. Rows with no Customer Name in
    column 0 are skipped — they're the trailing staff-only "Ordered by"
    entries. The "Ordered by" column on this sheet is NOT read — it's
    a misaligned alphabetical list, not actually keyed to the customer
    on its row.
    """
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = _norm(row[0]) if row else ""
        if not name:
            continue
        location = _norm(row[1]) if len(row) > 1 else ""
        out.append({"name": name, "location": location})
    return out


def _read_wholesale_names(ws):
    """Distinct names from column 0 of the WHOLESALE sheet, row 11 down.

    Preserves the first-seen casing for each name (matters for display).
    The "Wholesale Customer" label in the header row is skipped.
    """
    out = []
    seen = set()
    for row in ws.iter_rows(min_row=11, values_only=True):
        if not row:
            continue
        v = row[0]
        if v is None:
            continue
        name = str(v).strip()
        if not name:
            continue
        if name.lower() == "wholesale customer":
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(name)
    return out


def _read_contact_from_tab(ws):
    """Return the contact name from a customer tab's "Ordered by" header.

    Scans the first 8 rows for a cell whose text is "Ordered by"
    (case-insensitive, trailing colons ignored). Returns the first
    non-empty cell within 6 columns to the right of that label, or ``""``
    if there isn't one. The narrow 6-cell window stops far-right
    unrelated labels (TN100's header has a "w/c" date marker ~18
    columns past "Ordered by") from being mistaken for the contact.
    """
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        if not row:
            continue
        for i, cell in enumerate(row):
            if cell is None:
                continue
            text = str(cell).strip().lower().rstrip(":")
            if text != "ordered by":
                continue
            end = min(i + 7, len(row))
            for k in range(i + 1, end):
                v = row[k]
                if v is None:
                    continue
                s = str(v).strip()
                if s:
                    return s
            return ""
    return ""


def _build_contact_lookup(wb):
    """Return ``{sheet_title.lower(): contact}`` for every non-meta tab."""
    out = {}
    for title in wb.sheetnames:
        if title.strip().lower() in META_TABS:
            continue
        try:
            ws = wb[title]
        except KeyError:
            continue
        out[title.strip().lower()] = _read_contact_from_tab(ws)
    return out


class Command(BaseCommand):
    help = "Import customers from an order-sheet .xlsm (Customers + WHOLESALE tabs)."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the order-sheet workbook")
        parser.add_argument(
            "--department", default="Bakery",
            help="Department to attach the customers to (default: Bakery)")

    @transaction.atomic
    def handle(self, *args, **opts):
        wb = load_workbook(opts["path"], data_only=True)
        for required in ("Customers", "WHOLESALE"):
            if required not in wb.sheetnames:
                raise CommandError(f"workbook has no '{required}' sheet")

        dept, _ = Department.objects.get_or_create(name=opts["department"])

        customer_rows = _read_customers_tab(wb["Customers"])
        wholesale_names = _read_wholesale_names(wb["WHOLESALE"])
        wholesale_keys = {n.lower() for n in wholesale_names}
        contact_by_tab = _build_contact_lookup(wb)

        created = updated = preserved_manual = skipped_manual_entry = 0
        n_internal = n_wholesale = 0

        # Case-insensitive lookup of existing Customer rows by name so we
        # don't accidentally fork a "TEALS" / "Teals" pair on later runs.
        existing_by_key = {c.name.lower(): c for c in Customer.objects.all()}

        def _upsert(name, location, contact, is_wholesale):
            nonlocal created, updated, preserved_manual, skipped_manual_entry
            target_type = Customer.WHOLESALE if is_wholesale else Customer.INTERNAL
            key = name.lower()
            existing = existing_by_key.get(key)
            if existing is None:
                c = Customer.objects.create(
                    name=name, location=location, ordered_by=contact,
                    customer_type=target_type, is_type_manual=False,
                    is_manual_entry=False, department=dept,
                )
                existing_by_key[key] = c
                created += 1
                return c
            # Hand-created rows are off-limits — the operator owns them.
            if existing.is_manual_entry:
                skipped_manual_entry += 1
                return existing
            # is_type_manual protects every editable field (type, contact,
            # location) so operator edits survive re-import. Department is
            # still kept in sync so a re-deployed dept rename can land.
            existing.department = dept
            if existing.is_type_manual:
                preserved_manual += 1
            else:
                existing.customer_type = target_type
                existing.ordered_by = contact
                existing.location = location
            existing.save()
            updated += 1
            return existing

        seen_keys = set()

        for row in customer_rows:
            key = row["name"].lower()
            seen_keys.add(key)
            is_wholesale = key in wholesale_keys
            contact = contact_by_tab.get(key, "")
            _upsert(row["name"], row["location"], contact, is_wholesale)
            if is_wholesale:
                n_wholesale += 1
            else:
                n_internal += 1

        # Wholesale-only accounts: in WHOLESALE tab but not in Customers
        # tab. Some may still have their own order-form tab in the
        # workbook, so check the contact lookup before defaulting to "".
        for w_name in wholesale_names:
            key = w_name.lower()
            if key in seen_keys:
                continue
            seen_keys.add(key)
            contact = contact_by_tab.get(key, "")
            _upsert(w_name, "", contact, True)
            n_wholesale += 1

        self.stdout.write(self.style.SUCCESS(
            f"Imported customers into '{dept.name}':"))
        self.stdout.write(
            f"  {created} created, {updated} updated")
        self.stdout.write(
            f"  {n_internal} internal, {n_wholesale} wholesale")
        if preserved_manual:
            self.stdout.write(self.style.WARNING(
                f"  Preserved {preserved_manual} manually-edited row(s) "
                f"(type + contact + location untouched)"))
        if skipped_manual_entry:
            self.stdout.write(self.style.WARNING(
                f"  Skipped {skipped_manual_entry} hand-created customer(s) "
                f"(is_manual_entry=True)"))
