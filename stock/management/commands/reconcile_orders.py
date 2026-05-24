"""Reconciliation report for a historical order-sheet week.

Usage:
    python manage.py reconcile_orders data/historical/order_sheet_2026_03_30.xlsm

Re-parses the workbook (read-only), computes the sheet's per-customer
totals (including each row of the WHOLESALE tab as its own wholesale
customer), reads the imported totals from the database for the same
week, and prints a per-row table + grand-total verdict so each
imported week can be verified after the fact.

Read-only: the command never touches Order/OrderLine/HistoricalImport
rows — pure verification. Same conventions as the importer apply:
blank Price cell counts as £0 (sheet authoritative), wholesale
customers are matched to existing ``Customer`` rows by case-
insensitive trimmed name.
"""
from collections import OrderedDict
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError

from stock.models import Customer, Order, OrderLine
from stock.order_import import (
    META_TABS, WHOLESALE_TAB,
    iter_product_rows, iter_wholesale_rows,
    read_tab_dates, read_wholesale_dates,
    make_price_recovery, _resolve_unit_price,
)


def _sheet_total_for_tab(ws, *, recover_price=None):
    """Sum qty × price across all 7 days for a normal customer tab.

    Uses the same blank-price recovery chain as the importer
    (workbook Products tab → reference CSV → £0) when a recovery
    closure is passed, so the report's SHEET total reflects exactly
    what the importer would write. Returns ``(total, n_priced_rows)``.
    """
    total = Decimal("0")
    n_rows = 0
    for row in iter_product_rows(ws):
        price = _resolve_unit_price(
            row.get("price"), row.get("sage"), row.get("name"),
            recover_price=recover_price)
        row_had_qty = False
        for q in row["qtys"]:
            if q is None:
                continue
            row_had_qty = True
            total += q * price
        if row_had_qty:
            n_rows += 1
    return total, n_rows


def _wholesale_per_customer_totals(ws, *, recover_price=None):
    """For every wholesale customer in the WHOLESALE tab, return
    ``OrderedDict[name → sheet_total]`` preserving first-seen casing.

    Same recovery chain as :func:`_sheet_total_for_tab`. Customers
    whose week is entirely blank are omitted from the dict.
    """
    totals = OrderedDict()
    for row in iter_wholesale_rows(ws):
        name = row["customer"].strip()
        price = _resolve_unit_price(
            row.get("price"), row.get("sage"), row.get("name"),
            recover_price=recover_price)
        for q in row["qtys"]:
            if q is None:
                continue
            totals.setdefault(name, Decimal("0"))
            totals[name] += q * price
    return totals


def _db_total_for_customer(customer, week_dates):
    """Sum ``line.line_value`` for every OrderLine on this customer in
    the given week. ``line_value`` is the snapshot, so post-import
    catalogue drift doesn't affect the comparison."""
    total = Decimal("0")
    for line in OrderLine.objects.filter(
            order__customer=customer,
            order__order_date__in=week_dates):
        if line.line_value is not None:
            total += line.line_value
    return total


def _fmt_money(d):
    return f"£{d:>10,.2f}"


class Command(BaseCommand):
    help = ("Reconcile a historical order-sheet week: sheet totals vs "
            "imported totals, per customer + per wholesale customer.")

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            help="Path to a single historical .xlsm "
                 "(e.g. data/historical/order_sheet_2026_03_30.xlsm)")

    def handle(self, *args, **opts):
        from openpyxl import load_workbook
        path = opts["path"]
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except FileNotFoundError as e:
            raise CommandError(str(e))

        try:
            # Derive the week from the first non-meta, non-WHOLESALE
            # tab that parses cleanly; fall back to WHOLESALE's own
            # date row if every customer tab is malformed.
            week_dates = None
            for tab in wb.sheetnames:
                if tab.strip().lower() in META_TABS:
                    continue
                if tab == WHOLESALE_TAB:
                    continue
                try:
                    week_dates = read_tab_dates(wb[tab])
                    break
                except Exception:  # noqa: BLE001
                    continue
            if week_dates is None and WHOLESALE_TAB in wb.sheetnames:
                try:
                    week_dates = read_wholesale_dates(wb[WHOLESALE_TAB])
                except Exception:  # noqa: BLE001
                    pass
            if week_dates is None:
                raise CommandError(
                    f"{path}: could not derive a week from any tab")
            week_start = week_dates[0]

            # Use the same recovery chain the importer uses so the
            # report's SHEET totals reflect what was actually written
            # to the DB — not raw blank-is-£0 cell readings.
            recover_price = make_price_recovery(wb)

            # ---- customer tabs ----
            customer_rows = []
            empty_tabs = []
            meta_skipped = []
            unmatched_tabs = []
            for tab in wb.sheetnames:
                key = tab.strip().lower()
                if key in META_TABS:
                    meta_skipped.append(tab)
                    continue
                if tab == WHOLESALE_TAB:
                    continue  # handled separately below
                ws = wb[tab]
                try:
                    sheet_total, n_rows = _sheet_total_for_tab(
                        ws, recover_price=recover_price)
                except Exception as e:  # noqa: BLE001
                    customer_rows.append({
                        "name": tab, "sheet": None, "imported": None,
                        "diff": None, "status": f"PARSE-ERR ({e})"})
                    continue
                if n_rows == 0:
                    empty_tabs.append(tab)
                    continue
                cust = Customer.objects.filter(name__iexact=tab).first()
                if cust is None:
                    unmatched_tabs.append(tab)
                    customer_rows.append({
                        "name": tab, "sheet": sheet_total,
                        "imported": Decimal("0"),
                        "diff": -sheet_total,
                        "status": "NO CUSTOMER ROW"})
                    continue
                imp = _db_total_for_customer(cust, week_dates)
                diff = imp - sheet_total
                status = "OK" if abs(diff) < Decimal("0.01") else "MISMATCH"
                customer_rows.append({
                    "name": tab, "sheet": sheet_total,
                    "imported": imp, "diff": diff, "status": status})

            # ---- wholesale ----
            wholesale_rows = []
            wholesale_unmatched = []
            wholesale_present = WHOLESALE_TAB in wb.sheetnames
            ws_total_sheet = Decimal("0")
            ws_total_imported = Decimal("0")
            if wholesale_present:
                ws = wb[WHOLESALE_TAB]
                per_cust = _wholesale_per_customer_totals(
                    ws, recover_price=recover_price)
                for name, sheet_total in per_cust.items():
                    ws_total_sheet += sheet_total
                    cust = Customer.objects.filter(name__iexact=name).first()
                    if cust is None:
                        wholesale_unmatched.append(name)
                        wholesale_rows.append({
                            "name": name, "sheet": sheet_total,
                            "imported": Decimal("0"),
                            "diff": -sheet_total,
                            "status": "NO CUSTOMER ROW"})
                        continue
                    imp = _db_total_for_customer(cust, week_dates)
                    ws_total_imported += imp
                    diff = imp - sheet_total
                    status = "OK" if abs(diff) < Decimal("0.01") else "MISMATCH"
                    wholesale_rows.append({
                        "name": name, "sheet": sheet_total,
                        "imported": imp, "diff": diff, "status": status})

            # ---- render ----
            self._print_header(path, week_start)
            self._print_section("CUSTOMER TABS", customer_rows)
            if wholesale_present:
                self._print_section("WHOLESALE BREAKDOWN", wholesale_rows)
            else:
                self.stdout.write("")
                self.stdout.write("(no WHOLESALE tab in this workbook)")

            # Totals
            ct_sheet = sum((r["sheet"] for r in customer_rows
                            if r["sheet"] is not None), Decimal("0"))
            ct_imp = sum((r["imported"] for r in customer_rows
                          if r["imported"] is not None), Decimal("0"))
            grand_sheet = ct_sheet + ws_total_sheet
            grand_imp = ct_imp + ws_total_imported
            grand_diff = grand_imp - grand_sheet

            self._print_summary(
                customer_rows=customer_rows,
                wholesale_rows=wholesale_rows,
                empty_tabs=empty_tabs,
                meta_skipped=meta_skipped,
                unmatched_tabs=unmatched_tabs,
                wholesale_unmatched=wholesale_unmatched,
                ct_sheet=ct_sheet, ct_imp=ct_imp,
                ws_total_sheet=ws_total_sheet,
                ws_total_imported=ws_total_imported,
                grand_sheet=grand_sheet, grand_imp=grand_imp,
                grand_diff=grand_diff,
                wholesale_present=wholesale_present)
        finally:
            wb.close()

    # ---- formatting helpers ----

    def _print_header(self, path, week_start):
        self.stdout.write("=" * 78)
        self.stdout.write(
            f"RECONCILIATION REPORT  -  Historical week w/c "
            f"{week_start.isoformat()}")
        self.stdout.write(f"File: {path}")
        self.stdout.write("=" * 78)

    def _print_section(self, title, rows):
        self.stdout.write("")
        self.stdout.write(title)
        self.stdout.write(
            f"{'CUSTOMER / TAB':35s} {'SHEET':>12s} {'IMPORTED':>12s} "
            f"{'DIFF':>12s}  STATUS")
        self.stdout.write("-" * 78)
        for r in rows:
            name = r["name"][:34]
            sheet = _fmt_money(r["sheet"]) if r["sheet"] is not None else "         -"
            imp = _fmt_money(r["imported"]) if r["imported"] is not None else "         -"
            diff = _fmt_money(r["diff"]) if r["diff"] is not None else "         -"
            self.stdout.write(
                f"{name:35s} {sheet:>12s} {imp:>12s} {diff:>12s}  {r['status']}")

    def _print_summary(self, *, customer_rows, wholesale_rows,
                       empty_tabs, meta_skipped, unmatched_tabs,
                       wholesale_unmatched, ct_sheet, ct_imp,
                       ws_total_sheet, ws_total_imported,
                       grand_sheet, grand_imp, grand_diff,
                       wholesale_present):
        self.stdout.write("")
        self.stdout.write("=" * 78)
        self.stdout.write("SUMMARY")
        self.stdout.write("=" * 78)
        n_customer_rows = len([r for r in customer_rows if r["sheet"] is not None])
        n_mismatch = len(
            [r for r in customer_rows + wholesale_rows
             if r.get("status") == "MISMATCH"])
        self.stdout.write(
            f"  Customer tabs scanned:        {n_customer_rows} "
            f"({len(empty_tabs)} empty)")
        if wholesale_present:
            self.stdout.write(
                f"  Wholesale customers scanned:  {len(wholesale_rows)}")
        self.stdout.write(
            f"  Meta tabs skipped:            {len(meta_skipped)}")
        if unmatched_tabs:
            self.stdout.write(self.style.WARNING(
                f"  Customer tabs with no Customer row: {len(unmatched_tabs)} "
                f"({', '.join(unmatched_tabs)})"))
        if wholesale_unmatched:
            self.stdout.write(self.style.WARNING(
                f"  Wholesale names with no Customer row: "
                f"{len(wholesale_unmatched)} "
                f"({', '.join(wholesale_unmatched)})"))
        self.stdout.write("")
        self.stdout.write(
            f"  Customer-tabs sheet total:    {_fmt_money(ct_sheet)}")
        if wholesale_present:
            self.stdout.write(
                f"  Wholesale sheet total:        "
                f"{_fmt_money(ws_total_sheet)}")
        self.stdout.write(
            f"  GRAND TOTAL per sheet:        {_fmt_money(grand_sheet)}")
        self.stdout.write("")
        self.stdout.write(
            f"  Customer-tabs imported total: {_fmt_money(ct_imp)}")
        if wholesale_present:
            self.stdout.write(
                f"  Wholesale imported total:     "
                f"{_fmt_money(ws_total_imported)}")
        self.stdout.write(
            f"  GRAND TOTAL imported:         {_fmt_money(grand_imp)}")
        self.stdout.write("")
        self.stdout.write(
            f"  Difference:                   {_fmt_money(grand_diff)}")
        self.stdout.write("")
        if abs(grand_diff) < Decimal("0.01") and n_mismatch == 0:
            self.stdout.write(self.style.SUCCESS(
                "  RECONCILES (sheet == imported, per-row diffs all £0)"))
        else:
            self.stdout.write(self.style.ERROR(
                f"  MISMATCH (£{grand_diff:+.2f}, "
                f"{n_mismatch} per-row mismatch(es)) — see rows above"))
