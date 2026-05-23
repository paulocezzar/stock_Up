"""Import a customer's weekly orders from the bakery order sheet.

The order-sheet workbook has one customer per tab. Each tab is shaped
like the Garden Café form:

    row 4 col 1   "BAKERY ORDER FORM" (header marker)
    row 4 col 6   "Ordered by" value (operator contact)
    row 7         day-of-week labels ("Monday" .. "Sunday") at cols
                  3, 6, 9, 12, 15, 18, 21 (every day uses 3 columns:
                  Ordered / Sent / Comments)
    row 8         the 7 dates of the week at the same columns
    row 9         column-group headers (``SKU, Product, Price``,
                  then ``Ordered, Sent, Comments`` × 7, then
                  ``Total``)
    row 10+       product rows: SKU at col 0, name at col 1, price
                  at col 2, ordered qty per day at cols 3 / 6 / 9
                  / 12 / 15 / 18 / 21 — blank/zero cells skipped.
                  A blank row in the middle of the products section
                  is skipped (the sheet sometimes inserts spacing).
                  A totals row at the bottom (col 0 / col 1 blank,
                  numbers in the day columns) terminates the scan.

``import_orders_for_tab`` is the reusable entry-point used by both
the ``import_orders`` management command and the matching tests.
It's idempotent on ``(customer, date)``: every existing order for
that customer on each of the seven sheet-dates is wiped and rebuilt
from the workbook, so re-running on a deploy converges to the
sheet's current state. Lines whose product can't be resolved
(unknown Sage code AND no matching product name) are recorded in
``failures`` and skipped — one bad row never aborts the rest.
"""
from collections import OrderedDict
from decimal import Decimal, InvalidOperation
import datetime

from django.db import transaction

# Day-column offsets within each tab — 3-cell stride (Ordered /
# Sent / Comments). The "ordered" cell is the one we care about.
DAY_COL_ORDERED = [3, 6, 9, 12, 15, 18, 21]
DAY_LABELS = ["Monday", "Tuesday", "Wednesday", "Thursday",
              "Friday", "Saturday", "Sunday"]

# openpyxl row indices are 1-based. Mapped to the GARDEN CAFE layout
# (and the rest of the order-sheet tabs that follow the same template):
#   row 8  — day-of-week labels (Monday .. Sunday)
#   row 9  — the seven dates (Mon..Sun)
#   row 10 — column-group headers (SKU / Product / Price / Ordered ...)
#   row 11 — first product row
DATE_ROW = 9
PRODUCT_FIRST_ROW = 11


def _as_text(v):
    if v is None:
        return ""
    s = str(v).strip()
    if isinstance(v, float) and v.is_integer():
        s = str(int(v))
    return s


def _as_qty(v):
    """Parse a quantity cell. Blank / 0 / non-numeric → None (skip)."""
    if v is None or v == "":
        return None
    try:
        d = Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None
    if d <= 0:
        return None
    return d


def _date_from_cell(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    return None


def read_tab_dates(ws):
    """Return the 7 (Mon..Sun) order dates from row 8 of the tab.

    Raises ``ValueError`` if the row doesn't look like a 7-date strip
    (any column blank, or the days aren't 1-day apart Mon..Sun). The
    week-commencing Monday is always derived from the dates — we
    never read it from a header field.
    """
    rows = list(ws.iter_rows(min_row=DATE_ROW, max_row=DATE_ROW,
                             values_only=True))
    if not rows:
        raise ValueError("date row missing")
    row = rows[0]
    dates = []
    for col in DAY_COL_ORDERED:
        if col >= len(row):
            raise ValueError(f"date row too short — column {col} missing")
        d = _date_from_cell(row[col])
        if d is None:
            raise ValueError(f"date column {col} doesn't carry a date")
        dates.append(d)
    # Mon..Sun: each date is one day after the previous, and the first
    # is a Monday (weekday() == 0).
    if dates[0].weekday() != 0:
        raise ValueError(
            f"first date {dates[0]} is a {dates[0].strftime('%A')}, "
            "not a Monday")
    for i in range(1, 7):
        if dates[i] - dates[i - 1] != datetime.timedelta(days=1):
            raise ValueError(
                f"dates aren't consecutive: {dates[i-1]} -> {dates[i]}")
    return dates


def iter_product_rows(ws):
    """Yield ``{sage, name, price, qtys}`` dicts for each non-empty
    product row in the tab.

    ``qtys`` is a list of 7 Decimals-or-None aligned to Mon..Sun
    (cells with no qty come back as ``None``). Rows whose SKU and
    Product name are both blank are silently skipped (the sheet
    sometimes inserts spacing or a totals row above the data). The
    scan terminates the first time it sees a row whose product name
    is blank AND any day-Ordered column carries a number — the bakery
    totals row at the bottom of the sheet.
    """
    for row in ws.iter_rows(min_row=PRODUCT_FIRST_ROW, values_only=True):
        sku = _as_text(row[0]) if len(row) > 0 else ""
        name = _as_text(row[1]) if len(row) > 1 else ""
        # Totals row (no product name, but numbers in day columns) — stop.
        if not name:
            day_has_number = any(
                isinstance(row[c], (int, float)) and row[c] not in (None, 0)
                for c in DAY_COL_ORDERED if c < len(row))
            if day_has_number:
                return
            # Otherwise a blank spacer row — skip and keep scanning.
            continue
        price = _parse_price(row[2] if len(row) > 2 else None)
        qtys = []
        for col in DAY_COL_ORDERED:
            qtys.append(_as_qty(row[col]) if col < len(row) else None)
        yield {"sage": sku, "name": name, "price": price, "qtys": qtys}


def _parse_price(v):
    """Best-effort Decimal from a Price cell. Strips a leading "£" so
    string cells like "£6.25" round-trip. Returns ``None`` on blanks
    or unparseable values — historical lines may carry a name with no
    price, and the importer should still record them rather than
    silently dropping the row."""
    if v is None or v == "":
        return None
    if isinstance(v, str):
        v = v.strip().lstrip("£").strip()
        if not v:
            return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _build_product_lookups(dept):
    """Index this dept's SaleProducts by Sage no. and by name (both
    case-insensitive, whitespace-trimmed) so the importer's hot loop
    doesn't hammer the database."""
    from .models import SaleProduct
    by_sage = {}
    by_name = {}
    for sp in SaleProduct.objects.filter(department=dept):
        sage_key = (sp.sage_number or "").strip().lower()
        if sage_key and sage_key != "0":
            by_sage.setdefault(sage_key, sp)
        name_key = (sp.name or "").strip().lower()
        if name_key:
            by_name.setdefault(name_key, sp)
    return by_sage, by_name


def _match_product(row, by_sage, by_name):
    """Resolve a sheet row to a SaleProduct.

    Sage No. wins when present (it's the operator's authoritative
    code); exact-name match is the fallback so products without a
    Sage code on the sheet still link if their name matches a known
    SKU verbatim. Returns ``None`` when neither matches — the caller
    surfaces the row as a failure.
    """
    sage = (row.get("sage") or "").strip().lower()
    if sage and sage != "0":
        sp = by_sage.get(sage)
        if sp is not None:
            return sp
    name = (row.get("name") or "").strip().lower()
    if name:
        sp = by_name.get(name)
        if sp is not None:
            return sp
    return None


@transaction.atomic
def import_orders_for_tab(workbook, tab_name, customer, dept):
    """Import every product × day cell on ``tab_name`` for ``customer``.

    Replaces (idempotently) every existing order for ``customer`` on
    each of the seven sheet-dates: their OrderLines are wiped first,
    then re-created from the workbook. Orders left with no lines
    after the rebuild (e.g. a product row whose week is entirely
    blank) are deleted so the database mirrors the sheet exactly.

    Returns ``{dates, lines_imported, products_matched,
    products_unmatched, failures}``:
      * ``dates`` — the 7 Mon..Sun dates resolved from the sheet
      * ``lines_imported`` — the count of OrderLines now in place
      * ``products_matched`` — distinct SaleProducts that landed lines
      * ``products_unmatched`` — list of ``(sage, name)`` for rows
        whose SaleProduct couldn't be resolved
      * ``failures`` — free-text notes for malformed rows
    """
    from .models import Order, OrderLine
    if tab_name not in workbook.sheetnames:
        raise ValueError(f"workbook has no '{tab_name}' tab")
    ws = workbook[tab_name]
    dates = read_tab_dates(ws)
    by_sage, by_name = _build_product_lookups(dept)

    # Wipe existing lines for this customer × dates so re-running the
    # importer converges to the sheet. Done in one queryset to avoid
    # N+1 deletes. Orders themselves are kept (and recreated below if
    # missing) so a later qty_sent column has a stable row to write to.
    OrderLine.objects.filter(
        order__customer=customer,
        order__department=dept,
        order__order_date__in=dates,
    ).delete()

    lines_imported = 0
    products_matched = set()
    products_unmatched = []
    failures = []

    # Cache (date → Order) so the per-row loop never refetches.
    order_by_date = {
        o.order_date: o
        for o in Order.objects.filter(
            customer=customer, department=dept, order_date__in=dates)
    }

    for row in iter_product_rows(ws):
        sp = _match_product(row, by_sage, by_name)
        # Snapshot the sheet's name + price on EVERY line — even when
        # no current catalogue product matches. Historical lines
        # (discontinued SKUs like "Almond Pastry", "Hot Cross Buns")
        # need to keep their financial value; matched lines also keep
        # the price the customer was actually charged on the day,
        # which protects the total from later catalogue edits.
        sheet_name = (row.get("name") or "").strip()
        sheet_price = row.get("price")  # Decimal or None
        # Sheet is authoritative for imports: a BLANK Price cell means
        # £0, NOT "fall back to the linked SaleProduct's current price".
        # The fallback in OrderLine.save() is kept for the manual order
        # form (where the operator picks a product and we snapshot the
        # catalogue's current price). For sheet rows we always pass
        # something explicit so save()'s `unit_price is None` branch
        # never fires — even a row whose Price cell is empty contributes
        # £0 to the order total, matching the sheet exactly.
        unit_price = sheet_price if sheet_price is not None else Decimal("0")
        if sp is None:
            products_unmatched.append((row.get("sage") or "",
                                       sheet_name))
        for i, qty in enumerate(row["qtys"]):
            if qty is None:
                continue
            day = dates[i]
            order = order_by_date.get(day)
            if order is None:
                order = Order.objects.create(
                    customer=customer, department=dept,
                    order_date=day)
                order_by_date[day] = order
            OrderLine.objects.create(
                order=order, sale_product=sp,
                product_name=sheet_name,
                unit_price=unit_price,
                qty_ordered=qty)
            lines_imported += 1
            if sp is not None:
                products_matched.add(sp.pk)

    # Sweep orders that ended up with no lines after the rebuild —
    # mirrors the sheet exactly (a customer with no orders on a given
    # day has no Order row for it).
    Order.objects.filter(
        customer=customer, department=dept, order_date__in=dates,
        lines__isnull=True,
    ).delete()

    return {
        "dates": dates,
        "lines_imported": lines_imported,
        "products_matched": len(products_matched),
        "products_unmatched": products_unmatched,
        "failures": failures,
    }


# Workbook tabs that aren't order forms at all — never import as
# orders. WHOLESALE used to live here too (different layout) but it's
# now routed to ``import_wholesale_tab`` so we capture wholesale
# revenue: every other tab in this set really is just metadata or a
# production sheet that the importer should never look at.
META_TABS = {
    "start", "products", "customers", "customer lookup",
    "production", "starter", "daily production", "weekly production",
    "delivery note", "wholesale delivery note",
}

# The WHOLESALE tab is a per-row variant of the customer order form:
# one row per (wholesale customer, product), grouped by customer in
# vertical blocks. Routed to a dedicated parser because its column
# layout is shifted by one to make room for the customer-name column.
WHOLESALE_TAB = "WHOLESALE"

# Data-semantics version for the historical importer. Bump this any
# time a fix changes what the importer would write for the same input
# — the next deploy then re-imports every historical week whose
# ``HistoricalImport`` stamp is below this number, then settles back
# to skip-if-current. See ``HistoricalImport`` docstring for the
# trade-off this makes vs hand-edits.
#
# Changelog:
#   v1 — initial historical importer (pre-blank-price-fix; WHOLESALE
#        was a meta tab and its revenue was skipped).
#   v2 — blank Price cells snapshot as £0 (not catalogue fallback);
#        WHOLESALE routed through dedicated handler so each wholesale
#        customer gets first-class Orders.
HISTORICAL_IMPORT_VERSION = 2

# WHOLESALE layout: column 0 is the wholesale customer name (repeated
# on every product row of that customer's block), then SKU at 1,
# Product at 2, Price at 3, and Ordered/Sent/Comments triples per day
# starting at column 4 (one-stride right of the per-customer tabs).
WHOLESALE_NAME_COL = 0
WHOLESALE_SKU_COL = 1
WHOLESALE_PRODUCT_COL = 2
WHOLESALE_PRICE_COL = 3
WHOLESALE_DAY_COL_ORDERED = [4, 7, 10, 13, 16, 19, 22]
WHOLESALE_DATE_ROW = 9
WHOLESALE_PRODUCT_FIRST_ROW = 11


def read_wholesale_dates(ws):
    """Read the 7 (Mon..Sun) dates from the WHOLESALE tab's date row.

    Same shape check as :func:`read_tab_dates` (Monday start, one-day
    stride) but reads the wholesale-specific column offsets so the
    customer-name column doesn't shift the dates left by one.
    """
    rows = list(ws.iter_rows(min_row=WHOLESALE_DATE_ROW,
                             max_row=WHOLESALE_DATE_ROW,
                             values_only=True))
    if not rows:
        raise ValueError("WHOLESALE date row missing")
    row = rows[0]
    dates = []
    for col in WHOLESALE_DAY_COL_ORDERED:
        if col >= len(row):
            raise ValueError(
                f"WHOLESALE date row too short — column {col} missing")
        d = _date_from_cell(row[col])
        if d is None:
            raise ValueError(
                f"WHOLESALE date column {col} doesn't carry a date")
        dates.append(d)
    if dates[0].weekday() != 0:
        raise ValueError(
            f"first WHOLESALE date {dates[0]} is a "
            f"{dates[0].strftime('%A')}, not a Monday")
    for i in range(1, 7):
        if dates[i] - dates[i - 1] != datetime.timedelta(days=1):
            raise ValueError(
                f"WHOLESALE dates aren't consecutive: "
                f"{dates[i-1]} -> {dates[i]}")
    return dates


def iter_wholesale_rows(ws):
    """Yield ``{customer, sage, name, price, qtys}`` for every product
    row on the WHOLESALE tab.

    A row is kept iff both a customer name (col 0) and a product name
    (col 2) are present — that excludes blank spacers between customer
    blocks, the literal "Wholesale Customer" header label, leftover
    customer-only rows (a name with no product, see the
    TEALS-KITCHEN-only row in real sheets), and the totals row at the
    bottom (no name, no product). ``qtys`` is a 7-long Mon..Sun list of
    Decimals-or-None aligned to the day-ordered columns.
    """
    for row in ws.iter_rows(min_row=WHOLESALE_PRODUCT_FIRST_ROW,
                            values_only=True):
        if not row:
            continue
        cust = (_as_text(row[WHOLESALE_NAME_COL])
                if len(row) > WHOLESALE_NAME_COL else "")
        product = (_as_text(row[WHOLESALE_PRODUCT_COL])
                   if len(row) > WHOLESALE_PRODUCT_COL else "")
        if not cust or not product:
            continue
        if cust.strip().lower() == "wholesale customer":
            continue
        sku = (_as_text(row[WHOLESALE_SKU_COL])
               if len(row) > WHOLESALE_SKU_COL else "")
        price = _parse_price(row[WHOLESALE_PRICE_COL]
                             if len(row) > WHOLESALE_PRICE_COL else None)
        qtys = []
        for col in WHOLESALE_DAY_COL_ORDERED:
            qtys.append(_as_qty(row[col]) if col < len(row) else None)
        yield {"customer": cust, "sage": sku, "name": product,
               "price": price, "qtys": qtys}


def _build_product_lookups_for_dept_or_bakery(dept):
    """Same shape as :func:`_build_product_lookups` but defaults to the
    Bakery dept's catalogue when the caller has no specific dept (the
    wholesale path can encounter customers whose department field is
    None — their bakery products still live under Bakery)."""
    from .models import Department
    target = dept or Department.objects.filter(name="Bakery").first()
    return _build_product_lookups(target)


@transaction.atomic
def import_wholesale_tab(workbook):
    """Import the WHOLESALE tab, splitting it into per-wholesale-customer
    Orders that look first-class in the customer list / grid.

    Every wholesale customer must ALREADY EXIST as a ``Customer`` row
    (case-insensitive, trimmed name match). Unmatched names are
    returned in ``customer_unmatched`` for the operator to reconcile —
    auto-creating them would risk duplicating customers that exist
    under a slightly different spelling, which is far worse to clean
    up than a one-line report.

    Idempotent on every (wholesale customer × date) the tab touches:
    existing OrderLines for those days are wiped and rebuilt from the
    sheet, so a re-run converges and never leaves orphans. Blank
    Price cells snapshot as £0 (sheet authoritative, identical rule to
    :func:`import_orders_for_tab`).

    Returns:
        {
            "dates", "customers_imported",
            "lines_imported", "products_matched",
            "products_unmatched": [(sku, name)],
            "customer_unmatched": [sheet_name],
            "per_customer": {Customer.name: per-customer summary},
        }
    """
    from .models import Customer, Department, Order, OrderLine

    if WHOLESALE_TAB not in workbook.sheetnames:
        return {
            "dates": [],
            "customers_imported": 0,
            "lines_imported": 0,
            "products_matched": 0,
            "products_unmatched": [],
            "customer_unmatched": [],
            "per_customer": OrderedDict(),
            "skipped": True,
            "reason": f"workbook has no '{WHOLESALE_TAB}' tab",
        }

    ws = workbook[WHOLESALE_TAB]
    dates = read_wholesale_dates(ws)
    bakery = Department.objects.filter(name="Bakery").first()
    by_sage, by_name = _build_product_lookups_for_dept_or_bakery(bakery)

    # Group rows by wholesale customer name (preserving first-seen casing
    # for display + reporting). One pass over the sheet — no DB hits.
    grouped = OrderedDict()
    for row in iter_wholesale_rows(ws):
        key = row["customer"].strip().lower()
        if key not in grouped:
            grouped[key] = {"display_name": row["customer"].strip(),
                            "rows": []}
        grouped[key]["rows"].append(row)

    # Resolve each wholesale name to an existing Customer. Auto-create
    # is intentionally not an option here — the live sheet has a
    # handful of legitimate wholesale customers (BISHOPSTROW HOTEL is
    # the canonical example) that aren't yet in our Customers tab, and
    # we'd rather surface them than silently fork the master list.
    customer_unmatched = []
    resolved = OrderedDict()
    for key, block in grouped.items():
        cust = Customer.objects.filter(name__iexact=block["display_name"]).first()
        if cust is None:
            customer_unmatched.append(block["display_name"])
            continue
        resolved[key] = (cust, block["rows"])

    # Wipe existing wholesale lines for ONLY the resolved customers ×
    # the sheet's dates. Other (non-wholesale) orders on those dates
    # — e.g. the per-customer tabs already loaded — are untouched
    # because they belong to different Customer rows.
    resolved_customers = [cust for cust, _ in resolved.values()]
    if resolved_customers:
        OrderLine.objects.filter(
            order__customer__in=resolved_customers,
            order__order_date__in=dates,
        ).delete()

    # Cache (customer_pk, date) → Order
    order_cache = {
        (o.customer_id, o.order_date): o
        for o in Order.objects.filter(
            customer__in=resolved_customers, order_date__in=dates)
    }

    lines_imported = 0
    products_matched = set()
    products_unmatched = []
    per_customer = OrderedDict()

    for key, (customer, rows) in resolved.items():
        dept = customer.department or bakery
        if dept is None:
            # No fallback department — record as a customer-level
            # failure and keep going. Extremely unlikely (the seed
            # data always has a Bakery dept), but defensive.
            per_customer[customer.name] = {
                "lines_imported": 0, "value": Decimal("0"),
                "matched": 0, "unmatched": 0,
                "skipped": "no department"}
            continue

        cust_lines = 0
        cust_value = Decimal("0")
        cust_matched = 0
        cust_unmatched = 0

        for row in rows:
            sp = _match_product(row, by_sage, by_name)
            sheet_name = (row.get("name") or "").strip()
            sheet_price = row.get("price")
            unit_price = sheet_price if sheet_price is not None else Decimal("0")
            if sp is None:
                products_unmatched.append((row.get("sage") or "", sheet_name))

            for i, qty in enumerate(row["qtys"]):
                if qty is None:
                    continue
                day = dates[i]
                order = order_cache.get((customer.pk, day))
                if order is None:
                    order = Order.objects.create(
                        customer=customer, department=dept, order_date=day)
                    order_cache[(customer.pk, day)] = order
                OrderLine.objects.create(
                    order=order, sale_product=sp,
                    product_name=sheet_name,
                    unit_price=unit_price,
                    qty_ordered=qty)
                lines_imported += 1
                cust_lines += 1
                cust_value += qty * unit_price
                if sp is not None:
                    products_matched.add(sp.pk)
                    cust_matched += 1
                else:
                    cust_unmatched += 1

        per_customer[customer.name] = {
            "lines_imported": cust_lines,
            "value": cust_value.quantize(Decimal("0.01")),
            "matched": cust_matched,
            "unmatched": cust_unmatched,
        }

    # Sweep empty orders left behind for resolved customers (a
    # wholesale customer who had orders BEFORE the re-run but whose
    # tab block this run is entirely blank). Mirrors the same
    # convergence the per-customer importer does.
    if resolved_customers:
        Order.objects.filter(
            customer__in=resolved_customers, order_date__in=dates,
            lines__isnull=True,
        ).delete()

    return {
        "dates": dates,
        "customers_imported": sum(
            1 for v in per_customer.values() if v.get("lines_imported", 0) > 0),
        "lines_imported": lines_imported,
        "products_matched": len(products_matched),
        "products_unmatched": products_unmatched,
        "customer_unmatched": customer_unmatched,
        "per_customer": per_customer,
    }


def _resolve_week_dates(wb):
    """Find the 7-date Mon..Sun strip by trying each non-meta tab in turn.

    Historical workbooks have ~30 customer tabs, all sharing the same
    layout — any one with a valid date row tells us the week. We try in
    sheet order and return the first tab that parses cleanly. The
    WHOLESALE tab is skipped here because its date row sits at a
    different column offset; if every customer tab failed somehow the
    wholesale parser would still be the fallback.
    """
    for tab in wb.sheetnames:
        if tab.strip().lower() in META_TABS:
            continue
        if tab == WHOLESALE_TAB:
            continue
        try:
            return read_tab_dates(wb[tab])
        except Exception:  # noqa: BLE001 — try the next tab
            continue
    # Last resort: WHOLESALE's date row also carries the week.
    if WHOLESALE_TAB in wb.sheetnames:
        try:
            return read_wholesale_dates(wb[WHOLESALE_TAB])
        except Exception:  # noqa: BLE001
            pass
    return None


def import_historical_workbook(file_or_path, *, force=False):
    """Import an entire week's orders across every customer tab.

    Designed for the historical archive (``data/historical/*.xlsm``)
    where each file is one week — the deploy loops these and we want
    re-runs to be no-ops UNLESS the importer's data semantics have
    changed since the last import for this week.

    Version-gated idempotency:
      * if a ``HistoricalImport`` stamp exists for this week with
        ``import_version >= HISTORICAL_IMPORT_VERSION`` → skip
        (already imported under current importer; re-running would be
        wasted work and would clobber any hand-edits).
      * otherwise (no stamp OR stale stamp) → proceed; on success,
        upsert the stamp to the current version.
      * ``force=True`` overrides the gate but still updates the stamp.

    Bumping ``HISTORICAL_IMPORT_VERSION`` after a fix to the importer
    therefore forces a one-time re-import of every historical week on
    the next deploy, then settles back to skip-if-current — without
    an env flag or operator action. The cost is that any hand-edits
    to historical lines for that week are lost when the bump lands.

    Tab handling is non-fatal: a missing Customer, a malformed date
    row, or the wholesale tab having a different layout lands in
    ``failures`` and the rest of the workbook keeps going.

    Returns:
        {
            "path", "skipped", "reason"?, "week_start"?,
            "tabs_imported", "lines_imported",
            "products_matched", "products_unmatched_count",
            "import_version", "stamp_was"?,
            "per_tab": {tab: per-tab summary},
            "failures": [(tab_or_marker, reason)],
        }
    """
    from openpyxl import load_workbook
    from .models import Customer, Department, HistoricalImport

    wb = load_workbook(file_or_path, read_only=True, data_only=True)
    try:
        week_dates = _resolve_week_dates(wb)
        if week_dates is None:
            return {
                "path": str(file_or_path),
                "skipped": True,
                "reason": "could not derive a valid week from any tab",
                "tabs_imported": 0,
                "lines_imported": 0,
                "products_matched": 0,
                "products_unmatched_count": 0,
                "import_version": HISTORICAL_IMPORT_VERSION,
                "per_tab": OrderedDict(),
                "failures": [],
            }
        week_start = week_dates[0]

        stamp = HistoricalImport.objects.filter(week_start=week_start).first()
        stamp_was = stamp.import_version if stamp else None
        if not force and stamp is not None and \
                stamp.import_version >= HISTORICAL_IMPORT_VERSION:
            return {
                "path": str(file_or_path),
                "skipped": True,
                "reason": (f"w/c {week_start.isoformat()} already imported "
                           f"at v{stamp.import_version} (current v"
                           f"{HISTORICAL_IMPORT_VERSION}) — pass --force "
                           "to re-import"),
                "week_start": week_start,
                "tabs_imported": 0,
                "lines_imported": 0,
                "products_matched": 0,
                "products_unmatched_count": 0,
                "import_version": HISTORICAL_IMPORT_VERSION,
                "stamp_was": stamp_was,
                "per_tab": OrderedDict(),
                "failures": [],
            }

        per_tab = OrderedDict()
        failures = []
        wholesale_customer_unmatched = []
        bakery_fallback = Department.objects.filter(name="Bakery").first()
        for tab in wb.sheetnames:
            if tab.strip().lower() in META_TABS:
                continue
            try:
                if tab == WHOLESALE_TAB:
                    result = import_wholesale_tab(wb)
                    # Bubble up unmatched wholesale names so the
                    # caller can report them once at the top level.
                    wholesale_customer_unmatched.extend(
                        result.get("customer_unmatched", []))
                else:
                    customer = Customer.objects.filter(name__iexact=tab).first()
                    if customer is None:
                        failures.append(
                            (tab, "no Customer with that name — run "
                                  "import_customers first"))
                        continue
                    dept = customer.department or bakery_fallback
                    if dept is None:
                        failures.append(
                            (tab, "customer has no department and no Bakery "
                                  "fallback exists"))
                        continue
                    result = import_orders_for_tab(wb, tab, customer, dept)
            except Exception as e:  # noqa: BLE001 — keep going past a bad tab
                failures.append((tab, f"{type(e).__name__}: {e}"))
                continue
            # Tabs with no ordered cells (a blank customer this week)
            # come back with zero lines — keep them in per_tab anyway so
            # the per-file summary surfaces that they were scanned.
            per_tab[tab] = result

        # Stamp this week as imported at the current version so the
        # next deploy can skip it (until the next version bump).
        # update_or_create handles both first-import and version-bump
        # paths in one call; ``auto_now`` on imported_at refreshes
        # the audit timestamp automatically.
        HistoricalImport.objects.update_or_create(
            week_start=week_start,
            defaults={
                "import_version": HISTORICAL_IMPORT_VERSION,
                "file_path": str(file_or_path),
            },
        )

        return {
            "path": str(file_or_path),
            "skipped": False,
            "week_start": week_start,
            "tabs_imported": len(per_tab),
            "lines_imported": sum(r["lines_imported"] for r in per_tab.values()),
            "products_matched": sum(r["products_matched"] for r in per_tab.values()),
            "products_unmatched_count": sum(
                len(r["products_unmatched"]) for r in per_tab.values()),
            "import_version": HISTORICAL_IMPORT_VERSION,
            "stamp_was": stamp_was,
            "wholesale_customer_unmatched": wholesale_customer_unmatched,
            "per_tab": per_tab,
            "failures": failures,
        }
    finally:
        wb.close()


def import_orders(file_or_path, *, tabs=None):
    """Read one or more customer tabs from the order-sheet workbook.

    ``tabs`` defaults to ``["GARDEN CAFE"]`` — chunk 2 only imports the
    Garden Café tab so the whole pipeline can be verified end-to-end
    on a single customer before scaling. The ``WHOLESALE`` tab can be
    listed here too — it's routed to :func:`import_wholesale_tab` and
    its per-tab summary carries ``customer_unmatched`` for any
    wholesale customer name with no matching ``Customer`` row.

    Streams the workbook in ``read_only=True, data_only=True`` mode
    so a heavily formatted source doesn't blow the deploy worker's
    memory (the same defensive pattern the recipe importer uses).
    Tabs that fail to import (missing customer, malformed dates,
    etc.) are recorded in ``failures`` and don't abort the rest.
    """
    from openpyxl import load_workbook
    from .models import Customer, Department

    if tabs is None:
        tabs = ["GARDEN CAFE"]

    wb = load_workbook(file_or_path, read_only=True, data_only=True)
    try:
        results = OrderedDict()
        failures = []
        wholesale_customer_unmatched = []
        for tab in tabs:
            try:
                if tab == WHOLESALE_TAB:
                    results[tab] = import_wholesale_tab(wb)
                    wholesale_customer_unmatched.extend(
                        results[tab].get("customer_unmatched", []))
                    continue
                customer = Customer.objects.filter(name__iexact=tab).first()
                if customer is None:
                    failures.append(
                        (tab, "no Customer with that name — run "
                              "import_customers first"))
                    continue
                dept = customer.department or Department.objects.filter(
                    name="Bakery").first()
                if dept is None:
                    failures.append(
                        (tab, "customer has no department and no Bakery "
                              "fallback exists"))
                    continue
                results[tab] = import_orders_for_tab(wb, tab, customer, dept)
            except Exception as e:  # noqa: BLE001 — keep going past a bad tab
                failures.append((tab, f"{type(e).__name__}: {e}"))
                continue
        return {
            "tabs_processed": len(tabs),
            "tabs_imported": len(results),
            "per_tab": results,
            "wholesale_customer_unmatched": wholesale_customer_unmatched,
            "failures": failures,
        }
    finally:
        wb.close()
